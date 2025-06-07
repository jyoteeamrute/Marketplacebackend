import csv
import io
import json
from math import ceil
from rest_framework.pagination import PageNumberPagination
from django.db.models import Sum
from django.db.models import Case, When, Value, CharField, F
from rest_framework.throttling import UserRateThrottle
from django.utils.timesince import timesince
from rest_framework import pagination
from UserApp.serializers import *
from datetime import datetime, timedelta
from django.shortcuts import get_object_or_404
from rest_framework.exceptions import NotFound
from rest_framework import status
from datetime import datetime
from rest_framework.parsers import MultiPartParser, FormParser
from django.core.exceptions import ObjectDoesNotExist
from rest_framework.permissions import IsAuthenticated
import requests
from rest_framework.generics import ListAPIView,UpdateAPIView,DestroyAPIView
from django.core.files.storage import default_storage
import json 
from ProfessionalUser.reelupload import *
from  ProfessionalUser.signals import *
from drf_yasg import openapi    
from urllib.parse import urlencode
from rest_framework.parsers import MultiPartParser, FormParser
from ProfessionalUser.utils import generate_otp,send_email,send_sms
from django.core.cache import cache 
from django.conf import settings
from django.utils.dateparse import parse_date
from drf_yasg import openapi    
from drf_yasg.utils import swagger_auto_schema
from django.db import DatabaseError
from django.db.models import F, Q, ExpressionWrapper, DateTimeField, Func
from django.utils.timezone import now
from datetime import timedelta
from django.utils import timezone
from django.db.models.functions import Concat,Cast  
import openpyxl
from openpyxl.utils import get_column_letter
from .tasks import delete_s3_file
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django.db.models.functions import TruncDay, TruncMonth
import logging
import re
from datetime import datetime, timedelta, timezone ,date
from decimal import Decimal
from rest_framework.pagination import PageNumberPagination
import openpyxl
import requests
from django.conf import settings
from django.contrib.auth.hashers import check_password
from django.contrib.contenttypes.models import ContentType
from django.core.cache import cache
from django.core.exceptions import ObjectDoesNotExist
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import DatabaseError, IntegrityError
from django.db.models import (Case, CharField, DateTimeField,ExpressionWrapper, F, Func, Prefetch, Q, Sum,Value, When)
from django.db.models.functions import Cast, Concat, TruncDay, TruncMonth
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.timesince import timesince
from django.utils.timezone import now
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,TableStyle)
from rest_framework import generics, pagination, permissions, status
from rest_framework.exceptions import NotFound
from rest_framework.generics import DestroyAPIView, ListAPIView, UpdateAPIView
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.throttling import UserRateThrottle
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken
from Admin.models import SubscriptionPlan
from payment.models import *
from payment.serializers import *
from ProfessionalUser.models import *
from ProfessionalUser.productfunctions import *
from ProfessionalUser.serializers import *
from ProfessionalUser.signals import *
from ProfessionalUser.utils import generate_otp, send_email, send_sms
from UserApp.serializers import *
from .serializers import UpdatePasswordSerializer
from .tasks import delete_s3_file



logger = logging.getLogger(__name__)
class ProfessionalUserSignupView(generics.CreateAPIView):
    queryset = ProfessionalUser.objects.all()
    serializer_class = ProfessionalUserSignupSerializer
    
    @swagger_auto_schema(
        operation_summary="Register a Professional User",
        operation_description="Registers a new professional user along with company details and addresses.",
        request_body=ProfessionalUserSignupSerializer,
        responses={
            200: openapi.Response(
                description="Professional user registered successfully",
                examples={
                    "application/json": {
                        "message": "Professional user registered successfully",
                        "user": {
                            "email": "example@domain.com",
                            "phone": "+1234567890",
                            "company": {
                                "companyName": "Tech Corp",
                                "managerFullName": "John Doe",
                                "phoneNumber": "+9876543210",
                                "email": "manager@techcorp.com",
                                "siret": "123456789",
                                "sectorofActivity": "IT Services",
                                "iban": "DE89 3704 0044 0532 0130 00",
                                "vatNumber": "FR123456789"
                            },
                            "manual_address": {
                                "address1": "123 Street",
                                "address2": "Suite 4B",
                                "postalCode": "75001",
                                "lat": "48.8566",
                                "lang": "2.3522",
                                "city": "Paris",
                                "country": "France"
                            },
                            "automatic_address": {
                                "address1": "123 Street",
                                "address2": "Suite 4B",
                                "postalCode": "75001",
                                "lat": "48.8566",
                                "lang": "2.3522",
                                "city": "Paris",
                                "country": "France"
                            }
                        }
                    }
                }
            ),
            400: "Bad Request - Invalid input",
        },
    )
    def create(self, request):
        email = request.data.get("email")
        phone = request.data.get("phone")
        if ProfessionalUser.objects.filter(email=email).exists():
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "This Email is already registered. Please use a different email."
            }, status=200)
        if ProfessionalUser.objects.filter(phone=phone).exists():
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "This Phone number is already registered. Please use a different phone number."
            }, status=200)

        serializer = self.get_serializer(data=request.data)

        try:
            serializer.is_valid(raise_exception=True)
            user = serializer.save()
            on_new_professional_registered(user)
            refresh = RefreshToken()
            refresh.payload['email'] = user.email
            refresh.payload['user_type'] = "professional_user"
            access_token = str(refresh.access_token)

            def get_file_url(file_field):
                return file_field.url if file_field and hasattr(file_field, 'url') else None

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Professional user registered successfully",
                "access_token": access_token,
                "refresh_token": str(refresh),
                "user": {  
                    "email": user.email,
                    "phone": user.phone,
                    "id":user.id,
                    "subscription_status":user.subscription_status,
                    "is_subscription_active": user.subscription_active,
                    "is_verified": user.is_verified,
                    "Subscription": user.subscriptionplan,
                    "Category": list(user.categories.values_list('name', flat=True)) or None, 
                    "Subcategory": list(user.subcategories.values_list('name', flat=True)) or None,
                    "company": {
                        "companyName": user.company.companyName if user.company else None,
                        "managerFullName": user.company.managerFullName if user.company else None,
                        "phoneNumber": user.company.phoneNumber if user.company else None,
                        "email": user.company.email if user.company else None,
                        "siret": user.company.siret if user.company else None,
                        "sectorofActivity": user.company.sectorofActivity if user.company else None,
                        "vatNumber": user.company.vatNumber if user.company else None,
                        "iban": get_file_url(getattr(user.company, "iban", None)),
                        "kbiss": get_file_url(getattr(user.company, "kbiss", None)),
                        "identityCardFront": get_file_url(getattr(user.company, "identityCardFront", None)),
                        "identityCardBack": get_file_url(getattr(user.company, "identityCardBack", None)),
                        "proofOfAddress": get_file_url(getattr(user.company, "proofOfAddress", None)),
                        "profilePhoto": get_file_url(getattr(user.company, "profilePhoto", None)),
                        "coverPhotos": [get_file_url(photo) for photo in user.company.coverPhotos.all()]
                        if hasattr(user.company, "coverPhotos") and user.company.coverPhotos else None,
                    } if user.company else None,
                    "manual_address": {
                        "address1": user.manual_address.address1 if user.manual_address else None,
                        "address2": user.manual_address.address2 if user.manual_address else None,
                        "postalCode": user.manual_address.postalCode if user.manual_address else None,
                        "lat": user.manual_address.lat if user.manual_address else None,
                        "lang": user.manual_address.lang if user.manual_address else None,
                        "city": user.manual_address.city if user.manual_address else None,
                        "country": user.manual_address.country if user.manual_address else None,
                        "state": user.manual_address.state if user.manual_address else None

                    } if user.manual_address else None,
                    "automatic_address": {
                        "address1": user.automatic_address.address1 if user.automatic_address else None,
                        "address2": user.automatic_address.address2 if user.automatic_address else None,
                        "postalCode": user.automatic_address.postalCode if user.automatic_address else None,
                        "lat": user.automatic_address.lat if user.automatic_address else None,
                        "lang": user.automatic_address.lang if user.automatic_address else None,
                        "city": user.automatic_address.city if user.automatic_address else None,
                        "country": user.automatic_address.country if user.automatic_address else None
                    } if user.automatic_address else None
                }
            })

        except Exception as e:
            error_message = str(e).lower()

            if "siret" in error_message:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "This SIRET number is already registered. Please use a unique SIRET."
                }, status=200)

            if "vat" in error_message or "vatNumber" in error_message:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "This VAT number is invalid or already registered. Please use a valid VAT number."
                }, status=200)

            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Internal server.",
                "message": str(e)
            }, status=500)
class ProfessionalUserProfileView(generics.RetrieveAPIView):
    serializer_class = ProfessionalUserSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, *args, **kwargs):
        user = request.user
        try:
            professional_user = ProfessionalUser.objects.get(email=user.email)  
            serializer = self.get_serializer(professional_user)
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Professional user profile retrieved successfully.",
                "data": serializer.data
            }, status=status.HTTP_200_OK)
        
        except ProfessionalUser.DoesNotExist:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Professional user not found."
            }, status=status.HTTP_200_OK)
        
        
class ProfessionalUserLoginView(APIView):

    @swagger_auto_schema(
        operation_summary="Professional User Login",
        operation_description="Authenticate a professional user using email and password.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                'email': openapi.Schema(type=openapi.TYPE_STRING, format=openapi.FORMAT_EMAIL, description="User's email"),
                'password': openapi.Schema(type=openapi.TYPE_STRING, format=openapi.FORMAT_PASSWORD, description="User's password"),
            },
            required=['email', 'password'],
        ),
        responses={
            200: openapi.Response(
                description="Login successful",
                examples={  
                    "message": "User account logged in successfully",
                    "status": "success",
                    "access_token": "yourAccessToken",
                    "refresh_token": "yourRefreshToken"
                }
            ),
            401: openapi.Response(
                description="Unauthorized",
                examples={
                    "application/json": {
                        "message": "Invalid email or password"
                    }
                }
            ),
        }
    )

    def post(self, request):
        email = request.data.get('email')
        password = request.data.get('password')
        try:
            user = ProfessionalUser.objects.get(email=email)

            if user.is_deleted:
                return Response({
                    "statusCode": 403,
                    "status": False,
                    "message": "Your account is deleted. Please contact admin."
                }, status=status.HTTP_200_OK)

            if not check_password(password, user.password):
                return Response({"statusCode": 400, "status": False, "message": "Invalid email or password"}, status=status.HTTP_200_OK)
            
            verifications = [
                user.identityCardFront_status,
                user.identityCardBack_status,
                user.proofOfAddress_status,
                user.iban_status,
                user.kbiss_status,
            ]
            is_verified = all(verifications) and user.finalDocument_status == "approved"   
            user.last_login = timezone.now()
            user.save(update_fields=["last_login"]) 
            refresh = RefreshToken()
            refresh.payload['email'] = user.email
            refresh.payload['user_type'] = "professional_user"
            refresh.payload['professional_id'] = str(user.id)
            access_token = str(refresh.access_token)

            def get_file_url(file_field):
                return file_field.url if file_field and hasattr(file_field, 'url') else None
            
            if not user.identityCardFront_status:
                return generate_user_response(user, refresh, access_token, "identityCardFront User Documents are not verified")
            if not user.identityCardBack_status:
                return generate_user_response(user, refresh, access_token, "identityCardBack User Documents are not verified")
            if not user.proofOfAddress_status:
                return generate_user_response(user, refresh, access_token, "proofOfAddress User Documents are not verified")
            if not user.iban_status:
                return generate_user_response(user, refresh, access_token, "iban User Documents are not verified")
            if not user.kbiss_status:
                return generate_user_response(user, refresh, access_token, "kbiss User Documents are not verified")
            if user.finalDocument_status != "approved":
                return generate_user_response(user, refresh, access_token, "User Documents are not verified")
            if user.subscription_status.lower() == "trial" and not user.is_free_trial_active :
                subscription_message="your free trial has expired"
            elif user.subscription_status.lower() == "trial" and user.is_free_trial_active:
                subscription_message="your 3 month free trial is still active"
            elif user.subscription_status.lower() == "paid" and user.is_paid_subscription_active:
                subscription_message="your paid subscription is still active"
            elif user.subscription_status.lower() == "paid" and not user.is_paid_subscription_active:
                subscription_message="your paid subscription has expired please renew it or buy new one..."  
            else:
                subscription_message="your subscription is not active choose any subcription"
            if user.stripe_customer_id:
                customerId= user.stripe_customer_id
            else:
                customerId=""    
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Login successful",
                "refresh_token": str(refresh),
                "access_token": access_token,
                "role": str(user.role),
                "user": {
                    "email": user.email,
                    "phone": user.phone,
                    "id": user.id,
                    "customerId":customerId,
                    "subscription_status":user.subscription_status,
                    "is_subscription_active": user.subscription_active,
                    "subscription_message":subscription_message,
                    "is_free_trial_active":user.is_free_trial_active,
                    "is_paid_subscription_active":user.is_paid_subscription_active,
                    "finalDocument_status": user.finalDocument_status,
                    "Subscription": user.subscriptionplan.name if user.subscriptionplan else None,
                    "subscriptiontype": user.subscriptiontype.subscription_type if user.subscriptiontype else None,
                    "is_deleted":user.is_deleted,
                    "company": {
                        "companyID": user.company.id if user.company else None,
                        "companyName": user.company.companyName if user.company else None,
                        "managerFullName": user.company.managerFullName if user.company else None,
                        "phoneNumber": user.company.phoneNumber if user.company else None,
                        "email": user.company.email if user.company else None,
                        "siret": user.company.siret if user.company else None,
                        "sectorofActivity": user.company.sectorofActivity if user.company else None,
                        "vatNumber": user.company.vatNumber if user.company else None,
                        "iban": get_file_url(user.company.iban),
                        "kbiss": get_file_url(user.company.kbiss),
                        "identityCardFront": get_file_url(user.company.identityCardFront),
                        "identityCardBack": get_file_url(user.company.identityCardBack),
                        "proofOfAddress": get_file_url(user.company.proofOfAddress),
                        "profilePhoto": get_file_url(user.company.profilePhoto),
                        "coverPhotos": get_file_url(user.company.coverPhotos),
                    } if user.company else None,
                    "categoryID": [
                        {
                            "id": category.id,
                            "name": category.name,
                            "slug":category.slug
                        }
                        for category in user.categories.all()
                    ],
                    "subcategoriesID": [
                        {
                            "id": subcategory.id,
                            "name": subcategory.name,
                            "slug":subcategory.slug,
                            "category_id": subcategory.parentCategoryId.id if subcategory.parentCategoryId else None
                        }
                        for subcategory in user.subcategories.all()
                    ],
                    "category": {
                        "subscriptionPlan": user.subscriptionplan.name if user.subscriptionplan else None,
                        "limit": user.subscriptionplan.category_limit if user.subscriptionplan else None,
                    },
                    "Subcategory": {
                        "limit": user.subscriptionplan.subcategory_limit if user.subscriptionplan else None,
                    },
                    "manual_address": {
                        "address1": user.manual_address.address1 if user.manual_address else None,
                        "address2": user.manual_address.address2 if user.manual_address else None,
                        "postalCode": user.manual_address.postalCode if user.manual_address else None,
                        "lat": user.manual_address.lat if user.manual_address else None,
                        "lang": user.manual_address.lang if user.manual_address else None,
                        "city": user.manual_address.city if user.manual_address else None,
                        "country": user.manual_address.country if user.manual_address else None
                    } if user.manual_address else None,
                    "automatic_address": {
                        "address1": user.automatic_address.address1 if user.automatic_address else None,
                        "address2": user.automatic_address.address2 if user.automatic_address else None,
                        "postalCode": user.automatic_address.postalCode if user.automatic_address else None,
                        "lat": user.automatic_address.lat if user.automatic_address else None,
                        "lang": user.automatic_address.lang if user.automatic_address else None,
                        "city": user.automatic_address.city if user.automatic_address else None,
                        "country": user.automatic_address.country if user.automatic_address else None
                    } if user.automatic_address else None
                }
            })

        except ProfessionalUser.DoesNotExist:
            return Response({"statusCode": 400, "status": False, "message": "Invalid email or password"}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"statusCode": 500, "status": False, "message": "An error occurred while processing your request.", "message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class DocumentVerificationStatusView(APIView):
    permission_classes = [IsAuthenticated]  # Requires authentication

    def get(self, request):
        try:
            user = request.user  # Get the authenticated user
            professional_user = ProfessionalUser.objects.filter(email=user.email).first()
            document_status = {
                "identityCardFront_status": professional_user.identityCardFront_status,
                "identityCardBack_status": professional_user.identityCardBack_status,
                "proofOfAddress_status": professional_user.proofOfAddress_status,
                "iban_status": professional_user.iban_status,
                "kbiss_status": professional_user.kbiss_status,
                "finalDocument_status":professional_user.finalDocument_status
            }

            return Response({
                "status": True,
                "statusCode": 200,
                "message": "Document verification status retrieved successfully",
                "document_status": document_status
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "status": False,
                "statusCode": 500,
                "message": "An error occurred while fetching document verification status.",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class CompanyDetailsView(generics.RetrieveAPIView):
    permission_classes = [permissions.IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Get Professional User's Company Details",
        operation_description="Returns the company details of the authenticated professional user.",
        responses={
            200: CompanyDetailsSerializer(),
            400: openapi.Response(
                description="Bad Request",
                examples={"application/json": {"message": "Invalid user type"}}
            ),
            404: openapi.Response(
                description="Not Found",
                examples={"application/json": {"message": "No company details found"}}
            ),
        }
    )
    def get(self, request, *args, **kwargs):
        try:
            user = request.user  

            if not isinstance(user, ProfessionalUser):
                return Response(
                    {"statusCode": 400, "status": False, "message": "Invalid user type"},
                    status=status.HTTP_200_OK
                )
            company_details = user.company  
            if not company_details:
                return Response(
                    {"statusCode": 404, "status": False, "message": "No company details found"},
                    status=status.HTTP_200_OK
                )
            assigned_users = list(ProfessionalUser.objects.filter(company=company_details).values("id","email"))
            for user in assigned_users:
                user["professional_id"] = user.pop("id")
                
            serializer = CompanyDetailsSerializer(company_details, context={"request": request})
 
            return Response(
                {
                    "statusCode": 200,
                    "status": True,
                    "assigned_users": assigned_users,
                    "data": serializer.data
                },
                status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                {"statusCode": 500, "status": False, "message": "An unexpected error occurred", "details": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
class UpdateCompanyDetailsAPI(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request):
        try:
            user = request.user
            professional_user = ProfessionalUser.objects.filter(email=user.email).first()
            if not professional_user:
                return Response(
                    {"status": False, "statusCode": 403, "message": "Only a professional user can update this company."},
                    status=status.HTTP_200_OK
                )

            company = professional_user.company
            if not company:
                return Response(
                    {"statusCode": 404, "status": False, "message": "No company associated with this user."},
                    status=status.HTTP_200_OK
                )

            if company.id != professional_user.company.id:
                return Response(
                    {"status": False, "statusCode": 403, "message": "You are not authorized to update this company."},
                    status=status.HTTP_200_OK
                )

            request_data = request.data.copy()
            profile_photo = request.FILES.get("profilePhoto")
            if profile_photo:
                profile_path = f"profile_photos/{profile_photo.name}"
                company.profilePhoto = default_storage.save(profile_path, profile_photo)
            if "coverPhotos" in request.FILES:
                existing_photos = company.coverPhotos or []
                new_photos = request.FILES.getlist("coverPhotos")
                total_photos = len(existing_photos) + len(new_photos)

                if total_photos > 10:
                    return Response(
                        {"status": False, "statusCode": 400, "message": "You can upload a maximum of 10 cover photos in total (existing + new)."},
                        status=status.HTTP_200_OK
                    )
                new_photo_paths = [default_storage.save(f"coverPhotos/{photo.name}", photo) for photo in new_photos]
                company.coverPhotos = existing_photos + new_photo_paths
                company.save()
            if "facilities" in request_data:
                request_data["facilities"] = list(map(int, request_data["facilities"][0].split(',')))
            company_start_time = request_data.get("start_time")
            company_end_time = request_data.get("end_time")
            if company_start_time and company_end_time:
                try:
                    company.start_time = time.fromisoformat(company_start_time)
                    company.end_time = time.fromisoformat(company_end_time)
                    if company.start_time >= company.end_time:
                        return Response(
                            {"status": False, "statusCode": 400, "message": "start_time must be earlier than end_time"},
                            status=status.HTTP_200_OK
                        )
                    if "opening_hours" not in request_data or not request_data["opening_hours"]:
                        days = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
                        company.opening_hours = {
                            day: {
                                "start_time": company_start_time,
                                "end_time": company_end_time
                            } for day in days
                        }

                except ValueError:
                    return Response(
                        {"status": False, "statusCode": 400, "message": "Invalid time format. Use HH:MM."},
                        status=status.HTTP_200_OK
                    )
            if "opening_hours" in request_data:
                try:
                    opening_hours_str = request_data["opening_hours"]
                    if isinstance(opening_hours_str, list):
                        opening_hours_str = opening_hours_str[0]

                    updated_hours = json.loads(opening_hours_str)
                    if not isinstance(updated_hours, dict):
                        return Response(
                            {"status": False, "statusCode": 400, "message": "opening_hours must be a JSON object."},
                            status=status.HTTP_200_OK
                        )

                    existing_hours = company.opening_hours or {}

                    for day, hours in updated_hours.items():
                        if hours is None:
                            existing_hours.pop(day, None)
                        else:
                            start = hours.get("start_time")
                            end = hours.get("end_time")

                            time_pattern = re.compile(r"^\d{2}:\d{2}$")
                            if not time_pattern.match(start or "") or not time_pattern.match(end or ""):
                                return Response(
                                    {"status": False, "statusCode": 400, "message": f"Invalid time format for {day}. Expected HH:MM"},
                                    status=status.HTTP_200_OK
                                )

                            start_dt = datetime.datetime.strptime(start, "%H:%M")
                            end_dt = datetime.datetime.strptime(end, "%H:%M")
                            if start_dt >= end_dt:
                                return Response(
                                    {"status": False, "statusCode": 400, "message": f"For {day}, start_time must be earlier than end_time."},
                                    status=status.HTTP_200_OK
                                )

                            existing_hours[day] = {
                                "start_time": start,
                                "end_time": end
                            }

                    company.opening_hours = existing_hours

                except json.JSONDecodeError as e:
                    return Response(
                        {"status": False, "statusCode": 400, "message": "Invalid JSON format in opening_hours.", "details": str(e)},
                        status=status.HTTP_200_OK
                    )
            company.save()
            serializer = CompanyDetailsUpdateSerializer(
                company,
                data=request_data,
                partial=True,
                context={"request": request}
            )

            if serializer.is_valid():
                serializer.save()
                return Response(
                    {"statusCode": 200, "status": True, "message": "Company details updated successfully!", "data": serializer.data},
                    status=status.HTTP_200_OK
                )

            return Response(
                {"statusCode": 400, "status": False, "message": "Validation failed", "details": serializer.errors},
                status=status.HTTP_200_OK
            )

        except CompanyDetails.DoesNotExist:
            return Response(
                {"statusCode": 404, "status": False, "message": "Company not found!"},
                status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                {"statusCode": 500, "status": False, "message": "An unexpected error occurred", "details": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
class UpdateCoverPhotoAPI(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request):
        try:
            user = request.user
            professional_user = ProfessionalUser.objects.filter(email=user.email).first()
            if not professional_user:
                return Response(
                    {"status": False, "statusCode": 403, "message": "Only a professional user can update this company."},
                    status=status.HTTP_200_OK
                )

            company = professional_user.company
            if not company:
                return Response(
                    {"statusCode": 404, "status": False, "message": "No company associated with this user."},
                    status=status.HTTP_200_OK
                )

            selected_cover_photo = request.data.get("selected_cover_photo")
          
            if not selected_cover_photo:
                return Response(
                    {"status": False, "statusCode": 400, "message": "No cover photo provided!"},
                    status=status.HTTP_200_OK
                )
          
            
            if not isinstance(company.coverPhotos, list):
                company.coverPhotos = []
            if selected_cover_photo not in company.coverPhotos:
                return Response(
                    {"statusCode": 400, "status": False, "message": "Invalid cover photo selection!"},
                    status=status.HTTP_200_OK
                )
            company.selectedCoverPhoto = selected_cover_photo
            company.save()

            return Response(
                {"statusCode": 200, "status": True, "message": "Cover photo updated successfully!", "selectedCoverPhoto": selected_cover_photo},
                status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                {"statusCode": 500, "status": False, "message": "An unexpected error occurred", "details": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
class CompanyImageDeleteView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request):
        try:
            user = request.user
            professional_user = ProfessionalUser.objects.filter(email=user.email).first()
            if not professional_user:
                return Response(
                    {"status": False, "statusCode": 403, "message": "Only a professional user can delete images."},
                    status=status.HTTP_200_OK
                )
            company = professional_user.company
            if not company:
                return Response(
                    {"status": False, "statusCode": 403, "message": "You are not authorized to delete images."},
                    status=status.HTTP_200_OK
                )
            delete_coverPhotos = request.data.get("delete_coverPhotos", [])
            delete_on_site_ordering = request.data.get("delete_on_site_ordering", [])
            if not delete_coverPhotos and not delete_on_site_ordering:
                return Response(
                    {"status": False, "message": "No images specified for deletion"},
                    status=status.HTTP_200_OK
                )
            if delete_coverPhotos:
                existing_covers = company.coverPhotos or []  # Ensure it's a list
                company.coverPhotos = [photo for photo in existing_covers if photo not in delete_coverPhotos]

                for photo in delete_coverPhotos:
                    default_storage.delete(photo)  # Delete from storage
            if delete_on_site_ordering:
                existing_onsite = company.on_site_ordering or []  # Ensure it's a list
                company.on_site_ordering = [photo for photo in existing_onsite if photo not in delete_on_site_ordering]

                for photo in delete_on_site_ordering:
                    default_storage.delete(photo)  # Delete from storage

            company.save()

            return Response(
                {"status": True, "message": "Selected images deleted successfully"},
                status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                {"status": False, "message": "An error occurred", "details": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
class GetAddressView(APIView):    
    def post(self, request):
        latitude = request.data.get('latitude')
        longitude = request.data.get('longitude')

        if not latitude or not longitude:
            return Response({
                "message": "Latitude and longitude are required.",
                "statusCode": 400,
                "status": False
            }, status=status.HTTP_200_OK)

        GOOGLE_MAPS_API_KEY = settings.GOOGLE_MAPS_API_KEY
        geocoding_url = f"https://maps.googleapis.com/maps/api/geocode/json?latlng={latitude},{longitude}&key={GOOGLE_MAPS_API_KEY}"

        try:
            response = requests.get(geocoding_url)
            response_data = response.json()

            if response.status_code == 200 and response_data.get("status") == "OK":
                results = response_data.get("results", [])

                if results:
                    address_components = results[0].get("address_components", [])
                    formatted_address = results[0].get("formatted_address", "")

                    address_data = {
                        "address1": formatted_address,
                        "address2": "",
                        "postalCode": "",
                        "latitude": latitude,
                        "longitude": longitude,
                        "country": "",
                        "city": "",
                        "countryCode": "",
                        "countryShortName": "",
                        "dialCode": "",
                        "flag": "",
                        "formattedAddress": formatted_address
                    }
                    nearby_addresses = [] 

                    for component in address_components:
                        types = component.get("types", [])
                        if "country" in types:
                            address_data["country"] = component.get("long_name", "")
                            address_data["countryCode"] = component.get("short_name", "")
                        if "locality" in types:
                            address_data["city"] = component.get("long_name", "")
                        if "postal_code" in types:
                            address_data["postalCode"] = component.get("long_name", "")
                        if "sublocality" in types or "neighborhood" in types:
                            nearby_addresses.append(component.get("long_name", ""))
                        if "route" in types or "street_address" in types or "intersection" in types:
                            nearby_addresses.append(component.get("long_name", ""))
                    address_data["address2"] = ", ".join(set(nearby_addresses)) if nearby_addresses else ""
               
                    country_info = requests.get("https://restcountries.com/v3.1/all").json()
                    country_details = next((c for c in country_info if c.get("cca2") == address_data["countryCode"]), None)

                    if country_details:
                        dial_code = country_details.get("idd", {}).get("root", "") + \
                                    (country_details.get("idd", {}).get("suffixes", [""])[0] if country_details.get("idd", {}).get("suffixes") else "")
                        address_data["dialCode"] = dial_code
                        address_data["countryShortName"] = country_details.get("cca2", "")
                        address_data["flag"] = f"https://flagcdn.com/w320/{address_data['countryCode'].lower()}.png"

                    return Response({
                        "statusCode": 200,
                        "status": True,
                        "address": address_data
                    }, status=status.HTTP_200_OK)

            return Response({
                "message": "No address found for the given coordinates.",
                "statusCode": 400,
                "status": False
            }, status=status.HTTP_200_OK)

        except requests.exceptions.RequestException as e:
            return Response({
                "message": f"Failed to fetch address: {str(e)}",
                "statusCode": 500,
                "status": False
            }, status=status.HTTP_200_OK)



# ---------------------------------- Store reel -----------------------------------------------

# class StoreReelCreateAPIView(APIView):
    
#     permission_classes = [IsAuthenticated]
#     parser_classes = (MultiPartParser, FormParser)
    
#     def post(self, request, *args, **kwargs):
#         try:
#             video_file = request.FILES.get("video")
#             if not video_file:
#                 return Response({
#                     "statusCode": 400,
#                     "status": False,
#                     "message": "No video file provided."
#                 }, status=status.HTTP_200_OK)

            
#             max_file_size = 100 * 1024 * 1024  # 100MB
#             if video_file.size > max_file_size:
#                 return Response({
#                     "statusCode": 400,
#                     "status": False,
#                     "message": "Video file size exceeds the maximum allowed limit of 100MB."
#                 }, status=status.HTTP_200_OK)

            
#             allowed_formats = ["video/mp4", "video/avi", "video/quicktime"]
#             if video_file.content_type not in allowed_formats:
#                 return Response({
#                     "statusCode": 400,
#                     "status": False,
#                     "message": "Invalid video format. Only MP4, AVI, and MOV are allowed."
#                 }, status=status.HTTP_200_OK)
    
#             compressed_video = self.compress_video(video_file)
#             company_id = request.data.get("company_id")
#             category_id = request.data.get("category_id")  
#             title = request.data.get("title")

#             company = CompanyDetails.objects.filter(id=company_id).first()

           
#             user = request.user
#             professional = ProfessionalUser.objects.filter(email=request.user.email).first()
#             if not professional:
#                 return Response({
#                     "statusCode": 403,
#                     "status": False,
#                     "message": "Unauthorized access."
#                 }, status=status.HTTP_200_OK)

            
#             category = None
#             if category_id:
#                 category = Category.objects.filter(id=category_id).first()
#                 if not category:
#                     return Response({
#                         "statusCode": 400,
#                         "status": False,
#                         "message": "Invalid category."
#                     }, status=status.HTTP_200_OK)

                
#                 if not professional.categories.filter(id=category_id).exists():
#                     return Response({
#                         "statusCode": 403,
#                         "status": False,
#                         "message": "You can only select categories that you registered with."
#                     }, status=status.HTTP_200_OK)

           
#             reel = StoreReel.objects.create(
#                 video=compressed_video,
#                 company_id=company,
#                 category=category,  
#                 title=title
#             )

           
#             thumbnail_path = self.generate_thumbnail(reel.video)

            
#             if thumbnail_path:
#                 with open(thumbnail_path, "rb") as thumb_file:
#                     reel.thumbnail.save(os.path.basename(thumbnail_path), ContentFile(thumb_file.read()), save=True)

            
#             serialized_reel = StoreReelSerializer(reel, context={"request": request}).data

#             return Response({
#                 "statusCode": 200,
#                 "status": True,
#                 "message": "Reel uploaded successfully with thumbnail!",
#                 "data": serialized_reel
#             }, status=status.HTTP_200_OK)

#         except Exception as e:
#             return Response({
#                 "statusCode": 500,
#                 "status": False,
#                 "message": "An unexpected error occurred while uploading the reel.",
#                 "details": str(e)
#             }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
#     def generate_thumbnail(self, video):
#         """Extract the first frame of the video and save as a thumbnail."""
#         try:
#             temp_video_path = tempfile.NamedTemporaryFile(delete=False).name
#             with open(temp_video_path, 'wb') as temp_file:
#                 temp_file.write(video.read())

#             clip = VideoFileClip(temp_video_path)
#             frame = clip.get_frame(1) 

#             temp_thumb_path = tempfile.NamedTemporaryFile(delete=False, suffix=".jpg").name
#             cv2.imwrite(temp_thumb_path, cv2.cvtColor(frame, cv2.COLOR_RGB2BGR))

#             return temp_thumb_path
#         except Exception as e:
#             return None
        
#     def compress_video(self, video_file):
#         """Compress video without changing resolution."""
#         try:
#             import tempfile, os
#             from django.core.files.base import ContentFile
#             from moviepy.video.io.VideoFileClip  import VideoFileClip
 
#             # Save uploaded file temporarily
#             input_path = tempfile.NamedTemporaryFile(delete=False, suffix=".mp4").name
#             with open(input_path, 'wb') as f:
#                 f.write(video_file.read())
            
#             original_size_mb = os.path.getsize(input_path) / (1024 * 1024)
#             print(f"Original video size: {original_size_mb:.2f} MB")
#             # Load video
#             clip = VideoFileClip(input_path)
 
#             # Output path
#             output_path = tempfile.NamedTemporaryFile(delete=False, suffix=".mp4").name
           
 
#             # Compress without changing resolution
#             clip.write_videofile(
#                 output_path,
#                 bitrate="1000k",  # control compression rate
#                 codec="libx264",
#                 audio_codec="aac"
#             )
 
#             compressed_size_mb = os.path.getsize(output_path) / (1024 * 1024)
#             print(f"Compressed video size: {compressed_size_mb:.2f} MB")
 
#             # Read and return compressed video as Django file
#             with open(output_path, 'rb') as f:
#                 return ContentFile(f.read(), name=os.path.basename(video_file.name))
 
#         except Exception as e:
#             print(f"Compression error: {e}")
#             return video_file
    
# update reel title
class StoreReelUpdateTitleAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, *args, **kwargs):
        try:
            reel_id = request.data.get("reel_id")
            new_title = request.data.get("title")
            new_category_id = request.data.get("category")  # Ensure correct key
            new_subcategory_id = request.data.get("subcategory")
            isActive = request.data.get("isActive")
            if not reel_id:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Reel ID is required."
                }, status=status.HTTP_200_OK)
                
            if new_category_id is None:  # Explicitly check if category is missing
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Please select a category."
                }, status=status.HTTP_200_OK)
            
            reel = StoreReel.objects.filter(id=reel_id, is_deleted=False).first()
            if not reel:
                return Response({
                    "statusCode": 404,
                    "status": False,
                    "message": "Reel not found."
                }, status=status.HTTP_200_OK)
            user = request.user  
            professional_user = ProfessionalUser.objects.filter(email=user.email).first()

            if not professional_user:
                return Response({
                    "statusCode": 403,
                    "status": False,
                    "message": "Only professional users can update reels."
                }, status=status.HTTP_200_OK)
            user_categories = professional_user.categories.all()
            user_subcategories = professional_user.subcategories.all()
            if new_category_id:
                category = Category.objects.filter(id=new_category_id).first()

                if not category:
                    return Response({
                        "statusCode": 400,
                        "status": False,
                        "message": "Invalid category ID."
                    }, status=status.HTTP_200_OK)

                if category not in user_categories:
                    return Response({
                        "statusCode": 403,
                        "status": False,
                        "message": "You can only update reels to your registered category."
                    }, status=status.HTTP_200_OK)
                reel.category = category  
            if new_subcategory_id:
                subcategory = Subcategory.objects.filter(id=new_subcategory_id).first()
                if not subcategory:
                    return Response({
                        "statusCode": 400,
                        "status": False,
                        "message": "Invalid subcategory ID."
                    }, status=status.HTTP_200_OK)
                
                if subcategory not in user_subcategories:
                    return Response({
                        "statusCode": 403,
                        "status": False,
                        "message": "You can only update reels to your registered subcategory."
                    }, status=status.HTTP_200_OK)
                
                reel.subcategory = subcategory
            if new_title:
                reel.title = new_title
                
            if isActive is not None:
                reel.isActive = bool(isActive)

            reel.save()
            reel.refresh_from_db()  # Force refresh
            
            
            return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Reel updated successfully.",
                    "data": {
                        "id": reel.id,
                        "title": reel.title,
                        "type": "video",
                        "video": reel.video.url if reel.video else None,
                        "thumbnail": reel.thumbnail.url if reel.thumbnail else None,
                        "m3u8_url": reel.m3u8_url if hasattr(reel, "m3u8_url") else None,
                        "views": reel.views,
                        "likes": reel.likes,
                        "shares": reel.shares,
                        "comments": reel.comments,
                        "created_at": reel.created_at,
                        "isActive": reel.isActive,
                        "category": {
                            "id": reel.category.id if reel.category else None,
                            "name": reel.category.name if reel.category else None
                        } if reel.category else None,
                        "subcategory": {
                            "id": reel.subcategory.id if reel.subcategory else None,
                            "name": reel.subcategory.name if reel.subcategory else None
                        } if reel.subcategory else None,
                        "company": {
                            "id": reel.company_id.id if reel.company_id else None,
                            "name": reel.company_id.companyName if reel.company_id else None
                        } if reel.company_id else None
                    }
                }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred while updating the reel.",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        


class ReelLikeToggleAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        try:
            reel_id = request.data.get("reel_id")
            folder_id = request.data.get("folder_id")  # expecting folder ID now
 
            folder_id = request.data.get("folder_id")

            if not reel_id:
                return Response({
                    "statusCode": 400, "status": False, "message": "Reel ID is required"
                }, status=status.HTTP_200_OK)
 
            if not folder_id:
                return Response({
                    "statusCode": 400, "status": False, "message": "Folder ID is required"
                }, status=status.HTTP_200_OK)
            try:
                reel = StoreReel.objects.get(id=reel_id)
            except StoreReel.DoesNotExist:
                return Response({
                    "statusCode": 404, "status": False, "message": "Reel not found"
                }, status=status.HTTP_200_OK)
            try:
                folder = ReelFolder.objects.get(id=folder_id, user=request.user)
            except ReelFolder.DoesNotExist:
                return Response({
                    "statusCode": 404, "status": False, "message": "Folder not found"
                }, status=status.HTTP_200_OK)

            like, created = ReelLike.objects.get_or_create(user=request.user, reel=reel)
            if not created:
                already_saved = SavedReel.objects.filter(user=request.user, reel=reel, folder=folder).exists()

                if already_saved:
                    SavedReel.objects.filter(user=request.user, reel=reel, folder=folder).delete()

                    still_saved = SavedReel.objects.filter(user=request.user, reel=reel).exists()
                    if not still_saved:
                        like.delete()
                        reel.likes = max(0, reel.likes - 1)
                        message = "Reel unliked and removed from folder!"
                    else:
                        message = "Reel removed from folder."
                    response_data = {}

                else:
                    SavedReel.objects.filter(user=request.user, reel=reel).exclude(folder=folder).delete()
                    saved_reel, _ = SavedReel.objects.get_or_create(user=request.user, reel=reel, folder=folder)
                    message = "Reel moved to new folder!"
                    response_data = {
                        "like": ReelLikeSerializer(like, context={"request": request}).data,
                        "saved": SavedReelSerializer(saved_reel, context={"request": request}).data,
                    }

            else:
                professional = ProfessionalUser.objects.filter(company=reel.company_id).first()
                if professional:
                    notification_message = f"{request.user.username} liked your reel."
                    get_player_ids_by_professional_id(professional.id, notification_message)

                reel.likes += 1
                message = "Reel liked and saved successfully!"
                SavedReel.objects.filter(user=request.user, reel=reel).exclude(folder=folder).delete()
                saved_reel, _ = SavedReel.objects.get_or_create(user=request.user, reel=reel, folder=folder)

                response_data = {
                    "like": ReelLikeSerializer(like, context={"request": request}).data,
                    "saved": SavedReelSerializer(saved_reel, context={"request": request}).data,
                }
 
            reel.save(update_fields=["likes"])
 
            return Response({
                "statusCode": 200,
                "status": True,
                "message": message,
                "likes": reel.likes,
                "user_name": request.user.username,
                "data": response_data
            }, status=status.HTTP_200_OK)
 
        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ReelViewAPI(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, reel_id):
        user = request.user
        if not reel_id:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Reel ID is required"
            }, status=status.HTTP_200_OK)

        try:
            reel_id = int(reel_id)  # Ensure reel_id is a valid integer
        except ValueError:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Invalid Reel ID format"
            }, status=status.HTTP_200_OK)

        try:
            reel = StoreReel.objects.get(id=reel_id)
        except StoreReel.DoesNotExist:
            raise NotFound({
                "statusCode": 404,
                "status": False,
                "message": "Reel not found"
            }, status=status.HTTP_200_OK)

        try:
            ReelView.objects.create(user=user, reel=reel)
            StoreReel.objects.filter(id=reel.id).update(views=F("views") + 1)
            updated_reel = StoreReel.objects.get(id=reel.id)

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "View recorded successfully",
                "views_count": updated_reel.views  
            }, status=status.HTTP_200_OK)

        except IntegrityError:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Database integrity error occurred while saving view"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        except DatabaseError:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Database error occurred while updating view count"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        except ValidationError:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Invalid data format"
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Unexpected error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            


class ReelCommentCreateAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        reel_id = request.data.get("reel_id")
        comment_text = request.data.get("comment")
        parent_id = request.data.get("parent_id", None)

        if not reel_id or not comment_text:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Reel ID and comment are required"
            }, status=status.HTTP_200_OK)

        try:
            reel = StoreReel.objects.get(id=int(reel_id))
        except (StoreReel.DoesNotExist, ValueError):
            return Response({
                "statusCode": 404,
                "status": False,
                "message": "Reel not found"
            }, status=status.HTTP_200_OK)

        parent_comment = None
        if parent_id:
            try:
                parent_comment = ReelComment.objects.get(id=int(parent_id))
                if parent_comment.reel_id != reel.id:
                    return Response({
                        "statusCode": 400,
                        "status": False,
                        "message": "Parent comment does not belong to this reel"
                    }, status=status.HTTP_200_OK)
            except (ReelComment.DoesNotExist, ValueError):
                return Response({
                    "statusCode": 404,
                    "status": False,
                    "message": "Parent comment not found"
                }, status=status.HTTP_200_OK)

        try:
            mentioned_usernames = re.findall(r'@(\w+)', comment_text)
            mentioned_usernames = set(mentioned_usernames)

            if mentioned_usernames:
                user_content_type = ContentType.objects.get_for_model(user.__class__)
                follower_friendships = Friendship.objects.filter(
                    receiver_content_type=user_content_type,
                    receiver_object_id=user.id,
                    relationship_type='follow',
                    status='follow'
                )

                following_friendships = Friendship.objects.filter(
                    sender_content_type=user_content_type,
                    sender_object_id=user.id,
                    relationship_type='follow',
                    status='follow'
                )

                follower_ids = [
                    f.sender_object_id for f in follower_friendships
                    if f.sender_content_type.model_class() == user.__class__
                ]

                following_ids = [
                    f.receiver_object_id for f in following_friendships
                    if f.receiver_content_type.model_class() == user.__class__
                ]

                allowed_usernames = list(
                    Users.objects.filter(id__in=follower_ids + following_ids).values_list("username", flat=True)
                )
                for username in mentioned_usernames:
                    if username not in allowed_usernames:
                        comment_text = re.sub(f'@{username}', f'@{username} ', comment_text)

            new_comment = ReelComment.objects.create(
                user=user, reel=reel, comment=comment_text, parent=parent_comment
            )
            
            professional = ProfessionalUser.objects.filter(company=reel.company_id).first()
            notification_message = f"{request.user.username} commented on your reel."
            get_player_ids_by_professional_id(professional.id, notification_message)
            
            comment_data = ReelCommentSerializer(new_comment).data
            comment_data["comment_id"] = comment_data.pop("id")

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Comment added successfully",
                "user": UserSerializer(user).data,
                "data": comment_data
            }, status=status.HTTP_200_OK)

        except IntegrityError:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Database integrity error while saving comment"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Unexpected error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        


class ReelFolderDetailAPIView(APIView):
    def get(self, request, folder_id):
        lat = request.query_params.get('lat')
        lng = request.query_params.get('lang')

        try:
            folder = ReelFolder.objects.get(id=folder_id, user=request.user)
        except ReelFolder.DoesNotExist:
            return Response({
                "statusCode": 404,
                "status": False,
                "message": "Folder not found"
            }, status=status.HTTP_200_OK)

        saved_reels = SavedReel.objects.filter(folder=folder)
        reels_data = []

        for saved in saved_reels:
            reel = saved.reel
            company = reel.company_id if reel and reel.company_id else None

            address_obj = company.manual_address or company.automatic_address if company else None

            comp_lat = address_obj.lat if address_obj else None
            comp_lng = address_obj.lang if address_obj else None

            distance = calculate_distance(lat, lng, comp_lat, comp_lng) if lat and lng and comp_lat and comp_lng else None

            reels_data.append({
                "id": reel.id,
                "title": reel.title,
                "video": reel.video.url if reel.video else None,
                "thumbnail": reel.thumbnail.url if reel.thumbnail else None,
                "likes": reel.likes,
                "shares": reel.shares,
                "views": reel.views,
                "comments": reel.comments,
                "isActive": reel.isActive,
                "is_deleted": reel.is_deleted,
                "deleted_at": reel.deleted_at,
                "created_at": reel.created_at,
                "company": {
                    "id": company.id if company else None,
                    "name": company.companyName if company else None,
                    "average_rating": float(company.average_rating) if company else 0.0,
                    "address":address_obj.address1,
                    "address_lang":address_obj.lang,
                    "adress_lat":address_obj.lat,
                    "distance_in_km": distance
                } if company else None
            })

        return Response({
            "statusCode": 200,
            "status": True,
            "message": "Folder and reels fetched successfully",
            "data": {
                "id": folder.id,
                "name": folder.name,
                "reels": reels_data
            }
        }, status=status.HTTP_200_OK)
    

class ReelCommentListAPIView(APIView):

    def get(self, request, reel_id):
        try:
            reel = get_object_or_404(StoreReel, id=reel_id)
            comments = ReelComment.objects.filter(reel=reel, parent=None).order_by("-created_at")
            def serialize_comment(comment):
                elapsed_time = timesince(comment.created_at, now()).split(",")[0]
                is_like = request.user.is_authenticated and comment.likes.filter(id=request.user.id).exists()
                profile_image = None
                try:
                    privacy_settings = PrivacySetting.objects.get(user=comment.user)
                    if privacy_settings.identify_visibility != 'private':
                        profile_image = (
                            comment.user.profileImage.url 
                            if hasattr(comment.user, 'profileImage') and comment.user.profileImage 
                            else None
                        )
                except PrivacySetting.DoesNotExist:
                    profile_image = (
                        comment.user.profileImage.url 
                        if hasattr(comment.user, 'profileImage') and comment.user.profileImage 
                        else None
                    )

                comment_data = {
                    "id": comment.id,
                    "user": comment.user.id,
                    "username": comment.user.username,
                    "profile_image": profile_image,
                    "reel": comment.reel.id,
                    "comment": comment.comment,
                    "parent": comment.parent.id if comment.parent else None,
                    "created_at": comment.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                    "elapsed_time": elapsed_time,
                    "likes_count": comment.likes.count(),
                    "is_like": is_like,
                    "replies": [serialize_comment(reply) for reply in comment.replies.all()],
                }

                if hasattr(comment, 'is_store_reply'):
                    comment_data["is_store_reply"] = comment.is_store_reply

                return comment_data
            comment_list = [serialize_comment(comment) for comment in comments]

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Comments retrieved successfully",
                "data": comment_list
            }, status=status.HTTP_200_OK)

        except ValidationError as e:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": f"Validation error: {str(e)}"
            }, status=status.HTTP_200_OK)

        except StoreReel.DoesNotExist:
            return Response({
                "statusCode": 404,
                "status": False,
                "message": "Reel not found"
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Unexpected error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class LikeCommentAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            comment_id = request.data.get("comment_id")
            comment = get_object_or_404(ReelComment, id=comment_id)

            if request.user in comment.likes.all():
                comment.likes.remove(request.user)
                liked = False
            else:
                comment.likes.add(request.user)
                liked = True

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Comment liked" if liked else "Comment unliked",
                "likes_count": comment.likes.count()
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Unexpected error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class ReelCommentDeleteAPIView(APIView):
   
    permission_classes = [IsAuthenticated]

    def delete(self, request, comment_id):
        try:
           
            comment = get_object_or_404(ReelComment, id=comment_id)
           
            if request.user != comment.user and not request.user.is_staff:
                return Response({
                    "statusCode": 403,
                    "status": False,
                    "message": "You do not have permission to delete this comment"
                }, status=status.HTTP_200_OK)

            reel = comment.reel
            
            comment.delete()

            if reel:
                store_reel = StoreReel.objects.filter(id=reel.id).first()
                if store_reel and store_reel.comments is not None:
                    StoreReel.objects.filter(id=reel.id).update(comments=F("comments") - 1)

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Comment deleted successfully"
            }, status=status.HTTP_200_OK)

        except IntegrityError:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Database error while deleting the comment"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Unexpected error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




class ReelShareCreateAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        reel_id = request.query_params.get("reel_id")
        raw_ids = request.data.get("recipient_user_ids")
        share_method = request.data.get("share_method", "direct")
        if not reel_id:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Reel ID is required"
            }, status=status.HTTP_200_OK)
        recipient_user_ids = []
        if isinstance(raw_ids, str):
            recipient_user_ids = [int(x.strip()) for x in raw_ids.split(',') if x.strip().isdigit()]
        elif isinstance(raw_ids, list):
            recipient_user_ids = raw_ids
        if not recipient_user_ids:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "recipient_user_ids must be a list or comma-separated string of user IDs"
            }, status=status.HTTP_200_OK)

        reel = get_object_or_404(StoreReel, id=reel_id)
        successful_shares = []
        errors = []

        for recipient_id in recipient_user_ids:
            if str(user.id) == str(recipient_id):
                errors.append(f"Cannot share with yourself (User ID: {recipient_id})")
                continue

            try:
                recipient_user = Users.objects.get(id=recipient_id)
            except Users.DoesNotExist:
                errors.append(f"User ID {recipient_id} does not exist.")
                continue

            share_data = {
                "user": user.id,
                "reel": reel.id,
                "recipient_user": recipient_user.id,
                "share_method": share_method
            }

            serializer = ReelShareSerializer(data=share_data)
            if serializer.is_valid():
                serializer.save()
                successful_shares.append(serializer.data)
            else:
                errors.append({f"User ID {recipient_id}": serializer.errors})

        return Response({
            "statusCode": 200,
            "status": bool(successful_shares),
            "message": "Reel shared successfully" if successful_shares else "Failed to share reel",
            "data": successful_shares,
          
        }, status=status.HTTP_200_OK)
    


    
class GetStoreReelAPIView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Retrieve Store Reels",
        operation_description="Fetch all video reels for a specific company.",
        manual_parameters=[
            openapi.Parameter(
                "company_id",
                openapi.IN_PATH,
                description="ID of the company to fetch reels for",
                type=openapi.TYPE_INTEGER,
                required=True
            ),
        ],
        responses={
            200: openapi.Response(
                "Successful response",
                openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "message": openapi.Schema(type=openapi.TYPE_STRING, example="Company reels retrieved successfully!"),
                        "total_count": openapi.Schema(type=openapi.TYPE_INTEGER, example=5),
                        "data": openapi.Schema(
                            type=openapi.TYPE_ARRAY,
                            items=openapi.Schema(
                                type=openapi.TYPE_OBJECT,
                                properties={
                                    "id": openapi.Schema(type=openapi.TYPE_INTEGER, example=1),
                                    "company_id": openapi.Schema(type=openapi.TYPE_INTEGER, example=10),
                                    "video": openapi.Schema(type=openapi.TYPE_STRING, format="uri", example="https://example.com/media/reel.mp4"),
                                    "thumbnail": openapi.Schema(type=openapi.TYPE_STRING, format="uri", example="https://example.com/media/reel.jpg"),
                                    "description": openapi.Schema(type=openapi.TYPE_STRING, example="Promotional video"),
                                    "created_at": openapi.Schema(type=openapi.TYPE_STRING, format="date-time", example="2025-03-11T12:34:56Z"),
                                }
                            )
                        ),
                    }
                )
            ),
            400: openapi.Response("Bad Request - Missing or invalid company_id"),
            404: openapi.Response("Not Found - No reels found for this company"),
            500: openapi.Response("Internal Server Error"),
        },
    )
    def get(self, request, company_id=None):
        if company_id is None:
            return Response({"statusCode": 400, "status": False, "message": "Company ID is required."}, status=status.HTTP_200_OK)

        try:
            reels = StoreReel.objects.filter(company_id=company_id)
            if not reels.exists():
                return Response({"statusCode": 404, "status": False, "message": "No reels found for this company."}, status=status.HTTP_200_OK)

            serializer = StoreReelSerializer(reels, many=True, context={"request": request})
            
            return Response({
                "statusCode": 200, "status": True,
                "message": "Company reels retrieved successfully!",
                "total_count": reels.count(),
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"statusCode": 500, "status": False, "message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class StoreReelDeleteAPIView(APIView):
    permission_classes = [IsAuthenticated]
    
    @swagger_auto_schema(
        operation_summary="Delete a Store Reel",
        operation_description="Deletes a specific store reel by ID. Requires authentication.",
        manual_parameters=[
            openapi.Parameter(
                "pk",
                openapi.IN_PATH,
                description="ID of the Store Reel to delete",
                type=openapi.TYPE_INTEGER,
                required=True
            )
        ],
        responses={
            204: openapi.Response(
                description="Successful response",
                examples={
                    "application/json": {
                        "message": "Reel deleted successfully!"
                    }
                }
            ),
            404: openapi.Response(
                description="Reel Not Found",
                examples={
                    "application/json": {
                        "message": "Reel not found"
                    }
                }
            ),
            401: openapi.Response(
                description="Unauthorized - Bearer Token Required",
                examples={
                    "application/json": {
                        "detail": "Authentication credentials were not provided."
                    }
                }
            ),
            500: openapi.Response(
                description="Server Error",
                examples={
                    "application/json": {
                        "message": "An unexpected error occurred",
                        "details": "Error message here"
                    }
                }
            ),
        },
    )

    def delete(self, request, *args, **kwargs):
        try:
            reel_id = kwargs.get('pk')
            
            if not reel_id:
                return Response({"statusCode": 400, "status": False, "message": "Reel ID is required"}, 
                                status=status.HTTP_200_OK)
            try:
                instance = StoreReel.objects.get(pk=reel_id)
            except StoreReel.DoesNotExist:
                return Response({
                    "statusCode": 400, 
                    "status": False,
                    "message": "Reel not found"
                }, status=status.HTTP_200_OK)
            instance.delete()

            return Response({
                "statusCode": 200, "status":True,
                "message": "Reel deleted successfully!"
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 400, "status":False,
                "message": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class StoreImageUploadAPIView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        try:
            image_file = request.FILES.get("image")
            company_id = request.data.get("company_id")
            category_id = request.data.get("category_id")
            title = request.data.get("title")
            if not image_file:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "No image file provided."
                }, status=status.HTTP_200_OK)

            try:
                image_file = compress_image(image_file, quality=70)
            except Exception as e:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": f"Image compression failed: {str(e)}"
                }, status=status.HTTP_200_OK)

            company = CompanyDetails.objects.filter(id=company_id).first()
            if not company:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Invalid Company ID."
                }, status=status.HTTP_200_OK)
            category = None
            if category_id:
                category = Category.objects.filter(id=category_id).first()
                if not category:
                    return Response({
                        "statusCode": 400,
                        "status": False,
                        "message": "Invalid Category ID."
                    }, status=status.HTTP_200_OK)
            image_instance = StoreImage.objects.create(
                image=image_file,
                company_id=company,
                category=category,
                title=title
            )
            serialized_image = StoreImageSerializer(image_instance, context={"request": request}).data
            
            serialized_image["company"] = {
                "id": company.id,
                "name": company.companyName 
            }
            
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Image uploaded successfully!",
                "data": serialized_image
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred while uploading the image.",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR) 



class GetStoreMediaCategoryAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        company_id = request.query_params.get("company_id")
        if not company_id:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Company ID is required."
            }, status=status.HTTP_200_OK)

        try:
            company_id = int(company_id)  
        except ValueError:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Invalid Company ID. It must be an integer."
            }, status=status.HTTP_200_OK)

        company = get_object_or_404(CompanyDetails, id=company_id)

        reels = StoreReel.objects.filter(
            company_id=company_id,
            category__isnull=False
        ).exclude(category=None).order_by('-created_at')

        images = StoreImage.objects.filter(
            company_id=company_id,
            category__isnull=False
        ).exclude(category=None).order_by('-created_at')
        
        media_items = []
        current_time = timezone.now()  
        
        for reel in reels:
            thumbnail_url = None
            if reel.thumbnail:
                if str(reel.thumbnail).startswith("http"):
                    thumbnail_url = str(reel.thumbnail)
                else:
                    try:
                        if hasattr(reel.thumbnail, 'url'):
                            thumbnail_url = request.build_absolute_uri(reel.thumbnail.url)
                        else:
                            thumbnail_url = request.build_absolute_uri(str(reel.thumbnail))
                    except Exception:
                        thumbnail_url = None  
            media_items.append({
                "id": reel.id,
                "title": reel.title,
                "type": "video",
                "video": reel.video.url if reel.video else None,
                "thumbnail": thumbnail_url,
                "m3u8_url": reel.m3u8_url if reel.m3u8_url else None,
                "isActive": reel.isActive,
                "is_m3u8_generated": bool(reel.m3u8_url),
                "category": {
                    "id": reel.category.id if reel.category else None,
                    "name": reel.category.name if reel.category else None
                } if reel.category else None,
                "subcategory": {
                    "id": reel.subcategory.id if reel.subcategory else None,
                    "name": reel.subcategory.name if reel.subcategory else None
                } if reel.subcategory else None,
                "comments": reel.comments,
                "shares": reel.shares,
                "views": reel.views,
                "likes": reel.likes,  
                "created_at": reel.created_at if reel.created_at else current_time
            })

        for image in images:
            media_items.append({
                "id": image.id,
                "title": image.title,
                "type": "image",
                "image": image.image.url if image.image else None,
                "isActive": image.isActive,
                "category": {
                    "id": image.category.id if image.category else None,
                    "name": image.category.name if image.category else None
                } if image.category else None,
                "subcategory": {
                    "id": image.subcategory.id if image.subcategory else None,
                    "name": image.subcategory.name if image.subcategory else None
                } if image.subcategory else None,
                "created_at": image.created_at if image.created_at else current_time
            })

        
        media_items.sort(key=lambda x: x["created_at"], reverse=True)


        return Response({
            "statusCode": 200,
            "status": True,
            "message": "Store media retrieved successfully!",
            "total_media": len(media_items),
            "total_videos": len(reels),
            "total_images": len(images),
            "media": media_items
        }, status=status.HTTP_200_OK)
    


    
class GetStoreMediaAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        company_id = request.query_params.get("company_id")

   
        if not company_id:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Company ID is required."
            }, status=status.HTTP_200_OK)

        try:
            company_id = int(company_id)  
        except ValueError:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Invalid Company ID. It must be an integer."
            }, status=status.HTTP_200_OK)

        company = get_object_or_404(CompanyDetails, id=company_id)

        reels = StoreReel.objects.filter(company_id=company_id, category__isnull=True).order_by('-created_at')

        images = StoreImage.objects.filter(company_id=company_id, category__isnull=True).order_by('-created_at')

        media_items = []
        current_time = timezone.now()  
        
        for reel in reels:
            thumbnail_url = None
            if reel.thumbnail:
                if str(reel.thumbnail).startswith("http"):
                    thumbnail_url = str(reel.thumbnail)
                else:
                    try:
                        if hasattr(reel.thumbnail, 'url'):
                            thumbnail_url = request.build_absolute_uri(reel.thumbnail.url)
                        else:
                            thumbnail_url = request.build_absolute_uri(str(reel.thumbnail))
                    except Exception:
                        thumbnail_url = None  
            media_items.append({
                "id": reel.id,
                "title": reel.title,
                "type": "video",
                "video": request.build_absolute_uri(reel.video.url) if reel.video else None,
                "thumbnail": thumbnail_url,
                "m3u8_url": reel.m3u8_url if reel.m3u8_url else None,
                "is_m3u8_generated": bool(reel.m3u8_url),
                "isActive": reel.isActive,
                "category": {
                    "id": reel.category.id if reel.category else None,
                    "name": reel.category.name if reel.category else None
                } if reel.category else None,
                "subcategory": {
                    "id": reel.subcategory.id if reel.subcategory else None,
                    "name": reel.subcategory.name if reel.subcategory else None
                } if reel.subcategory else None,
                "comments": reel.comments,
                "shares": reel.shares,
                "views": reel.views,
                "likes": reel.likes,  
                "created_at": reel.created_at if reel.created_at else current_time
            })

        for image in images:
            media_items.append({
                "id": image.id,
                "title": image.title,
                "type": "image",
                "image": image.image.url if image.image else None,
                "isActive": image.isActive,
                "category": {
                    "id": image.category.id if image.category else None,
                    "name": image.category.name if image.category else None
                } if image.category else None,
                "subcategory": {
                    "id": image.subcategory.id if image.subcategory else None,
                    "name": image.subcategory.name if image.subcategory else None
                } if image.subcategory else None,
                "created_at": image.created_at if image.created_at else current_time
            })
        media_items.sort(key=lambda x: x["created_at"], reverse=True)


        return Response({
            "statusCode": 200,
            "status": True,
            "message": "Store media retrieved successfully!",
            "total_media": len(media_items),
            "total_videos": len(reels),
            "total_images": len(images),
            "media": media_items
        }, status=status.HTTP_200_OK)
class StoreImageDeleteAPIView(APIView):
    permission_classes = [IsAuthenticated]
    @swagger_auto_schema(
        operation_summary="Delete Store Image",
        operation_description="Deletes a specific store image. Requires authentication.",
        security=[{"Bearer": []}],  
        manual_parameters=[
            openapi.Parameter(
                "pk",
                openapi.IN_PATH,
                description="ID of the Image to delete",
                type=openapi.TYPE_INTEGER,
                required=True
            )
        ],
        responses={
            204: openapi.Response(
                description="Image deleted successfully",
                examples={
                    "application/json": {
                        "message": "Image deleted successfully!"
                    }
                }
            ),
            400: openapi.Response(
                description="Bad Request",
                examples={
                    "application/json": {
                        "message": "Invalid request data"
                    }
                }
            ),
            401: openapi.Response(
                description="Unauthorized - Bearer Token Required",
                examples={
                    "application/json": {
                        "detail": "Authentication credentials were not provided."
                    }
                }
            ),
            404: openapi.Response(
                description="Image Not Found",
                examples={
                    "application/json": {
                        "message": "Image not found"
                    }
                }
            ),
            500: openapi.Response(
                description="Server Error",
                examples={
                    "application/json": {
                        "message": "An unexpected error occurred",
                        "details": "Error message here"
                    }
                }
            ),
        },
    )

    def delete(self, request, *args, **kwargs):
        try:
           
            image_id = kwargs.get('pk')
            if not image_id:
                return Response(
                    {"statusCode": 400, "status": False, "message": "Image ID is required."},
                    status=status.HTTP_200_OK
                )
            try:
                instance = StoreImage.objects.get(pk=image_id)
            except StoreImage.DoesNotExist:
                return Response({"statusCode": 400, "status":False,
                    "message": "Image not found"
                }, status=status.HTTP_200_OK)

            instance.delete()

            return Response({"statusCode": 200, "status":True,"message": "Image deleted successfully!"}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"statusCode": 400, "status":False,"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
      

class StoreImageUpdateTitleAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, *args, **kwargs):
        try:
            image_id = request.data.get("image_id")
            new_title = request.data.get("title")
            new_category_id = request.data.get("categories")  # Ensure correct key
            new_subcategory_id = request.data.get("subcategories")
            isActive = request.data.get("isActive")
            if not image_id:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Image ID is required."
                }, status=status.HTTP_200_OK)
                
            if new_category_id is None:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Please select a category."
                }, status=status.HTTP_200_OK)
            
         
            image = get_object_or_404(StoreImage, id=image_id)
            user = request.user  
            professional_user = ProfessionalUser.objects.filter(email=user.email).first()

            if not professional_user:
                return Response({
                    "statusCode": 403,
                    "status": False,
                    "message": "Only professional users can update store images."
                }, status=status.HTTP_200_OK)

          
            
            user_category_ids = professional_user.categories.values_list('id', flat=True)
            user_subcategory_ids = professional_user.subcategories.values_list('id', flat=True)
            category = Category.objects.filter(id=new_category_id).first()
            if not category or category.id not in user_category_ids:
                return Response({
                    "statusCode": 403,
                    "status": False,
                    "message": "You can only update store images to your registered category."
                }, status=status.HTTP_200_OK)
            subcategory = None
            if new_subcategory_id:  # ✅ Subcategory is optional now
                subcategory = Subcategory.objects.filter(id=new_subcategory_id).first()
                if not subcategory or subcategory.id not in user_subcategory_ids:
                    return Response({
                        "statusCode": 403,
                        "status": False,
                        "message": "You can only update store images to your registered subcategory."
                    }, status=status.HTTP_200_OK)
           
            image.category = category  
            image.subcategory = subcategory
            if new_title:
                image.title = new_title
            
            if isActive is not None:
                image.isActive = bool(isActive)
            
            image.save()
            image.refresh_from_db()  
            
            response_data = {
                "id": image.id,
                "title": image.title,
                "type": "image",
                "isActive": image.isActive,  
                "image": image.image.url if image.image else None,
                "created_at": image.created_at,
                "category": {
                    "id": image.category.id if image.category else None,
                    "name": image.category.name if image.category else None
                } if image.category else None,
                "subcategory": {
                    "id": image.subcategory.id if image.subcategory else None,
                    "name": image.subcategory.name if image.subcategory else None
                } if image.subcategory else None
            }

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Store Image updated successfully.",
                "data": response_data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred while updating the store image.",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




class AddEventAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
           

            company_id = request.data.get('company_id')
            if not company_id:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Company ID is required."
                }, status=status.HTTP_200_OK)

            try:
                company = CompanyDetails.objects.get(id=int(company_id))
         
            except CompanyDetails.DoesNotExist:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Invalid company ID."
                }, status=status.HTTP_200_OK)
            event_address_str = request.data.get("eventAddress")
            event_address = None
            if event_address_str:
                try:
                    event_address = json.loads(event_address_str)
                except json.JSONDecodeError:
                    return Response({
                        "statusCode": 400,
                        "status": False,
                        "message": "Invalid JSON format for eventAddress."
                    }, status=status.HTTP_200_OK)

            event_data = {
                "company": company.id,
                "eventTitle": request.data.get("eventTitle"),
                "eventImage": request.FILES.get("eventImage"),
                "startDate": request.data.get("startDate"),
                "endDate": request.data.get("endDate"),
                "startTime": request.data.get("startTime"),
                "endTime": request.data.get("endTime"),
                "description": request.data.get("description"),
                "eventAddress": event_address,  #  Pass dictionary instead of string
            }
          

            serializer = EventSerializer(data=event_data, context={'request': request})
            if serializer.is_valid():
                serializer.save()
                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Event created successfully!",
                    "event": serializer.data
                }, status=status.HTTP_200_OK)

            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Invalid data",
                "details": serializer.errors
            }, status=status.HTTP_200_OK)

        except IntegrityError as e:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Database integrity error. Please check your input values.",
                "details": str(e)
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Something went wrong",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class EventPagination(PageNumberPagination):
    page_size_query_param = 'page_size'
class GetEventsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, event_id=None):
        try:
            user = request.user  # Get the logged-in user
            professional_user = ProfessionalUser.objects.filter(email=user.email).first()

            if not professional_user:
                return Response({
                    "statusCode": 403,
                    "status": False,
                    "message": "Only professional users can access events."
                }, status=status.HTTP_200_OK)

            company = professional_user.company  # Get the company of the logged-in user
            if not company:
                return Response({
                    "statusCode": 404,
                    "status": False,
                    "message": "No company associated with this user."
                }, status=status.HTTP_200_OK)

            current_time = now()  # Get current datetime
            search_query = request.query_params.get('search', '')
            events_qs = StoreEvent.objects.filter(company=company).annotate(
                startDatetime=ExpressionWrapper(
                    Func(F("startDate"), F("startTime"), function="TIMESTAMP"),
                    output_field=DateTimeField()
                )
            )

            if event_id:
                event = events_qs.filter(id=event_id).first()
                if not event:
                    return Response({
                        "statusCode": 404,
                        "status": False,
                        "message": "Event not found for this company."
                    }, status=status.HTTP_200_OK)

                serializer = EventSerializer(event)
                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Successfully fetched the Event Details",
                    "event": serializer.data
                }, status=status.HTTP_200_OK)
            
            if search_query:
                events_qs = events_qs.filter(
                    Q(eventTitle__icontains=search_query) |
                    Q(description__icontains=search_query)
                )
            events = events_qs.order_by("-createdAt")

            if not events:
                return Response({
                    "statusCode": 200,
                    "status": False,
                    "message": "No events found for this company."
                }, status=status.HTTP_200_OK)
            events_data = [
                {
                    "id": event.id,
                    "company": event.company.id if event.company else None,
                    "eventTitle": event.eventTitle,
                    "eventImage": event.eventImage.url if event.eventImage else None,
                    "startDate": event.startDate.strftime('%Y-%m-%d') if event.startDate else None,
                    "endDate": event.endDate.strftime('%Y-%m-%d') if event.endDate else None,
                    "startTime": event.startTime.strftime('%H:%M:%S') if event.startTime else None,
                    "endTime": event.endTime.strftime('%H:%M:%S') if event.endTime else None,
                    "description": event.description,
                    "eventAddress": {
            "address1": event.eventAddress.address1 if event.eventAddress else None,
            "address2": event.eventAddress.address2 if event.eventAddress else None,
            "postalCode": event.eventAddress.postalCode if event.eventAddress else None,
            "lat": event.eventAddress.lat if event.eventAddress else None,
            "lang": event.eventAddress.lang if event.eventAddress else None,
            "city": event.eventAddress.city if event.eventAddress else None,
            "country": event.eventAddress.country if event.eventAddress else None
        } if event.eventAddress else None,

                    "created_at": event.createdAt
                }
                for event in events
            ]

            paginator = EventPagination()
            paginated_events = paginator.paginate_queryset(events_data, request)

            return paginator.get_paginated_response({
                "statusCode": 200,
                "status": True,
                "message": "Successfully fetched the Event Details",
                "total_count": len(events_data),
                "events": paginated_events
            })


        except ValidationError as ve:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Invalid input",
                "details": ve.message_dict
            }, status=status.HTTP_200_OK)

        except StoreEvent.DoesNotExist:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Event not found",
                "details": f"No event found with ID {event_id}."
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)





class EventsUpdateAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request):
        try:
          
            event_id = request.data.get("event_id")
       

            if not event_id:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Event ID is required"
                }, status=status.HTTP_200_OK)

            try:
                event_id = int(event_id)  # Convert to integer
            except ValueError:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Invalid event ID format"
                }, status=status.HTTP_200_OK)

            event = get_object_or_404(StoreEvent, id=event_id)
            event.eventTitle = request.data.get("eventTitle", event.eventTitle)
            event.startDate = request.data.get("startDate", event.startDate)
            event.endDate = request.data.get("endDate", event.endDate)
            event.startTime = request.data.get("startTime", event.startTime)
            event.endTime = request.data.get("endTime", event.endTime)
            event.description = request.data.get("description", event.description)

            if "eventImage" in request.FILES:
                event.eventImage = request.FILES["eventImage"]
            event_address = request.data.get("eventAddress")  
            if event_address:
                try:
                    event_address = json.loads(event_address)  #  Convert JSON string to dictionary
                    event.eventAddress.address1 = event_address.get("address1", event.eventAddress.address1)
                    event.eventAddress.address2 = event_address.get("address2", event.eventAddress.address2)
                    event.eventAddress.postalCode = event_address.get("postalCode", event.eventAddress.postalCode)
                    event.eventAddress.lat = event_address.get("lat", event.eventAddress.lat)
                    event.eventAddress.lang = event_address.get("lang", event.eventAddress.lang)
                    event.eventAddress.city = event_address.get("city", event.eventAddress.city)
                    event.eventAddress.country = event_address.get("country", event.eventAddress.country)
                    event.eventAddress.save()
                except json.JSONDecodeError:
                    return Response({
                        "statusCode": 400,
                        "status": False,
                        "message": "Invalid JSON format for eventAddress"
                    }, status=status.HTTP_200_OK)

            event.save()
            serializer = EventSerializer(event)

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Event updated successfully",
                "event": serializer.data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class DeleteEventAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, event_id=None):
        try:
            if not event_id:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Event ID is required"
                }, status=status.HTTP_200_OK)
            event = get_object_or_404(StoreEvent, id=event_id)
            event.delete()

            return Response({
                "statusCode": 200,
                "status": True,
                "message": f"Event with ID {event_id} deleted successfully"
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class VerifyEmailView(APIView):
    @swagger_auto_schema(
        operation_summary="Verify Email Availability",
        operation_description="Checks if the provided email is available for registration.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["email"],
            properties={
                "email": openapi.Schema(type=openapi.TYPE_STRING, format="email", description="Email to check"),
            },
        ),
        responses={
            200: openapi.Response(
                description="Email is available",
                examples={"application/json": {"message": "Email is available", "statusCode": 200, "status": True}},
            ),
            400: openapi.Response(
                description="Email already exists or missing email",
                examples={"application/json": {"message": "Email address already exists", "statusCode": 400, "status": False}},
            ),
            500: openapi.Response(
                description="Internal Server Error",
                examples={"application/json": {"message": "Internal Server Error", "statusCode": 500, "status": False}},
            ),
        },
    )
    def post(self, request):
        try:
            email = request.data.get("email")

            if not email:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Email is required.",
                })

            if ProfessionalUser.objects.filter(email=email).exists():
                return Response({
                "statusCode": 400,
                "status": False,
                "message": "Email address already exists",
                },status=status.HTTP_200_OK,)

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Email is available",
            },status=status.HTTP_200_OK,)

        except Exception as e:
            return Response({
                   "statusCode": 500,
                "status": False,
                "message": "Internal Server Error",
            })
class SendOTPView(APIView):
    @swagger_auto_schema(
        operation_summary="Send OTP to Email or Phone",
        operation_description="Sends a one-time password (OTP) to the provided email or phone number for verification.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                "email": openapi.Schema(type=openapi.TYPE_STRING, format="email", description="User's email address (optional if phone is provided)"),
                "phone": openapi.Schema(type=openapi.TYPE_STRING, description="User's phone number (optional if email is provided)"),
            },
            required=[],
        ),
        responses={
            200: openapi.Response(
                
                description="OTP sent successfully",
                examples={"application/json": {"message": "OTP sent to email", "status": "success", "otp": "123456"}},
            ),
            400: openapi.Response(
                description="Bad Request - Missing required fields",
                examples={"application/json": {"statusCode": 400, "status": "message", "message": "Email or phone is required."}},
            ),
            404: openapi.Response(
                description="User Not Found",
                examples={"application/json": {"statusCode": 404, "status": "message", "message": "User not found."}},
            ),
            500: openapi.Response(
                description="Internal Server Error",
                examples={"application/json": {"statusCode": 500, "status": "message", "message": "An error occurred."}},
            ),
        },
    )
    def post(self, request):
        try:
            email = request.data.get("email")
            phone = request.data.get("phone")

            if not email and not phone:
                return Response({"statusCode": 400, "status": "message", "message": "Email or phone is required."})

            response_data = {}

  
            if email:
                user = ProfessionalUser.objects.filter(email=email).first()
                if not user:
                    return Response({"statusCode": 404, "status": "message", "message": "User not found."},status=status.HTTP_200_OK,)

                otp = generate_otp()
                user.set_otp(otp) 
                cache.set(f"otp_{email}", otp, timeout=300)

                email_sent = send_email(email=email, data={"otp": otp})  
                if email_sent:
                    return Response({"statusCode": 200,"status": True,"message": "OTP sent to email", "status": "success", "otp": otp})
                else:
                    return Response({   "statusCode": 400,"status": False ,"message": "Failed to send OTP via email", "status": "message"})

          
            if phone:
                user = ProfessionalUser.objects.filter(phone=phone).first() 
                if not user:
                    return Response({"statusCode": 404, "status": False, "message": "User not found." },status=status.HTTP_200_OK,)

                otp = generate_otp()
                user.set_otp(otp)
                cache.set(f"otp_{phone}", otp, timeout=300)

           
                sms_sent, sms_response = send_sms(phone, otp)  
                if sms_sent:
                     return Response({"statusCode": 200,"status": True,"message": "OTP sent to phone", "status": "success", "otp": otp},status=status.HTTP_200_OK,)
                else:
                     return Response({"statusCode": 400,"status": False ,"message": "Failed to send OTP via SMS", "status": "message"})  

            return Response({"statusCode": 200, "status": True, "data": response_data},status=status.HTTP_200_OK,)

        except Exception as e:
            return Response({"statusCode": 500, "status":False, "message": str(e)})
class VerifyOTPView(APIView):
    @swagger_auto_schema(
        operation_summary="Verify OTP for Email or Phone",
        operation_description="Verifies the one-time password (OTP) sent to the provided email or phone.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                "email": openapi.Schema(type=openapi.TYPE_STRING, format="email", description="User's email address (optional if phone is provided)"),
                "phone": openapi.Schema(type=openapi.TYPE_STRING, description="User's phone number (optional if email is provided)"),
                "otp": openapi.Schema(type=openapi.TYPE_STRING, description="The OTP received via email or SMS"),
            },
            required=["otp"],
        ),
        responses={
            200: openapi.Response(
                description="User verified successfully",
                examples={"application/json": {"status": "success", "message": "User verified successfully"}},
            ),
            400: openapi.Response(
                description="Bad Request - Missing required fields",
                examples={"application/json": {"status": "message", "message": "Email/Phone and OTP are required."}},
            ),
            404: openapi.Response(
                description="User Not Found",
                examples={"application/json": {"status": "message", "message": "User not found."}},
            ),
            401: openapi.Response(
                description="Invalid OTP",
                examples={"application/json": {"status": "message", "message": "Invalid OTP"}},
            ),
        },
    )
    def post(self, request):
        email = request.data.get('email')
        phone = request.data.get('phone')
        otp = request.data.get('otp')

        if not otp or (not email and not phone):
            return Response({
                        "statusCode": 400,
                "status": False,
                "message": "Email/Phone and OTP are required."
            }, status=200)

        try:
          
            if email:
                user = ProfessionalUser.objects.get(email=email)
            elif phone:
                user = ProfessionalUser.objects.get(phone=phone)
            else:
                return Response({
                            "statusCode": 400,
                "status": False,
                    "message": "Invalid email or phone format."
                })
        except ProfessionalUser.DoesNotExist:
            return Response({
                        "statusCode": 400,
                "status": False,
                "message": "User not found."
            },status=status.HTTP_200_OK,)

        if user.verify_otp(otp): 
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "User verified successfully"
            },status=status.HTTP_200_OK,)
        else:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Invalid OTP"
            })
class ResendOTPView(APIView):
    @swagger_auto_schema(
        operation_summary="Resend OTP via Email or Phone",
        operation_description="Allows users to request a new OTP to be sent to their registered email or phone.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                "email": openapi.Schema(type=openapi.TYPE_STRING, format="email", description="User's email address (optional if phone is provided)"),
                "phone": openapi.Schema(type=openapi.TYPE_STRING, description="User's phone number (optional if email is provided)"),
            },
            required=[],
        ),
        responses={
            200: openapi.Response(
                description="OTP resent successfully",
                examples={"application/json": {"message": "OTP resent successfully", "statusCode": 200, "status": True, "otp": "123456"}},
            ),
            400: openapi.Response(
                description="Bad Request - Missing required fields",
                examples={"application/json": {"status": False, "message": "Email or phone is required."}},
            ),
            404: openapi.Response(
                description="User Not Found",
                examples={"application/json": {"status": False, "message": "User not found."}},
            ),
            429: openapi.Response(
                description="Too Many Requests - OTP Generation Blocked",
                examples={"application/json": {"statusCode": 429, "status": False, "message": "OTP generation blocked. Try again in 5 minutes."}},
            ),
            500: openapi.Response(
                description="Internal Server Error - Failed to send OTP",
                examples={"application/json": {"statusCode": 500, "status": False, "message": "Failed to resend OTP."}},
            ),
        },
    )

    def post(self, request):
        try:
            email = request.data.get("email")
            phone = request.data.get("phone")

            if not email and not phone:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Email or phone is required."
                }, status=status.HTTP_200_OK)

            user = None
            if email:
                user = ProfessionalUser.objects.filter(email=email).first()
            elif phone:
                user = ProfessionalUser.objects.filter(phone=phone).first()

            if not user:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "User not found."
                }, status=status.HTTP_200_OK)

           

            otp = generate_otp()
            user.set_otp(otp)

            cache_key = f"otp_{email or phone}"
            cache.set(cache_key, otp, timeout=300)  

            otp_sent = True  # Simulate OTP sending logic

            if otp_sent:
                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "OTP resent successfully",
                    "otp": otp  
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    "statusCode": 500,
                    "status": False,
                    "message": "Failed to resend OTP."
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred.",
                "message": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class UpdateProfessionalUserView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request):
        user = request.user  
        if 'email' in request.data:
            return Response({
                "message": "Invalid request data",
                "statusCode": 400,
                "status": False,
                "error": "You cannot update the email."
            }, status=status.HTTP_200_OK)

        if 'userName' in request.data:
            return Response({
                "message": "Invalid request data",
                "statusCode": 400,
                "status": False,
                "error": "You cannot update the username."
            }, status=status.HTTP_200_OK)
        def parse_csv_string(value):
            try:
                return [int(x.strip()) for x in value.split(",") if x.strip().isdigit()] if isinstance(value, str) else []
            except ValueError:
                return []
        category_ids = parse_csv_string(request.data.get("categories", ""))
        subcategory_ids = parse_csv_string(request.data.get("subcategories", ""))

        serializer = UpdateProfessionalUserSerializer(user, data=request.data, partial=True, context={"request": request})
        if serializer.is_valid():
            with transaction.atomic():  #  Ensure Atomic Updates
                user = serializer.save()
                if category_ids:
                    user.categories.set(Category.objects.filter(id__in=category_ids))
                if subcategory_ids:
                    user.subcategories.set(Subcategory.objects.filter(id__in=subcategory_ids))

                user.save()
                
                try:
                    company = user.company  
                    if category_ids:
                        company.categories.set(Category.objects.filter(id__in=category_ids))
                    if subcategory_ids:
                        company.subcategories.set(Subcategory.objects.filter(id__in=subcategory_ids))
                    company.save()
                except CompanyDetails.DoesNotExist:
                    pass  # company not created yet
            updated_categories = list(user.categories.values("id", "name", "slug"))
            updated_subcategories = list(user.subcategories.values("id", "name", "slug"))
            subscription_status = getattr(user, 'subscription_status', '').lower()
            is_free_trial_active = getattr(user, 'is_free_trial_active', False)
            is_paid_subscription_active = getattr(user, 'is_paid_subscription_active', False)
            is_subscription_active = getattr(user, 'subscription_active', False)
            customerId = getattr(user, 'stripe_customer_id', "")

            if subscription_status == "trial" and not is_free_trial_active:
                subscription_message = "Your free trial has expired"
            elif subscription_status == "trial" and is_free_trial_active:
                subscription_message = "Your 3-month free trial is still active"
            elif subscription_status == "paid" and is_paid_subscription_active:
                subscription_message = "Your paid subscription is still active"
            elif subscription_status == "paid" and not is_paid_subscription_active:
                subscription_message = "Your paid subscription has expired. Please renew or buy a new one."
            else:
                subscription_message = "Your subscription is not active. Choose a subscription plan."
 
            user_data = serializer.data

            user_data["customerId"] = customerId
            user_data["subscription_status"] = subscription_status
            user_data["is_subscription_active"] = is_subscription_active
            user_data["subscription_message"] = subscription_message
            user_data["is_free_trial_active"] = is_free_trial_active
            user_data["is_paid_subscription_active"] = is_paid_subscription_active

            return Response({
                "message": "User updated successfully",
                "statusCode": 200,
                "status": True,
                "user": user_data,  # updated serializer data
            }, status=status.HTTP_200_OK)


        if serializer.errors:
            error_messages = next(iter(serializer.errors.values()))
            return Response({
                "message": "Invalid request data",
                "statusCode": 400,
                "status": False,
                "error": error_messages[0] if isinstance(error_messages, list) else error_messages
            }, status=status.HTTP_200_OK)

        return Response({
            "message": "Something went wrong",
            "statusCode": 500,
            "status": False
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

class ForgotPasswordAPIView(APIView):
    @swagger_auto_schema(
        operation_summary="Forgot Password",
        operation_description="Allows users to request an OTP for password reset via email or phone.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                "email": openapi.Schema(type=openapi.TYPE_STRING, format="email", description="User's registered email (optional if phone is provided)"),
                "phone": openapi.Schema(type=openapi.TYPE_STRING, description="User's registered phone number (optional if email is provided)"),
            },
            required=[],
        ),
        responses={
            200: openapi.Response(
                description="OTP sent successfully",
                examples={
                    "application/json": {
                        "status": True,
                        "statusCode": 200,
                        "message": "OTP sent to email",
                        "otp": "123456"
                    }
                }
            ),
            400: openapi.Response(
                description="Invalid request",
                examples={
                    "application/json": {
                        "status": False,
                        "statusCode": 400,
                        "message": "Invalid request",
                        "errors": {"email": ["This field is required"]}
                    }
                }
            ),
        },
    )

    def post(self, request):
        try:
            serializer = ForgotPasswordSerializer(data=request.data)
            
            if serializer.is_valid():
                otp = serializer.save()
                user = serializer.validated_data.get("user")

                otp_sent_to = "email" if user.email else "phone"
                message = f"OTP sent to {otp_sent_to}"

                return Response({
                    "status": True,
                    "statusCode": 200,
                    "message": message,
                    "otp": otp  
                }, status=status.HTTP_200_OK)

            return Response({
                "status": False,
                "statusCode": 400,
                "message": "User not found",
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "status": False,
                "statusCode": 500,
                "message": "An error occurred while processing your request.",
                "message": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ResetPasswordAPIView(APIView):
    @swagger_auto_schema(
        operation_summary="Reset Password",
        operation_description="Allows users to reset their password and receive new access and refresh tokens.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                "email": openapi.Schema(type=openapi.TYPE_STRING, format="email", description="Registered email address"),
                "otp": openapi.Schema(type=openapi.TYPE_STRING, description="OTP received via email or SMS"),
                "newPassword": openapi.Schema(type=openapi.TYPE_STRING, description="New password to be set"),
                "confirm_password": openapi.Schema(type=openapi.TYPE_STRING, description="Confirm new password"),
            },
            required=["email", "otp", "newPassword", "confirm_password"],
        ),
        responses={
            200: openapi.Response(
                description="Password reset successful",
                examples={
                    "application/json": {
                        "message": "Password reset successful",
                        "access_token": "eyJhbGciOiJIUzI1NiIsInR...",
                        "refresh_token": "eyJhbGciOiJIUzI1NiIsInR..."
                    }
                }
            ),
            400: openapi.Response(
                description="Invalid request",
                examples={
                    "application/json": {
                        "message": "Invalid request",
                        "errors": {"otp": ["OTP is incorrect or expired"]}
                    }
                }
            ),
            404: openapi.Response(
                description="User not found",
                examples={"application/json": {"message": "User with this email not found"}},
            ),
        },
    )
    def post(self, request):
        try:
            serializer = ResetPasswordSerializer(data=request.data)

            if serializer.is_valid():
                tokens = serializer.save()
                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Password reset successful",
                    "access_token": tokens["access_token"],
                    "refresh_token": tokens["refresh_token"]
                }, status=status.HTTP_200_OK)

            return Response({
                "statusCode": 400,
                "status": False,
                "messasge": "Invalid or expired OTP."
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class UpdatePasswordAPIView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Update Password",
        operation_description="Allows authenticated users to update their password.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                "old_password": openapi.Schema(type=openapi.TYPE_STRING, description="Current password of the user"),
                "new_password": openapi.Schema(type=openapi.TYPE_STRING, description="New password to be set"),
                "confirm_password": openapi.Schema(type=openapi.TYPE_STRING, description="Confirm new password"),
            },
            required=["old_password", "new_password", "confirm_password"],
        ),
        responses={
            200: openapi.Response(
                description="Password updated successfully",
                examples={"application/json": {"message": "Password updated successfully"}},
            ),
            400: openapi.Response(
                description="Bad Request - Password update failed",
                examples={"application/json": {"message": "Password update failed", "errors": {"new_password": ["New password must be at least 8 characters."]}}},
            ),
            401: openapi.Response(
                description="Unauthorized - Authentication required",
                examples={"application/json": {"detail": "Authentication credentials were not provided."}},
            ),
        },
        security=[{"Bearer": []}],
    )

    def put(self, request):
        try:
            serializer = UpdatePasswordSerializer(data=request.data, context={"request": request})

            if serializer.is_valid():
                serializer.save()
                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Password updated successfully"
                }, status=status.HTTP_200_OK)
            error_messages = next(iter(serializer.errors.values()))
            if isinstance(error_messages, list) and error_messages:
                error_message = error_messages[0]
            else:
                error_message = str(error_messages)

            return Response({
                "statusCode": 400,
                "status": False,
                "message": f" {error_message}"
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




class LogoutAPIView(APIView):
    permission_classes = [IsAuthenticated] 
    @swagger_auto_schema(
        operation_summary="Logout User",
        operation_description="Logs out an authenticated user by blacklisting the provided refresh token.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                "refresh_token": openapi.Schema(
                    type=openapi.TYPE_STRING, description="User's refresh token"
                )
            },
            required=["refresh_token"],
        ),
        responses={
            200: openapi.Response(
                description="Logout successful",
                examples={
                    "application/json": {
                        "message": "Logout successful"
                    }
                }
            ),
            400: openapi.Response(
                description="Bad request - Missing or Invalid Refresh Token",
                examples={
                    "application/json": {
                        "message": "Invalid token or token already expired"
                    }
                }
            ),
            401: openapi.Response(
                description="Unauthorized - Bearer Token Required",
                examples={
                    "application/json": {
                        "detail": "Authentication credentials were not provided."
                    }
                }
            )
        },
    ) 

    def post(self, request):
        try:
            refresh_token = request.data.get("refresh_token")
            if not refresh_token:
                return Response({"message": "Refresh token is required"}, status=status.HTTP_200_OK)
            
            token = RefreshToken(refresh_token)
            token.blacklist() 
            return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Logout successful"
                }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                    "statusCode": 400,
                    "status":False,
                    "message":  "Invalid token or token already expired"
                },status=status.HTTP_200_OK)
        



class CreateProductAPIView(APIView):
    
    permission_classes = [IsAuthenticated] 
    parser_classes = [MultiPartParser, FormParser]  
        
    def post(self, request):
        
        
        """Create a new product with image uploads"""
        try:
        
            professional_user = ProfessionalUser.objects.get(email=request.user.email)  
            company = professional_user.company
            
            if not company:
                return Response(
                    {"statusCode": 403, "status": False, "message": "User is not linked to any company"},
                    status=status.HTTP_200_OK,
                )
                
        except ProfessionalUser.DoesNotExist:
            return Response(
                {"statusCode": 400, "status": False, "message": "User not found"},
            status=status.HTTP_200_OK,
            )
            
        allowed_categories = professional_user.categories.values_list("id", flat=True)
        allowed_subcategories = professional_user.subcategories.values_list("id", flat=True)
        
        category_id = request.data.get("categoryId")
        subcategory_id = request.data.get("subCategoryId")

        if isinstance(category_id, list):
            category_id_list = category_id
        else:
            category_id_list = [int(id_.strip()) for id_ in str(category_id).split(',')]

        if isinstance(subcategory_id, list):
            subcategory_id_list = subcategory_id
        else:
            subcategory_id_list = [int(id_.strip()) for id_ in str(subcategory_id).split(',')]

        if not any(cat_id in allowed_categories for cat_id in category_id_list):
            return Response(
                {"statusCode": 400, "status": False, "message": "Invalid category selection"},
                status=status.HTTP_200_OK,
            )

        if not any(subcat_id in allowed_subcategories for subcat_id in subcategory_id_list):
            return Response(
                {"statusCode": 400, "status": False, "message": "Invalid subcategory selection"},
                status=status.HTTP_200_OK,
            )

            
        gallery_images = []

        # if "galleryImage" in request.FILES:
        #     for image in request.FILES.getlist("galleryImage"):
        #         saved_path = default_storage.save(f"gallery_images/{image.name}", image)
        #         uploaded_url = default_storage.url(saved_path)
        #         gallery_images.append(uploaded_url)
        
        if "galleryImage" in request.FILES:
            for image in request.FILES.getlist("galleryImage"):
                compressed = compress_image(image)
                saved_path = default_storage.save(f"gallery_images/{image.name}", compressed)
                uploaded_url = default_storage.url(saved_path)
                gallery_images.append(uploaded_url)

        product_data = request.data.copy()
        product_data["galleryImage"] = json.dumps(gallery_images)
        product_data["petAllowed"] = request.data.get("petAllowed", False)
        data = request.data.copy()  
        pet_type_raw = data.get("petType", None)
        cruise_facility_raw = request.data.get("cruiseFacility", None)
        room_facility_raw = request.data.get("roomFacility", None)
       
       
            
        try:
            if cruise_facility_raw:
                if isinstance(cruise_facility_raw, str):
                    cruise_facility_parsed = json.loads(cruise_facility_raw)
                elif isinstance(cruise_facility_raw, list):
                    cruise_facility_parsed = cruise_facility_raw
                else:
                    cruise_facility_parsed = [int(cruise_facility_raw)]

                if not isinstance(cruise_facility_parsed, list):
                    raise ValueError("cruiseFacility must be a list")
                
                product_data.setlist("cruiseFacility", cruise_facility_parsed)
        except (json.JSONDecodeError, ValueError) as e:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": {"cruiseFacility": ["Invalid format for cruiseFacility.", str(e)]}
            }, status=status.HTTP_200_OK)
      
        try:
            room_facility_parsed = get_roomFacility(room_facility_raw)
         
        except ValueError as e:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": {"roomFacility": [str(e)]}
            }, status=status.HTTP_200_OK)

        product_data.setlist("roomFacility", room_facility_parsed)
       
        try:
            if pet_type_raw:
                if isinstance(pet_type_raw, str):
                    pet_type_parsed = json.loads(pet_type_raw)
                elif isinstance(pet_type_raw, list):
                    pet_type_parsed = pet_type_raw
                else:
                    raise ValueError("petType must be a JSON array string or a list")

                if not isinstance(pet_type_parsed, list):
                    raise ValueError("petType must be a list")

                data["petType"] = pet_type_parsed  
        except (json.JSONDecodeError, ValueError) as e:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": {"petType": ["Invalid format for petType.", str(e)]}
            }, status=status.HTTP_200_OK)
        product_data["cruiseName"] = request.data.get("cruiseName", None)
        
        start_address_data = request.data.get("startAddress")
        end_address_data = request.data.get("endAddress")

        if start_address_data:
            try:
                if isinstance(start_address_data, str):
                    start_address_data = json.loads(start_address_data)
                
                start_address_serializer = AddressSerializer(data=start_address_data)
                if start_address_serializer.is_valid():
                    start_address_instance = start_address_serializer.save()
                else:
                    return Response({
                        "statusCode": 400,
                        "status": False,
                        "message": {"startAddress": start_address_serializer.errors}
                    }, status=status.HTTP_200_OK)
            except (json.JSONDecodeError, ValueError) as e:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": {"startAddress": [str(e)]}
                }, status=status.HTTP_200_OK)
        else:
            start_address_instance = None

        if end_address_data:
            try:
                if isinstance(end_address_data, str):
                    end_address_data = json.loads(end_address_data)

                end_address_serializer = AddressSerializer(data=end_address_data)
                if end_address_serializer.is_valid():
                    end_address_instance = end_address_serializer.save()
                else:
                    return Response({
                        "statusCode": 400,
                        "status": False,
                        "message": {"endAddress": end_address_serializer.errors}
                    }, status=status.HTTP_200_OK)
            except (json.JSONDecodeError, ValueError) as e:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": {"endAddress": [str(e)]}
                }, status=status.HTTP_200_OK)
        else:
            end_address_instance = None
        product_data.update({
            "startAddress": start_address_instance.id if start_address_instance else None,
            "endAddress": end_address_instance.id if end_address_instance else None,
        })
        
        
        serializer = ProductcreateSerializer(data=product_data)
        
        if serializer.is_valid():
            product = serializer.save(company=company)

            rooms_data = []
            rooms_json = request.data.get("rooms", None)
            tickets_data_raw = request.data.get("concertTicket", None)
            night_club_raw = request.data.get("nightClubTicket", None)
            amusements_club_raw = request.data.get("amusementsTicket", None)
            

            try:
                created_tickets = []
                if tickets_data_raw:
                    created_tickets = create_tickets_concert(tickets_data_raw, product)
                

                created_tickets_nightClub = []
                if night_club_raw:
                    created_tickets_nightClub = create_tickets_nightclub(night_club_raw, product)

                created_tickets_amusements = []
                if amusements_club_raw:
                    created_tickets_amusements = create_tickets_amusements(amusements_club_raw, product)
                    
                rooms_raw = []
                if rooms_json:
                    rooms_raw = json.loads(rooms_json)

                for index, room in enumerate(rooms_raw, start=1):
                    room_id = f"ROOM-{product.id}-{index}"
                    room_type = room.get("roomType", "")
                    if not room_type:
                        continue  # or handle error

                    CruiseRoom.objects.create(
                        product=product,
                        room_id=room_id,
                        roomType=room_type,
                        roomQuantity=room.get("roomQuantity", 1),
                        roomPrice=room.get("roomPrice", 0),
                        adults=room.get("adults", 0),
                    )

                    rooms_data.append({
                        "roomType": room_type,
                        "roomQuantity": room.get("roomQuantity", 1),
                        "roomPrice": room.get("roomPrice", 0),
                        "adults": room.get("adults", 0),
                        "room_id": room_id,
                    })

            except json.JSONDecodeError as e:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": f"Invalid rooms format. Ensure it's valid JSON. :{str(e)}"
                }, status=status.HTTP_200_OK)


            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Product created successfully",
                "data": {
                    "id": product.id,
                    "company": product.company.id if product.company else None,
                    "folder": product.folder.id if product.folder else None,
                    "categoryId": product.categoryId.id if product.categoryId else None,
                    "subCategoryId": product.subCategoryId.id if product.subCategoryId else None,
                    "productname": product.productname,
                    "productType": product.productType,
                    "description": product.description,
                    "totalEmployees": product.totalEmployees,
                    "priceOnsite": str(product.priceOnsite) if product.priceOnsite else None,
                    "priceClickAndCollect": str(product.priceClickAndCollect) if product.priceClickAndCollect else None,
                    "priceDelivery": str(product.priceDelivery) if product.priceDelivery else None,
                    "quantity": product.quantity,
                    "deliveryMethod": product.deliveryMethod,
                    "deliveryPricePerGram": str(product.deliveryPricePerGram) if product.deliveryPricePerGram else None,
                    "preparationDateTime": product.preparationDateTime,
                    "availabilityDateTime": product.availabilityDateTime,
                    "onDelivery": product.onDelivery,
                    "onsite": product.onsite,
                    "clickandCollect": product.clickandCollect,
                    "serviceTime": product.serviceTime,
                    "basePrice": str(product.basePrice) if product.basePrice else None,
                    "nonRestaurant": product.nonRestaurant,
                    "delivery": str(product.delivery) if product.delivery else None,
                    "keywords": product.keywords,
                    "ProductImage": request.build_absolute_uri(product.ProductImage.url) if product.ProductImage else None,
                    "galleryImage": product.galleryImage,
                    "isActive": product.isActive,
                    "created_at": product.created_at,
                    "updated_at": product.updated_at,
                    "discount": product.discount,
                    "onhome": product.onhome,
                    "duration": product.duration,
                    "rooms": rooms_data,
                    "startAddress": {
                        "address1": product.startAddress.address1 if product.startAddress else None,
                        "address2": product.startAddress.address2 if product.startAddress else None,
                        "postalCode": product.startAddress.postalCode if product.startAddress else None,
                        "lat": product.startAddress.lat if product.startAddress else None,
                        "lang": product.startAddress.lang if product.startAddress else None,
                        "city": product.startAddress.city if product.startAddress else None,
                        "country": product.startAddress.country if product.startAddress else None,
                    },
                    "endAddress": {
                        "address1": product.endAddress.address1 if product.endAddress else None,
                        "address2": product.endAddress.address2 if product.endAddress else None,
                        "postalCode": product.endAddress.postalCode if product.endAddress else None,
                        "lat": product.endAddress.lat if product.endAddress else None,
                        "lang": product.endAddress.lang if product.endAddress else None,
                        "city": product.endAddress.city if product.endAddress else None,
                        "country": product.endAddress.country if product.endAddress else None,
                    },
                    "concertTickets": created_tickets,
                    "nightClubTickets":created_tickets_nightClub,
                    "amusementParkTickets":created_tickets_amusements,
                    "petAllowed": product.petAllowed,
                    "petType": product.petType,
                    "cruiseName": product.cruiseName,
                    "vatRate": product.vatRate,
                    "promotionalPrice": str(product.promotionalPrice) if product.promotionalPrice else None,
                    "average_rating": product.average_rating,
                    "total_ratings": product.total_ratings,
                    "folder": product.folder,
                    "created_at": product.created_at,
                    "updated_at": product.updated_at,
                    "cruiseFacility": [{"id": cf.id, "name": cf.name} for cf in product.cruiseFacility.all()],
                    "roomFacility": [{"id": rf.id, "name": rf.name} for rf in product.roomFacility.all()],
                    "smokingAllowed": product.smokingAllowed,
                    "noofMembers": product.noofMembers,
                    "view": product.roomview,
                }
            }, status=status.HTTP_200_OK)

        else:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Invalid product data",
                "errors": serializer.errors,
            }, status=status.HTTP_200_OK)


class LiveSearchThrottle(UserRateThrottle):
    rate = '15/second'  # Adjust if needed

class ProductPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100
    
    def get_paginated_response(self, data):
        return Response({
            'count': self.page.paginator.count,
            'next': self.get_next_link(),
            'previous': self.get_previous_link(),
            'statusCode': 200,
            'status': True,
            'message': 'Company products retrieved successfully',
            'data': data
        })
    
class CompanyProductListAPIView(ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ProductSerializer
    throttle_classes = [LiveSearchThrottle]
    pagination_class = ProductPagination 
    
    def get_queryset(self):
        try:
            professional_user = ProfessionalUser.objects.get(email=self.request.user)
            company = professional_user.company

            if not company:
                raise ValidationError("User is not linked to any company")

            product_type = self.request.query_params.get("productType")
            product_name = self.request.query_params.get("product_name")

            queryset = Product.objects.filter(company=company)

            if product_type:
                normalized_type = product_type.strip().capitalize()

                valid_types = ["Product", "Services", "Ticket"]
                if normalized_type not in valid_types:
                    raise ValidationError("Invalid product type. Allowed values: Product, Services, Ticket")

                queryset = queryset.filter(productType=normalized_type)


            if product_name:
                queryset = queryset.filter(productname__icontains=product_name.strip())
                
            queryset = queryset.order_by('-created_at')
            
            return queryset

        except ProfessionalUser.DoesNotExist:
            raise ValidationError("Professional user not found")
        except ObjectDoesNotExist:
            raise ValidationError("Requested product does not exist")
        except Exception as e:
            raise ValidationError(f"An unexpected error occurred: {str(e)}")

    def list(self, request, *args, **kwargs):
        try:
            queryset = self.get_queryset()
            page = self.paginate_queryset(queryset)
            products = []

            for product in page:
                product_data = ProductSerializer(product).data

                if product.categoryId:
                    product_data['category'] = {
                        "id": product.categoryId.id,
                        "name": product.categoryId.name,
                        "slug": product.categoryId.slug,
                    }
                    
                if product.subCategoryId:
                    product_data['subCategory'] = {
                        "id": product.subCategoryId.id,
                        "name": product.subCategoryId.name,
                        "slug": product.subCategoryId.slug,
                    }
                    
                if product.startAddress:
                    product_data['startAddress'] = {
                        "address1": product.startAddress.address1,
                        "address2": product.startAddress.address2,
                        "postalCode": product.startAddress.postalCode,
                        "lat": product.startAddress.lat,
                        "lang": product.startAddress.lang,
                        "city": product.startAddress.city,
                        "country": product.startAddress.country,
                    }
                
                if product.endAddress:
                    product_data['endAddress'] = {
                        "address1": product.endAddress.address1,
                        "address2": product.endAddress.address2,
                        "postalCode": product.endAddress.postalCode,
                        "lat": product.endAddress.lat,
                        "lang": product.endAddress.lang,
                        "city": product.endAddress.city,
                        "country": product.endAddress.country,
                    }
                    
                rooms = CruiseRoom.objects.filter(product=product)
                if rooms.exists():
                    rooms_data = CruiseRoomSerializer(rooms, many=True).data
                    product_data['rooms'] = rooms_data
                    product_data['type'] = 'cruise'
                
                concertTickets = TicketsConcert.objects.filter(product=product)
                if concertTickets.exists():
                    concertTickets_data = TicketsConcertSerializer(concertTickets, many=True).data
                    product_data['concertTickets'] = concertTickets_data
                    product_data['type'] = 'concert'
                
                nightclub = NightClubTicket.objects.filter(product=product)
                if nightclub.exists():
                    nightclub_data = NightClubTicketSerializer(nightclub, many=True).data
                    product_data['nightclub'] = nightclub_data
                    product_data['type'] = 'nightclub'
                                
                amusement_tickets = TicketsAmusementPark.objects.filter(product=product)
                if amusement_tickets.exists():
                    amusement_data = TicketsAmusementParkSerializer(amusement_tickets, many=True).data
                    product_data['amusements'] = amusement_data
                    product_data['type'] = 'amusement'
              
                products.append(product_data)
            return self.get_paginated_response(products)

        except ValidationError as ve:
            return Response(
                {"statusCode": 400, "status": False, "message": str(ve)},
                status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"statusCode": 500, "status": False, "message": f"Server error: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR)





class CompanyProductDetailAPIViewss(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, id, *args, **kwargs):
        try:
            professional_user = ProfessionalUser.objects.get(email=request.user)
            company = professional_user.company

            if not company:
                raise ValidationError("User is not linked to any company")
            product = Product.objects.get(id=id, company=company)
            product_data = ProductSerializer(product).data
            if product.startAddress:
                product_data['startAddress'] = {
                    "address1": product.startAddress.address1,
                    "address2": product.startAddress.address2,
                    "postalCode": product.startAddress.postalCode,
                    "lat": product.startAddress.lat,
                    "lang": product.startAddress.lang,
                    "city": product.startAddress.city,
                    "country": product.startAddress.country,
                }

            if product.endAddress:
                product_data['endAddress'] = {
                    "address1": product.endAddress.address1,
                    "address2": product.endAddress.address2,
                    "postalCode": product.endAddress.postalCode,
                    "lat": product.endAddress.lat,
                    "lang": product.endAddress.lang,
                    "city": product.endAddress.city,
                    "country": product.endAddress.country,
                }
            rooms = CruiseRoom.objects.filter(product=product)
            rooms_data = CruiseRoomSerializer(rooms, many=True).data
            product_data['rooms'] = rooms_data

            return Response(
                {
                    "statusCode": 200,
                    "status": True,
                    "message": "Product retrieved successfully",
                    "data": product_data
                },
                status=200,
            )

        except ProfessionalUser.DoesNotExist:
            return Response(
                {"statusCode": 400, "status": False, "message": "Professional user not found"},
                status=200,
            )
        except Product.DoesNotExist:
            return Response(
                {"statusCode": 400, "status": False, "message": "Product not found"},
                status=200,
            )
        except ObjectDoesNotExist:
            return Response(
                {"statusCode": 400, "status": False, "message": "Product does not exist"},
                status=200,
            )
        except Exception as e:
            return Response(
                {"statusCode": 500, "status": False, "message": f"Server error: {str(e)}"},
                status=500,
            )
class DeleteTicketAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, *args, **kwargs):
        ticket_type = request.data.get("ticket_type") 
        ticket_id = request.data.get("ticket_id")

        if not ticket_type or not ticket_id:
            return Response({
                "status": False,
                "message": "ticket_type and ticket_id are required."
            }, status=status.HTTP_200_OK)

        try:
            if ticket_type == "cruise":
                obj = CruiseRoom.objects.get(room_id=ticket_id)
            elif ticket_type == "concert":
                obj = TicketsConcert.objects.get(id=ticket_id)
            elif ticket_type == "nightclub":
                obj = NightClubTicket.objects.get(id=ticket_id)
            elif ticket_type == "amusement":
                obj = TicketsAmusementPark.objects.get(id=ticket_id)
            else:
                return Response({
                    "status": False,
                    "statusCode": 400,
                    "message": "Invalid ticket_type. Must be one of: cruise, concert, nightclub, amusement."
                }, status=status.HTTP_200_OK)

            obj.delete()

            return Response({
                "status": True,
                "statusCode": 200,
                "message": f"{ticket_type.capitalize()} ticket with ID {ticket_id} deleted successfully."
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "status": False,
                "statusCode": 400,
                "message": str(e)
            }, status=status.HTTP_200_OK)

class GetWithoutFolderProductListAPIView(ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ProductSerializer

    def get_queryset(self):
        try:
            professional_user = ProfessionalUser.objects.get(email=self.request.user)
            company = professional_user.company

            if not company:
                raise ValidationError("User is not linked to any company")

            product_id = self.request.query_params.get("product_id")
            product_type = self.request.query_params.get("productType")
            queryset = Product.objects.filter(company=company, folder__isnull=True)

            if product_id:
                if not product_id.isdigit():
                    raise ValidationError("Invalid product ID format")
                queryset = queryset.filter(id=int(product_id))
            
            if product_type:
                queryset = queryset.filter(productType__iexact=product_type)

            return queryset

        except ProfessionalUser.DoesNotExist:
            raise ValidationError("Professional user not found")
        except ObjectDoesNotExist:
            raise ValidationError("Requested product does not exist")
        except Exception as e:
            raise ValidationError(f"An unexpected error occurred: {str(e)}")

    def list(self, request, *args, **kwargs):
        try:
            queryset = self.get_queryset()
            serializer = self.get_serializer(queryset, many=True)

            return Response(
                {
                    "statusCode": 200,
                    "status": True,
                    "message": "Company products retrieved successfully",
                    "data": serializer.data,
                },
                status=status.HTTP_200_OK,
            )

        except ValidationError as ve:
            return Response(
                {"statusCode": 400, "status": False, "message": str(ve)},
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"statusCode": 500, "status": False, "message": f"Server error: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )



class CategoryProductListAPIView(ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ProductSerializer

    def get_queryset(self):
        """Fetch products linked to the user's company, filtered by category and product type if provided."""
        try:
            professional_user = ProfessionalUser.objects.get(email=self.request.user)
            company = professional_user.company

            if not company:
                raise ValidationError("User is not linked to any company")

            category_id = self.request.query_params.get("category_id")
            product_type = self.request.query_params.get("productType")  # Get product type from params

            queryset = Product.objects.filter(company=company)

            if category_id:
                if not category_id.isdigit():
                    raise ValidationError("Invalid category ID format")
                queryset = queryset.filter(categoryId=int(category_id))

            if product_type:
                valid_types = ["Product", "Services", "Ticket"]
                if product_type not in valid_types:
                    raise ValidationError("Invalid product type. Allowed values: Product, Services, Ticket")
                queryset = queryset.filter(productType=product_type)  # Filter by product type

            return queryset

        except ProfessionalUser.DoesNotExist:
            raise ValidationError("Professional user not found")
        except ObjectDoesNotExist:
            raise ValidationError("Requested product does not exist")
        except Exception as e:
            raise ValidationError(f"An unexpected error occurred: {str(e)}")

    def list(self, request, *args, **kwargs):
        """Return products grouped by category, filtered by product type if provided."""
        try:
            queryset = self.get_queryset()
            s3_base_url = settings.AWS_S3_CUSTOM_DOMAIN if hasattr(settings, 'AWS_S3_CUSTOM_DOMAIN') else ""
            if not s3_base_url.startswith("https://"):
                s3_base_url = f"https://{s3_base_url}"

            category_id = self.request.query_params.get("category_id")
            category_map = {}

            for product in queryset:
                cat_id = product.categoryId.id if product.categoryId else None
                cat_name = product.categoryId.name if product.categoryId else "Uncategorized"

                if cat_id not in category_map:
                    category_map[cat_id] = {
                        "category_id": cat_id,
                        "category_name": cat_name,
                        "total_products": 0,
                        "products": []
                    }

                category_map[cat_id]["total_products"] += 1
                category_map[cat_id]["products"].append(self.serialize_product(product, s3_base_url))

            response_data = list(category_map.values())

            if category_id and not response_data:
                return Response(
                    {
                        "statusCode": 404,
                        "status": False,
                        "message": "No products found for the given category",
                        "categories": [],
                    },
                    status=status.HTTP_200_OK,
                )

            return Response(
                {
                    "statusCode": 200,
                    "status": True,
                    "message": "Company products retrieved successfully",
                    "categories": response_data,
                },
                status=status.HTTP_200_OK,
            )

        except ValidationError as ve:
            return Response(
                {"statusCode": 400, "status": False, "message": str(ve)},
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"statusCode": 500, "status": False, "message": f"Server error: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def serialize_product(self, product, s3_base_url):
        """Serialize product details."""
        return {
        "id": product.id,
        "productname": product.productname,
        "productType": product.productType,
        "description": product.description,
        "priceOnsite": float(product.priceOnsite) if product.priceOnsite is not None else 0.0,
        "priceClickAndCollect": float(product.priceClickAndCollect) if product.priceClickAndCollect is not None else 0.0,
        "priceDelivery": float(product.priceDelivery) if product.priceDelivery is not None else 0.0,
        "quantity": int(product.quantity) if product.quantity is not None else 0,
        "vatRate": float(product.vatRate) if product.vatRate is not None else 0.0,
        "promotionalPrice": float(product.promotionalPrice) if product.promotionalPrice is not None else 0.0,
        "availabilityDateTime": product.availabilityDateTime.isoformat() if product.availabilityDateTime else None,
        "product_image": f"{s3_base_url}/{product.ProductImage}" if product.ProductImage else None,
        "gallery_images": [f"{s3_base_url}/{img}" for img in product.galleryImage] if product.galleryImage else [],
        "isActive": product.isActive,
        "discount":product.discount,
        "subcategory": {
            "subcategory_id": product.subCategoryId.id if product.subCategoryId else None,
            "subcategory_name": product.subCategoryId.name if product.subCategoryId else None
        }
    }






class UpdateProductAPIView(UpdateAPIView):
    
    permission_classes = [IsAuthenticated]
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    lookup_field = "id"
    parser_classes = [MultiPartParser, FormParser] 

    def update(self, request, *args, **kwargs):
         
        product = self.get_object()

        try:
            professional_user = ProfessionalUser.objects.get(email=request.user.email)
            company = professional_user.company
            if not company:
                return Response(
                    {"statusCode": 403, "status": False, "message": "User is not linked to any company"},
                    status=status.HTTP_200_OK,
                )
        except ProfessionalUser.DoesNotExist:
            return Response(
                {"statusCode": 400, "status": False, "message": "User not found"},
                status=status.HTTP_200_OK,
            )

        if product.company != company:
            return Response(
                {"statusCode": 403, "status": False, "message": "You are not authorized to update this product"},
                status=status.HTTP_200_OK,
            )

        product_data = request.data.copy()
        allowed_categories = professional_user.categories.values_list("id", flat=True)
        allowed_subcategories = professional_user.subcategories.values_list("id", flat=True)
        category_id = request.data.get("categoryId", product.categoryId.id if product.categoryId else None)

        subcategory_id = request.data.get("subCategoryId", product.subCategoryId.id if product.subCategoryId else None)
        try:
            category_id = int(category_id) if category_id is not None else None
            subcategory_id = int(subcategory_id) if subcategory_id is not None else None
        except ValueError:
            return Response(
                {"statusCode": 400, "status": False, "message": "Invalid category or subcategory ID"},
                status=status.HTTP_200_OK,
            )

        if category_id and category_id not in allowed_categories:
            return Response(
                {"statusCode": 400, "status": False, "message": "Invalid category selection"},
                status=status.HTTP_200_OK,
            )

        if subcategory_id and subcategory_id not in allowed_subcategories:
            return Response(
                {"statusCode": 400, "status": False, "message": "Invalid subcategory selection"},
                status=status.HTTP_200_OK,
            )
        title = request.data.get("productname", "").strip()
        if title and Product.objects.filter(company=company, productname=title).exclude(id=product.id).exists():
            return Response(
                {"statusCode": 400, "status": False, "message": "A product with this name already exists for your company"},
                status=status.HTTP_200_OK,
            )

        folder_id = request.data.get("folder") or getattr(product, "folderId", None)
        if folder_id:   
            try:
                folder = CategoryFolder.objects.get(id=folder_id)
                if folder.productType != product.productType:
                    return Response(
                        {
                            "statusCode": 400,
                            "status": False,
                            "message": "Product can only be added to a folder with the same product type",
                        },
                        status=status.HTTP_200_OK,
                    )
            except CategoryFolder.DoesNotExist:
                return Response(
                    {
                        "statusCode": 400,
                        "status": False,
                        "message": "Selected folder does not exist",
                    },
                    status=status.HTTP_200_OK,
                )
        try:
            existing_gallery_images = json.loads(product.galleryImage) if isinstance(product.galleryImage, str) else (product.galleryImage or [])
        except json.JSONDecodeError:
            existing_gallery_images = []
            
        all_gallery_images = existing_gallery_images
        
        if "galleryImage" in request.FILES:
            new_gallery_images = []
            for image in request.FILES.getlist("galleryImage"):
                saved_path = default_storage.save(f"gallery_images/{image.name}", image)
                uploaded_url = default_storage.url(saved_path)
                new_gallery_images.append(uploaded_url)
            all_gallery_images = existing_gallery_images + new_gallery_images
            if len(all_gallery_images) > 30:
                return Response(
                    {"statusCode": 400, "status": False, "message": "Maximum 30 images allowed in gallery"},
                    status=status.HTTP_200_OK,
                )

            product_data["galleryImage"] = json.dumps(all_gallery_images)
        else:
            product_data["galleryImage"] = json.dumps(existing_gallery_images)  # Keep old images



        product_data["petAllowed"] = request.data.get("petAllowed", product.petAllowed)
        product_data["cruiseName"] = request.data.get("cruiseName", product.cruiseName)
       
       
        if "startAddress" in request.data:
                start_address_data = request.data["startAddress"]
                if isinstance(start_address_data, str):
                    start_address_data = json.loads(start_address_data)
                
                start_address = product.startAddress  # Existing Address instance
                address_serializer = AddressSerializer(start_address, data=start_address_data, partial=True)
                if address_serializer.is_valid():
                    address_serializer.save()
                else:
                    return Response({
                        "statusCode": 400,
                        "status": False,
                        "message": "Invalid startAddress data",
                        "errors": address_serializer.errors
                    }, status=status.HTTP_200_OK)

        if "endAddress" in request.data:
            end_address_data = request.data["endAddress"]
            if isinstance(end_address_data, str):
                end_address_data = json.loads(end_address_data)

            end_address = product.endAddress  # Existing Address instance
            address_serializer = AddressSerializer(end_address, data=end_address_data, partial=True)
            if address_serializer.is_valid():
                address_serializer.save()
            else:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Invalid endAddress data",
                    "errors": address_serializer.errors
                }, status=status.HTTP_200_OK)
                
        
               
        start_address_data = AddressSerializer(product.startAddress).data if product.startAddress else None
        end_address_data = AddressSerializer(product.endAddress).data if product.endAddress else None
        
        product_data.update({
            "startAddress": product.startAddress.id if product.startAddress else None,
            "endAddress": product.endAddress.id if product.endAddress else None,
        })
        data = request.data.copy()

        try:
            parse_json_field(data, "keywords")
            parse_json_field(data, "artistName")
            parse_json_field(data, "bandName")
        except ValueError as e:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": {str(e).split()[0]: [str(e)]}
            }, status=status.HTTP_200_OK)
        
            
        def parse_time_safe(val):
            try:
                return time.fromisoformat(val)
            except:
                return None

        start_time = request.data.get("startTime")
        end_time = request.data.get("endTime")

        if start_time:
            parsed_start = parse_time_safe(start_time)
            if parsed_start:
                product_data["startTime"] = parsed_start
            else:
                product_data["startTime"] = product.startTime

        if end_time:
            parsed_end = parse_time_safe(end_time)
            if parsed_end:
                product_data["endTime"] = parsed_end
            else:
                product_data["endTime"] = product.endTime
        product_data = request.data.copy()

        for field in ["cruiseFacility", "roomFacility"]:
            success, error_response = parse_list_field_and_set(product_data, field)
            if not success:
                return error_response
        serializer = self.get_serializer(product, data=product_data, partial=True)
        
        if serializer.is_valid():
            product = serializer.save(company=company) 
            
            try:
                update_result = update_tickets(request, product)
                if isinstance(update_result, Response):
                    return update_result  # return error from update_tickets
                room_result = update_rooms(request, product)
                if isinstance(room_result, Response):
                    return room_result
                    
                print("=================================================================")       
                return Response(
                    {
                        "statusCode": 200,
                        "status": True,
                        "message": "Product updated successfully",
                        "data": {
                            "company": company.id,
                            "categoryId": int(category_id) if category_id else None,
                            "subCategoryId": int(subcategory_id) if subcategory_id else None,
                            "productType": product.productType,
                            "description": product.description,
                            "ProductImage": request.build_absolute_uri(product.ProductImage.url) if product.ProductImage else None,
                            "galleryImages": all_gallery_images,
                            "availabilityDateTime": product.availabilityDateTime,
                            "keywords": product.keywords,
                            "artistName": product.artistName,
                            "bandName": product.bandName,
                            "startTime": product.startTime,
                            "endTime": product.endTime,
                            "rooms": room_result,
                            
                            "startAddress": {
                                "address1": product.startAddress.address1 if product.startAddress else None,
                                "address2": product.startAddress.address2 if product.startAddress else None,
                                "postalCode": product.startAddress.postalCode if product.startAddress else None,
                                "lat": product.startAddress.lat if product.startAddress else None,
                                "lang": product.startAddress.lang if product.startAddress else None,
                                "city": product.startAddress.city if product.startAddress else None,
                                "country": product.startAddress.country if product.startAddress else None,
                            } if product.startAddress else None,
                            "endAddress": {
                                "address1": product.endAddress.address1 if product.endAddress else None,
                                "address2": product.endAddress.address2 if product.endAddress else None,
                                "postalCode": product.endAddress.postalCode if product.endAddress else None,
                                "lat": product.endAddress.lat if product.endAddress else None,
                                "lang": product.endAddress.lang if product.endAddress else None,
                                "city": product.endAddress.city if product.endAddress else None,
                                "country": product.endAddress.country if product.endAddress else None,
                            } if product.endAddress else None,
                            "petAllowed": product.petAllowed,
                            "petType": product.petType,
                            "cruiseName": product.cruiseName,
                            "vatRate": product.vatRate,
                            "promotionalPrice": str(product.promotionalPrice) if product.promotionalPrice else None,
                            "average_rating": product.average_rating,
                            "total_ratings": product.total_ratings,
                            "folder": product.folder,
                            "cruiseFacility": list(product.cruiseFacility.all().values_list('id', flat=True)),
                            "roomFacility": list(product.roomFacility.all().values_list('id', flat=True)),
                            "concert_tickets": update_result["concert_tickets"],
                            "nightclub_tickets": update_result["nightclub_tickets"],
                            "amusement_tickets": update_result["amusements_tickets"],
                            "created_at": product.created_at,
                            "updated_at": product.updated_at
                        },
                    },
                    status=status.HTTP_200_OK,
                )

                    
                return Response(
                    {
                        "statusCode": 200,
                        "status": True,
                        "message": "Product updated successfully",
                        "data": serializer.data,
                    }
                )
            except  Exception as e:
                return Response(
                    {"statusCode": 500, "status": False, "message": f"Internal Server Error:{str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

        return Response(
            {"statusCode": 400, "status": False, "message": serializer.errors},
            status=status.HTTP_200_OK,
        )


class DeleteProductAPIView(DestroyAPIView):
    permission_classes = [IsAuthenticated]
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    lookup_field = "id"

    def delete(self, request, *args, **kwargs):
        product = self.get_object()

        try:
            professional_user = ProfessionalUser.objects.get(email=request.user.email)
            company = professional_user.company
            if not company:
                return Response(
                    {"statusCode": 403, "status": False, "message": "User is not linked to any company"},
                    status=status.HTTP_200_OK,
                )
        except ProfessionalUser.DoesNotExist:
             return Response(
                {"statusCode": 400, "status":False, "message": "User not found"},
                status=status.HTTP_200_OK,
            )

        if product.company != company:
            return Response(
                {"statusCode": 403, "status": False, "message": "You are not authorized to delete this product"},
                status=status.HTTP_200_OK,
            )

        product.delete()
        return Response(
            {"statusCode": 200, "status":True, "message": "Product deleted successfully"},
            status=status.HTTP_200_OK,
        )
class CreateCategoryFolderAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Create a new folder with allowed categories for Professional Users"""
        user = request.user
        professional_user = ProfessionalUser.objects.get(email=request.user.email)
        if not professional_user:
            return Response({
                "statusCode": 403,
                "status": False,
                "message": "Only Professional Users can create folders."
            }, status=status.HTTP_200_OK)
        folder_name = request.data.get("name")
        category_input = request.data.get("categories", "")
        product_type = request.data.get("productType")
        valid_product_types = [choice[0] for choice in CategoryFolder.PRODUCT_TYPE_CHOICES]
        if not folder_name:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Folder name is required."
            }, status=status.HTTP_200_OK)
        
        if CategoryFolder.objects.filter(professionalUser=professional_user, name__iexact=folder_name).exists():
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Folder name already exists. Please choose a different name."
            }, status=status.HTTP_200_OK)

        if not product_type or product_type not in valid_product_types:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": f"Invalid or missing productType. Choose one of {', '.join(valid_product_types)}."
            }, status=status.HTTP_200_OK)
        try:
            selected_category_ids = [int(cat_id) for cat_id in category_input.split(",") if cat_id.strip()]
        except ValueError:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Invalid category format. Use comma-separated values like '1,2,3'."
            }, status=status.HTTP_200_OK)
        allowed_category_ids = set(professional_user.categories.values_list("id", flat=True))
        invalid_categories = [cat_id for cat_id in selected_category_ids if cat_id not in allowed_category_ids]

        if invalid_categories:
            return Response({
                "statusCode": 403,
                "status": False,
                "message": "You can only select categories assigned to you during registration."
            }, status=status.HTTP_200_OK)
        folder = CategoryFolder.objects.create(
            professionalUser=professional_user,
            name=folder_name,
            productType=product_type
        )
        folder.categories.set(selected_category_ids)
        serializer = CategoryFolderSerializer(folder)

        return Response({
            "statusCode": 200,
            "status": True,
            "message": "Folder created successfully.",
            "data": serializer.data
        }, status=status.HTTP_200_OK)
class UpdateCategoryFolderAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, folder_id):
        
        user = request.user

        try:
            professional_user = ProfessionalUser.objects.get(email=user.email)
        except ProfessionalUser.DoesNotExist:
            return Response({
                "statusCode": 403,
                "status": False,
                "message": "Only Professional Users can update folders."
            }, status=status.HTTP_200_OK)

        try:
            folder = CategoryFolder.objects.get(id=folder_id, professionalUser=professional_user)
        except CategoryFolder.DoesNotExist:
            return Response({
                "statusCode": 404,
                "status": False,
                "message": "Folder not found."
            }, status=status.HTTP_200_OK)

        folder_name = request.data.get("name")
        category_input = request.data.get("categories", "")
        product_type = request.data.get("productType")
        valid_product_types = [choice[0] for choice in CategoryFolder.PRODUCT_TYPE_CHOICES]
        if folder_name:
            if CategoryFolder.objects.filter(
                professionalUser=professional_user,
                name__iexact=folder_name
            ).exclude(id=folder.id).exists():
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Folder name already exists. Please choose a different name."
                }, status=status.HTTP_200_OK)
            folder.name = folder_name
        if product_type:
            if product_type not in valid_product_types:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": f"Invalid productType. Choose one of {', '.join(valid_product_types)}."
                }, status=status.HTTP_200_OK)
            folder.productType = product_type
        if category_input:
            try:
                selected_category_ids = [int(cat_id) for cat_id in category_input.split(",") if cat_id.strip()]
            except ValueError:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Invalid category format. Use comma-separated values like '1,2,3'."
                }, status=status.HTTP_200_OK)

            allowed_category_ids = set(professional_user.categories.values_list("id", flat=True))
            invalid_categories = [cat_id for cat_id in selected_category_ids if cat_id not in allowed_category_ids]

            if invalid_categories:
                return Response({
                    "statusCode": 403,
                    "status": False,
                    "message": "You can only select categories assigned to you during registration."
                }, status=status.HTTP_200_OK)

            folder.categories.set(selected_category_ids)
        folder.save()
        category_ids = [cat.id for cat in folder.categories.all()]
        category_names = [cat.name for cat in folder.categories.all()]

        return Response({
            "statusCode": 200,
            "status": True,
            "message": "Folder updated successfully.",
            "data": {
                "folder_id": folder.id,
                "folder_name": folder.name,
                "productType": folder.productType,
                "categories": {
                    "category_id": category_ids,
                    "category_name": category_names
                }
            }
        }, status=status.HTTP_200_OK)
    

class StandardResultsSetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100
class GetProductsByFolderNameAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            folder_name = request.query_params.get("folder_name", "").strip()
            
            professional_user = get_object_or_404(ProfessionalUser, email=request.user.email)

            s3_base_url = settings.AWS_S3_CUSTOM_DOMAIN if hasattr(settings, 'AWS_S3_CUSTOM_DOMAIN') else ""
            if not s3_base_url.startswith("https://"):
                s3_base_url = f"https://{s3_base_url}"
            if folder_name:
                try:
                    folder = CategoryFolder.objects.get(name__iexact=folder_name, professionalUser=professional_user)
                except ObjectDoesNotExist:
                    return Response({
                        "statusCode": 404,
                        "status": False,
                        "message": f"Folder '{folder_name}' not found or you do not have access."
                    }, status=status.HTTP_200_OK)

                products = Product.objects.filter(folder=folder)
                paginator = StandardResultsSetPagination()
                paginated_products = paginator.paginate_queryset(products, request)
                product_data = self.serialize_products(paginated_products, s3_base_url)

                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": f"Products in '{folder.name}' retrieved successfully.",
                    "folder": {
                        "folder_id": folder.id,
                        "folder_name": folder.name,
                        "productType": folder.productType,
                        "total_products": products.count(),
                        "products": product_data
                    }
                }, status=status.HTTP_200_OK)
            folders = CategoryFolder.objects.filter(professionalUser=professional_user)
            all_folders_data = []

            for folder in folders:
                products = Product.objects.filter(folder=folder)
                product_data = self.serialize_products(products, s3_base_url)

                all_folders_data.append({
                    "folder_id": folder.id,
                    "folder_name": folder.name,
                    "productType": folder.productType,
                    "total_products": products.count(),
                    "products": product_data
                })

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "All folders with their products retrieved successfully.",
                "folders": all_folders_data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred.",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def serialize_products(self, products, s3_base_url):
        product_list = []
        for product in products:
            product_list.append({
                "id": product.id,
                "productname": product.productname,
                "productType": product.productType,
                "description": product.description,
                "priceOnsite": float(product.priceOnsite) if product.priceOnsite is not None else 0.0,
                "priceClickAndCollect": float(product.priceClickAndCollect) if product.priceClickAndCollect is not None else 0.0,
                "priceDelivery": float(product.priceDelivery) if product.priceDelivery is not None else 0.0,
                "promotionalPrice": float(product.promotionalPrice) if product.promotionalPrice is not None else 0.0,
                "vatRate": float(product.vatRate) if product.vatRate is not None else 0.0,
                "availabilityDateTime": product.availabilityDateTime.isoformat() if product.availabilityDateTime else None,
                "product_image": f"{s3_base_url}/{product.ProductImage}" if product.ProductImage else None,
                "gallery_images": [f"{img}" for img in product.galleryImage] if product.galleryImage else [],
                "isActive": product.isActive,
                "category": {
                    "category_id": product.categoryId.id if product.categoryId else None,
                    "category_name": product.categoryId.name if product.categoryId else None
                },
                "subcategory": {
                    "subcategory_id": product.subCategoryId.id if product.subCategoryId else None,
                    "subcategory_name": product.subCategoryId.name if product.subCategoryId else None
                }
            })
        return product_list
    


class GetAllFoldersWithProductCountAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            product_type = request.query_params.get("productType", "").strip()
            exclude_folder_name = request.query_params.get("excludeFolderName", "").strip()
            folder_name = request.query_params.get("folderName", "").strip()
            try:
                page = int(request.GET.get("page", 1))
            except (ValueError, TypeError):
                page = 1
            
            try:
                page_size = int(request.GET.get("page_size", 10))
            except (ValueError, TypeError):
                page_size = 10
            
            # Validate product type if provided
            if product_type and product_type not in ["Product", "Services", "Ticket"]:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Invalid product type."
                }, status=status.HTTP_200_OK)

            professional_user = get_object_or_404(ProfessionalUser, email=request.user.email)

            # Filter folders by professional user, and also by productType if provided
            folders = CategoryFolder.objects.filter(professionalUser=professional_user)
            if product_type:
                folders = folders.filter(productType=product_type)

            if exclude_folder_name:
                folders = folders.exclude(name=exclude_folder_name)
            
            if folder_name:
                folders = folders.filter(name=folder_name)
            
            if not folders.exists():
                return Response({
                    "statusCode": 404,
                    "status": False,
                    "message": "No folders found for this user."
                }, status=status.HTTP_200_OK)

            folder_data = []
            for folder in folders:
                products = Product.objects.filter(folder=folder)
                if product_type:
                    products = products.filter(productType=product_type)

                folder_data.append({
                    "folder_id": folder.id,
                    "folder_name": folder.name,
                    "productType": folder.productType,
                    "categories": {
                        "category_id": [cat.id for cat in folder.categories.all()],
                        "category_name": [cat.name for cat in folder.categories.all()]
                    },
                    "product_count": products.count()
                })
            
            total_items = len(folder_data)
            total_pages = (total_items + page_size - 1) // page_size
            start = (page - 1) * page_size
            end = start + page_size
            paginated_data = folder_data[start:end]

            # Previous and next page URLs
            base_url = request.build_absolute_uri(request.path)
            query_params = request.GET.copy()

            def build_url(page_number):
                query_params["page"] = page_number
                return f"{base_url}?{query_params.urlencode()}"

            previous_page = build_url(page - 1) if page > 1 else None
            next_page = build_url(page + 1) if end < total_items else None

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Folders retrieved successfully.",
                "total_count": total_items,
                "current_page": page,
                "page_size": page_size,
                "total_pages": total_pages,
                "previous_page": previous_page,
                "next_page": next_page,
                "folders": paginated_data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred.",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        





class UserCampaignListAPIView(APIView):
    permission_classes = [IsAuthenticated]
    @swagger_auto_schema(
        operation_summary="List User Campaigns",
        operation_description="Returns a list of all campaigns linked to the authenticated ProfessionalUser's company.",
        responses={
            200: openapi.Response(
                description="Campaign list retrieved successfully",
                examples={
                    "application/json": [
                        {
                            "id": 1,
                            "title": "Summer Marketing Campaign",
                            "description": "A campaign for summer sales boost.",
                            "start_date": "2025-06-01",
                            "end_date": "2025-09-01",
                        },
                        {
                            "id": 2,
                            "title": "Winter Discount Campaign",
                            "description": "A campaign focused on winter discounts.",
                            "start_date": "2025-12-01",
                            "end_date": "2026-01-31",
                        },
                    ]
                },
            ),
            403: openapi.Response(
                description="Forbidden - User not linked to any company",
                examples={
                    "application/json": {
                        "statusCode": 403,
                        "status": "message",
                        "message": "User is not linked to any company"
                    }
                },
            ),
            401: openapi.Response(
                description="Unauthorized - Bearer Token Required",
                examples={
                    "application/json": {
                        "detail": "Authentication credentials were not provided."
                    }
                },
            ),
        },
    )

    def get(self, request):
        professional_user = ProfessionalUser.objects.get(email=self.request.user.email)
        company = professional_user.company  
        campaigns = Campaign.objects.filter(company=company)  
        
        serializer = CampaignSerializer(campaigns, many=True)  
        return Response(serializer.data, status=status.HTTP_200_OK)


class CreateCampaignAPIView(APIView):
    permission_classes = [IsAuthenticated]
    @swagger_auto_schema(
        operation_summary="Create a Campaign",
        operation_description="Allows authenticated Professional Users to create a new campaign for their company.",
        request_body=CampaignSerializer,
        responses={
            200: openapi.Response(
                description="Campaign created successfully",
                examples={
                    "application/json": {
                        "id": 1,
                        "title": "Black Friday Campaign",
                        "description": "A marketing campaign for Black Friday deals.",
                        "start_date": "2025-11-25",
                        "end_date": "2025-11-30",
                        "company": 3
                    }
                },
            ),
            400: openapi.Response(
                description="Bad Request - Validation errors",
                examples={
                    "application/json": {
                        "title": ["This field is required."],
                        "start_date": ["Invalid date format."]
                    }
                },
            ),
            403: openapi.Response(
                description="Forbidden - Only Professional Users can create campaigns",
                examples={
                    "application/json": {
                        "message": "Only Professional Users can create campaigns"
                    }
                },
            ),
            401: openapi.Response(
                description="Unauthorized - Bearer Token Required",
                examples={
                    "application/json": {
                        "detail": "Authentication credentials were not provided."
                    }
                },
            ),
        },
    )


    def post(self, request):
        try:
            professional_user = get_object_or_404(ProfessionalUser, email=request.user.email)
            company = professional_user.company   
            serializer = CampaignSerializer(data=request.data, context={"request": request})
            if serializer.is_valid():
                serializer.save(company=company)  # Save with user & company
                return Response(serializer.data)
            
            return Response(serializer.errors, status=status.HTTP_200_OK)

        except ProfessionalUser.DoesNotExist:
            return Response(
                {"message": "Only Professional Users can create campaigns"},
                status=status.HTTP_200_OK
            )


class RetrieveCampaignAPIView(APIView):
    permission_classes = [IsAuthenticated]
    @swagger_auto_schema(
        operation_summary="Retrieve a Campaign",
        operation_description="Fetch details of a specific campaign by its ID for authenticated Professional Users.",
        manual_parameters=[
            openapi.Parameter(
                "pk",
                openapi.IN_PATH,
                description="ID of the campaign to retrieve",
                type=openapi.TYPE_INTEGER,
                required=True,
            )
        ],
        responses={
            200: openapi.Response(
                description="Campaign retrieved successfully",
                examples={
                    "application/json": {
                        "id": 1,
                        "title": "Black Friday Campaign",
                        "description": "A marketing campaign for Black Friday deals.",
                        "start_date": "2025-11-25",
                        "end_date": "2025-11-30",
                        "company": 3
                    }
                },
            ),
            404: openapi.Response(
                description="Not Found - Campaign does not exist",
                examples={
                    "application/json": {
                        "detail": "Not found."
                    }
                },
            ),
            401: openapi.Response(
                description="Unauthorized - Bearer Token Required",
                examples={
                    "application/json": {
                        "detail": "Authentication credentials were not provided."
                    }
                },
            ),
        },
    )

    def get(self, request, pk):
        professional_user = get_object_or_404(ProfessionalUser, email=request.user.email)
        company = professional_user.company  
        campaign = get_object_or_404(Campaign, pk=pk, company=company)
        
        serializer = CampaignSerializer(campaign)
        return Response(serializer.data, status=status.HTTP_200_OK)


class UpdateCampaignAPIView(APIView):
    permission_classes = [IsAuthenticated]
    @swagger_auto_schema(
        operation_summary="Update a Campaign",
        operation_description="Allows authenticated Professional Users to update a campaign by its ID.",
        manual_parameters=[
            openapi.Parameter(
                "pk",
                openapi.IN_PATH,
                description="ID of the campaign to update",
                type=openapi.TYPE_INTEGER,
                required=True,
            )
        ],
        request_body=CampaignSerializer,
        responses={
            200: openapi.Response(
                description="Campaign updated successfully",
                examples={
                    "application/json": {
                        "id": 1,
                        "title": "Updated Campaign Title",
                        "description": "Updated campaign description",
                        "start_date": "2025-11-25",
                        "end_date": "2025-11-30",
                        "company": 3
                    }
                },
            ),
            400: openapi.Response(
                description="Bad Request - Validation errors",
                examples={
                    "application/json": {
                        "title": ["This field is required."]
                    }
                },
            ),
            404: openapi.Response(
                description="Not Found - Campaign does not exist",
                examples={
                    "application/json": {
                        "detail": "Not found."
                    }
                },
            ),
            401: openapi.Response(
                description="Unauthorized - Bearer Token Required",
                examples={
                    "application/json": {
                        "detail": "Authentication credentials were not provided."
                    }
                },
            ),
        },
    )

    def put(self, request, pk):
        professional_user = get_object_or_404(ProfessionalUser, email=request.user.email)

        campaign = get_object_or_404(Campaign, pk=pk, company=professional_user.company)  
        serializer = CampaignSerializer(campaign, data=request.data, partial=True, context={"request": request})

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_200_OK)

class DeleteCampaignAPIView(APIView):
    permission_classes = [IsAuthenticated]
    @swagger_auto_schema(
        operation_summary="Delete a Campaign",
        operation_description="Allows authenticated Professional Users to delete a campaign by its ID.",
        manual_parameters=[
            openapi.Parameter(
                "pk",
                openapi.IN_PATH,
                description="ID of the campaign to delete",
                type=openapi.TYPE_INTEGER,
                required=True,
            )
        ],
        responses={
            204: openapi.Response(
                description="Campaign deleted successfully",
                examples={
                    "application/json": {
                        "message": "Campaign deleted successfully"
                    }
                },
            ),
            404: openapi.Response(
                description="Not Found - Campaign does not exist",
                examples={
                    "application/json": {
                        "detail": "Not found."
                    }
                },
            ),
            401: openapi.Response(
                description="Unauthorized - Bearer Token Required",
                examples={
                    "application/json": {
                        "detail": "Authentication credentials were not provided."
                    }
                },
            ),
        },
    )


    def delete(self, request, pk):
        professional_user = get_object_or_404(ProfessionalUser, email=request.user.email)
        company = professional_user.company  
        campaign = get_object_or_404(Campaign, pk=pk, company=company)
        campaign.delete()

        return Response({"message": "Campaign deleted successfully"}, status=status.HTTP_204_NO_CONTENT)




class UserPromotionListAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            professional_user = ProfessionalUser.objects.get(email=request.user.email)

            if not professional_user or not professional_user.company:
                return Response({
                    "statusCode": 401,
                    "status": False,
                    "message": "Authenticated user is not associated with a professional company."
                }, status=status.HTTP_200_OK)

            company = professional_user.company
            search_query = request.query_params.get('search', '')

            page = int(request.query_params.get('page', 1))
            page_size = int(request.query_params.get('page_size', 10))
            offset = (page - 1) * page_size

            promotions_qs = Promotions.objects.filter(company=company)

            if search_query:
                promotions_qs = promotions_qs.filter(
                    Q(promotionName__icontains=search_query) 
                )

            promotions_qs = promotions_qs.order_by('-created_at')
            total = promotions_qs.count()

            paginated_qs = promotions_qs[offset:offset + page_size]
            serializer = PromotionsSerializer(paginated_qs, many=True)

            total_pages = math.ceil(total / page_size)

            # Build next/previous URLs manually
            base_url = request.build_absolute_uri('?')
            next_page = page + 1 if page < total_pages else None
            prev_page = page - 1 if page > 1 else None

            def build_page_url(p):
                if p:
                    return f"{request.path}?page={p}&page_size={page_size}&search={search_query}"
                return None

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Promotions retrieved successfully.",
                "total": total,
                "page": page,
                "total_pages": total_pages,
                "next": build_page_url(next_page),
                "previous": build_page_url(prev_page),
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        except ProfessionalUser.DoesNotExist:
            return Response({
                "statusCode": 401,
                "status": False,
                "message": "The authenticated user does not have a professional user account."
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred.",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    


class CreatePromotionsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        user = request.user

        try:
            professional_user = ProfessionalUser.objects.get(email=user.email)
            company = professional_user.company
        except ProfessionalUser.DoesNotExist:
            logger.error(f"User {user.email} does not have a professional company associated.")
            return Response({
                "statusCode": 401,
                "status": False,
                "message": "Authenticated user is not associated with a professional company."
            }, status=status.HTTP_200_OK)

        if not company:
            logger.error(f"User {user.email} is not linked to any company.")
            return Response({
                "statusCode": 401,
                "status": False,
                "message": "Authenticated user is not linked to any company."
            }, status=status.HTTP_200_OK)

        product_ids_raw = request.data.get("product_ids", [])
        product_service_type = request.data.get("product_service_type", None)
        if not product_ids_raw:
            logger.error("Product IDs are empty or not provided.")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Product IDs must be a non-empty list."
            }, status=status.HTTP_200_OK)
        if isinstance(product_ids_raw, str):
            try:
                product_ids = json.loads(product_ids_raw)
            except json.JSONDecodeError:
                logger.error("Invalid JSON format for product_ids.")
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Invalid JSON format for product_ids."
                }, status=status.HTTP_200_OK)
        else:
            product_ids = product_ids_raw
        if not isinstance(product_ids, list) or not product_ids:
            logger.error("Product IDs must be a non-empty list.")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Product IDs must be a non-empty list."
            }, status=status.HTTP_200_OK)
        valid_products = Product.objects.filter(id__in=product_ids, company=company)

        if not valid_products.exists():
            logger.error(f"No valid products found for company {company.id} with product IDs {product_ids}.")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "No valid products found for this company."
            }, status=status.HTTP_200_OK)
        product_types = list(valid_products.values_list('productType', flat=True).distinct())
        logger.debug(f"Product types of selected products: {product_types}")
        if product_service_type:
            if len(product_types) != 1 or product_types[0].lower() != product_service_type.lower():
                logger.error(f"Mismatch between provided product_service_type '{product_service_type}' and actual product types: {product_types}")
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "All selected products must have the same product_service_type, which must match the provided type."
                }, status=status.HTTP_200_OK)
        promo_data = request.data.copy()
        promo_data.setlist("productId", [str(product.id) for product in valid_products])
        promo_data["company"] = str(company.id)
        promo_data.pop("product_ids", None)

        try:
            serializer = PromotionsSerializer(data=promo_data)
            serializer.is_valid(raise_exception=True)
            promotion = serializer.save()
            logger.info(f"Promotion created successfully for company {company.id} with product IDs {product_ids}.")
            response_data = PromotionsSerializer(promotion).data
            response_data["productId"] = [product.id for product in valid_products]  # Ensure it's a list of integers

            return Response({
                "statusCode": 201,
                "status": True,
                "message": "Promotion created successfully.",
                "data": response_data
            }, status=status.HTTP_201_CREATED)

        except serializers.ValidationError as ve:
            logger.error(f"Validation error: {ve.detail}")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Validation error.",
                "details": ve.detail
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"An unexpected error occurred: {str(e)}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred.",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)





class RetrievePromotionAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
        
            user = request.user
            professional_user = ProfessionalUser.objects.get(email=user.email)
            company = professional_user.company
            promotion = Promotions.objects.filter(pk=pk, company=company).first()

            if not promotion:
                return Response({
                    "statusCode": 404,
                    "status": False,
                    "message": "Promotion not found or does not belong to your company.",
                    "details": "Ensure the promotion exists and belongs to the authenticated user's company."
                }, status=status.HTTP_200_OK)
            serializer = PromotionsSerializer(promotion)

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Promotion retrieved successfully.",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        except ProfessionalUser.DoesNotExist:
            return Response({
                "statusCode": 404,
                "status": False,
                "message": "Professional user not found.",
                "details": "The authenticated user does not have a corresponding professional user."
            }, status=status.HTTP_200_OK)

        except Promotions.DoesNotExist:
            return Response({
                "statusCode": 404,
                "status": False,
                "message": "Promotion not found.",
                "details": "No promotion found matching the provided ID."
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred.",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UpdatePromotionAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, pk):
        try:
            professional_user = get_object_or_404(ProfessionalUser, email=request.user.email)
            promotion = get_object_or_404(Promotions, pk=pk, company=professional_user.company)

            serializer = PromotionsSerializer(promotion, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Promotion updated successfully.",
                    "data": serializer.data
                }, status=status.HTTP_200_OK)

            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Validation error.",
                "details": serializer.errors
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred.",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class DeletePromotionAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def delete(self, request, pk):  # Accepting pk
        try:
            professional_user = ProfessionalUser.objects.filter(email=request.user.email).first()

            if not professional_user:
                return Response({
                    "detail": "Authentication error: User not found in both ProfessionalUser and UserApp"
                }, status=status.HTTP_200_OK)

            user_company = professional_user.company
            promotion = Promotions.objects.filter(id=pk, company=user_company).first()

            if not promotion:
                return Response({
                    "statusCode": 404,
                    "status": False,
                    "message": "Promotion not found."
                }, status=status.HTTP_200_OK)
            promotion.delete()

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Promotion deleted successfully."
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred.",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class GetInvoiceListView(APIView):
   
   permission_classes = [IsAuthenticated]
    
   def get(self, request, id=None):
        
        if id:
            try:
                invoice = Invoice.objects.get(id=id)  # Corrected variable name
                serializer = InvoiceSerializer(invoice)
                return Response({"message": "Invoice found", "data": serializer.data}, status=status.HTTP_200_OK)
            except Invoice.DoesNotExist:
                return Response({"message": "Invoice not found"}, status=status.HTTP_200_OK)
        
        invoices = Invoice.objects.all()
        serializer = InvoiceSerializer(invoices, many=True)
        
        response_data = {
            "total_count": invoices.count(),
            "invoices": serializer.data
        }
        return Response(response_data, status=status.HTTP_200_OK)
class UpdateInvoiceListView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request):
        
        companyStore_id = request.data.get("companyStore_id")
        invoice_id = request.data.get("id")  # Assuming 'id' refers to invoice_id

        if not companyStore_id or not invoice_id:
            return Response({"message": "Both companyStore ID and Invoice ID are required."}, status=status.HTTP_200_OK)

        try:
            invoice = Invoice.objects.get(id=invoice_id, companyStore=companyStore_id)
        except Invoice.DoesNotExist:
            return Response({"message": "Invoice not found for the given companyStore ID."}, status=status.HTTP_200_OK)

        serializer = InvoiceSerializer(invoice, data=request.data, partial=True)

        if serializer.is_valid():
            updated_invoice = serializer.save()
            if any(field in request.data for field in ["quantity", "unit_price", "tax_rate"]):
                updated_invoice.total_amount = (updated_invoice.quantity * updated_invoice.unit_price) + (
                    (updated_invoice.quantity * updated_invoice.unit_price) * updated_invoice.tax_rate / 100)
                updated_invoice.save()

            return Response({"message": "Invoice updated successfully", "data": serializer.data}, status=status.HTTP_200_OK)

        return Response(serializer.errors, status=status.HTTP_200_OK)
class DeleteInvoiceListView(APIView):
    def delete(self, request, id=None):
        
        invoice_id = id or request.data.get("id")
        
        if not invoice_id:
            return Response({"message": "Invoice ID is required"}, status=status.HTTP_200_OK)
        
        try:
            invoice = Invoice.objects.get(id=invoice_id)
            invoice.delete()
            return Response({"message": "Invoice deleted successfully"}, status=status.HTTP_200_OK)
        except Invoice.DoesNotExist:
            return Response({"message": "Invoice not found"}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class UserPromocodeListAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            user = request.user

            if not hasattr(user, 'company') or not user.company:
                return Response({
                    "statusCode": 401,
                    "status": False,
                    "message": "Authenticated user is not associated with any company."
                }, status=status.HTTP_200_OK)

            company = user.company
            search_query = request.query_params.get('search', '').strip()

            promocodes_qs = Promocode.objects.filter(company=company)

            if search_query:
                promocodes_qs = promocodes_qs.filter(
                    Q(promocode__icontains=search_query) |
                    Q(description__icontains=search_query)
                )

            promocodes_qs = promocodes_qs.order_by('-id')

            # Pagination logic
            page = int(request.query_params.get('page', 1))
            page_size = int(request.query_params.get('page_size', 10))
            total = promocodes_qs.count()
            total_pages = ceil(total / page_size)

            offset = (page - 1) * page_size
            paginated_qs = promocodes_qs[offset:offset + page_size]

            serializer = PromocodeSerializer(paginated_qs, many=True)

            # Build next/previous links
            def build_url(p):
                if p < 1 or p > total_pages:
                    return None
                base = request.path
                return f"{base}?page={p}&page_size={page_size}&search={search_query}"

            next_url = build_url(page + 1) if page < total_pages else None
            prev_url = build_url(page - 1) if page > 1 else None

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Promocodes retrieved successfully.",
                "total": total,
                "page": page,
                "total_pages": total_pages,
                "next": next_url,
                "previous": prev_url,
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred.",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class CreatePromocodeAPIView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, *args, **kwargs):
        logger.debug("POST /create-promocode/ called.")
        try:
            user = request.user
            logger.debug(f"Authenticated user: {user}")
            logger.debug(f"User company attribute: {getattr(user, 'company', None)}")
            company = getattr(user, 'company', None)
            if company is None:
                logger.warning("User is not associated with a company.")
                return Response({
                    "statusCode": 401,
                    "status": False,
                    "message": "User does not belong to a company."
                }, status=status.HTTP_200_OK)

            logger.debug(f"Company ID: {company.id}")
            data = request.data.copy()
            data["company"] = company.id
            logger.debug(f"Incoming data: {data}")
            promocode_value = data.get("promocode")
            if Promocode.objects.filter(promocode=promocode_value, company=company).exists():
                logger.info(f"Duplicate promocode attempted: {promocode_value}")
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "A promocode with this code already exists for your company."
                }, status=status.HTTP_200_OK)
            try:
                start_datetime = datetime.datetime.strptime(data.get("startDateTime"), "%d/%m/%Y %H:%M:%S")
                end_datetime = datetime.datetime.strptime(data.get("endDateTime"), "%d/%m/%Y %H:%M:%S")
                data["startDateTime"] = start_datetime
                data["endDateTime"] = end_datetime
                logger.debug(f"Parsed start: {start_datetime}, end: {end_datetime}")
            except Exception as date_err:
                logger.error(f"Date parsing error: {date_err}")
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Invalid datetime format.",
                    "details": str(date_err)
                }, status=status.HTTP_200_OK)
            serializer = PromocodeSerializer(data=data, context={"request": request})
            serializer.is_valid(raise_exception=True)
            promocode_instance = serializer.save()
            logger.info(f"Promocode created: {promocode_instance.id}")

            return Response({
                "statusCode": 201,
                "status": True,
                "message": "Promocode created successfully.",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        except serializers.ValidationError as val_err:
            logger.warning(f"Validation Error: {val_err.detail}")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Validation Error.",
                "details": val_err.detail
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.exception("Unexpected error during promocode creation")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Internal Server Error.",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class CheckPromocodeAvailabilityAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        user = request.user
        if not hasattr(user, "company") or user.company is None:
            return Response({
                "statusCode": 401,
                "status": False,
                "message": "User does not belong to a company."
            }, status=status.HTTP_200_OK)

        company = user.company
        partial_code = request.query_params.get("promocode", "").strip()

        if not partial_code:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Promocode input is required."
            }, status=status.HTTP_200_OK)

        
        exists = Promocode.objects.filter(
            company=company,
            promocode__startswith=partial_code
        ).exists()

        return Response({
            "statusCode": 200,
            "status": True,
            "message": "Not applicable" if exists else "Applicable",
            "is_applicable": not exists
        }, status=status.HTTP_200_OK)

    

class UpdatePromocodeAPIView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]  # For image upload support

    def put(self, request, promocode_id, *args, **kwargs):
        try:
            user = request.user
            if not hasattr(user, "company") or user.company is None:
                return Response({
                    "statusCode": 401,
                    "status": False,
                    "message": "User does not belong to a company."
                }, status=status.HTTP_200_OK)

            company = user.company
            try:
                promocode = Promocode.objects.get(id=promocode_id, company=company)
            except Promocode.DoesNotExist:
                return Response({
                    "statusCode": 404,
                    "status": False,
                    "message": "Promocode not found for this company."
                }, status=status.HTTP_200_OK)
            data = request.data.copy()
            serializer = PromocodeSerializer(promocode, data=data, context={"request": request}, partial=True)  # Allow partial update
            serializer.is_valid(raise_exception=True)
            serializer.save()

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Promocode updated successfully.",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Internal Server Error.",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class DeletePromocodeAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, promocode_id, *args, **kwargs):
        
        promocode_id = promocode_id
        
        if not promocode_id:
            return Response({"message": "Promocode ID is required"}, status=status.HTTP_200_OK)

        try:
            user = request.user
            if not hasattr(user, "company") or user.company is None:
                return Response({
                    "statusCode": 401,
                    "status": False,
                    "message": "User does not belong to a company."
                }, status=status.HTTP_200_OK)

            company = user.company

            promocode = Promocode.objects.get(id=promocode_id, company=company)
            promocode.delete()
            
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Promocode deleted successfully."
            }, status=status.HTTP_200_OK)

        except Promocode.DoesNotExist:
            return Response({
                "statusCode": 404,
                "status": False,
                "message": "Promocode not found for this company"
                }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Internal Server Error.",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class LiveSearchThrottle(UserRateThrottle):
    rate = '15/second'  # Prevents excessive API calls

class GetInventoryView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    throttle_classes = [LiveSearchThrottle]
    
    def get(self, request):
        """Retrieve all inventories for the user's company, optionally filtered by product ID and stock status."""
        try:
            user_company = request.user.company

            product_id = request.query_params.get("product_id")
            stock_status_filter = request.query_params.get("stock_status")
            search_query = request.query_params.get("search_query", "").strip().lower()
            inventory_query = Inventory.objects.filter(product__company=user_company)
            if product_id:
                inventory_query = inventory_query.filter(product_id=product_id)
               
            
            if search_query:
                search_query = search_query.lower()  # Ensure lowercase comparison
                inventory_query = inventory_query.filter(
                    Q(product__productname__icontains=search_query)  # Search by product name
                )
            if stock_status_filter:
                stock_status_filter = stock_status_filter.strip().lower()
                if stock_status_filter == "critical stock":
                    inventory_query = inventory_query.filter(stock_quantity__lte=5)
                elif stock_status_filter == "low stock":
                    inventory_query = inventory_query.filter(stock_quantity__lte=F('low_stock_threshold'))
                elif stock_status_filter == "medium stock":
                    inventory_query = inventory_query.filter(
                        stock_quantity__gt=F('low_stock_threshold'),
                        stock_quantity__lte=F('medium_stock_threshold')
                    )
                elif stock_status_filter == "in stock":
                    inventory_query = inventory_query.filter(stock_quantity__gt=F('medium_stock_threshold'))
            inventories = inventory_query.annotate(
                stock_status=Case(
                    When(stock_quantity__lte=5, then=Value("Critical Stock")),
                    When(stock_quantity__lte=F('low_stock_threshold'), then=Value("Low Stock")),
                    When(stock_quantity__gt=F('low_stock_threshold'), stock_quantity__lte=F('medium_stock_threshold'), then=Value("Medium Stock")),
                    default=Value("In Stock"),
                    output_field=CharField(),
                ),
                low_stock_alert=Case(
                    When(stock_quantity__lte=F('low_stock_threshold'), 
                        then=Concat(Value("⚠ Low stock alert! "), 
                                    Cast(F('stock_quantity'), output_field=CharField()),  
                                    Value(" in stock."))),
                    default=Value(None),
                    output_field=CharField(),
                ),
                medium_stock_alert=Case(
                    When(stock_quantity__gt=F('low_stock_threshold'), stock_quantity__lte=F('medium_stock_threshold'),  
                        then=Concat(Value("⚠ Medium stock alert! "), 
                                    Cast(F('stock_quantity'), output_field=CharField()),  
                                    Value(" in stock."))),
                    default=Value(None),
                    output_field=CharField(),
                ),
                immediate_attention_alert=Case(
                    When(stock_quantity__lte=5, then=Value("⚠ 5 items require immediate attention!")),
                    default=Value(None),
                    output_field=CharField(),
                )
            )
            if not inventories.exists():
                return Response({
                    "statusCode": 200,
                    "status": False,
                    "message": "No inventory found for this company or product.",
                    "inventory": []
                }, status=status.HTTP_200_OK)
            total_products = inventories.count()
            total_low_stock = inventories.filter(stock_status="Low Stock").count()
            total_medium_stock = inventories.filter(stock_status="Medium Stock").count()
            total_in_stock = inventories.filter(stock_status="In Stock").count()
            total_critical_stock = inventories.filter(stock_status="Critical Stock").count()

            paginator = pagination.PageNumberPagination()
            paginator.page_size = 20
            paginated_inventories = paginator.paginate_queryset(inventories, request)
            inventory_data = [
                {
                    "id": inventory.id,
                      "company": {
                        "id": inventory.company.id if inventory.company else None,
                        "name": inventory.company.companyName if inventory.company else "Unknown Company"
                    },
                    "product": {
                        "id": inventory.product.id if inventory.product else None,
                        "name": inventory.product.productname if inventory.product else "Unknown Product"
                    },
                    "stock_quantity": inventory.stock_quantity,
                    "medium_stock_threshold": inventory.medium_stock_threshold,
                    "low_stock_threshold": inventory.low_stock_threshold,
                    "last_updated": inventory.last_updated,
                    "stock_status": inventory.stock_status,
                    "low_stock_alert": inventory.low_stock_alert,
                    "medium_stock_alert": inventory.medium_stock_alert,
                    "immediate_attention_alert": inventory.immediate_attention_alert
                }
                for inventory in paginated_inventories
            ]

            return paginator.get_paginated_response({
                "statusCode": 200,
                "status": True,
                "message": "Inventory retrieved successfully.",
                "total_products": total_products,
                "total_in_stock": total_in_stock,
                "total_medium_stock": total_medium_stock,
                "total_low_stock": total_low_stock,
                "total_critical_stock": total_critical_stock,
                "inventory": inventory_data
            })

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred.",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

 
class UpdateInventoryView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def put(self, request):
        try:
            inventory_id = request.query_params.get("inventory_id")  
            
            if not inventory_id:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Inventory ID is required."
                }, status=status.HTTP_200_OK)
            
            user_company = request.user.company
            inventory = get_object_or_404(Inventory, id=inventory_id, product__company=user_company)
            new_stock_quantity = request.data.get("stock_quantity")
            if new_stock_quantity is not None:
                inventory.stock_quantity = int(new_stock_quantity)
                inventory.save(update_fields=["stock_quantity"])  
                if inventory.product:
                    inventory.product.quantity = inventory.stock_quantity
                    inventory.product.save(update_fields=["quantity"]) 
            serializer = InventorySerializer(inventory, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                response_data = serializer.data
                response_data["company_id"] = user_company.id
                response_data["product"] = {
                    "id": inventory.product.id if inventory.product else None,
                    "name": inventory.product.productname if inventory.product else "Unknown Product",
                }

                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Inventory Updated Successfully",
                    "data": response_data,
                }, status=status.HTTP_200_OK)

            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Invalid data.",
                "errors": serializer.errors
            }, status=status.HTTP_200_OK)

        except ValidationError as e:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Validation error.",
                "errors": str(e)
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred.",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class DeleteInventoryView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def delete(self, request):
        """Delete inventory for a product in the user's company"""
        try:
            
            inventory_id = request.query_params.get("inventory_id")
            
            if not inventory_id:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Inventory ID is required."
                }, status=status.HTTP_200_OK)
                
            user_company = request.user.company
            inventory = get_object_or_404(Inventory, id=inventory_id, product__company=user_company)
            inventory.delete()

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Inventory deleted successfully."
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred.",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
class RespondToReviewView(generics.UpdateAPIView):      
    permission_classes = [IsAuthenticated]
    
    def put(self, request, review_id):
        try:
            review = Review.objects.get(id=review_id)
            professional_user = ProfessionalUser.objects.get(email=request.user.email)
            if review.product.company != professional_user.company:
                return Response({"statusCode": 403, "status": False, "message": "You can only respond to reviews for your company's products."},
                                status=status.HTTP_200_OK)

            response_text = request.data.get("response", "").strip()
            if not response_text:
                return Response({"statusCode": 400, "status": False, "message": "Response cannot be empty."},
                                status=status.HTTP_200_OK)
            review.response = response_text
            review.responded_by = request.user
            review.responded_at = now()
            review.save()

            return Response({"statusCode": 200, "status": True, "message": "Response added successfully."},
                            status=status.HTTP_200_OK)
        except Review.DoesNotExist:
            return Response({"statusCode": 404, "status": False, "message": "Review not found."},
                            status=status.HTTP_200_OK)
        except ProfessionalUser.DoesNotExist:
            return Response({"statusCode": 403, "status": False, "message": "You are not a verified professional user."},
                        status=status.HTTP_200_OK)


class ReviewDetailView(generics.RetrieveAPIView):
    """Fetch a single review by ID"""
    permission_classes = [IsAuthenticated]
    queryset = Review.objects.all()
    serializer_class = ReviewSerializer

    def get(self, request, review_id):
        try:
            review = Review.objects.get(id=review_id)
            serializer = ReviewSerializer(review)
            return Response({
                "statusCode": 200,
                "status": True,
                "data": serializer.data
            })
        except Review.DoesNotExist:
            return Response({
                "statusCode": 404,
                "status": False,
                "message": "Review not found"
            }, status=200)
class CompanyReviewCreateView(generics.CreateAPIView):
    serializer_class = CompanyReviewSerializer
    permission_classes = [permissions.IsAuthenticated]  

    def create(self, request, *args, **kwargs):
        try:
            if not isinstance(request.user, Users):  
                return Response({
                    "statusCode": 401,
                    "status": False,
                    "message": "Unauthorized: Only normal users can submit reviews."
                }, status=status.HTTP_200_OK)

            company_id = request.data.get("company")
            if not company_id:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Company ID is required."
                }, status=status.HTTP_200_OK)

            try:
                company = CompanyDetails.objects.get(id=company_id)
            except CompanyDetails.DoesNotExist:
                return Response({
                    "statusCode": 404,
                    "status": False,
                    "message": "Company not found."
                }, status=status.HTTP_200_OK)

            if CompanyReview.objects.filter(company=company, user=request.user).exists():
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "You have already reviewed this company."
                }, status=status.HTTP_200_OK)
            serializer = self.get_serializer(data=request.data, context={"request": request})
            serializer.is_valid(raise_exception=True)
            serializer.save(company=company)  # Let the serializer handle the user

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Review submitted successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class CustomPagination(PageNumberPagination):
    page_size = 20 
    page_size_query_param = 'page_size'
    max_page_size = 100

class AllReviewListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        professional = request.user

        if not isinstance(professional, ProfessionalUser):
            logger.warning(f"Unauthorized access attempt by user ID {professional.id} - Not a ProfessionalUser.")
            return Response({
                "statusCode": 403,
                "status": False,
                "message": "Access denied: Only professional users can view reviews."
            }, status=status.HTTP_200_OK)

        company = getattr(professional, "company", None)
        if not company:
            logger.warning(f"Professional user ID {professional.id} has no associated company.")
            return Response({
                "statusCode": 404,
                "status": False,
                "message": "No company associated with this professional user."
            }, status=status.HTTP_200_OK)
        cache_key = f"all_reviews_company_{company.id}"
        cached_data = cache.get(cache_key)
        if cached_data:
            return Response(cached_data, status=status.HTTP_200_OK)

        try:
            products = Product.objects.filter(company=company).only('id', 'productname')
            product_reviews = Review.objects.filter(product__in=products).select_related('user', 'product').only(
                'id', 'rating', 'comment', 'created_at', 'user__username', 'product__productname'
            )
            company_reviews = CompanyReview.objects.filter(company=company).select_related('user').only(
                'id', 'rating', 'review_text', 'created_at', 'user__username'
            )

            all_reviews = []

            for review in product_reviews:
                all_reviews.append({
                    "type": "product",
                    "id":review.id,
                    "username": review.user.username,
                    "userId": review.user.id, 
                    "profile_img":review.user.profileImage.url if review.user.profileImage else None,
                    "product_name": review.product.productname,
                    "rating": review.rating,
                    "comment": review.comment,
                    "created_at": review.created_at,
                    "created_at_str": review.created_at.strftime("%Y-%m-%d %H:%M:%S")
                })

            for review in company_reviews:
                all_reviews.append({
                    "type": "company",
                    "id":review.id,
                    "username": review.user.username,
                    "userId": review.user.id, 
                    "profile_img":review.user.profileImage.url if review.user.profileImage else None,
                    "product_name": None,
                    "rating": float(review.rating),
                    "comment": review.review_text,
                    "created_at": review.created_at,
                    "created_at_str": review.created_at.strftime("%Y-%m-%d %H:%M:%S")
                })

            all_reviews.sort(key=lambda x: x["created_at"], reverse=True)
            for review in all_reviews:
                del review["created_at"]
                
            paginator = CustomPagination()
            paginated_reviews = paginator.paginate_queryset(all_reviews, request)

            response_data = {
                "statusCode": 200,
                "status": True,
                "message": "All reviews fetched successfully (latest first).",
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
                "total_pages": paginator.page.paginator.num_pages,
                "total_review": len(all_reviews),
                "reviews": paginated_reviews
            }
            cache.set(cache_key, response_data, timeout=300)

            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error while fetching reviews for user ID {professional.id}: {str(e)}", exc_info=True)
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An error occurred while fetching reviews."
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class ReviewersListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        professional = request.user

        if not isinstance(professional, ProfessionalUser):
            logger.warning(f"Unauthorized access attempt by user ID {professional.id} - Not a ProfessionalUser.")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Access denied: Only professional users can view reviews."
            }, status=status.HTTP_200_OK)

        company = getattr(professional, "company", None)
        if not company:
            company = CompanyDetails.objects.filter(userName=professional).first()

        if not company:
            logger.warning(f"Professional user ID {professional.id} has no associated company.")
            return Response({
                "statusCode": 404,
                "status": False,
                "message": "No company associated with this professional user."
            }, status=status.HTTP_200_OK)

        product_ids = Product.objects.filter(company=company).values_list('id', flat=True)

        product_review_user_ids = Review.objects.filter(product_id__in=product_ids).values_list('user_id', flat=True)
        company_review_user_ids = CompanyReview.objects.filter(company=company).values_list('user_id', flat=True)

        all_user_ids = set(product_review_user_ids).union(company_review_user_ids)

        users_qs = Users.objects.filter(id__in=all_user_ids).values('id', 'username', 'profileImage')
        users_list = []

        for user in users_qs:
            users_list.append({
                "id": user["id"],
                "username": user["username"],
                "profileImage": user.get("profileImage") or None  # ensures null if missing
            })

        return Response({
            "statusCode": 200,
            "status": True,
            "message": "User profiles fetched successfully",
            "users": users_list
        }, status=status.HTTP_200_OK)




class LocalFolderCreateAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            folder_name = request.data.get("name")

            if not folder_name:
                return Response({
                    "statusCode": 400, "status": False, "message": "Folder name is required"
                }, status=status.HTTP_200_OK)
            s3_client = boto3.client(
                "s3",
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME,
            )
            if ReelFolder.objects.filter(name=folder_name, user=request.user).exists():
                return Response({
                    "statusCode": 400, "status": False, "message": "Folder already exists"
                }, status=status.HTTP_200_OK)
            folder_path = f"uploads/{folder_name}/"  # Trailing slash to signify a folder
            s3_client.put_object(
                Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                Key=folder_path,
            )
            folder = ReelFolder.objects.create(user=request.user, name=folder_name)

            return Response({
                "statusCode": 201, "status": True, "message": "Folder created successfully",
                "folder": {
                    "id": folder.id,
                    "name": folder.name,
                    "path": f"http:/{settings.MEDIA_URL}{folder_path}"
                }
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({
                "statusCode": 500, "status": False, "message": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        


class CreateDeliveryServiceAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        try:
            professional_user = ProfessionalUser.objects.get(email=request.user.email)
            company = professional_user.company
            if not company:
                return Response(
                    {"statusCode": 403, "status": False, "message": "User is not linked to any company"},
                    status=status.HTTP_200_OK
                )
        except ProfessionalUser.DoesNotExist:
            return Response(
                {"statusCode": 400, "status": False, "message": "User not found"},
                status=status.HTTP_200_OK
            )

        services_data = request.data.get("services", [])

        if not services_data:
            return Response(
                {"statusCode": 400, "status": False, "message": "Please select at least one service"},
                status=status.HTTP_200_OK
            )

        created_services = []
        for service in services_data:
            service_type = service.get("service_type", "").strip().lower()
            is_enabled = service.get("is_enabled", False)
            delivery_fee = service.get("delivery_fee")
            minimum_order_amount = service.get("minimum_order_amount")
            travel_fee_per_km = service.get("travel_fee_per_km")

            if not service_type:
                return Response(
                    {"statusCode": 400, "status": False, "message": "Service type is required"},
                    status=status.HTTP_200_OK
                )
            if DeliveryService.objects.filter(company=company, service_type=service_type).exists():
                return Response(
                    {
                        "statusCode": 400,
                        "status": False,
                        "message": f"{service_type.capitalize()} service already exists. You can only update it."
                    },
                    status=status.HTTP_200_OK
                )

            try:
                delivery_service = DeliveryService.objects.create(
                    company=company,
                    service_type=service_type,
                    is_enabled=is_enabled,
                    delivery_fee=delivery_fee if service_type == "catering" else None,
                    minimum_order_amount=minimum_order_amount if service_type == "catering" else None,
                    travel_fee_per_km=travel_fee_per_km if service_type == "home_services" else None
                )
                created_services.append(DeliveryServiceSerializer(delivery_service).data)

            except IntegrityError:
                return Response(
                    {
                        "statusCode": 400,
                        "status": False,
                        "message": f"{service_type.capitalize()} service already exists for this company"
                    },
                    status=status.HTTP_200_OK
                )
        
class UpdateDeliveryServiceAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def put(self, request):
        try:
            professional_user = ProfessionalUser.objects.get(email=request.user.email)
            company = professional_user.company
            if not company:
                return Response(
                    {"statusCode": 403, "status": False, "message": "User is not linked to any company"},
                    status=status.HTTP_200_OK
                )
        except ProfessionalUser.DoesNotExist:
            return Response(
                {"statusCode": 400, "status": False, "message": "User not found"},
                status=status.HTTP_200_OK
            )

        services_data = request.data.get("services", [])

        if not services_data:
            return Response(
                {"statusCode": 400, "status": False, "message": "Please select at least one service to update"},
                status=status.HTTP_200_OK
            )

        updated_services = []
        for service in services_data:
            service_type = service.get("service_type", "").strip()
            is_enabled = service.get("is_enabled", False)
            delivery_fee = service.get("delivery_fee")
            minimum_order_amount = service.get("minimum_order_amount")
            travel_fee_per_km = service.get("travel_fee_per_km")

            if not service_type:
                return Response(
                    {"statusCode": 400, "status": False, "message": "Service type is required"},
                    status=status.HTTP_200_OK
                )

            try:
                delivery_service = DeliveryService.objects.get(company=company, service_type=service_type)
            except DeliveryService.DoesNotExist:
                return Response(
                    {"statusCode": 400, "status": False, "message": f"{service_type} service does not exist"},
                    status=status.HTTP_200_OK
                )
            if service_type == "catering":
                if delivery_fee is not None:
                    delivery_service.delivery_fee = delivery_fee
                if minimum_order_amount is not None:
                    delivery_service.minimum_order_amount = minimum_order_amount

            elif service_type == "home_services":
                if travel_fee_per_km is not None:
                    delivery_service.travel_fee_per_km = travel_fee_per_km

            delivery_service.is_enabled = is_enabled
            delivery_service.save()
            updated_services.append(DeliveryServiceSerializer(delivery_service).data)

        return Response(
            {
                "statusCode": 200,
                "status": True,
                "message": "Delivery services updated successfully",
                "data": updated_services,
            },
            status=status.HTTP_200_OK
        )

class GetDeliveryServiceAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            
            professional_user = ProfessionalUser.objects.get(email=request.user.email)
            company = professional_user.company

            if not company:
                return Response(
                    {"statusCode": 403, "status": False, "message": "User is not linked to any company"},
                    status=status.HTTP_200_OK
                )
        except ProfessionalUser.DoesNotExist:
            return Response(
                {"statusCode": 400, "status": False, "message": "User not found"},
                status=status.HTTP_200_OK
            )
        delivery_services = DeliveryService.objects.filter(company=company)
        if not delivery_services.exists():
            return Response(
                {
                    "statusCode": 404,
                    "status": False,
                    "message": "No delivery services found for this company",
                    "company": {
                        "id": company.id,
                        "name": company.companyName
                    },
                    "data": []
                },
                status=status.HTTP_200_OK
            )
        serialized_data = DeliveryServiceSerializer(delivery_services, many=True).data

        return Response(
            {
                "statusCode": 200,
                "status": True,
                "message": "Delivery services retrieved successfully",
                "data": serialized_data
            },
            status=status.HTTP_200_OK
        )
class PerformanceMetricsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            professional_user = ProfessionalUser.objects.get(email=request.user.email)

            try:
                company = professional_user.company
            except AttributeError:
                return Response({"error": "Company details not found for this user"}, status=200)

            total_visits = company.total_visits
            all_orders = Order.objects.filter(company=company)
            completed_orders = all_orders.filter(orderStatus='Fulfilled', is_paid=True)

            total_completed_orders = completed_orders.count()
            total_orders = all_orders.count()
            completed_percentage = (total_completed_orders / total_orders * 100) if total_orders > 0 else 0
            completed_percentage = round(completed_percentage, 2)
            total_sales = completed_orders.aggregate(total_sales=Sum('total_price'))['total_sales'] or 0.00

            return Response({
                'statusCode':200,
                'status':True,
                'message':'data retrived successfully',
                "user_details": {
                    "id": professional_user.id,
                    "email": professional_user.email,
                    "role": str(professional_user.role)
                },
                "company_details": {
                    "id": company.id,
                    "company_name": company.companyName,
                    "manager_name": company.managerFullName,
                    "email": company.email,
                    "phone": company.phoneNumber
                },
                "performance_metrics": {
                    "total_visits": total_visits,
                    "total_completed_orders": total_completed_orders,
                    "total_orders": total_orders,
                    "completed_order_percentage": completed_percentage,
                    "total_sales": str(total_sales)
                }
            })

        except ProfessionalUser.DoesNotExist:
            return Response({"error": "Only Professional Users can access this data"}, status=200)
        



def get_percentage(part, total):
    try:
        return round((float(part) / float(total)) * 100, 2) if total > 0 else 0.0
    except (ZeroDivisionError, ValueError, TypeError):
        return 0.0


def make_naive_aware(dt):
    if timezone.is_naive(dt):
        return timezone.make_aware(dt)
    return dt

class OrderPerformanceView(APIView):
    permission_classes = [IsAuthenticated]  

    def get(self, request):
        try:
            professional_user = ProfessionalUser.objects.get(email=request.user.email)
        except ProfessionalUser.DoesNotExist:
            return Response({"error": "Professional user not found"}, status=200)
        
        try:
            company = professional_user.company
        except AttributeError:
            return Response({"error": "Company details not found for this user"}, status=200)

        today = timezone.now()
        orders = Order.objects.filter(is_paid=True, company=company)
        month_param = request.query_params.get("month")
        year_param = request.query_params.get("year")

        if month_param:
            try:
                month = int(month_param)
                year = int(year_param) if year_param else today.year

                if 1 <= month <= 12:
                    start_of_month = timezone.make_aware(datetime(year, month, 1))
                    if month == 12:
                        end_of_month = timezone.make_aware(datetime(year + 1, 1, 1))
                    else:
                        end_of_month = timezone.make_aware(datetime(year, month + 1, 1))
                    orders = orders.filter(created_at__gte=start_of_month, created_at__lt=end_of_month)
                else:
                    return Response({"error": "Invalid month. Must be between 1 and 12."}, status=200)
            except ValueError:
                return Response({"error": "Invalid month or year. Must be numeric."}, status=200)
        
        year_param = request.query_params.get("year")
        if year_param:
            try:
                year = int(year_param)
                start_of_year = timezone.make_aware(datetime(year, 1, 1))
                end_of_year = timezone.make_aware(datetime(year + 1, 1, 1))
                orders = orders.filter(created_at__gte=start_of_year, created_at__lt=end_of_year)
            except ValueError:
                return Response({"error": "Invalid year. Must be numeric."}, status=200)

        total_orders = orders.count()

        def format_data(count):
            return {"value": count, "dataPointText": str(count)}

        order_type_mapping = {
            'onsite': 'onsite',
            'click_collect': 'clickcollect',
            'delivery': 'delivery',
        }
        start_of_week = make_naive_aware((today - timedelta(days=today.weekday())).replace(hour=0, minute=0, second=0, microsecond=0))
        weekly_graph_data = {key: [{"value": 0, "dataPointText": "0"}] for key in order_type_mapping.values()}
        weekly_totals = {key: 0 for key in order_type_mapping.values()}

        for i in range(7):
            day = start_of_week + timedelta(days=i)
            start_day = make_naive_aware(day.replace(hour=0, minute=0, second=0, microsecond=0))
            end_day = make_naive_aware((day + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0))
            for db_type, key in order_type_mapping.items():
                count = orders.filter(created_at__gte=start_day, created_at__lt=end_day, order_type=db_type).count()
                weekly_totals[key] += count
                weekly_graph_data[key].append(format_data(count))
        monthly_graph_data = {key: [{"value": 0, "dataPointText": "0"}] for key in order_type_mapping.values()}
        monthly_totals = {key: 0 for key in order_type_mapping.values()}

        for i in range(4):
            start_week = make_naive_aware((today - timedelta(weeks=3 - i)).replace(hour=0, minute=0, second=0, microsecond=0))
            end_week = start_week + timedelta(days=7)
            for db_type, key in order_type_mapping.items():
                count = orders.filter(created_at__gte=start_week, created_at__lt=end_week, order_type=db_type).count()
                monthly_totals[key] += count
                monthly_graph_data[key].append(format_data(count))
        yearly_graph_data = {key: [{"value": 0, "dataPointText": "0"}] for key in order_type_mapping.values()}
        yearly_totals = {key: 0 for key in order_type_mapping.values()}

        for i in range(1, 13):
            start_month = make_naive_aware(today.replace(month=i, day=1, hour=0, minute=0, second=0, microsecond=0))
            if i == 12:
                end_month = make_naive_aware(today.replace(year=today.year + 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0))
            else:
                end_month = make_naive_aware(today.replace(month=i + 1, day=1, hour=0, minute=0, second=0, microsecond=0))

            for db_type, key in order_type_mapping.items():
                count = orders.filter(created_at__gte=start_month, created_at__lt=end_month, order_type=db_type).count()
                yearly_totals[key] += count
                yearly_graph_data[key].append(format_data(count))

        return Response({
            "statusCode": 200,
            "status": True,
            "message": "Order performance data retrieved successfully",
            "total_orders": total_orders,
            "weekly_graph_data": weekly_graph_data,
            "monthly_graph_data": monthly_graph_data,
            "yearly_graph_data": yearly_graph_data,
            "weekly_totals": weekly_totals,
            "monthly_totals": monthly_totals,
            "yearly_totals": yearly_totals
        })

class RevenuePerformanceView(APIView):
    permission_classes = [IsAuthenticated]  

    def get(self, request):
        try:
            professional_user = ProfessionalUser.objects.get(email=request.user.email)
        except ProfessionalUser.DoesNotExist:
            return Response({"error": "Professional user not found"}, status=200)
        try:
                company = professional_user.company
        except AttributeError:
                return Response({"error": "Company details not found for this user"}, status=200)
        today = timezone.now()
        orders = Order.objects.filter(is_paid=True, company=company)
        month_param = request.query_params.get("month")
        year_param = request.query_params.get("year")
        if month_param:
            try:
                month = int(month_param)
                year = int(year_param) if year_param else today.year

                if 1 <= month <= 12:
                    start_of_month = timezone.make_aware(datetime(year, month, 1))
                    if month == 12:
                        end_of_month = timezone.make_aware(datetime(year + 1, 1, 1))
                    else:
                        end_of_month = timezone.make_aware(datetime(year, month + 1, 1))
                    orders = orders.filter(created_at__gte=start_of_month, created_at__lt=end_of_month)
                else:
                    return Response({"error": "Invalid month. Must be between 1 and 12."}, status=200)
            except ValueError:
                return Response({"error": "Invalid month or year. Must be numeric."}, status=200)
        if year_param:
            try:
                year = int(year_param)
                start_of_year = timezone.make_aware(datetime(year, 1, 1))
                end_of_year = timezone.make_aware(datetime(year + 1, 1, 1))
                orders = orders.filter(created_at__gte=start_of_year, created_at__lt=end_of_year)
            except ValueError:
                return Response({"error": "Invalid year. Must be numeric."}, status=200)

        total_revenue = orders.aggregate(total=Sum("total_price"))["total"] or 0

        def format_data(amount):
            return {"value": float(amount), "dataPointText": str(round(amount, 2))}

        order_type_mapping = {
            'onsite': 'onsite',
            'click_collect': 'clickcollect',
            'delivery': 'delivery',
        }
        start_of_week = make_naive_aware((today - timedelta(days=today.weekday())).replace(hour=0, minute=0, second=0, microsecond=0))
        weekly_graph_data = {key: [{"value": 0, "dataPointText": "0"}] for key in order_type_mapping.values()}
        weekly_totals = {key: 0 for key in order_type_mapping.values()}

        for i in range(7):
            day = start_of_week + timedelta(days=i)
            start_day = make_naive_aware(day.replace(hour=0, minute=0, second=0, microsecond=0))
            end_day = make_naive_aware((day + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0))
            for db_type, key in order_type_mapping.items():
                amount = orders.filter(created_at__gte=start_day, created_at__lt=end_day, order_type=db_type).aggregate(total=Sum("total_price"))["total"] or 0
                weekly_totals[key] += amount
                weekly_graph_data[key].append(format_data(amount))
        monthly_graph_data = {key: [{"value": 0, "dataPointText": "0"}] for key in order_type_mapping.values()}
        monthly_totals = {key: 0 for key in order_type_mapping.values()}

        for i in range(4):
            start_week = make_naive_aware((today - timedelta(weeks=3 - i)).replace(hour=0, minute=0, second=0, microsecond=0))
            end_week = start_week + timedelta(days=7)
            for db_type, key in order_type_mapping.items():
                amount = orders.filter(created_at__gte=start_week, created_at__lt=end_week, order_type=db_type).aggregate(total=Sum("total_price"))["total"] or 0
                monthly_totals[key] += amount
                monthly_graph_data[key].append(format_data(amount))
        yearly_graph_data = {key: [{"value": 0, "dataPointText": "0"}] for key in order_type_mapping.values()}
        yearly_totals = {key: 0 for key in order_type_mapping.values()}

        for i in range(1, 13):
            start_month = make_naive_aware(today.replace(month=i, day=1, hour=0, minute=0, second=0, microsecond=0))
            if i == 12:
                end_month = make_naive_aware(today.replace(year=today.year + 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0))
            else:
                end_month = make_naive_aware(today.replace(month=i + 1, day=1, hour=0, minute=0, second=0, microsecond=0))

            for db_type, key in order_type_mapping.items():
                amount = orders.filter(created_at__gte=start_month, created_at__lt=end_month, order_type=db_type).aggregate(total=Sum("total_price"))["total"] or 0
                yearly_totals[key] += amount
                yearly_graph_data[key].append(format_data(amount))

        return Response({
            "statusCode": 200,
            "status": True,
            "message": "Revenue performance data retrieved successfully",
            "total_revenue": round(total_revenue, 2),
            "weekly_graph_data": weekly_graph_data,
            "monthly_graph_data": monthly_graph_data,
            "yearly_graph_data": yearly_graph_data,
            "weekly_totals": {k: round(v, 2) for k, v in weekly_totals.items()},
            "monthly_totals": {k: round(v, 2) for k, v in monthly_totals.items()},
            "yearly_totals": {k: round(v, 2) for k, v in yearly_totals.items()},
        })
    
    
class OrderChartView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            professional_user = ProfessionalUser.objects.get(email=request.user.email)
        except ProfessionalUser.DoesNotExist:
            return Response({"error": "Professional user not found"}, status=200)

        today = timezone.now()
        yesterday = today - timedelta(days=1)
        start_today = today.replace(hour=0, minute=0, second=0, microsecond=0)
        start_yesterday = yesterday.replace(hour=0, minute=0, second=0, microsecond=0)
        months_param = request.query_params.get('months')
        try:
            if months_param:
                months_to_check = [int(m.strip()) for m in months_param.split(',') if m.strip().isdigit() and 1 <= int(m.strip()) <= 12]
                if not months_to_check:
                    return Response({"error": "Invalid months parameter"}, status=200)
            else:
                months_to_check = list(range(1, 13))
        except ValueError:
            return Response({"error": "Invalid months format. Use comma-separated numbers."}, status=200)
        today_orders_qs = Order.objects.filter(
            professional_user=professional_user,
            is_paid=True,
            created_at__gte=start_today
        )
        yesterday_orders_qs = Order.objects.filter(
            professional_user=professional_user,
            is_paid=True,
            created_at__gte=start_yesterday,
            created_at__lt=start_today
        )

        today_orders = today_orders_qs.count()
        yesterday_orders = yesterday_orders_qs.count()
        today_percent_change = get_percentage(today_orders - yesterday_orders, yesterday_orders)
        today_sales = today_orders_qs.aggregate(total=Sum('total_price'))['total'] or 0.0
        yesterday_sales = yesterday_orders_qs.aggregate(total=Sum('total_price'))['total'] or 0.0
        today_sales_percent_change = get_percentage(today_sales - yesterday_sales, yesterday_sales)
        monthly_data = {}
        total_sales = 0
        total_orders = 0
        statuses = ['created', 'accepted', 'paid', 'pending', 'on hold', 'completed', 'cancelled']

        for m in months_to_check:
            try:
                month_name = datetime(today.year, m, 1).strftime('%B')
                start_date = timezone.make_aware(datetime(today.year, m, 1))
                if m == 12:
                    end_date = timezone.make_aware(datetime(today.year + 1, 1, 1))
                else:
                    end_date = timezone.make_aware(datetime(today.year, m + 1, 1))

                orders = Order.objects.filter(
                    professional_user=professional_user,
                    is_paid=True,
                    created_at__gte=start_date,
                    created_at__lt=end_date
                )

                data = {status: orders.filter(orderStatus=status).count() for status in statuses}
                revenue = orders.aggregate(sales=Sum('total_price'))['sales'] or 0.0
                count = orders.count()

                total_sales += float(revenue)
                total_orders += count

                monthly_data[month_name] = data
            except Exception:
                continue

        return Response({
            "status": True,
            "message": "Data fetched successfully",
            "today": {
                "orders": today_orders,
                "percent_change": today_percent_change,
                "sales": float(today_sales),
                "sales_percent_change": today_sales_percent_change
            },
            "summary": {
                "total_sales": float(total_sales),
                "number_of_orders": total_orders
            },
            "monthly_data": monthly_data
        })


# -----------------------------------dashboard records -------------------------------------


class DashboardIncomeSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        current_date = now().date()
        current_time = now()

        try:
            professional_user = ProfessionalUser.objects.get(id=request.user.id)
            company = professional_user.company

            if not company:
                return Response({
                    "status": False,
                    "message": "No company associated with this professional user."
                }, status=404)

            # ----------------------------
            # GET REQUEST PARAMS
            # ----------------------------
            filter_type = request.GET.get("filter_type", "today")
            date_type = request.GET.get("date_type", "day")
            start_date = request.GET.get("start_date")
            end_date = request.GET.get("end_date")
            fulfilled_status = request.GET.get("fulfilled_status")

            # ----------------------------
            # PAID ORDERS FILTERING
            # ----------------------------
            income_filter = {
                "is_paid": True,
                "company": company
            }
            if filter_type == "today":
                income_filter["created_at__date"] = current_date
            if fulfilled_status:
                income_filter["orderStatus"] = fulfilled_status
            paid_orders = Order.objects.filter(**income_filter)
            income_total = paid_orders.aggregate(total=Sum("total_price"))["total"] or 0
            paid_order_count = paid_orders.count()
            avg_ticket_size = income_total / paid_order_count if paid_order_count > 0 else 0

            # ----------------------------
            # TOTAL ORDER COUNT
            # ----------------------------
            all_orders_filter = {"company": company}
            if filter_type == "today":
                all_orders_filter["created_at__date"] = current_date

            all_orders = Order.objects.filter(**all_orders_filter)
            all_order_count = all_orders.count()

            # ----------------------------
            # INCOME GRAPH DATA
            # ----------------------------
            trunc_func = TruncDay if date_type == "day" else TruncMonth
            graph_filter = {
                "is_paid": True,
                "company": company
            }
            if start_date and end_date:
                graph_filter["created_at__date__range"] = [start_date, end_date]
            graph_data = (
                Order.objects.filter(**graph_filter)
                .annotate(period=trunc_func("created_at"))
                .values("period")
                .annotate(total_income=Sum("total_price"))
                .order_by("period")
            )
            income_graph = [
                {
                    "date": entry["period"].strftime("%Y-%m-%d"),
                    "total_income": round(entry["total_income"], 2)
                }
                for entry in graph_data
            ]

            # ----------------------------
            # TODAY ORDER GRAPH
            # ----------------------------
            today_order_graph = []
            if filter_type == "today":
                today_order_values = paid_orders.order_by("created_at").values_list("total_price", flat=True)
                today_order_graph = [
                    {f"order{i+1}": round(price, 2)} for i, price in enumerate(today_order_values)
                ]

            # ----------------------------
            # UNREAD MESSAGES
            # ----------------------------
            new_messages_count = Message.objects.filter(
                recipient=request.user,
                is_read=False
            ).count()



            # Filter active campaigns for the company
            active_ads = AdvertiseCampaign.objects.filter(
                company=company,
                startDateTime__lte=current_time,
                endDateTime__gte=current_time
            )

            # Count total active paid ads
            total_paid_ads = active_ads.count()

            # Calculate today's income from paid ads
            paid_ads_income_today = Decimal(0)

            for ad in active_ads:
                if ad.bid_type == 'cpc':
                    paid_ads_income_today += Decimal(ad.today_clicks) * ad.max_bid
                elif ad.bid_type == 'cpm':
                    paid_ads_income_today += Decimal(ad.today_impressions) / 1000 * ad.max_bid

            # ----------------------------
            # ACTIVE PROMOTIONS
            # ----------------------------
            active_promotions = Promotions.objects.filter(
                company=company,
                startDateTime__lte=current_time,
                endDateTime__gte=current_time
            ).count()


            # ----------------------------
            # MENU COUNT
            # ----------------------------
            menu_count = 0
            if company.on_site_ordering:
                try:
                    menu_list = json.loads(company.on_site_ordering) if isinstance(company.on_site_ordering, str) else company.on_site_ordering
                    if isinstance(menu_list, list):
                        menu_count = len(menu_list)
                except Exception:
                    menu_count = 0

            # ----------------------------
            # BOOKINGS COUNT & INCOME
            # ----------------------------
            
            today = date.today()

            def get_booking_stats(model):
                # Booking count: all created today
                count = model.objects.filter(created_at__date=today,is_paid=True,).count()

                # Booking income: is_paid=True and booking_date=today
                price_field = 'price' if 'price' in [f.name for f in model._meta.fields] else 'total_price'
                income = model.objects.filter(
                    is_paid=True,
                    created_at__date=today
                ).aggregate(total=Sum(price_field))["total"] or 0

                return count, income


            total_booking_count = 0
            today_booking_income = 0

            for model in [
                RoomBooking, eventBooking, experienceBooking,
                slotBooking, aestheticsBooking, relaxationBooking,artandcultureBooking
            ]:
                count, income = get_booking_stats(model)
                total_booking_count += count
                today_booking_income += income

            # ----------------------------
            # FINAL RESPONSE
            # ----------------------------
            return Response({
                "status": True,
                "message": "Dashboard income summary retrieved successfully",
                "data": {
                    "total_orders_today" if filter_type == "today" else "total_orders": paid_order_count,
                    "today_orders_income" if filter_type == "today" else "total_income": round(income_total, 2),
                    "total_reserverstions_today": total_booking_count,
                    "today_reservations_income": round(today_booking_income, 2),
                    "net_income_today" if filter_type == "today" else "net_income": round((income_total + today_booking_income), 2),
                    "total_income_graph": income_graph,
                    "today_order_graph": today_order_graph,
                    "new_messages": new_messages_count,
                    "active_promotions_today": active_promotions,
                    # "Paid_Advertising": str(round(paid_ads_income_today, 2)),  # e.g., "56.78"
                    # "total_paid_ads": total_paid_ads,

                    "Paid_Advertising": "11K",
                    "menu": menu_count,
                }
            })
        except ProfessionalUser.DoesNotExist:
            return Response({
                "status": False,
                "message": "Professional user not found."
            }, status=404)


CURRENCY_SYMBOLS = {
    "EUR": "€","USD": "$","INR": "₹",
    "GBP": "£","JPY": "¥","CAD": "C$",}
    

class ProfessionalUserTransactionView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            user = request.user
            professional_user = ProfessionalUser.objects.filter(email=user.email).first()
            if not professional_user:
                return Response({
                    "status": False,
                    "statusCode": 400,
                    "message": "Professional user not found."
                }, status=status.HTTP_400_BAD_REQUEST)

            transactions = ProfessionalUserTransactionLog.objects.filter(paid_to=professional_user).order_by('-created_at')[:10]

            data = []
            total_balance = Decimal('0.00')

            for txn in transactions:
                currency = txn.payment.currency if hasattr(txn.payment, "currency") and txn.payment.currency else "EUR"
                symbol = CURRENCY_SYMBOLS.get(currency.upper(), "")
                paid_by_user = txn.paid_by
                amount = txn.amount
                txn_type = txn.status  # "credited" or "debited"
                sign = "+" if txn_type == "credited" else "-"
                
                if txn_type == "credited":
                    total_balance += amount
                else:
                    total_balance -= amount

                data.append({
                    "id": txn.id,
                    "user": {
                        "id": paid_by_user.id,
                        "name": f"{paid_by_user.firstName} {paid_by_user.lastName}",
                        "email": paid_by_user.email
                    },
                    "order_id": txn.order.order_id if txn.order else None,
                    "currency": currency,
                    "currency_symbol": symbol,
                    "amount": str(amount),
                    "base_price": str(txn.base_price),
                    "discount": str(txn.discount),
                    "tax": str(txn.tax),
                    "payment_mode": txn.payment_mode,
                    "product_summary": txn.product_summary,
                    "status": txn_type,
                    "transaction_symbol": sign,
                    "created_at": txn.created_at,
                    "transactionId":txn.transaction_id
                })

            return Response({
                "status": True,
                "statusCode": 200,
                "message": "Transaction logs fetched successfully.",
                "data": data,
                "total_balance": str(total_balance)
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "status": False,
                "statusCode": 500,
                "message": f"Error: {str(e)}",
                "data": [],
                "total_balance": "0.00"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


from django.http import HttpResponse
from reportlab.pdfgen import canvas


class ExportDataAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        professional_user = request.user

        from_date = parse_date(request.data.get('from_date'))
        to_date = parse_date(request.data.get('to_date'))
        export_format = request.data.get('export_format', 'CSV').upper()  # Default to CSV

        if not from_date or not to_date:
            return Response({
                "status": False,
                "message": "Please provide both from_date and to_date in the format YYYY-MM-DD."
            }, status=status.HTTP_200_OK)
        queryset = ProfessionalUserTransactionLog.objects.filter(
            paid_to=professional_user,
            created_at__date__gte=from_date,
            created_at__date__lte=to_date
        )
        if export_format == "CSV":
            return self.export_csv(queryset)
        elif export_format == "EXCEL":
            return self.export_excel(queryset)
        elif export_format == "PDF":
            return self.export_pdf(queryset)
        else:
            return Response({
                "status": False,
                "message": "Invalid export_format. Choose CSV, EXCEL, or PDF."
            }, status=status.HTTP_200_OK)

    def export_csv(self, queryset):

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow([
            "ID", "Status", "Amount", "Order ID", "User ID", 
            "Username", "User Email", "Transaction ID", 
            "Payment Mode", "Created At"
        ])
        for payment in queryset:
            writer.writerow([
                payment.id,
                payment.status,
                str(payment.amount),
                payment.order.id if payment.order else "",
                payment.paid_by.id if payment.paid_by else "",
                payment.paid_by.username if payment.paid_by else "",
                payment.paid_by.email if payment.paid_by else "",
                payment.transaction_id,
                payment.payment_mode,
                payment.created_at.strftime("%Y-%m-%d %H:%M:%S")
            ])

        buffer.seek(0)
        response = Response(buffer.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=transactions.csv'
        return response


    def export_excel(self, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        headers = [
            "ID", "Status", "Amount", "Order ID", "User ID",
            "Username", "User Email", "Transaction ID",
            "Payment Mode", "Created At"
        ]
        ws.append(headers)

        for payment in queryset:
            ws.append([
                payment.id,
                payment.status,
                str(payment.amount),
                payment.order.id if payment.order else "",
                payment.paid_by.id if payment.paid_by else "",
                payment.paid_by.username if payment.paid_by else "",
                payment.paid_by.email if payment.paid_by else "",
                payment.transaction_id,
                payment.payment_mode,
                payment.created_at.strftime("%Y-%m-%d %H:%M:%S")
            ])
        for i, column_cells in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value) or "") for cell in column_cells)
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=transactions.xlsx'
        wb.save(response)
        return response

    def export_pdf(self, queryset):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=transactions.pdf'

        p = canvas.Canvas(response, pagesize=letter)
        width, height = letter
        x_margin = 50
        y_margin = 750
        line_height = 15

        headers = [
            "ID", "Status", "Amount", "Order ID", "User ID",
            "Username", "User Email", "Transaction ID",
            "Payment Mode", "Created At"
        ]
        for i, header in enumerate(headers):
            p.drawString(x_margin + i*70, y_margin, header)

        y = y_margin - line_height

        for payment in queryset:
            row = [
                str(payment.id),
                str(payment.status),
                str(payment.amount),
                str(payment.order.id) if payment.order else "",
                str(payment.paid_by.id) if payment.paid_by else "",
                payment.paid_by.username if payment.paid_by else "",
                payment.paid_by.email if payment.paid_by else "",
                payment.transaction_id,
                payment.payment_mode,
                payment.created_at.strftime("%Y-%m-%d %H:%M:%S")
            ]

            for i, cell in enumerate(row):
                p.drawString(x_margin + i*70, y, cell)

            y -= line_height
            if y < 50:
                p.showPage()
                y = y_margin

        p.save()
        return response
    
"""

class ExportDataAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        professional_user = request.user

        from_date = parse_date(request.data.get('from_date'))
        to_date = parse_date(request.data.get('to_date'))
        export_types = request.data.get('export_types', [])  # e.g., ["Transactions", "Revenue", "Fees", "Tax"]
        export_format = request.data.get('export_format', 'JSON').upper()

        if not from_date or not to_date:
            return Response({
                "status": False,
                "message": "Please provide both from_date and to_date in YYYY-MM-DD format."
            }, status=status.HTTP_200_OK)

        valid_types = {"TRANSACTIONS", "REVENUE", "FEES", "TAX"}
        requested_types = set([t.upper() for t in export_types])

        if not requested_types.issubset(valid_types):
            return Response({
                "status": False,
                "message": f"Invalid export_types. Valid options: {', '.join(valid_types)}."
            }, status=status.HTTP_200_OK)

        queryset = UserPayment.objects.filter(
            professional_user=professional_user,
            status='succeeded',
            created_at__date__gte=from_date,
            created_at__date__lte=to_date
        )

        if 'TRANSACTIONS' not in requested_types:
            type_filters = Q()
            if 'REVENUE' in requested_types:
                type_filters |= Q(type='Revenue')
            if 'FEES' in requested_types:
                type_filters |= Q(type='Fees')
            if 'TAX' in requested_types:
                type_filters |= Q(type='Tax')
            queryset = queryset.filter(type_filters)

        if export_format == "JSON":
            return self.export_json(queryset)
        elif export_format == "CSV":
            return self.export_csv(queryset)
        elif export_format == "EXCEL":
            return self.export_excel(queryset)
        elif export_format == "PDF":
            return self.export_pdf(queryset)
        else:
            return Response({
                "status": False,
                "message": "Invalid export_format. Use JSON, CSV, EXCEL, or PDF."
            }, status=status.HTTP_200_OK)

    def export_json(self, queryset):
        serialized_data = UserPaymentTransactionSerializer(queryset, many=True).data
        return Response({
            "status": True,
            "statusCode": 200,
            "message": "Exported data successfully.",
            "data": serialized_data
        }, status=status.HTTP_200_OK)

    def export_csv(self, queryset):
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(["ID", "Amount", "Transaction Type", "Created At"])

        for payment in queryset:
            writer.writerow([
                payment.id,
                str(payment.amount),
                payment.type,
                payment.created_at.strftime("%Y-%m-%d %H:%M:%S")
            ])

        buffer.seek(0)
        file_content = ContentFile(buffer.getvalue().encode('utf-8'))
        file_path = f"exports/transactions_{now().strftime('%Y%m%d%H%M%S')}.csv"
        default_storage.save(file_path, file_content)
        delete_s3_file.apply_async(args=[file_path], eta=now() + timedelta(hours=24))

        download_url = default_storage.url(file_path)
        return Response({
            "status": True,
            "statusCode": 200,
            "message": "CSV Exported successfully.",
            "download_url": download_url
        }, status=status.HTTP_200_OK)

    def export_excel(self, queryset):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "ExportedData"

        headers = ["ID", "Amount", "Transaction Type", "Created At"]
        sheet.append(headers)

        for payment in queryset:
            sheet.append([
                payment.id,
                float(payment.amount),
                payment.type,
                payment.created_at.strftime("%Y-%m-%d %H:%M:%S")
            ])

        for col in sheet.columns:
            max_length = max((len(str(cell.value)) for cell in col), default=10)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        file_content = ContentFile(buffer.getvalue())
        file_path = f"exports/transactions_{now().strftime('%Y%m%d%H%M%S')}.xlsx"
        default_storage.save(file_path, file_content)
        delete_s3_file.apply_async(args=[file_path], eta=now() + timedelta(hours=24))

        download_url = default_storage.url(file_path)
        return Response({
            "status": True,
            "statusCode": 200,
            "message": "Excel Exported successfully.",
            "download_url": download_url
        }, status=status.HTTP_200_OK)

    def export_pdf(self, queryset):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                                rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)

        elements = []
        data = [["ID", "Amount", "Transaction Type", "Created At"]]  # Header row
        for payment in queryset:
            data.append([
                str(payment.id),
                str(payment.amount),
                payment.type,
                payment.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        num_cols = len(data[0])
        table_width = 750  # adjust based on page size
        col_width = table_width / num_cols

        table = Table(data, colWidths=[col_width] * num_cols)
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # grid for every cell
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#DCE6F1")])
        ])
        table.setStyle(style)

        elements.append(table)
        doc.build(elements)

        buffer.seek(0)

        file_content = ContentFile(buffer.getvalue())
        file_path = f"exports/transactions_{now().strftime('%Y%m%d%H%M%S')}.pdf"
        default_storage.save(file_path, file_content)
        delete_s3_file.apply_async(args=[file_path], eta=now() + timedelta(hours=24))

        download_url = default_storage.url(file_path)
        return Response({
            "status": True,
            "statusCode": 200,
            "message": "PDF Exported successfully in full excel-like grid format!",
            "download_url": download_url
        }, status=status.HTTP_200_OK)
"""


# ---offers  views---

class CreateOfferAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, *args, **kwargs):
        try:
            serializer = OfferSerializer(data=request.data)
            if serializer.is_valid():
                offer = serializer.save()
                return Response({
                    'status': True,
                    "statusCode": 201,
                    'message': 'Offer created successfully.',
                    'data': serializer.data
                }, status=status.HTTP_201_CREATED)
            else:
                return Response({
                    'status': False,
                    "statusCode": 400,
                    'message': 'Validation failed.',
                    'errors': serializer.errors
                }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Offer creation failed: {str(e)}", exc_info=True)
            return Response({
                'status': False,
                "statusCode": 500,
                'message': 'Something went wrong while creating the offer.',
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class OfferListAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        try:
            # Handle search
            search_query = request.query_params.get('search', '').strip()
            offers_qs = Offer.objects.all()

            if search_query:
                offers_qs = offers_qs.filter(
                    Q(title__icontains=search_query) |
                    Q(description__icontains=search_query)
                )

            offers_qs = offers_qs.order_by('-id')

            # Pagination logic
            page = int(request.query_params.get('page', 1))
            page_size = int(request.query_params.get('page_size', 10))
            total = offers_qs.count()
            total_pages = ceil(total / page_size)

            offset = (page - 1) * page_size
            paginated_qs = offers_qs[offset:offset + page_size]

            serializer = OfferSerializer(paginated_qs, many=True)

            # Build next/previous URLs
            def build_page_url(p):
                if p:
                    return f"{request.path}?page={p}&page_size={page_size}&search={search_query}"
                return None

            next_url = build_page_url(page + 1) if page < total_pages else None
            previous_url = build_page_url(page - 1) if page > 1 else None

            return Response({
                "status": True,
                "statusCode": 200,
                "message": "Offer list retrieved successfully.",
                "total": total,
                "page": page,
                "total_pages": total_pages,
                "next": next_url,
                "previous": previous_url,
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error fetching offers: {str(e)}", exc_info=True)
            return Response({
                "status": False,
                "statusCode": 500,
                "message": "Something went wrong while fetching offers.",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)       


class UpdateOfferAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, offer_id, *args, **kwargs):
        try:
            offer = Offer.objects.get(id=offer_id)
            serializer = OfferSerializer(offer, data=request.data, partial=False)  # Full update
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'status': True,
                    "statusCode": 200,
                    'message': 'Offer updated successfully.',
                    'data': serializer.data
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    'status': False,
                    "statusCode": 400,
                    'message': 'Validation failed.',
                    'errors': serializer.errors
                }, status=status.HTTP_200_OK)

        except Offer.DoesNotExist:
            return Response({
                'status': False,
                "statusCode": 404,
                'message': 'Offer not found.',
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                "statusCode":500,
                'message': 'Something went wrong while updating the offer.',
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    def patch(self, request, offer_id, *args, **kwargs):
        try:
            offer = Offer.objects.get(id=offer_id)
            serializer = OfferSerializer(offer, data=request.data, partial=True)  # Partial update
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'status': True,
                    'message': 'Offer updated successfully.',
                    'data': serializer.data
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    'status': False,
                    'message': 'Validation failed.',
                    'errors': serializer.errors
                }, status=status.HTTP_200_OK)

        except Offer.DoesNotExist:
            return Response({
                'status': False,
                'message': 'Offer not found.',
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': 'Something went wrong while updating the offer.',
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)    
        
class DeleteOfferAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, offer_id, *args, **kwargs):
        try:
            offer = Offer.objects.get(id=offer_id)
            offer_data = OfferSerializer(offer).data  # Serialize the offer data before deletion
            offer.delete()  # Delete the offer
            return Response({
                'status': True,
                'statusCode': 200,
                'message': 'Offer deleted successfully.',
                'deleted_data': offer_data  # Include the deleted offer data
            }, status=status.HTTP_200_OK)  # Changed to 200 OK, as 204 shouldn't return content

        except Offer.DoesNotExist:
            return Response({
                'status': False,
                'statusCode': 404,
                'message': 'Offer not found.',
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'statusCode': 500,
                'message': 'Something went wrong while deleting the offer.',
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        



class EmployeeCreateView(generics.CreateAPIView):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        try:
            user = request.user
            professional_user = get_object_or_404(ProfessionalUser, email=user.email)
            company = professional_user.company
            data = request.data.copy()
            data['company'] = company.id  # Assign company ID to data
            email = data.get('email')
            if Employee.objects.filter(email=email).exists():
                return Response({
                    "statusCode": 200,
                    "status": False,
                    "message": "An employee with this email already exists.",
                   
                }, status=status.HTTP_200_OK)

            serializer = self.get_serializer(data=data)
            if serializer.is_valid():
                self.perform_create(serializer)
                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Employee created successfully",
                    "data": serializer.data
                }, status=status.HTTP_200_OK)  # 201 CREATED is better
            else:
                return Response({
                    "statusCode": 200,
                    "status": False,
                    "message": "Failed to create employee",
                    "errors": serializer.errors
                }, status=status.HTTP_200_OK)  # 400 BAD REQUEST is better

        except Exception as e:
            return Response({
                'status': False,
                'statusCode': 500,
                'message': 'Something went wrong while creating the employee.',
                'error': str(e)
            }, status=status.HTTP_200_OK)
from django.core.paginator import EmptyPage, Paginator


from django.core.paginator import Paginator, EmptyPage

class EmployeeListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        user = request.user
        search_query = request.query_params.get('search', '')
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 6))

        if user.company:
            queryset = Employee.objects.filter(company=user.company).order_by('-created_at')
        else:
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "No company associated with the user.",
                "data": []
            })
        if search_query:
            queryset = queryset.filter(
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(email__icontains=search_query)
            )
        paginator = Paginator(queryset, page_size)
        try:
            paginated_data = paginator.page(page)
        except EmptyPage:
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "No more data available.",
                "data": [],
                "pagination": {
                    "current_page": page,
                    "total_pages": paginator.num_pages,
                    "total_items": paginator.count,
                    "has_next": False,
                    "has_previous": paginator.num_pages > 0,
                }
            })

        serializer = EmployeeSerializer(paginated_data.object_list, many=True)

        return Response({
            "statusCode": 200,
            "status": True,
            "message": "Employee list fetched successfully",
            "data": serializer.data,
            "pagination": {
                "current_page": page,
                "total_pages": paginator.num_pages,
                "total_items": paginator.count,
                "has_next": paginated_data.has_next(),
                "has_previous": paginated_data.has_previous(),
            }
        })    


    
class EmployeeDetailView(generics.RetrieveAPIView):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    permission_classes = [IsAuthenticated]

    def retrieve(self, request, *args, **kwargs):
        employee = self.get_object()
        serializer = self.get_serializer(employee)
        return Response({
            "status": True,
            "statusCode": 200,
            "message": "Employee details fetched successfully",
            "data": serializer.data
        })

class UpdateEmployeeLastNameView(generics.UpdateAPIView):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    permission_classes = [IsAuthenticated]

    def update(self, request, *args, **kwargs):
        employee = self.get_object()
        serializer = self.get_serializer(employee, data=request.data, partial=True)  # allow partial update
        data = request.data.copy()
        new_email = data.get('email')
        if new_email and Employee.objects.filter(email__iexact=new_email).exclude(id=employee.id).exists():
            return Response({
                "statusCode": 200,
                "status": False,
                "message": "An employee with this email already exists.",
                
            }, status=status.HTTP_200_OK)

        if serializer.is_valid():
            self.perform_update(serializer)
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Employee updated successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)
        else:
            return Response({
                "statusCode": 200,
                "status": False,
                "message": "Failed to update employee",
                "errors": serializer.errors
            }, status=status.HTTP_200_OK)


class EmployeeDeleteByIdView(generics.DestroyAPIView):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    permission_classes = [IsAuthenticated]

    def destroy(self, request, *args, **kwargs):
        try:
            employee = self.get_object()
            self.perform_destroy(employee)
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Employee deleted successfully"
            }, status=status.HTTP_200_OK)

        except NotFound:
            return Response({
                "statusCode": 200,
                "status": False,
                "message": "Employee not found"
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 200,
                "status": False,
                "message": "An error occurred while deleting the employee",
                "error": str(e)
            }, status=status.HTTP_200_OK)





class FeedbackProfessionalUser(APIView):
    permission_classes = [IsAuthenticated]  # Require authentication

    def post(self, request):
        serializer = FeedbackProfessionalSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(user=request.user)  # Set user from token
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Thanks for your feedback! We appreciate it.",
                "data": serializer.data
            }, status=status.HTTP_200_OK)
        return Response({
            "statusCode": 500,
            "status": False,
            "errors": serializer.errors
        }, status=status.HTTP_200_OK)
            


# -------------------------rise Ticket----------------------------

from Admin.models import SupportTicket
from Admin.serializers import SubscriptionSerializer, SupportTicketSerializer


class CreateSupportTicketView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user

        if not user.is_authenticated:
            return Response({
                "status": False,
                "statusCode": 401,
                "message": "Authentication required."
            }, status=status.HTTP_200_OK)
        if user._meta.app_label == "ProfessionalUser":
            user_type = "professionaluser"
        else:
            user_type = "user"

        ticket = SupportTicket.objects.create(
            ticket_category=request.data.get("ticket_category", "other"),
            subject=request.data.get("subject"),
            description=request.data.get("description"),
            documents=request.FILES.get("documents"),
            type_of_user=user_type,
            status='open',
            created_by_user_id=user.id
        )
        on_support_ticket_created(ticket)

        serializer = SupportTicketSerializer(ticket)
        return Response({
            "status": True,
            "statusCode": 201,
            "message": "Support ticket created successfully.",
            "data": serializer.data
        }, status=status.HTTP_201_CREATED)

class SellerTicketPagination(PageNumberPagination):
    page_size = 100  # default page size
    page_size_query_param = 'page_size'
    max_page_size = 100

class SellerTicketListAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user

        tickets = SupportTicket.objects.filter(
            type_of_user="professionaluser",
            created_by_user_id=user.id,
            is_deleted=False
        )
        ticket_id = request.query_params.get('ticket_id')
        status_param = request.query_params.get('status')
        ticket_category = request.query_params.get('ticket_category')

        if ticket_id:
            tickets = tickets.filter(ticket_id__icontains=ticket_id)
        if status_param:
            tickets = tickets.filter(status__iexact=status_param)
        if ticket_category:
            tickets = tickets.filter(ticket_category__icontains=ticket_category)

        if not tickets.exists():
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "No tickets found",
                "data": [],
                "count": 0,
                "next": None,
                "previous": None
            }, status=status.HTTP_200_OK)

        paginator = SellerTicketPagination()
        paginated_tickets = paginator.paginate_queryset(tickets, request)
        serializer = SupportTicketSerializer(paginated_tickets, many=True)

        return paginator.get_paginated_response({
            "statusCode": 200,
            "status": True,
            "message": "Tickets fetched successfully",
            "data": serializer.data
        })
# -----------------------------refresh token-----------------------------


from rest_framework_simplejwt.tokens import RefreshToken, TokenError


class RefreshProfessionalUserAccessTokenView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_description="Refresh access token for a Professional User using a valid refresh token.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["refresh"],
            properties={
                "refresh": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description="The refresh token required to generate a new access token"
                )
            }
        ),
        responses={
            200: openapi.Response(
                description="Access token refreshed successfully for Professional User",
                examples={
                    "application/json": {
                        "status": True,
                        "access_token": "new_access_token_here"
                    }
                }
            ),
            400: openapi.Response(
                description="Invalid or expired refresh token",
                examples={
                    "application/json": {
                        "status": False,
                        "message": "Invalid or expired refresh token"
                    }
                }
            )
        }
    )
    def post(self, request):
        ip = request.META.get('REMOTE_ADDR')
        user_agent = request.META.get('HTTP_USER_AGENT', 'unknown')
        refresh_token = request.data.get("refresh_token") 

        if not refresh_token:
            logger.warning(f"[Token Refresh - ProfessionalUser] Missing refresh token | IP: {ip}, User-Agent: {user_agent}")
            return Response(
                {"statusCode": 400, "status": False, "message": "Refresh token is required"},
                status=status.HTTP_200_OK
            )

        try:
            refresh = RefreshToken(refresh_token)  
            new_access_token = str(refresh.access_token) 
            logger.info(f"[Token Refresh - ProfessionalUser] Token refreshed successfully | IP: {ip}")
            return Response({
                "statusCode": 200, 
                "status": True,
                "access_token": new_access_token
            }, status=status.HTTP_200_OK)

        except TokenError:
            logger.warning(f"[Token Refresh - ProfessionalUser] Invalid or expired refresh token | IP: {ip}, User-Agent: {user_agent}")
            return Response(
                {"statusCode": 400, "status": False, "message": "Invalid or expired refresh token"},
                status=status.HTTP_200_OK
            )

        except Exception as e:
            logger.error(f"[Token Refresh - ProfessionalUser] Unexpected error: {str(e)} | IP: {ip}, User-Agent: {user_agent}")
            return Response(
                {"statusCode": 500, "status": False, "message": "An unexpected error occurred", "error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    


class GenerateInvoiceAPIView(APIView):
    def post(self, request):
        try:
            data = request.data
            pdf_url = self.generate_invoice_pdf(data)
            return Response({'pdf_url': pdf_url}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_200_OK)

    def generate_invoice_pdf(self, data):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=18
        )
        elements = []

        invoice_no = data.get('invoice_no', f"INV-{uuid.uuid4().hex[:6].upper()}")
        invoice_date = data.get('invoice_date', now().strftime("%Y-%m-%d"))

        styles = getSampleStyleSheet()
        elements.append(Paragraph(f"Invoice No: {invoice_no}", styles['Heading2']))
        elements.append(Paragraph(f"Invoice Date: {invoice_date}", styles['Normal']))
        elements.append(Spacer(1, 12))

        bill_to = data.get('customer_name', 'N/A')
        payment_details = data.get('payment_details', 'N/A')

        billing_table = Table([
            ["Bill To", "Payment Details"],
            [bill_to, payment_details]
        ], colWidths=[270, 270])
        billing_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black)
        ]))
        elements.append(billing_table)
        elements.append(Spacer(1, 20))

        quantity = float(data.get('quantity', 0))
        unit_price = float(data.get('unit_price', 0))
        total = quantity * unit_price

        item_data = [["Item", "Quantity", "Unit Price", "Total"]]
        item_data.append([
            data.get('product_service', ''),
            str(quantity),
            f"{unit_price:.2f}",
            f"{total:.2f}"
        ])
        item_data.append(["", "", "Grand Total", f"{data.get('total_amount', total):.2f}"])

        table = Table(item_data, colWidths=[200, 100, 100, 100])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor("#DCE6F1")]),
            ('BACKGROUND', (-2, -1), (-1, -1), colors.HexColor("#F2F2F2")),
            ('FONTNAME', (-2, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        filename = f"exports/invoice_{now().strftime('%Y%m%d%H%M%S')}.pdf"
        file_content = ContentFile(buffer.getvalue())
        default_storage.save(filename, file_content)
        delete_s3_file.apply_async(args=[filename], eta=now() + timedelta(hours=24))

        return default_storage.url(filename)


class NotificationListView(APIView):
    permission_classes = [IsAuthenticated]

    def get_queryset(self, user, notification_type=None, start_date=None, end_date=None):
        
        queryset = Notification.objects.filter(user=user).order_by('-created_at')
        if notification_type:
            queryset = queryset.filter(notification_type=notification_type)
        if start_date:
            try:
                queryset = queryset.filter(created_at__gte=start_date)
            except ValueError:
                raise ValidationError("Invalid start date format. Please use 'YYYY-MM-DD'.")
        if end_date:
            try:
                queryset = queryset.filter(created_at__lte=end_date)
            except ValueError:
                raise ValidationError("Invalid end date format. Please use 'YYYY-MM-DD'.")
        
        return queryset

    def get(self, request, *args, **kwargs):
        
        try:
            user = request.user  # The user is already authenticated due to permission class
            
            notification_type = request.query_params.get('notification_type', None)
            start_date = request.query_params.get('start_date', None)
            end_date = request.query_params.get('end_date', None)
            queryset = self.get_queryset(user, notification_type, start_date, end_date)
            serializer = NotificationSerializer(queryset, many=True)
            return Response({
                'statusCode': 200,
                'status': True,
                'message': 'Notifications fetched successfully',
                'data': serializer.data
            }, status=status.HTTP_200_OK)
        
        except ValidationError as e:
            logger.error(f"Validation error: {str(e)}")
            return Response({
                'statusCode': 400,
                'status': False,
                'message': str(e)
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Unexpected error: {str(e)}")
            return Response({
                'statusCode': 500,
                'status': False,
                'message': f"Unexpected error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            
class NotificationDeleteView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, *args, **kwargs):
        try:
            notification_id = request.query_params.get('notification_id', None)

            if notification_id:
                notification = Notification.objects.filter(id=notification_id, user=request.user).first()
                if not notification:
                    return Response({
                        'statusCode': 404,
                        'status': False,
                        'message': 'Notification not found or does not belong to the user.'
                    }, status=status.HTTP_200_OK)

                notification.delete()
                message = 'Notification deleted successfully.'

            else:
                Notification.objects.filter(user=request.user).delete()
                message = 'All notifications deleted successfully.'

            return Response({
                'statusCode': 200,
                'status': True,
                'message': message
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Notification deletion error for user {request.user.id}: {str(e)}")
            return Response({
                'statusCode': 500,
                'status': False,
                'message': f'An error occurred while deleting notification(s): {str(e)}.'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    



class PreviewSubscriptionChangeView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            logger.debug("Received request for PreviewSubscriptionChangeView")
            user = request.user
            professional = get_object_or_404(
                ProfessionalUser.objects.select_related("subscriptiontype")
                .prefetch_related("categories", "subcategories"),
                email=user.email
            )

            new_plan_id = request.data.get("plan_id")
            if not new_plan_id:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "plan_id is required."
                }, status=status.HTTP_200_OK)

            new_plan = get_object_or_404(SubscriptionPlan, id=new_plan_id)
            current_plan = professional.subscriptiontype

            if not current_plan:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "User has no current subscription plan."
                }, status=status.HTTP_200_OK)
            plan_order = {
                "Silver": 1,
                "Gold": 2,
                "Premium": 3
            }
            current_price = current_plan.price or Decimal('0.00')
            new_price = new_plan.price or Decimal('0.00')
            current_features = set(current_plan.features) if current_plan.features else set()
            new_features = set(new_plan.features) if new_plan.features else set()
            start_date = professional.updated_at or professional.created_at
            days_used = (now() - start_date).days
            days_total = 30
            days_remaining = max(0, days_total - days_used)
            daily_rate = current_price / Decimal(days_total)
            remaining_value = round(daily_rate * days_remaining, 2)
            final_price = max(Decimal('0.00'), round(new_price - remaining_value, 2))
            category_diff = new_plan.subscription.category_limit - current_plan.subscription.category_limit
            subcategory_diff = new_plan.subscription.subcategory_limit - current_plan.subscription.subcategory_limit
            lost_features = current_features - new_features
            gained_features = new_features - current_features
            current_plan_tier = plan_order.get(current_plan.subscription.name, 0)
            new_plan_tier = plan_order.get(new_plan.subscription.name, 0)

            if current_plan_tier == new_plan_tier:
                message = "Your selected plan is the same as your current plan."
                status_change = "Samilar_plan"
            elif new_plan_tier > current_plan_tier:
                message = "You're upgrading to a higher plan. Updated categories and subcategories as per new plan limit"
                status_change = "upgrade"
            else:
                message = (
                    "Warning:"
                    "You may lose access to some features, and all your existing data "
                    "(including store, media, products, orders, etc.) may be permanently deleted."
                    "First you have to set category as per new plan "
                )
                status_change = "downgrade"
            feature_change_message = None
            if status_change == "upgrade":
                benefits = []
                if category_diff > 0:
                    benefits.append(f"+{category_diff} category slots")
                if subcategory_diff > 0:
                    benefits.append(f"+{subcategory_diff} subcategory slots")
                if gained_features:
                    benefits.append(f"New features: {', '.join(gained_features)}")

                feature_change_message = "You will gain: " + ", ".join(benefits) if benefits else \
                    "No additional benefits in this upgrade."

            elif status_change == "downgrade":
                losses = []
                if category_diff < 0:
                    losses.append(f"{abs(category_diff)} category")
                if subcategory_diff < 0:
                    losses.append(f"{abs(subcategory_diff)} subcategory")
                if lost_features:
                    losses.append(f"Lost features: {', '.join(lost_features)}")

                feature_change_message = "You will lose: " + ", ".join(losses) if losses else \
                    "No major feature loss in this downgrade."
            new_expiry_date = start_date + timedelta(days=30)
            extend_day = None

            if remaining_value > new_price:
                overpayment = remaining_value - new_price
                new_daily_rate = new_price / Decimal(30)
                extra_days = int(overpayment / new_daily_rate)
                new_expiry_date = now() + timedelta(days=30 + extra_days)
                extend_day = f"Your new plan is extended until {new_expiry_date.isoformat()} based on the remaining value. Total days extended: {extra_days}"

            return Response({
                "statusCode": 200,
                "status": True,
                "message": message,
                "plan_status": status_change,
                "current_plan": {
                    "name": current_plan.subscription.name,
                    "price": float(current_price),
                    "type": current_plan.subscription_type,
                    "features": list(current_features),
                    "category_limit": current_plan.subscription.category_limit,
                    "subcategory_limit": current_plan.subscription.subcategory_limit,
                    "end_date": new_expiry_date.isoformat()
                },
                "new_plan": {
                    "name": new_plan.subscription.name,
                    "price": float(new_price),
                    "type": new_plan.subscription_type,
                    "features": list(new_features),
                    "category_limit": new_plan.subscription.category_limit,
                    "subcategory_limit": new_plan.subscription.subcategory_limit
                },
                "feature_change_message": feature_change_message,
                "remaining_value": float(remaining_value),
                "final_price_to_pay": float(final_price),
                "new_expiry_date":new_expiry_date,
                "extend_day": extend_day
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Unexpected error: {str(e)}", exc_info=True)
            return Response({
                "statusCode": 400,
                "status": False,
                "message": f"Error occurred: {str(e)}"
            }, status=status.HTTP_200_OK)

from Admin.serializers import CategorySerializer, SubcategorySerializer


class GetCategorySelectionLimitsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, plan_id=None):
        try:
            if not plan_id:
                return Response({
                    "message": "Plan ID is required",
                    "statusCode": 400,
                    "status": False
                }, status=status.HTTP_200_OK)
            user = request.user
            try:
                subscription_plan = SubscriptionPlan.objects.get(id=plan_id)
            except SubscriptionPlan.DoesNotExist:
                return Response({
                    "message": "Subscription plan not found",
                    "statusCode": 404,
                    "status": False
                }, status=status.HTTP_200_OK)

            category_limit = subscription_plan.subscription.category_limit
            subcategory_limit = subscription_plan.subscription.subcategory_limit
            user_categories = user.categories.all()
            user_subcategories = user.subcategories.all()

            user_categories_serialized = CategorySerializer(user_categories, many=True).data
            user_subcategories_serialized = SubcategorySerializer(user_subcategories, many=True).data
            all_categories = Category.objects.filter(is_active=True, is_deleted=False)
            all_categories_serialized = []
            for category in all_categories:
                subcategories = category.subcategories.filter(is_active=True, is_deleted=False)
                category_data = CategorySerializer(category).data
                category_data['subcategories'] = SubcategorySerializer(subcategories, many=True).data
                all_categories_serialized.append(category_data)

            return Response({
                "message": "Fetched user and available categories successfully",
                "statusCode": 200,
                "status": True,
                "data": {
                    "category_limit": category_limit,
                    "subcategory_limit": subcategory_limit,
                    "current_categories": user_categories_serialized,
                    "current_subcategories": user_subcategories_serialized,
                    "all_categories": all_categories_serialized
                }
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "message": f"An unexpected error occurred: {str(e)}",
                "statusCode": 500,
                "status": False
            }, status=status.HTTP_200_OK)


class CruiseFacilityCreateView(APIView):
    def post(self, request):
        serializer = CruiseFacilitySerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "statusCode": 201,
                "status": True,
                "message": "Cruise Facility created successfully",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)
        return Response({
            "statusCode": 400,
            "status": False,
            "message": "Error in data",
            "errors": serializer.errors
        }, status=status.HTTP_200_OK)


class CruiseFacilityListView(APIView):
    def get(self, request):
        facilities = CruiseFacility.objects.all()
        serializer = CruiseFacilitySerializer(facilities, many=True)
        if serializer.data:
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Cruise Facilities fetched successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)
        return Response({
            "statusCode": 404,
            "status": False,
            "message": "No facilities found.",
            "data": []
        }, status=status.HTTP_200_OK)
class FacilityBulkCreateView(APIView):
    def post(self, request, *args, **kwargs):
        try:
            facility_names = request.POST.getlist('name')  # Fetch facility names
            icons = request.FILES.getlist('icon')          # Fetch icons
            if not facility_names or not icons or len(facility_names) != len(icons):
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Facilities name and icons must be provided with the same count",
                }, status=status.HTTP_200_OK)

            facility_objects = []
            for name, icon in zip(facility_names, icons):
                if not name.strip():
                    return Response({
                        "statusCode": 400,
                        "status": False,
                        "message": "Facility name cannot be empty."
                    }, status=status.HTTP_200_OK)
                facility_objects.append(CruiseFacility(name=name.strip(), icon=icon))
            created_facilities = CruiseFacility.objects.bulk_create(facility_objects)
            created_data = CruiseFacility.objects.filter(
                name__in=[f.name for f in created_facilities]
            ).values("id", "name", "icon")

            return Response({
                "statusCode": 201,
                "status": True,
                "message": "Facilities created successfully",
                "data": list(created_data)
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)        

class RoomTypeListAPIView(APIView):
    """
    API to get the list of room type choices for Cruise Rooms.
    """
    def get(self, request):
        try:
            room_types = [{"id": choice[0], "value": choice[1]} for choice in CruiseRoom.ROOM_TYPE_CHOICES]
            
            return Response({
                "message": "Room types fetched successfully",
                "statusCode": 200,
                "status": True,
                "data": room_types
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "message": "Failed to fetch room types",
                "statusCode": 500,
                "status": False,
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        


logger = logging.getLogger(__name__)

class MySpecificOrderTicketsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            user = request.user
            professional_user = ProfessionalUser.objects.filter(email=user.email).first()
            
            if not professional_user:
                return Response({
                    "message": "Unauthorized: Not a professional user",
                    "statusCode": 403,
                    "status": False,
                    "data": []
                }, status=status.HTTP_200_OK)
            
            try:
                order_ct = ContentType.objects.get_for_model(Order)
                roombooking_ct = ContentType.objects.get_for_model(RoomBooking)
                eventbooking_ct = ContentType.objects.get_for_model(eventBooking)
                expbooking_ct = ContentType.objects.get_for_model(experienceBooking)
                slotbooking_ct = ContentType.objects.get_for_model(slotBooking)
                aesthbooking_ct = ContentType.objects.get_for_model(aestheticsBooking)
                relaxbooking_ct = ContentType.objects.get_for_model(relaxationBooking)
                artcultbooking_ct = ContentType.objects.get_for_model(artandcultureBooking)
            except ContentType.DoesNotExist:
                return Response({
                    "message": "One or more content types could not be resolved",
                    "statusCode": 500,
                    "status": False,
                    "data": []
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            # Get all booking IDs for filtering
            order_ids = Order.objects.filter(company=professional_user.company).values_list('id', flat=True)
            roombooking_ids = RoomBooking.objects.filter(company=professional_user.company).values_list('id', flat=True)
            eventbooking_ids = eventBooking.objects.filter(company=professional_user.company).values_list('id', flat=True)
            expbooking_ids = experienceBooking.objects.filter(company=professional_user.company).values_list('id', flat=True)
            slotbooking_ids = slotBooking.objects.filter(company=professional_user.company).values_list('id', flat=True)
            aestheticbooking_ids = aestheticsBooking.objects.filter(company=professional_user.company).values_list('id', flat=True)
            relaxbooking_ids = relaxationBooking.objects.filter(company=professional_user.company).values_list('id', flat=True)
            artculbooking_ids = artandcultureBooking.objects.filter(company=professional_user.company).values_list('id', flat=True)

            # Base tickets
            tickets = SupportTicket.objects.filter(
                specific_order=True
            ).filter(
                Q(content_type=order_ct, object_id__in=order_ids) |
                Q(content_type=roombooking_ct, object_id__in=roombooking_ids) |
                Q(content_type=eventbooking_ct, object_id__in=eventbooking_ids) |
                Q(content_type=expbooking_ct, object_id__in=expbooking_ids) |
                Q(content_type=slotbooking_ct, object_id__in=slotbooking_ids) |
                Q(content_type=aesthbooking_ct, object_id__in=aestheticbooking_ids) |
                Q(content_type=relaxbooking_ct, object_id__in=relaxbooking_ids) |
                Q(content_type=artcultbooking_ct, object_id__in=artculbooking_ids)
            ).order_by('-created_at')

            # Optional search filters
            ticket_id = request.query_params.get('ticket_id')
            subject = request.query_params.get('subject')
            ticket_category = request.query_params.get('ticket_category')
            status_param = request.query_params.get('status')

            if ticket_id:
                tickets = tickets.filter(ticket_id__icontains=ticket_id)
            if subject:
                tickets = tickets.filter(subject__icontains=subject)
            if ticket_category:
                tickets = tickets.filter(ticket_category__icontains=ticket_category)
            if status_param:
                tickets = tickets.filter(status__iexact=status_param)

            # Paginate
            paginator = SellerTicketPagination()
            paginated_tickets = paginator.paginate_queryset(tickets, request)
            serialized_data = SupportTicketSerializer(paginated_tickets, many=True)

            return paginator.get_paginated_response({
                "message": "Tickets fetched successfully",
                "statusCode": 200,
                "status": True,
                "data": serialized_data.data
            })

        except Exception as e:
            return Response({
                "message": f"An unexpected error occurred: {str(e)}",
                "statusCode": 500,
                "status": False,
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


            
class GetSubscriptionListView(APIView):
    permission_classes = [IsAuthenticated]  # 🔐 Auth required

    def get(self, request):
        try:
            subscription_type = request.GET.get("subscription_type", None)
            plan_queryset = SubscriptionPlan.objects.all()
            if subscription_type:
                if subscription_type not in ["Monthly", "Annual"]:
                    return Response(
                        {
                            "statusCode": 400,
                            "status": False,
                            "message": "Invalid subscription_type. Use 'Monthly' or 'Annual'.",
                        },
                        status=status.HTTP_200_OK,
                    )
                plan_queryset = plan_queryset.filter(subscription_type=subscription_type)
            subscriptions = Subscription.objects.prefetch_related(
                Prefetch("plans", queryset=plan_queryset)
            )

            if subscription_type:
                subscriptions = subscriptions.filter(plans__subscription_type=subscription_type).distinct()
            else:
                subscriptions = subscriptions.all()
            avg_discount = SubscriptionPlan.objects.filter(
                annualPlan__isnull=False, price__isnull=False
            ).annotate(
                discount_percentage=((12 * F("price") - F("annualPlan")) / (12 * F("price")) * 100)
            ).aggregate(Avg("discount_percentage"))["discount_percentage__avg"]

            total_discount = round(avg_discount, 2) if avg_discount else 0
            paginator = PageNumberPagination()
            paginated_data = paginator.paginate_queryset(subscriptions, request, view=self)

            serialized_paginated_data = SubscriptionSerializer(paginated_data, many=True).data

            if subscription_type:
                for subscription in serialized_paginated_data:
                    subscription["plans"] = [
                        plan for plan in subscription.get("plans", [])
                        if plan.get("subscription_type") == subscription_type
                    ]

            response = {
                "statusCode": 200,
                "status": True,
                "total_discount": total_discount,
                "message": "Subscriptions retrieved successfully",
                "data": serialized_paginated_data,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
                "total_pages": paginator.page.paginator.num_pages if paginator.page else 1,
                "total_items": paginator.page.paginator.count if paginator.page else len(serialized_paginated_data),
            }

            return Response(response, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response(
                {
                    "statusCode": 500,
                    "status": False,
                    "message": f"An error occurred: {str(e)}"
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )



class RoomFacilityListView(APIView):
    def get(self, request, *args, **kwargs):  # <-- FIX HERE
        try:
            facilities = RoomFacility.objects.all()
            serializer = RoomFacilitySerializer(facilities, many=True)

            if serializer.data:
                return Response(
                    {
                        "statusCode": 200,
                        "status": True,
                        "message": "Room Facilities retrieved successfully",
                        "data": serializer.data
                    },
                    status=status.HTTP_200_OK
                )
            else:
                return Response(
                    {
                        "statusCode": 404,
                        "status": False,
                        "message": "No Room Facilities found",
                        "data": []
                    },
                    status=status.HTTP_200_OK
                )

        except Exception as e:
            return Response(
                {
                    "statusCode": 500,
                    "status": False,
                    "message": f"An error occurred: {str(e)}"
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )




class StandardResultsSetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


class UnifiedBookingListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        search_query = request.GET.get("search", "")
        status_filter = request.GET.get("status", "")
        page = int(request.GET.get("page", 1))
        page_size = int(request.GET.get("page_size", 10))

        base_url = request.build_absolute_uri(request.path)

        try:
            professional_user = ProfessionalUser.objects.get(email=user.email)
        except ProfessionalUser.DoesNotExist:
            return Response({
                "status": False,
                "status_code": 404,
                "message": "Professional user not found.",
                "data": []
            }, status=status.HTTP_200_OK)

        company = professional_user.company
        response_data = []

        def should_include_booking(booking_user, booking_status):
            customer_name = f"{booking_user.firstName} {booking_user.lastName}".lower()
            if search_query and search_query.lower() not in customer_name:
                return False
            if status_filter and booking_status.lower() != status_filter.lower():
                return False
            return True

        room_bookings = RoomBooking.objects.filter(company=company,is_paid=True).select_related('product', 'room')
        for booking in room_bookings:
            if not should_include_booking(booking.user, booking.booking_status):
                continue
            response_data.append({
                "type": "room",
                 "user": {
        "id": booking.user.id,
        "order_id":booking.id,
        "customer_name": f"{booking.user.firstName} {booking.user.lastName}",
        "email": booking.user.email,
        "phone":booking.user.phone,
        "address": (
    f"{booking.user.manualAddress.address1} {booking.user.manualAddress.address2} "
    f"{booking.user.manualAddress.city} {booking.user.manualAddress.country}"
    if booking.user.manualAddress else ""
),
        
        
    },"company": {
        "id": company.id,
        "name": company.companyName,
        
    },
               
                 
                "products": [
            {
                "price": str(booking.product.promotionalPrice),  
                "product": ProductSerializer(booking.product).data
            }],
                "category":{
                    "cat_slug":booking.product.categoryId.slug,
                    "cat_id":booking.product.categoryId.id,
                    "cat_name":booking.product.categoryId.name
                },
                "subcategory":{
                    "subcat_slug":booking.product.subCategoryId.slug,
                    "subcat_id":booking.product.subCategoryId.id,
                    "subcat_name":booking.product.subCategoryId.name
                },

                "room_data": {
                    "id": booking.room.id,
                    "room_type": booking.room.roomType
                } if booking.room else None,
                "booking_createdat":booking.created_at.strftime("%B %d, %Y"),
                "booking_createdat_raw": booking.created_at.isoformat(),
                "total_products_ordered": booking.room_quantity,
                "is_paid":booking.is_paid,
                "status":booking.booking_status,
                "subtotal": booking.product.promotionalPrice,
                "discount": booking.product.discount,
                "total_price": booking.total_price,
                 "booking_details":{
                      "booking_id":booking.booking_id,
                      "booking_time":booking.checkin_date,
                      "checkout_date":booking.checkout_date,
                      "room_quantity":booking.room_quantity,

            }
            }),
        event_bookings = eventBooking.objects.filter(company=company,is_paid=True).select_related('ticket_id')
        for booking in event_bookings:
            if not should_include_booking(booking.user, booking.status):
                continue
            try:
                ticket_data = json.loads(booking.ticket_type) if isinstance(booking.ticket_type, str) else booking.ticket_type
            except Exception:
                ticket_data = []
         

            response_data.append({
                "type": "event",
                      "user": {
        "id": booking.user.id,
        "order_id":booking.id,
        "customer_name": f"{booking.user.firstName} {booking.user.lastName}",
        "email": booking.user.email,
        "phone":booking.user.phone,
        "address": (
    f"{booking.user.manualAddress.address1} {booking.user.manualAddress.address2} "
    f"{booking.user.manualAddress.city} {booking.user.manualAddress.country}"
    if booking.user.manualAddress else ""
),
        "status":booking.status,
        
    },"company": {
        "id": company.id,
        "name": company.companyName,
        
    },
                 
                "products": [
            {
                "price": str(booking.ticket_id.promotionalPrice),  
                "product": ProductSerializer(booking.ticket_id).data
            }],
            
                "category":{
                    "cat_slug":booking.ticket_id.categoryId.slug,
                    "cat_id":booking.ticket_id.categoryId.id,
                    "cat_name":booking.ticket_id.categoryId.name
                },
                "subcategory":{
                    "subcat_slug":booking.ticket_id.subCategoryId.slug,
                    "subcat_id":booking.ticket_id.subCategoryId.id,
                    "subcat_name":booking.ticket_id.subCategoryId.name
                },

             "ticket":{
                       "Ticket":booking.ticket_type,
                       "price":booking.price

                }if booking.ticket_id else None,

                "id": booking.id,
                "number_of_people":booking.number_of_people,
                "end_date": booking.end_date,
                "booking_createdat": booking.created_at.strftime("%B %d, %Y"),
                "booking_createdat_raw": booking.created_at.isoformat(),
                "status":booking.status, 
                "price":booking.price,
                "is_paid":booking.is_paid,
               "total_products_ordered": sum(item["quantity"] for item in ticket_data) ,
                "subtotal": booking.ticket_id.promotionalPrice,
                "discount": booking.ticket_id.discount,
                "total_price": booking.price,
                "booking_details": EventBookingSerializer(booking).data 

                
            })
        experience_bookings = experienceBooking.objects.filter(company=company,is_paid=True).select_related('ticket_id')
        for booking in experience_bookings:
            if not should_include_booking(booking.user, booking.status):
                continue
            try:
                ticket_data = json.loads(booking.ticket_type) if isinstance(booking.ticket_type, str) else booking.ticket_type
            except Exception:
                ticket_data = []
            response_data.append({
                "type": "experience",
                    "user": {
        "id": booking.user.id,
        "order_id":booking.id,
        "customer_name": f"{booking.user.firstName} {booking.user.lastName}",
        "email": booking.user.email,
        "phone":booking.user.phone,
        "address": (
    f"{booking.user.manualAddress.address1} {booking.user.manualAddress.address2} "
    f"{booking.user.manualAddress.city} {booking.user.manualAddress.country}"
    if booking.user.manualAddress else ""
),
       
        
    },"company": {
        "id": company.id,
        "name": company.companyName,
        
    },
                 
                "products": [
            {
                "price": str(booking.ticket_id.promotionalPrice),  
                "product": ProductSerializer(booking.ticket_id).data
            }],
                "category":{
                    "cat_slug":booking.ticket_id.categoryId.slug,
                    "cat_id":booking.ticket_id.categoryId.id,
                    "cat_name":booking.ticket_id.categoryId.name
                },
                "subcategory":{
                    "subcat_slug":booking.ticket_id.subCategoryId.slug,
                    "subcat_id":booking.ticket_id.subCategoryId.id,
                    "subcat_name":booking.ticket_id.subCategoryId.name
                },
                "ticket":{
                       "Ticket":booking.ticket_type
                }if booking.ticket_id else None,

                "id": booking.id,
                "number_of_people":booking.number_of_people,
                "booking_createdat": booking.created_at.strftime("%B %d, %Y"),
                "booking_createdat_raw": booking.created_at.isoformat(),
                "end_date": booking.end_date,
                "status":booking.status, 
                "price":booking.price,
                "is_paid":booking.is_paid,
                "subtotal": booking.ticket_id.promotionalPrice,
                "discount": booking.ticket_id.discount,
                "total_price": booking.price,
                "booking_details": ExperienceBookingSerializer(booking).data 
            })
        slot_bookings = slotBooking.objects.filter(company=company,is_paid=True).select_related('Product')
        for booking in slot_bookings:
            if not should_include_booking(booking.user, booking.status):
                continue
          
            response_data.append({
                "type": "experience",
                    "user": {
        "id": booking.user.id,
        "order_id":booking.id,
        "customer_name": f"{booking.user.firstName} {booking.user.lastName}",
        "email": booking.user.email,
        "phone":booking.user.phone,
        "address": (
    f"{booking.user.manualAddress.address1} {booking.user.manualAddress.address2} "
    f"{booking.user.manualAddress.city} {booking.user.manualAddress.country}"
    if booking.user.manualAddress else ""
),
       
        
    },"company": {
        "id": company.id,
        "name": company.companyName,
        
    },
                 
                "products": [
            {
                "price": str(booking.Product.promotionalPrice),  
                "product": ProductSerializer(booking.Product).data
            }],
            
                "category":{
                    "cat_slug":booking.Product.categoryId.slug,
                    "cat_id":booking.Product.categoryId.id,
                    "cat_name":booking.Product.categoryId.name
                },
                "subcategory":{
                    "subcat_slug":booking.Product.subCategoryId.slug,
                    "subcat_id":booking.Product.subCategoryId.id,
                    "subcat_name":booking.Product.subCategoryId.name
                },
            

                "id": booking.id,
                "number_of_people":booking.number_of_people,
                "booking_createdat": booking.created_at.strftime("%B %d, %Y"),
                "booking_createdat_raw": booking.created_at.isoformat(),
                "status":booking.status, 
                "price":booking.price,
                "is_paid":booking.is_paid,
                "subtotal": booking.Product.promotionalPrice,
                "discount": booking.Product.discount,
                "total_price": booking.price,
                "booking_details": SlotBookingSerializer(booking).data 
            })

        aesthetic_bookings = aestheticsBooking.objects.filter(company=company,is_paid=True).select_related('Product')
        for booking in aesthetic_bookings:
            if not should_include_booking(booking.user, booking.status):
                continue
           
            response_data.append({
                "type": "experience",
                    "user": {
        "id": booking.user.id,
        "order_id":booking.id,
        "customer_name": f"{booking.user.firstName} {booking.user.lastName}",
        "email": booking.user.email,
        "phone":booking.user.phone,
        "address": (
    f"{booking.user.manualAddress.address1} {booking.user.manualAddress.address2} "
    f"{booking.user.manualAddress.city} {booking.user.manualAddress.country}"
    if booking.user.manualAddress else ""
),
       
        
    },"company": {
        "id": company.id,
        "name": company.companyName,
        
    },
                 
                "products": [
            {
                "price": str(booking.Product.promotionalPrice),  
                "product": ProductSerializer(booking.Product).data
            }],
                "category":{
                    "cat_slug":booking.Product.categoryId.slug,
                    "cat_id":booking.Product.categoryId.id,
                    "cat_name":booking.Product.categoryId.name
                },
                "subcategory":{
                    "subcat_slug":booking.Product.subCategoryId.slug,
                    "subcat_id":booking.Product.subCategoryId.id,
                    "subcat_name":booking.Product.subCategoryId.name
                },
            

                "id": booking.id,
                "number_of_people":booking.number_of_people,
                "booking_createdat": booking.created_at.strftime("%B %d, %Y"),
                "booking_createdat_raw": booking.created_at.isoformat(),
                "status":booking.status, 
                "price":booking.price,
                "is_paid":booking.is_paid,
                "subtotal": booking.Product.promotionalPrice,
                "discount": booking.Product.discount,
                "total_price": booking.price,
                "booking_details": AestheticsBookingSerializer(booking).data 
            })
       
        relaxation_bookings = relaxationBooking.objects.filter(company=company,is_paid=True).select_related('Product')
        for booking in relaxation_bookings:
            if not should_include_booking(booking.user, booking.status):
                continue
           
            response_data.append({
                "type": "experience",
                    "user": {
        "id": booking.user.id,
        "order_id":booking.id,
        "customer_name": f"{booking.user.firstName} {booking.user.lastName}",
        "email": booking.user.email,
        "phone":booking.user.phone,
        "address": (
    f"{booking.user.manualAddress.address1} {booking.user.manualAddress.address2} "
    f"{booking.user.manualAddress.city} {booking.user.manualAddress.country}"
    if booking.user.manualAddress else ""
),
       
        
    },"company": {
        "id": company.id,
        "name": company.companyName,
        
    },
                 
                "products": [
            {
                "price": str(booking.Product.promotionalPrice),  
                "product": ProductSerializer(booking.Product).data
            }],
                "category":{
                    "cat_slug":booking.Product.categoryId.slug,
                    "cat_id":booking.Product.categoryId.id,
                    "cat_name":booking.Product.categoryId.name
                },
                "subcategory":{
                    "subcat_slug":booking.Product.subCategoryId.slug,
                    "subcat_id":booking.Product.subCategoryId.id,
                    "subcat_name":booking.Product.subCategoryId.name
                },
            

                "id": booking.id,
                "number_of_people":booking.number_of_people,
                "booking_createdat": booking.created_at.strftime("%B %d, %Y"),
                "booking_createdat_raw": booking.created_at.isoformat(),
                "status":booking.status, 
                "price":booking.price,
                "is_paid":booking.is_paid,
                "subtotal": booking.Product.promotionalPrice,
                "discount": booking.Product.discount,
                "total_price": booking.price,
                "booking_details": RelaxationBookingSerializer(booking).data 
            })
       
        artandculture_bookings = artandcultureBooking.objects.filter(company=company,is_paid=True).select_related('Product')
        for booking in artandculture_bookings:
            if not should_include_booking(booking.user, booking.status):
                continue
           
            response_data.append({
                "type": "experience",
                    "user": {
        "id": booking.user.id,
        "order_id":booking.id,
        "customer_name": f"{booking.user.firstName} {booking.user.lastName}",
        "email": booking.user.email,
        "phone":booking.user.phone,
        "address": (
    f"{booking.user.manualAddress.address1} {booking.user.manualAddress.address2} "
    f"{booking.user.manualAddress.city} {booking.user.manualAddress.country}"
    if booking.user.manualAddress else ""
),
       
        
    },"company": {
        "id": company.id,
        "name": company.companyName,
        
    },
                 
                "products": [
            {
                "price": str(booking.Product.promotionalPrice),  
                "product": ProductSerializer(booking.Product).data
            }],
                "category":{
                    "cat_slug":booking.Product.categoryId.slug,
                    "cat_id":booking.Product.categoryId.id,
                    "cat_name":booking.Product.categoryId.name
                },
                "subcategory":{
                    "subcat_slug":booking.Product.subCategoryId.slug,
                    "subcat_id":booking.Product.subCategoryId.id,
                    "subcat_name":booking.Product.subCategoryId.name
                },
            

                "id": booking.id,
                "number_of_people":booking.number_of_people,
                "booking_createdat": booking.created_at.strftime("%B %d, %Y"),
                "booking_createdat_raw": booking.created_at.isoformat(),
                "status":booking.status, 
                "price":booking.price,
                "is_paid":booking.is_paid,
                "subtotal": booking.Product.promotionalPrice,
                "discount": booking.Product.discount,
                "total_price": booking.price,
                "booking_details": ArtandCultureBookingSerializer(booking).data 
            })
        response_data.sort(key=lambda x: x["booking_createdat_raw"], reverse=True)
        paginator = Paginator(response_data, page_size)
        paginated_data = paginator.get_page(page)
        serialized_page = list(paginated_data)     

        def build_url(page_number):
            query_params = request.GET.copy()
            query_params["page"] = page_number
            return f"{base_url}?{urlencode(query_params)}"

        next_url = build_url(paginated_data.next_page_number()) if paginated_data.has_next() else None
        prev_url = build_url(paginated_data.previous_page_number()) if paginated_data.has_previous() else None

        for item in serialized_page:
            item.pop("booking_createdat_raw", None)
        return Response({
            "status": True,
            "status_code": 200,
            "message": "All bookings fetched successfully.",
            "next": next_url,
            "previous": prev_url,
            "data": response_data,
            "total_pages": paginator.num_pages,
            "current_page": page,
            "total_items": paginator.count,
            "data": serialized_page
        }, status=status.HTTP_200_OK)




class CancelOrderByProfessionalView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, order_id):
        serializer = CancelOrderSerializer(data=request.data)
        if serializer.is_valid():
            reason = serializer.validated_data['cancel_reasons']
            try:
                professional_user = request.user
            except ProfessionalUser.DoesNotExist:
                return Response(
                    {"statusCode": 400,
                        "status": False, "message": "Authenticated user is not a professional user."},
                    status=status.HTTP_200_OK
                )
            try:
                order = Order.objects.get(order_id=order_id)
            except Order.DoesNotExist:
                return Response(
                    {   "statusCode": 400,
                        "status": False, "message": "Order not found or not associated with this professional user."},
                    status=status.HTTP_200_OK
                )
            if order.orderStatus == "cancelled":
                return Response(
                    {   "statusCode": 400,
                        "status": False, "message": "Order is already cancelled."},
                    status=status.HTTP_200_OK
                )
            order.orderStatus = "cancelled"
            order.cancel_reasons = reason
            order.cancel_by = "professionaluser"
            order.save()

            return Response({ "statusCode": 200,
                        "status": True, "message": "Order cancelled successfully."}, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_200_OK)

class ResolveTicketView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, ticket_id):
        user = request.user
        logger.info(f"User {user.email} attempting to resolve ticket {ticket_id}")

        try:
            professional_user = ProfessionalUser.objects.get(email=user.email)
            ticket = SupportTicket.objects.get(ticket_id=ticket_id, specific_order=True)
            if ticket.order and ticket.order.professional_user_id == professional_user.id:
                
                refund_done = request.data.get('refund_done', False)

                if refund_done:
                    ticket.status = 'resolved'
                    ticket.resolved_at = datetime.now()
                    ticket.save()

                    return Response({
                        "message": f"Ticket {ticket.ticket_id} resolved,Refund succesful .",
                        "statusCode": 200,
                        "status": True,
                        "status":ticket.status
                    }, status=status.HTTP_200_OK)
                else:
                    
                    return Response({
                        "message": f"Ticket {ticket.ticket_id} disputed,Refund Unsuccesful .",
                        
                        "statusCode": 400,
                        "status": False,
                        "status":ticket.status
                    }, status=status.HTTP_200_OK)

            else:
                return Response({
                    "detail": "You are not authorized to manage this ticket.",
                    "status": False
                }, status=status.HTTP_200_OK)

        except SupportTicket.DoesNotExist:
            return Response({
                "message": "Ticket not found.",
                "statusCode": 404,
                "status": False
            }, status=status.HTTP_200_OK)

        except ProfessionalUser.DoesNotExist:
            return Response({
                "message": "Professional user not found.",
                "statusCode": 404,
                "status": False
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error resolving ticket: {str(e)}")
            return Response({
                "message": "Something went wrong.",
                "error": str(e),
                "statusCode": 500,
                "status": False
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UpdateBookingStatusAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        booking_type = request.data.get('type')
        booking_id = request.data.get('id')
        new_status = request.data.get('status')

        if not all([booking_type, booking_id, new_status]):
            return Response({
                  "statusCode":400,
                "status": False,
                "message": "Missing required fields: 'type', 'id', 'status'."
            }, status=status.HTTP_200_OK)

        model_map = {
            'room': (RoomBooking, 'booking_status'),
            'event': (eventBooking, 'status'),
            'experience': (experienceBooking, 'status'),
        }

        model_tuple = model_map.get(booking_type.lower())
        if not model_tuple:
            return Response({
                "statusCode":400,
                "status": False,
                "message": "Invalid booking type."
            }, status=status.HTTP_200_OK)

        model, status_field = model_tuple

        try:
            booking = model.objects.get(id=booking_id)
        except model.DoesNotExist:
            return Response({
                "statusCode":400,
                "status": False,
                "message": "Booking not found."
            }, status=status.HTTP_200_OK)
        try:
            valid_statuses = [choice[0] for choice in booking._meta.get_field(status_field).choices]
        except Exception:
            return Response({
                "statusCode":400,
                "status": False,
                "message": f"Model '{booking_type}' has no field named '{status_field}'."
            }, status=status.HTTP_200_OK)

        if new_status not in valid_statuses:
            return Response({
                "statusCode":400,
                "status": False,
                "message": f"Invalid status. Allowed: {valid_statuses}"
            }, status=status.HTTP_200_OK)

        setattr(booking, status_field, new_status)
        booking.save()

        return Response({
            "statusCode":200,
            "status": True,
            "message": f"{booking_type.capitalize()} booking status updated to '{new_status}'."
        }, status=status.HTTP_200_OK)
        
class OrderBookingIconsCreateView(APIView):
    def post(self, request, *args, **kwargs):
        name = request.data.get('name')
        icon = request.FILES.get('icon')

        if not name:
            return Response({"error": "Name is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            booking_icon = OrderBookingIcons.objects.create(name=name, icon=icon)
            return Response({
                "id": booking_icon.id,
                "name": booking_icon.name,
                "icon": booking_icon.icon.url if booking_icon.icon else None
            }, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        


class SoftDeleteProfessionalUserView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request):
        try:
            user = request.user
            professional_user = ProfessionalUser.objects.get(email=user.email)
            professional_user.is_deleted = True
            professional_user.deleted_at = timezone.now()
            professional_user.save()
            return Response({"detail": "Account soft-deleted successfully."}, status=status.HTTP_200_OK)
        except ProfessionalUser.DoesNotExist:
            return Response({"detail": "Professional user not found."}, status=status.HTTP_404_NOT_FOUND)