from rest_framework.views import APIView
from rest_framework.generics import ListAPIView
from django.shortcuts import get_object_or_404
from rest_framework.response import Response
from rest_framework import status,generics, permissions,viewsets
from rest_framework.permissions import AllowAny,IsAdminUser
from UserApp.models import Users
from UserApp.serializers import AdminUserCreateSerializer, AdminUserUpdateSerializer, UserListSerializer
from .serializers import *
from django.contrib.auth.tokens import default_token_generator
from django.contrib.auth import get_user_model
from rest_framework.pagination import PageNumberPagination
from drf_yasg.utils import swagger_auto_schema
from .utils import SwaggerView
from drf_yasg import openapi
from django.db.models import Q
from django_filters.rest_framework import DjangoFilterBackend
from datetime import datetime, timedelta
from django.db.models.functions import TruncDay, TruncMonth, TruncYear
from django.db.models import Count
from django.utils.timezone import now
from django.utils import timezone 
from django.http import Http404
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views import View
import json
from collections import defaultdict
from ProfessionalUser.models import ProfessionalUser, Order, OrderItem
from django.http import JsonResponse
from django.db.models import Prefetch
from django.db import transaction
from django.db.utils import IntegrityError
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.permissions import IsAuthenticated ,BasePermission 
from .models import RolePermissions,Coupon
from django.urls import reverse
from urllib.parse import urlparse,parse_qs
from rest_framework.parsers import MultiPartParser, FormParser
from .permissions import IsAdminAccount
from django.core.exceptions import ObjectDoesNotExist, ValidationError
from rest_framework_simplejwt.exceptions import TokenError  # Correct import
from django.db.models import F, Avg
from UserApp.serializers import *
from drf_yasg.utils import swagger_auto_schema
from .serializers import AdminUserLoginSerializer
from django.core.exceptions import FieldError
import random
import string
from rest_framework.filters import SearchFilter, OrderingFilter
import logging
from django.db import DatabaseError
from rest_framework import status, filters
from ProfessionalUser.serializers import CompanyDetailsUpdateSerializer,ProfessionalUserSerializer,OrderSerializer,EmployeeSerializer,ProductSerializer
from django.db.models import Count, Q, Sum, F, FloatField
from payment.models import Payment, UserPayment, Card
from payment.serializers import PaymentSerializer, UserPaymentSerializer, CardSerializer
from rest_framework_simplejwt.tokens import RefreshToken
import csv
import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import openpyxl

wb = openpyxl.Workbook()
logger = logging.getLogger(__name__)
User = get_user_model()

class IsAdministrator(BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.role and request.user.role.name == "administrator"

class CustomPagination(PageNumberPagination):
    """Custom Pagination for Professional Users"""
    page_size = 10  # Number of results per page
    page_size_query_param = "page_size"
    max_page_size = 100  # Set a reasonable limit

class AdminLoginAPIView(APIView):
    @swagger_auto_schema(request_body=AdminUserLoginSerializer)
    def post(self, request):
        ip = request.META.get('REMOTE_ADDR')
        user_agent = request.META.get('HTTP_USER_AGENT', 'unknown')
        logger.info(f"[Admin Login Attempt] IP: {ip}, User-Agent: {user_agent}, Data: {request.data}")

        serializer = AdminUserLoginSerializer(data=request.data)

        if serializer.is_valid():
            validated_data = serializer.validated_data
            user_data = validated_data["user"]
            email = user_data["email"]
            logger.info(f"[Admin Login Success] Email: {email}, IP: {ip}")

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Login successful",
                "data": {
                    "refresh_token": validated_data["refresh_token"],
                    "access_token": validated_data["access_token"],
                    "email": user_data["email"],
                    "user_type": user_data["user_type"],
                    "admin_id": user_data["id"],
                    "role": user_data["role"],
                    "name": user_data["name"],
                }
            }, status=status.HTTP_200_OK)

        logger.warning(f"[Admin Login Failed] IP: {ip}, Errors: {serializer.errors}")
        return Response({
            "statusCode": 400,
            "status": False,
            "message": "Invalid credentials",
            "errors": serializer.errors
        }, status=status.HTTP_200_OK)


class AdminProfileView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Retrieve the profile details of the authenticated admin user"""
        ip = request.META.get('REMOTE_ADDR')
        user_agent = request.META.get('HTTP_USER_AGENT', 'unknown')
        user = request.user  

        try:
            if not user.is_authenticated:
                logger.warning(f"[Unauthorized Admin Profile Access] IP: {ip}, User-Agent: {user_agent}")
                return Response({
                    "statusCode": 401,
                    "status": False,
                    "message": "Unauthorized access"
                }, status=status.HTTP_200_OK)

            profile_data = {
                "id": user.id,
                "name": user.name,
                "email": user.email,
                "mobile_number": user.mobile,  
                "is_active": user.is_active,
                "is_staff": user.is_staff,
            }

            logger.info(f"[Admin Profile Retrieved] Admin ID: {user.id}, Email: {user.email}, IP: {ip}")
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Admin profile retrieved successfully",
                "profile": profile_data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"[Admin Profile Error] IP: {ip}, User-Agent: {user_agent}, Error: {str(e)}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Internal server error"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

class RefreshAccessTokenView(APIView):
    permission_classes = [AllowAny]  

    @swagger_auto_schema(
        operation_description="Refresh access token using a valid refresh token.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["refresh"],
            properties={
                "refresh": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description="The refresh token required to generate a new access token"
                )
            }
        ),
        responses={
            200: openapi.Response(
                description="Access token refreshed successfully",
                examples={
                    "application/json": {
                        "status": True,
                        "access_token": "new_access_token_here"
                    }
                }
            ),
            400: openapi.Response(
                description="Invalid or expired refresh token",
                examples={
                    "application/json": {
                        "status": False,
                        "message": "Invalid or expired refresh token"
                    }
                }
            )
        }
    )
    def post(self, request):
        ip = request.META.get('REMOTE_ADDR')
        user_agent = request.META.get('HTTP_USER_AGENT', 'unknown')
        refresh_token = request.data.get("refresh_token") 

        if not refresh_token:
            logger.warning(f"[Token Refresh] Missing refresh token | IP: {ip}, User-Agent: {user_agent}")
            return Response(
                {"statusCode": 400, "status": False, "message": "Refresh token is required"},
                status=status.HTTP_200_OK
            )

        try:
            refresh = RefreshToken(refresh_token)  
            new_access_token = str(refresh.access_token) 
            logger.info(f"[Token Refresh] Token refreshed successfully | IP: {ip}")
            return Response({
                "statusCode": 200, 
                "status": True,
                "access_token": new_access_token
            }, status=status.HTTP_200_OK)

        except TokenError:
            logger.warning(f"[Token Refresh] Invalid or expired refresh token | IP: {ip}, User-Agent: {user_agent}")
            return Response(
                {"statusCode": 400, "status": False, "message": "Invalid or expired refresh token"},
                status=status.HTTP_200_OK
            )

        except Exception as e:
            logger.error(f"[Token Refresh] Unexpected error: {str(e)} | IP: {ip}, User-Agent: {user_agent}")
            return Response(
                {"statusCode": 500, "status": False, "message": "An unexpected error occurred", "error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class AdminForgotPasswordAPIView(APIView):
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated]  

    @swagger_auto_schema(
        operation_description="Send password reset email to admin and return reset link in response.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["email"],
            properties={
                "email": openapi.Schema(type=openapi.TYPE_STRING, format="email", description="Admin email address")
            },
        ),
        responses={
            200: openapi.Response(description="Password reset email sent"),
            400: openapi.Response(description="Bad request"),
            404: openapi.Response(description="Admin not found"),
        }
    )
    def post(self, request):
        """Send password reset email to admin and return reset link in response"""
        ip = request.META.get('REMOTE_ADDR')
        user_agent = request.META.get('HTTP_USER_AGENT', 'unknown')

        try:
            email = request.data.get("email")
            logger.info(f"[Forgot Password] Password reset request received for email: {email} | IP: {ip}")

            if not email:
                logger.warning(f"[Forgot Password] Email is missing in request | IP: {ip}")
                return Response(
                    {"statusCode": 400, "status": False, "message": "Email is required"},
                    status=status.HTTP_200_OK
                )

            try:
                user = User.objects.get(email=email, is_staff=True)  
            except User.DoesNotExist:
                logger.warning(f"[Forgot Password] Admin not found with email: {email} | IP: {ip}")
                return Response(
                    {"statusCode": 404, "status": False, "message": "Admin not found"},
                    status=status.HTTP_200_OK
                )

            token = default_token_generator.make_token(user)
            reset_url = request.build_absolute_uri(reverse("reset-password"))
            full_reset_link = f"{reset_url}?token={token}&user_id={user.pk}"

            try:
                send_mail(
                    subject="Admin Password Reset Request",
                    message=f"Click the link to reset your password: {full_reset_link}",
                    from_email="your-email@gmail.com",
                    recipient_list=[email],
                    fail_silently=False,
                )
                logger.info(f"[Forgot Password] Reset email sent to: {email} | IP: {ip}")
            except Exception as e:
                logger.error(f"[Forgot Password] Failed to send reset email to {email}: {e} | IP: {ip}")
                return Response(
                    {"statusCode": 500, "status": False, "message": "Failed to send email", "error": str(e)},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Password reset email sent",
                "reset_link": full_reset_link
            }, status=status.HTTP_200_OK)           

        except Exception as e:
            logger.exception(f"[Forgot Password] Unexpected error: {e} | IP: {ip}")
            return Response(
                {"statusCode": 500, "status": False, "message": "An unexpected error occurred", "error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class AdminResetPasswordAPIView(APIView):
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated]  

    @swagger_auto_schema(
        operation_description="Reset admin password using a reset link token.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["reset_link", "new_password"],
            properties={
                "reset_link": openapi.Schema(type=openapi.TYPE_STRING, description="Password reset link"),
                "new_password": openapi.Schema(type=openapi.TYPE_STRING, description="New password")
            },
        ),
        responses={
            200: openapi.Response(description="Password reset successful"),
            400: openapi.Response(description="Bad request"),
            401: openapi.Response(description="Unauthorized"),
        }
    )
    def post(self, request):
        """Reset admin password using token from payload"""
        ip = request.META.get('REMOTE_ADDR')
        user_agent = request.META.get('HTTP_USER_AGENT', 'unknown')

        try:
            reset_link = request.data.get("reset_link")  
            new_password = request.data.get("new_password")

            logger.info(f"[Reset Password] Reset request received | IP: {ip} | User-Agent: {user_agent}")

            if not reset_link or not new_password:
                logger.warning(f"[Reset Password] Missing reset_link or new_password | IP: {ip}")
                return Response(
                    {"statusCode": 400, "status": False, "message": "Reset link and new password are required"},
                    status=status.HTTP_200_OK
                )

            parsed_url = urlparse(reset_link)
            query_params = parse_qs(parsed_url.query)

            user_id = query_params.get("user_id", [None])[0]
            token = query_params.get("token", [None])[0]

            if not user_id or not token:
                logger.warning(f"[Reset Password] Invalid reset link format | IP: {ip}")
                return Response(
                    {"statusCode": 400, "status": False, "message": "Invalid reset link"},
                    status=status.HTTP_200_OK
                )

            try:
                user_id = int(user_id)
            except ValueError:
                logger.warning(f"[Reset Password] Invalid user ID format: {user_id} | IP: {ip}")
                return Response(
                    {"statusCode": 400, "status": False, "message": "Invalid user ID"},
                    status=status.HTTP_200_OK
                )

            try:
                user = User.objects.get(pk=user_id, is_staff=True)
            except User.DoesNotExist:
                logger.warning(f"[Reset Password] User not found: ID={user_id} | IP: {ip}")
                return Response(
                    {"statusCode": 400, "status": False, "message": "Invalid user"},
                    status=status.HTTP_200_OK
                )

            if not default_token_generator.check_token(user, token):
                logger.warning(f"[Reset Password] Token invalid or expired for user ID={user_id} | IP: {ip}")
                return Response(
                    {"statusCode": 400, "status": False, "message": "Invalid or expired token"},
                    status=status.HTTP_200_OK
                )

            user.set_password(new_password)
            user.save()

            logger.info(f"[Reset Password] Password successfully reset for user ID={user_id} | IP: {ip}")
            return Response(
                {"statusCode": 200, "status": True, "message": "Password reset successful"},
                status=status.HTTP_200_OK
            )

        except Exception as e:
            logger.exception(f"[Reset Password] Unexpected error occurred | IP: {ip} | Error: {str(e)}")
            return Response(
                {"statusCode": 500, "status": False, "message": "An unexpected error occurred", "error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class AdminLogoutView(APIView):
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated]  

    @swagger_auto_schema(
        operation_description="Logs out an admin by blacklisting the provided refresh token.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["refresh_token"],
            properties={
                "refresh_token": openapi.Schema(type=openapi.TYPE_STRING, description="The JWT refresh token to be blacklisted")
            },
        ),
        responses={
            200: openapi.Response(description="Successfully logged out"),
            400: openapi.Response(description="Bad request"),
            401: openapi.Response(description="Unauthorized"),
        }
    )
    def post(self, request):
        ip = request.META.get("REMOTE_ADDR")
        user_agent = request.META.get("HTTP_USER_AGENT", "unknown")
        user = request.user

        try:
            refresh_token = request.data.get("refresh_token")  

            logger.info(f"[Admin Logout] Logout request received | Admin: {user.email} | IP: {ip} | User-Agent: {user_agent}")

            if not refresh_token:
                logger.warning(f"[Admin Logout] Missing refresh token | Admin: {user.email} | IP: {ip}")
                return Response(
                    {"statusCode": 400, "status": False, "message": "Refresh token is required"},
                    status=status.HTTP_200_OK
                )

            token = RefreshToken(refresh_token)
            token.blacklist()  

            logger.info(f"[Admin Logout] Refresh token successfully blacklisted | Admin: {user.email} | IP: {ip}")
            return Response(
                {"statusCode": 200, "status": True, "message": "Successfully logged out"},
                status=status.HTTP_200_OK
            )

        except TokenError as e:
            logger.warning(f"[Admin Logout] Invalid or expired token | Admin: {user.email} | IP: {ip} | Error: {str(e)}")
            return Response(
                {"statusCode": 400, "status": False, "message": "Invalid or expired refresh token"},
                status=status.HTTP_200_OK
            )

        except Exception as e:
            logger.exception(f"[Admin Logout] Unexpected error during logout | Admin: {user.email} | IP: {ip} | Error: {str(e)}")
            return Response(
                {"statusCode": 500, "status": False, "message": "An unexpected error occurred", "error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class AdminUserListView(APIView):
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated, IsAdministrator]
    
    def get(self, request):
        """Retrieve all admin users."""
        ip = request.META.get("REMOTE_ADDR")
        user = request.user
        user_agent = request.META.get("HTTP_USER_AGENT", "unknown")

        try:
            logger.info(f"[Admin User List] Request by {user.email} | IP: {ip} | User-Agent: {user_agent}")

            users = AdminUser.objects.all()
            serializer = AdminUserSerializer(users, many=True)
            paginator = PageNumberPagination()
            paginated_data = paginator.paginate_queryset(serializer.data, request)

            response = {
                "statusCode": 200,
                "status": True,
                "message": "User Get Successfully",
                "data": paginated_data,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
                "total_pages": paginator.page.paginator.num_pages,
                "total_items": paginator.page.paginator.count,
            }

            logger.info(f"[Admin User List] Successfully returned admin user list | Count: {len(paginated_data)}")
            return Response(response)

        except Exception as e:
            logger.exception(f"[Admin User List] Internal server error | Admin: {user.email} | Error: {str(e)}")
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)


class AdminCreateUserAPIView(APIView):
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated, IsAdministrator]

    @SwaggerView(AdminUserSerializer)
    def post(self, request, *args, **kwargs):
        ip = request.META.get("REMOTE_ADDR")
        user = request.user
        user_agent = request.META.get("HTTP_USER_AGENT", "unknown")

        logger.info(f"[Admin Create User] Create request by {user.email} | IP: {ip} | User-Agent: {user_agent}")

        serializer = AdminUserSerializer(data=request.data)
        if serializer.is_valid():
            new_user = serializer.save()

            logger.info(f"[Admin Create User] New admin user created | Created by: {user.email} | New Admin: {new_user.email}")

            response_data = {
                "statusCode": 200,
                "status": True,
                "message": "User Created Successfully",
                "data": serializer.data
            }
            return Response(response_data, status=status.HTTP_201_CREATED)
        
        logger.warning(f"[Admin Create User] Invalid data submitted | Admin: {user.email} | Errors: {serializer.errors}")

        response_data = {
            "statusCode": 400,
            "status": False,
            "message": "Invalid Data",
            "errors": serializer.errors
        }
        return Response(response_data, status=status.HTTP_200_OK)


        
 
class AdminUserDetailAPIView(APIView):
    """API to get, update, delete user by ID"""
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated, IsAdministrator]
    
    def get_object(self, pk):
        """Helper method to get user object by ID"""
        try:
            return AdminUser.objects.get(pk=pk)
        except AdminUser.DoesNotExist:
            return None

    def get(self, request, pk):
        """Retrieve a user by ID"""
        ip = request.META.get("REMOTE_ADDR")
        admin_user = request.user
        user_agent = request.META.get("HTTP_USER_AGENT", "unknown")

        try:
            logger.info(f"[Admin GET User] Admin: {admin_user.email} | IP: {ip} | User-Agent: {user_agent} | Target ID: {pk}")
            user = self.get_object(pk)
            if user:
                serializer = AdminUserSerializer(user)
                response = {
                    "statusCode": 200,
                    "status": True,
                    "message": "User Found Successfully",
                    "data": serializer.data
                }
                logger.info(f"[Admin GET User] Found user {user.email} by ID {pk}")
                return Response(response)

            logger.warning(f"[Admin GET User] User ID {pk} not found by admin {admin_user.email}")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "User Not Found !!"
            })

        except Exception as e:
            logger.exception(f"[Admin GET User] Error retrieving user ID {pk} | Admin: {admin_user.email}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            })

    @SwaggerView(AdminUserSerializer)
    def put(self, request, pk):
        """Update a user by ID"""
        ip = request.META.get("REMOTE_ADDR")
        admin_user = request.user
        user_agent = request.META.get("HTTP_USER_AGENT", "unknown")

        try:
            logger.info(f"[Admin UPDATE User] Admin: {admin_user.email} | IP: {ip} | User-Agent: {user_agent} | Target ID: {pk}")
            user = self.get_object(pk)
            if user:
                serializer = AdminUserSerializer(user, data=request.data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    logger.info(f"[Admin UPDATE User] Updated user ID {pk} successfully by admin {admin_user.email}")
                    return Response({
                        "statusCode": 200,
                        "status": True,
                        "message": "User Updated Successfully",
                        "data": serializer.data
                    })
                else:
                    logger.warning(f"[Admin UPDATE User] Validation failed | Admin: {admin_user.email} | Errors: {serializer.errors}")
                    return Response({
                        "statusCode": 400,
                        "status": False,
                        "message": serializer.errors
                    })
            
            logger.warning(f"[Admin UPDATE User] User ID {pk} not found | Admin: {admin_user.email}")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "User Not Found !!"
            })

        except Exception as e:
            logger.exception(f"[Admin UPDATE User] Exception while updating user ID {pk} | Admin: {admin_user.email}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            })
    
    def delete(self, request, pk=None):
        """Delete a user by ID"""
        ip = request.META.get("REMOTE_ADDR")
        admin_user = request.user
        user_agent = request.META.get("HTTP_USER_AGENT", "unknown")

        try:
            logger.info(f"[Admin DELETE User] Admin: {admin_user.email} | IP: {ip} | User-Agent: {user_agent} | Target ID: {pk}")
            user = self.get_object(pk)
            if user:
                user_email = user.email
                user.delete()
                logger.info(f"[Admin DELETE User] Deleted user {user_email} (ID: {pk}) by admin {admin_user.email}")
                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "User deleted successfully"
                })
            else:
                logger.warning(f"[Admin DELETE User] User ID {pk} not found by admin {admin_user.email}")
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "User not found"
                })

        except Exception as e:
            logger.exception(f"[Admin DELETE User] Error deleting user ID {pk} | Admin: {admin_user.email}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {e}"
            })
  
class AdminListAdministratorsAPIView(APIView):
    """Retrieve all users with role Administrator"""
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated, IsAdministrator]

    def get(self, request):
        ip = request.META.get("REMOTE_ADDR")
        admin_user = request.user
        user_agent = request.META.get("HTTP_USER_AGENT", "unknown")

        try:
            logger.info(f"[Admin FETCH Administrators] Admin: {admin_user.email} | IP: {ip} | User-Agent: {user_agent}")
            users = AdminUser.objects.filter(role__name="administrator") 

            if users.exists():
                serializer = AdminUserSerializer(users, many=True)
                logger.info(f"[Admin FETCH Administrators] {users.count()} administrator(s) retrieved successfully.")
                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Administrator users retrieved successfully",
                    "data": serializer.data
                }, status=status.HTTP_200_OK)

            logger.info("[Admin FETCH Administrators] No administrators found.")
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "No administrator users found",
                "data": []
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.exception(f"[Admin FETCH Administrators] Error occurred | Admin: {admin_user.email}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {e}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

import math

class AdminListUsersAPIView(APIView):
    """Retrieve all users with role User, with optional search functionality"""
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, IsAdministrator]

    def get(self, request):
        ip = request.META.get("REMOTE_ADDR")
        admin_user = request.user
        user_agent = request.META.get("HTTP_USER_AGENT", "unknown")
        search_query = request.query_params.get('search', None)

        try:
            logger.info(f"[Admin FETCH Users] Admin: {admin_user.email} | IP: {ip} | User-Agent: {user_agent}")
            users = Users.objects.filter(role__name="user")

            if search_query:
                users = users.filter(
                    username__icontains=search_query
                ) | users.filter(
                    email__icontains=search_query
                ) | users.filter(
                    phone__icontains=search_query
                ) | users.filter(
                    firstName__icontains=search_query
                ) | users.filter(
                    lastName__icontains=search_query
                ) | users.filter(
                    countryCode__icontains=search_query
                )

            if users.exists():
                page_size = request.query_params.get('page_size')
                if page_size:
                    paginator = CustomPagination()
                    page = paginator.paginate_queryset(users, request)
                    serializer = AdminProUserSerializer(page, many=True)

                    total_items = paginator.page.paginator.count
                    page_size_val = paginator.get_page_size(request) or paginator.page_size
                    total_pages = ceil(total_items / page_size_val)

                    return Response({
                        "statusCode": 200,
                        "status": True,
                        "message": "Users retrieved successfully",
                        "data": serializer.data,
                        "next": paginator.get_next_link(),
                        "previous": paginator.get_previous_link(),
                        "total_pages": total_pages,
                        "total_items": total_items
                    }, status=status.HTTP_200_OK)

                serializer = AdminProUserSerializer(users, many=True)
                logger.info(f"[Admin FETCH Users] {users.count()} user(s) retrieved successfully.")
                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Users retrieved successfully",
                    "data": serializer.data,
                    "next": None,
                    "previous": None,
                    "total_pages": 1,
                    "total_items": users.count()
                }, status=status.HTTP_200_OK)

            logger.info("[Admin FETCH Users] No users found.")
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "No users found",
                "data": [],
                "next": None,
                "previous": None,
                "total_pages": 0,
                "total_items": 0
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.exception(f"[Admin FETCH Users] Error occurred | Admin: {admin_user.email}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {e}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class AdminListProfessionalsAPIView(APIView):
    """Retrieve all Admin created Professional_users with role Professional"""
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated, IsAdministrator]

    def get(self, request):
        ip = request.META.get("REMOTE_ADDR")
        user_agent = request.META.get("HTTP_USER_AGENT", "unknown")
        admin_user = request.user

        try:
            logger.info(f"[Admin FETCH Professionals] Admin: {admin_user.email} | IP: {ip} | User-Agent: {user_agent}")
            users = AdminUser.objects.filter(role__name="professionaluser")  

            if users.exists():
                serializer = AdminUserSerializer(users, many=True)
                logger.info(f"[Admin FETCH Professionals] {users.count()} professional user(s) retrieved.")
                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Professional users retrieved successfully",
                    "data": serializer.data
                }, status=status.HTTP_200_OK)

            logger.info("[Admin FETCH Professionals] No professional users found.")
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "No professional users found",
                "data": []
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.exception(f"[Admin FETCH Professionals] Error occurred | Admin: {admin_user.email}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {e}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
import math

class ProfessionalUserListView(generics.ListAPIView):
    """
    Fetch all Professional Users with pagination, search, and error handling
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    serializer_class = ProfessionalUserListSerializer
    pagination_class = CustomPagination

    def get_queryset(self):
        """Filter and search professional users dynamically"""
        search_query = self.request.query_params.get("search", None)
        queryset = ProfessionalUser.objects.all()

        if search_query:
            queryset = queryset.filter(
                Q(company__companyName__icontains=search_query) |
                Q(userName__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(phone__icontains=search_query) |
                Q(kbiss_status__icontains=search_query) |
                Q(iban_status__icontains=search_query) |
                Q(identityCardFront_status__icontains=search_query) |
                Q(identityCardBack_status__icontains=search_query) |
                Q(proofOfAddress_status__icontains=search_query) |
                Q(finalDocument_status__icontains=search_query)
            )
        return queryset

    def list(self, request, *args, **kwargs):
        try:
            queryset = self.get_queryset()
            page = self.paginate_queryset(queryset)

            if page is not None:
                serialized_data = self.get_serializer(page, many=True).data

                # Access pagination instance and metadata
                paginator = self.paginator
                total_items = paginator.page.paginator.count
                page_size_val = paginator.get_page_size(request) or paginator.page_size
                total_pages = math.ceil(total_items / page_size_val)

                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Professional Users retrieved successfully",
                    "data": serialized_data,
                    "next": paginator.get_next_link(),
                    "previous": paginator.get_previous_link(),
                    "total_pages": total_pages,
                    "total_items": total_items
                }, status=status.HTTP_200_OK)

            # If no pagination applied
            serialized_data = self.get_serializer(queryset, many=True).data
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Professional Users retrieved successfully",
                "data": serialized_data,
                "next": None,
                "previous": None,
                "total_pages": 1,
                "total_items": len(serialized_data)
            }, status=status.HTTP_200_OK)

        except FieldError as fe:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": f"Invalid search field: {str(fe)}",
                "data": []
            }, status=status.HTTP_200_OK)

        except DatabaseError as db_error:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Database error occurred. Please try again later.",
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Unexpected error: {str(e)}",
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# new based on final status approved or rejected/pending
from math import ceil
class ProfessionalUserListViewNew(ListAPIView):
    serializer_class = ProfessionalUserSerializer
    pagination_class = CustomPagination

    def get_queryset(self):
        queryset = ProfessionalUser.objects.all().order_by('-created_at')


        # Filter by finalDocument_status
        status_filter = self.request.query_params.get("finalDocument_status", "").lower()
        if status_filter == "approved":
            queryset = queryset.filter(finalDocument_status="approved")
        elif status_filter:
            queryset = queryset.filter(finalDocument_status__in=["pending", "rejected"])

        # Search functionality
        search_query = self.request.query_params.get("search", "").strip()
        if search_query:
            queryset = queryset.filter(
                Q(userName__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(company__companyName__icontains=search_query) |
                Q(role__name__icontains=search_query)
            )

        # Filter by subscription plan name
        subscription_plan = self.request.query_params.get("subscriptionplan", "").strip()
        if subscription_plan:
            queryset = queryset.filter(subscriptionplan__name__icontains=subscription_plan)

        # Filter by subscription type
        subscription_type = self.request.query_params.get("subscriptiontype", "").strip()
        if subscription_type:
            queryset = queryset.filter(subscriptiontype__subscription_type__icontains=subscription_type)

        return queryset



    def list(self, request, *args, **kwargs):
        try:
            queryset = self.get_queryset()

            if not queryset.exists():
                return Response({
                    "statusCode": 404,
                    "status": False,
                    "message": "No professional users found matching the criteria",
                    "data": [],
                    "next": None,
                    "previous": None,
                    "total_pages": 0,
                    "total_items": 0
                }, status=status.HTTP_404_NOT_FOUND)

            paginator = self.pagination_class()
            page = paginator.paginate_queryset(queryset, request, view=self)
            serializer = self.get_serializer(page, many=True)

            total_items = paginator.page.paginator.count
            total_pages = ceil(total_items / paginator.get_page_size(request))

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Professional users retrieved successfully",
                "data": serializer.data,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
                "total_pages": total_pages,
                "total_items": total_items
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred while fetching users",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class GetProfessionalUserAPIView(APIView):
    """Retrieve a professional user's details by ID"""

    def get(self, request, user_id):
        try:
            professional_user = ProfessionalUser.objects.get(id=user_id)
            serialized_user = ProfessionalUserListSerializer(
                professional_user, context={"request": request}
            ).data

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Professional user details retrieved successfully",
                "user_details": serialized_user
            }, status=status.HTTP_200_OK)

        except ProfessionalUser.DoesNotExist:
            logger.warning(f"Professional user with ID {user_id} not found.")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Professional user not found"
            }, status=status.HTTP_200_OK)

        except ValidationError as e:
            logger.warning(f"Validation error while fetching professional user: {e.messages}")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Invalid data",
                "errors": e.messages
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Unexpected error in GetProfessionalUserAPIView: {str(e)}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def get_time_difference(last_login):
    """Calculate time difference and return a readable string."""
    if not last_login:
        return "Never Logged In"
    
    delta = now() - last_login
    if delta < timedelta(minutes=1):
        return "Just now"
    elif delta < timedelta(hours=1):
        return f"{int(delta.seconds / 60)} min ago"
    elif delta < timedelta(days=1):
        return f"{int(delta.seconds / 3600)} h ago"
    elif delta < timedelta(days=30):
        return f"{delta.days} days ago"
    elif delta < timedelta(days=365):
        return f"{int(delta.days / 30)} months ago"
    else:
        return f"{int(delta.days / 365)} years ago"

class AdminUserLoginDetailsView(generics.ListAPIView):
    """
    API View for Admin to view user login details with search and filtering.
    """
    queryset = Users.objects.filter(is_active=True).order_by('-last_login')
    serializer_class = UserLoginDetailSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['gender', 'is_active']
    search_fields = ['username', 'email', 'phone']

    def get_queryset(self):
        """Apply 'is_currently_logged_in' filter if present."""
        queryset = super().get_queryset()
        is_currently_logged_in = self.request.query_params.get('is_currently_logged_in')
        
        if is_currently_logged_in is not None:
            is_currently_logged_in = is_currently_logged_in.lower() == 'true'
            if is_currently_logged_in:
                queryset = queryset.filter(last_login__gte=now() - timedelta(minutes=30))
            else:
                queryset = queryset.exclude(last_login__gte=now() - timedelta(minutes=30))
                
        return queryset

    def list(self, request, *args, **kwargs):
        try:
            queryset = self.filter_queryset(self.get_queryset())
            page = self.paginate_queryset(queryset)

            # Pagination details
            paginator = self.paginator
            total_pages = paginator.page.paginator.num_pages if paginator.page.paginator.num_pages > 0 else 1
            total_items = paginator.page.paginator.count

            if page is not None:
                serializer = self.get_serializer(page, many=True)
                user_data = serializer.data

                for user in user_data:
                    last_login = user.get("last_login")
                    if last_login:
                        last_login = datetime.strptime(last_login, "%Y-%m-%dT%H:%M:%S.%f%z")
                    user["last_login_formatted"] = get_time_difference(last_login)
                    user["is_currently_logged_in"] = last_login is not None and (now() - last_login) < timedelta(minutes=30)

                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "User login details retrieved successfully",
                    "next": paginator.get_next_link(),
                    "previous": paginator.get_previous_link(),
                    "total_pages": total_pages,
                    "total_items": total_items,
                    "user_details": user_data
                }, status=status.HTTP_200_OK)

            serializer = self.get_serializer(queryset, many=True)
            user_data = serializer.data

            for user in user_data:
                last_login = user.get("last_login")
                if last_login:
                    last_login = datetime.strptime(last_login, "%Y-%m-%dT%H:%M:%S.%f%z")
                user["last_login_formatted"] = get_time_difference(last_login)
                user["is_currently_logged_in"] = last_login is not None and (now() - last_login) < timedelta(minutes=30)

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "User login details retrieved successfully",
                "next": None,
                "previous": None,
                "total_pages": 1,
                "total_items": len(user_data),
                "user_details": user_data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"An error occurred: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

class AdminProfessionalUserLoginDetailsView(generics.ListAPIView):
    """
    API View for Admin to view professional user login details with search and filtering.
    """
    queryset = ProfessionalUser.objects.filter().order_by('-last_login')
    serializer_class = ProfessionalUserLoginDetailSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['kbiss_status', 'iban_status', 'is_verified', 'subscription_active']
    search_fields = ['userName', 'email', 'phone']

    def get_queryset(self):
        """Apply 'is_currently_logged_in' filter if present."""
        queryset = super().get_queryset()
        is_currently_logged_in = self.request.query_params.get('is_currently_logged_in')
        
        if is_currently_logged_in is not None:
            is_currently_logged_in = is_currently_logged_in.lower() == 'true'
            if is_currently_logged_in:
                queryset = queryset.filter(last_login__gte=now() - timedelta(minutes=30))
            else:
                queryset = queryset.exclude(last_login__gte=now() - timedelta(minutes=30))
                
        return queryset

    def list(self, request, *args, **kwargs):
        try:
            queryset = self.filter_queryset(self.get_queryset())
            page = self.paginate_queryset(queryset)

            # Pagination details
            paginator = self.paginator
            total_pages = paginator.page.paginator.num_pages if paginator.page.paginator.num_pages > 0 else 1
            total_items = paginator.page.paginator.count

            if page is not None:
                serializer = self.get_serializer(page, many=True)
                user_data = serializer.data

                for user in user_data:
                    last_login = user.get("last_login")
                    if last_login:
                        last_login = datetime.strptime(last_login, "%Y-%m-%dT%H:%M:%S.%f%z")
                    user["last_login_formatted"] = get_time_difference(last_login)
                    user["is_currently_logged_in"] = last_login is not None and (now() - last_login) < timedelta(minutes=30)

                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Professional user login details retrieved successfully",
                    "next": paginator.get_next_link(),
                    "previous": paginator.get_previous_link(),
                    "total_pages": total_pages,
                    "total_items": total_items,
                    "user_details": user_data
                }, status=status.HTTP_200_OK)

            serializer = self.get_serializer(queryset, many=True)
            user_data = serializer.data

            for user in user_data:
                last_login = user.get("last_login")
                if last_login:
                    last_login = datetime.strptime(last_login, "%Y-%m-%dT%H:%M:%S.%f%z")
                user["last_login_formatted"] = get_time_difference(last_login)
                user["is_currently_logged_in"] = last_login is not None and (now() - last_login) < timedelta(minutes=30)

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Professional user login details retrieved successfully",
                "next": None,
                "previous": None,
                "total_pages": 1,
                "total_items": len(user_data),
                "user_details": user_data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"An error occurred: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UserAppAPIView(APIView):
    """Retrieve a general user’s details by ID"""

    def get(self, request, user_id):
        """Fetch user details by their ID"""
        try:
            user = Users.objects.get(id=user_id)
            serialized_user = AdminProUserSerializer(user).data

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "User details retrieved successfully",
                "user_details": serialized_user
            }, status=status.HTTP_200_OK)

        except Users.DoesNotExist:
            logger.warning(f"User with ID {user_id} not found.")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "user not found"
            }, status=status.HTTP_200_OK)

        except ValidationError as e:
            logger.warning(f"Validation error for user ID {user_id}: {e.messages}")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Invalid data",
                "errors": e.messages
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Unexpected error in UserAppAPIView: {str(e)}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
       


class UpdateUserByIdAPIView(APIView):
    def put(self, request, user_id):
        try:
            user = get_object_or_404(Users, id=user_id)
            previous_status = user.is_active  # Store the previous status
            serializer = UserUpdateSerializer(user, data=request.data, partial=True)

            if serializer.is_valid():
                serializer.save()

                # Check if 'is_active' status has been changed
                if 'is_active' in request.data and request.data['is_active'] != previous_status:
                    warning_message = request.data.get('warning_message', "User status has been changed. Please review the account settings.")
                    
                    # Save the warning message in the model if provided
                    user.warning_message = warning_message
                    user.save()

                    return Response({
                        "message": "User updated successfully",
                        "statusCode": 200,
                        "status": True,
                        "data": serializer.data,
                        "warning": warning_message
                    }, status=status.HTTP_200_OK)

                return Response({
                    "message": "User updated successfully",
                    "statusCode": 200,
                    "status": True,
                    "data": serializer.data
                }, status=status.HTTP_200_OK)

            else:
                return Response({
                    "message": "Validation failed",
                    "statusCode": 400,
                    "status": False,
                    "errors": serializer.errors
                }, status=status.HTTP_200_OK)

        except Users.DoesNotExist:
            return Response({
                "message": "User not found",
                "statusCode": 404,
                "status": False
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Unexpected error updating user: {str(e)}", exc_info=True)
            return Response({
                "message": "An unexpected error occurred",
                "statusCode": 500,
                "status": False,
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class AdminRoleCreateAPIView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, IsAdministrator]

    @SwaggerView(RoleSerializer)
    def post(self, request):
        try:
            serializer = RoleSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                response = {
                    "statusCode": 200,
                    "status": True,
                    "message": "Role Created Successfully",
                    "data": serializer.data
                }
                return Response(response)
            else:
                logger.warning(f"Role creation failed due to validation error: {serializer.errors}")
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": serializer.errors
                })
        except Exception as e:
            logger.error(f"Unexpected error while creating role: {str(e)}")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": f"Internal server error {e}"
            })

    def get(self, request):
        """Retrieve all admin roles with search capabilities."""
        try:
            search_query = request.query_params.get('search')
            roles = Role.objects.all()

            if search_query:
                roles = roles.filter(name__icontains=search_query)

            serializer = RoleSerializer(roles, many=True)
            paginator = PageNumberPagination()
            paginated_data = paginator.paginate_queryset(serializer.data, request)

            response = {
                "statusCode": 200,
                "status": True,
                "message": "Role Get Successfully",
                "data": paginated_data,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
                "total_pages": paginator.page.paginator.num_pages,
                "total_items": paginator.page.paginator.count,
            }
            return Response(response)
        except Exception as e:
            logger.error(f"Unexpected error while fetching roles: {str(e)}")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": f"Internal server error {e}"
            })



# class AdminUserCreateView(generics.CreateAPIView):# this vews is not in use
#     queryset = Users.objects.all()
#     serializer_class = AdminUserCreateSerializer
#     permission_classes = [IsAdminAccount]  # Use custom permission

#     def create(self, request, *args, **kwargs):
#         serializer = self.get_serializer(data=request.data)
#         serializer.is_valid(raise_exception=True)
#         user = serializer.save()
        
#         return Response( {
#                 "statusCode": 200,
#                 "status": True,
#                 "message": "user created successfully"
#             })

#     @SwaggerView(RoleSerializer)
#     def put(self, request, pk):
#         """Update a role by ID"""
#         try:
#             role = self.get_object(pk)
#             if role:
#                 serializer = RoleSerializer(role,data=request.data, partial=True)
#                 if serializer.is_valid():
#                     serializer.save()
#                     response = {"statusCode":200,
#                                 "status":True,
#                                 "message":"Role Updated Successfully",
#                                 "data":serializer.data}
#                     return Response(response)
#                 else:
#                     return Response({"statusCode": 400, "status": False, "message": serializer.errors})
#             response = {"statusCode":400,
#                         "status":False,
#                         "message":"Role Not Found !!"
#                         }
#             return Response(response)
        
#         except Exception as e:
#             response = {
#                 "statusCode": 500,
#                 "status": False,
#                 "message":"Server Error"
#             }


# # Admin Update User API
# class AdminUserUpdateView(generics.UpdateAPIView):
#     queryset = Users.objects.all()
#     serializer_class = AdminUserUpdateSerializer
#     permission_classes = [IsAdminUser]  # Only Admins can update users
#     lookup_field = 'id'  # Identify user by ID

#     def update(self, request, *args, **kwargs):
#         partial = kwargs.pop('partial', False)
#         instance = self.get_object()
#         serializer = self.get_serializer(instance, data=request.data, partial=partial)
#         serializer.is_valid(raise_exception=True)
#         serializer.save()
        

    
#     def delete(self, request, pk=None):
#         """Delete a role by ID or all roles"""
#         try:
#             role = self.get_object(pk)
#             if role:
#                 role.delete()
#                 response = {
#                     "statusCode": 200,
#                     "status": True,
#                     "message": "Role deleted successfully"
#                 }
#                 return Response(response)
#             else:
#                 response = {
#                     "statusCode": 400,
#                     "status": False,
#                     "message": "Role not found"
#                 }
#                 return Response(response)
        
#         except Exception as e:
#             response = {
#                 "statusCode": 500,
#                 "status": False,
#                 "message": f"Internal server error: {e}"
#             }
#             return Response(response)

    
#     def get(self, request):
#         """Retrieve all admin roles."""
#         try:
#             roles = Role.objects.all()
#             serializer = RoleSerializer(roles,many=True)
#             paginator = PageNumberPagination()
#             paginated_data = paginator.paginate_queryset(serializer.data, request)
#             response = {"statusCode":200,
#                 "status":True,
#                 "message":"Role Get Successfully",
#                 "data":paginated_data,
#                 "next": paginator.get_next_link(),
#                 "previous": paginator.get_previous_link(),
#                 "total_pages": paginator.page.paginator.num_pages,
#                 "total_items": paginator.page.paginator.count,
#             }
#             return Response(response)
#         except Exception as e:
#             response = {
#                 "statusCode": 500,
#                 "status": False,
#                 "message": f"Internal server error {e}"
#             }
#             return Response(response)
 
class AdminRoleDetailAPIView(APIView):
    """API to get, update, delete role by ID & delete all roles"""

    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, IsAdministrator]

    def get_object(self, pk):
        """Helper method to get role object by ID"""
        try:
            return Role.objects.get(pk=pk)
        except Role.DoesNotExist:
            return None

    def get(self, request, pk):
        """Retrieve a role by ID"""
        try:
            role = self.get_object(pk)
            if role:
                serializer = RoleSerializer(role)
                response = {
                    "statusCode": 200,
                    "status": True,
                    "message": "Role Found Successfully",
                    "data": [serializer.data]  # Wrap in array of objects
                }
                return Response(response)
            logger.warning(f"Role with ID {pk} not found.")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Role Not Found !!"
            })
        except Exception as e:
            logger.error(f"Error retrieving role with ID {pk}: {str(e)}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            })

    @SwaggerView(RoleSerializer)
    def put(self, request, pk):
        """Update a role by ID"""
        try:
            role = self.get_object(pk)
            if role:
                serializer = RoleSerializer(role, data=request.data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    response = {
                        "statusCode": 200,
                        "status": True,
                        "message": "Role Updated Successfully",
                        "data": [serializer.data]  # Wrap in array of objects
                    }
                    return Response(response)
                logger.warning(f"Validation failed while updating role {pk}: {serializer.errors}")
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": serializer.errors
                })
            logger.warning(f"Role with ID {pk} not found for update.")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Role Not Found !!"
            })
        except Exception as e:
            logger.error(f"Error updating role with ID {pk}: {str(e)}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            })

    def delete(self, request, pk=None):
        """Delete a role by ID"""
        try:
            role = self.get_object(pk)
            if role:
                role.delete()
                logger.info(f"Role with ID {pk} deleted successfully.")
                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Role deleted successfully"
                })
            logger.warning(f"Role with ID {pk} not found for deletion.")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Role not found"
            })
        except Exception as e:
            logger.error(f"Error deleting role with ID {pk}: {str(e)}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {e}"
            })
    
class MenuListCreateAPIView(APIView):
    """API for listing all menus and creating a new menu"""
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated,IsAdministrator]   
    def get(self, request):
        """Get a list of all menus"""
        try:
            menus = Menu.objects.all()
            serializer = MenuSerializer(menus,many=True)
            paginator = PageNumberPagination()
            paginated_data = paginator.paginate_queryset(serializer.data, request)
            response = {"statusCode":200,
                "status":True,
                "message":"Menus Get Successfully",
                "data":paginated_data,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
                "total_pages": paginator.page.paginator.num_pages,
                "total_items": paginator.page.paginator.count,
            }
            return Response(response)
        except Exception as e:
            response = {
                "statusCode": 400,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)
    @SwaggerView(MenuSerializer)
    def post(self, request):
        """Create a new menu"""
        
        try:
            serializer = MenuSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                response = {"statusCode":200,
                            "status":True,
                            "message":"Menu Created Successfully",
                            "data":serializer.data}
            else:
                return Response({"statusCode": 400, "status": False, "message": serializer.errors})
            return Response(response)
        except Exception as e:
            response = {
                "statusCode": 400,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)

class MenuDetailAPIView(APIView):
    """API for getting, updating, and deleting a menu by ID"""

    def get_object(self, pk):
        """Helper method to get menu object by ID"""
        try:
            return Menu.objects.get(pk=pk)
        except Menu.DoesNotExist:
            return None
   
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated,IsAdministrator]   
    def get(self, request, pk):
        """Retrieve a menu by ID"""
        try:
            menu = self.get_object(pk)
            if menu:
                serializer = MenuSerializer(menu)
                response = {"statusCode":200,
                            "status":True,
                            "message":"Menu Found Successfully",
                            "data":serializer.data}
                return Response(response)
            response = {"statusCode":400,
                        "status":False,
                        "message":"Menu Not Found !!"
                        }
            return Response(response)
        
        except Exception as e:
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)
        
    @SwaggerView(MenuSerializer) 
    def put(self, request, pk):
        """Update a menu by ID"""
        try:
            menu = self.get_object(pk)
            if menu:
                serializer = MenuSerializer(menu,data=request.data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    response = {"statusCode":200,
                                "status":True,
                                "message":"Menu Updated Successfully",
                                "data":serializer.data}
                    return Response(response)
                else:
                    return Response({"statusCode": 500, "status": False, "message": serializer.errors})
            response = {"statusCode":400,
                        "status":False,
                        "message":"Menu Not Found !!"
                        }
            return Response(response)
        
        except Exception as e:
            response = {
                "statusCode": 400,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)

    def delete(self, request, pk):
        """Delete a menu by ID"""
        try:
            menu = self.get_object(pk)
            if menu:
                menu.delete()
                response_data = {
                    "statusCode": 200,
                    "status": True,
                    "message": "Menu Deleted successfully"
                }
                return Response(response_data)
            else:
                response_data = {
                    "statusCode": 400,
                    "status": False,
                    "message": "Menu Not Found"
                }
                return Response(response_data)
        except Exception as e:
            response_data = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            }
            return Response(response_data)
    
    
class SubmenuListCreateAPIView(APIView):
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated,IsAdministrator]  
    def get(self, request):
        """Get a list of all submenus"""
        try:
            submenus = Submenu.objects.all()
            serializer = SubmenuSerializer(submenus,many=True)
            paginator = PageNumberPagination()
            paginated_data = paginator.paginate_queryset(serializer.data, request)
            response = {"statusCode":200,
                "status":True,
                "message":"Submenus Get Successfully",
                "data":paginated_data,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
                "total_pages": paginator.page.paginator.num_pages,
                "total_items": paginator.page.paginator.count,
            }
            return Response(response)
        except Exception as e:
            print(f"error's {e}")
            response = {
                "statusCode": 400,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)
        
    @SwaggerView(SubmenuSerializer)
    def post(self, request):
        """Create a new submenu"""
        try:
            serializer = SubmenuSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                response = {"statusCode":200,
                            "status":True,
                            "message":"Submenu Created Successfully",
                            "data":serializer.data}
            else:
                return Response({"statusCode": 400, "status": False, "message": serializer.errors})
            return Response(response)
        except Exception as e:
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)


class SubmenuDetailAPIView(APIView):
    def get_object(self, pk):
        """Retrieve submenu object or return 400"""
        try:
            return Submenu.objects.get(pk=pk, is_deleted=False)
        except Submenu.DoesNotExist:
            return None
        
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated,IsAdministrator]  
    def get(self, request, pk):
        """Get submenu by ID"""
        try:
            submenu = self.get_object(pk)
            if submenu:
                serializer = SubmenuSerializer(submenu)
                response = {"statusCode":200,
                            "status":True,
                            "message":"Submenu Found Successfully",
                            "data":serializer.data}
                return Response(response)
            response = {"statusCode":400,
                        "status":False,
                        "message":"Submenu Not Found !!"
                        }
            return Response(response)
        
        except Exception as e:
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)
        
    @SwaggerView(SubmenuSerializer)
    def put(self, request, pk):
        """Update submenu by ID"""

        try:
            submenu = self.get_object(pk)
            if submenu:
                serializer = SubmenuSerializer(submenu,data=request.data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    response = {"statusCode":200,
                                "status":True,
                                "message":"Submenu Updated Successfully",
                                "data":serializer.data}
                    return Response(response)
                else:
                    return Response({"statusCode": 400, "status": False, "message": serializer.errors})
            response = {"statusCode":400,
                        "status":False,
                        "message":"Submenu Not Found !!"
                        }
            return Response(response)
        
        except Exception as e:
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)

    def delete(self, request, pk):
        """Delete submenu by ID"""
        try:
            submenu = self.get_object(pk)
            if submenu:
                submenu.delete()
                response_data = {
                    "statusCode": 200,
                    "status": True,
                    "message": "Submenu Deleted successfully"
                }
                return Response(response_data)
            else:
                response_data = {
                    "statusCode": 400,
                    "status": False,
                    "message": "Submenu Not Found"
                }
                return Response(response_data)
        except Exception as e:
            response_data = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            }
            return Response(response_data)
        
        
        
class ModuleListCreateAPIView(APIView):
    """API for listing all Module and creating a new menu"""
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated,IsAdministrator]  
    def get(self, request):
        """Get a list of all Module"""
        try:
            module = Module.objects.all()
            serializer = ModuleSerializer(module,many=True)
            paginator = PageNumberPagination()
            paginated_data = paginator.paginate_queryset(serializer.data, request)
            response = {"statusCode":200,
                "status":True,
                "message":"Module Get Successfully",
                "data":paginated_data,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
                "total_pages": paginator.page.paginator.num_pages,
                "total_items": paginator.page.paginator.count,
            }
            return Response(response)
        except Exception as e:
            response = {
                "statusCode": 400,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)
        
    @SwaggerView(ModuleSerializer)
    def post(self, request):
        """Create a new menu"""
        
        try:
            serializer = ModuleSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                response = {"statusCode":200,
                            "status":True,
                            "message":"Module Created Successfully",
                            "data":serializer.data}
            else:
                return Response({"statusCode": 400, "status": False, "message": serializer.errors})
            return Response(response)
        except Exception as e:
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)
        

class ModuleDetailAPIView(APIView):
    """API for retrieving, updating, and deleting a module by ID"""

    def get_object(self, pk):
        try:
            return Module.objects.get(pk=pk)
        except Module.DoesNotExist:
            return None
        
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated,IsAdministrator]  
    def get(self, request, pk):
        """Retrieve a module by ID"""
        try:
            module = self.get_object(pk)
            if module:
                serializer = ModuleSerializer(module)
                response = {"statusCode":200,
                            "status":True,
                            "message":"Module Found Successfully",
                            "data":serializer.data}
                return Response(response)
            response = {"statusCode":404,
                        "status":False,
                        "message":"Module Not Found !!"
                        }
            return Response(response)
        
        except Exception as e:
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)
        
    @SwaggerView(ModuleSerializer)
    def put(self, request, pk):
        """Update a module by ID"""
    
        try:
            module = self.get_object(pk)
            if module:
                serializer = ModuleSerializer(module,data=request.data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    response = {"statusCode":200,
                                "status":True,
                                "message":"Module Updated Successfully",
                                "data":serializer.data}
                    return Response(response)
                else:
                    return Response({"statusCode": 400, "status": False, "message": serializer.errors})
            response = {"statusCode":400,
                        "status":False,
                        "message":"Module Not Found !!"
                        }
            return Response(response)
        
        except Exception as e:
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)

    def delete(self, request, pk):
        """Delete a module by ID"""
           
        try:
            module = self.get_object(pk)
            if module:
                module.delete()
                response_data = {
                    "statusCode": 200,
                    "status": True,
                    "message": "Module Deleted successfully"
                }
                return Response(response_data)
            else:
                response_data = {
                    "statusCode": 400,
                    "status": False,
                    "message": "Module Not Found"
                }
                return Response(response_data)
        except Exception as e:
            response_data = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            }
            return Response(response_data)
        
        
class ModuleStatusDropdownAPIView(APIView):
    """API for fetching module status dropdown options"""
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated,IsAdministrator]  
    def get(self, request):
        try:
            dropdown_options = ['active', 'inactive', 'pending'] 
            return Response({"dropdownOptions": dropdown_options})
        except Exception as e:
            return Response({"error": str(e)}, status=500)        


class SaleListCreateAPIView(APIView):
    """API for listing all Sales and creating a new Sale"""
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated,IsAdministrator]  
    def get(self, request):
        """Get a list of all sales"""
            
        try:
            sales = Sale.objects.all()
            serializer = SaleSerializer(sales,many=True)
            paginator = PageNumberPagination()
            paginated_data = paginator.paginate_queryset(serializer.data, request)
            response = {"statusCode":200,
                "status":True,
                "message":"Sales Get Successfully",
                "data":paginated_data,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
                "total_pages": paginator.page.paginator.num_pages,
                "total_items": paginator.page.paginator.count,
            }
            return Response(response)
        except Exception as e:
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)
        
    @SwaggerView(SaleSerializer)
    def post(self, request):
        """Create a new sale"""
        try:
            serializer = SaleSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                response = {"statusCode":200,
                            "status":True,
                            "message":"Sale Created Successfully",
                            "data":serializer.data}
            else:
                return Response({"statusCode": 500, "status": False, "message": serializer.errors})
            return Response(response)
        except Exception as e:
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)
            
            
class SaleDetailAPIView(APIView):
    """API for retrieving, updating, and deleting a sale by ID"""

    def get_object(self, pk):
        try:
            return Sale.objects.get(pk=pk)
        except Sale.DoesNotExist:
            return None
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated,IsAdministrator]   
    def get(self, request, pk):
        """Retrieve a sale by ID"""
        try:
            sale = self.get_object(pk)
            if sale:
                serializer = SaleSerializer(sale)
                response = {"statusCode":200,
                            "status":True,
                            "message":"Sale Found Successfully",
                            "data":serializer.data}
                return Response(response)
            response = {"statusCode":400,
                        "status":False,
                        "message":"Sale Not Found !!"
                        }
            return Response(response)
        
        except Exception as e:
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)
    
    @SwaggerView(SaleSerializer)
    def put(self, request, pk):
        """Update a sale by ID"""    
        try:
            sale = self.get_object(pk)
            if sale:
                serializer = SaleSerializer(sale,data=request.data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    response = {"statusCode":200,
                                "status":True,
                                "message":"Sale Updated Successfully",
                                "data":serializer.data}
                    return Response(response)
                else:
                    return Response({"statusCode": 500, "status": False, "message": serializer.errors})
            response = {"statusCode":400,
                        "status":False,
                        "message":"Sale Not Found !!"
                        }
            return Response(response)
        
        except Exception as e:
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)


    def delete(self, request, pk):
        """Delete a sale by ID"""
        try:
            sale = self.get_object(pk)
            if sale:
                sale.delete()
                response_data = {
                    "statusCode": 200,
                    "status": True,
                    "message": "Sale Deleted successfully"
                }
                return Response(response_data)
            else:
                response_data = {
                    "statusCode": 400,
                    "status": False,
                    "message": "Sale Not Found"
                }
                return Response(response_data)
        except Exception as e:
            response_data = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            }
            return Response(response_data)
        

class SubmoduleListCreateAPIView(APIView):
    """API for listing all submodules and creating a new submodule"""
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated,IsAdministrator]  
    def get(self, request):
        """Get a list of all submodules"""
        try:
            submodules = Submodule.objects.all()
            serializer = SubmoduleSerializer(submodules,many=True)
            paginator = PageNumberPagination()
            paginated_data = paginator.paginate_queryset(serializer.data, request)
            response = {"statusCode":200,
                "status":True,
                "message":"Submodules Get Successfully",
                "data":paginated_data,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
                "total_pages": paginator.page.paginator.num_pages,
                "total_items": paginator.page.paginator.count,
            }
            return Response(response)
        except Exception as e:
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)
    
    @SwaggerView(SubmoduleSerializer)
    def post(self, request):
        """Create a new submodule"""
        try:
            serializer = SubmoduleSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                response = {"statusCode":200,
                            "status":True,
                            "message":"Submodule Created Successfully",
                            "data":serializer.data}
            else:
                return Response({"statusCode": 400, "status": False, "message": serializer.errors})
            return Response(response)
        except Exception as e:
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)


class SubmoduleDetailAPIView(APIView):
    """API for retrieving, updating, and deleting a submodule by ID"""

    def get_object(self, pk):
        try:
            return Submodule.objects.get(pk=pk)
        except Submodule.DoesNotExist:
            return None
        
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated,IsAdministrator]       
    def get(self, request, pk):
        """Retrieve a submodule by ID"""
        try:
            submodule = self.get_object(pk)
            if submodule:
                serializer = SubmoduleSerializer(submodule)
                response = {"statusCode":200,
                            "status":True,
                            "message":"Submodule Found Successfully",
                            "data":serializer.data}
                return Response(response)
            response = {"statusCode":400,
                        "status":False,
                        "message":"Submodule Not Found !!"
                        }
            return Response(response)
        
        except Exception as e:
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)
        
    @SwaggerView(SubmoduleSerializer)
    def put(self, request, pk):
        """Update a submodule by ID"""
        try:
            submodule = self.get_object(pk)
            if submodule:
                serializer = SubmoduleSerializer(submodule,data=request.data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    response = {"statusCode":200,
                                "status":True,
                                "message":"Submodule Updated Successfully",
                                "data":serializer.data}
                    return Response(response)
                else:
                    return Response({"statusCode": 500, "status": False, "message": serializer.errors})
            response = {"statusCode":400,
                        "status":False,
                        "message":"Submodule Not Found !!"
                        }
            return Response(response)
        
        except Exception as e:
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)

    def delete(self, request, pk):
        """Delete a submodule by ID"""
        try:
            submodule = self.get_object(pk)
            if submodule:
                submodule.delete()
                response_data = {
                    "statusCode": 200,
                    "status": True,
                    "message": "Submodule Deleted successfully"
                }
                return Response(response_data)
            else:
                response_data = {
                    "statusCode": 400,
                    "status": False,
                    "message": "Submodule Not Found"
                }
                return Response(response_data)
        except Exception as e:
            response_data = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            }
            return Response(response_data)


# Logger setup for error tracking


class LanguageListCreateAPIView(APIView):
    """API for listing all languages and creating a new language"""

    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)

    def get_permissions(self):
        """Set different permissions for GET and POST requests"""
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated(), IsAdministrator()]

    def get(self, request):
        """Get a list of all languages with pagination"""
        try:
            languages = Language.objects.all().order_by("id")
            paginator = PageNumberPagination()
            paginator.page_size = 30  # Adjust page size as needed
            
            paginated_data = paginator.paginate_queryset(languages, request)
            serialized_data = LanguageSerializer(paginated_data, many=True).data if paginated_data else LanguageSerializer(languages, many=True).data

            response = {
                "statusCode": 200,
                "status": True,
                "message": "Language list retrieved successfully",
                "data": serialized_data,
                "next": paginator.get_next_link() if paginated_data else None,
                "previous": paginator.get_previous_link() if paginated_data else None,
                "total_pages": paginator.page.paginator.num_pages if paginated_data else 1,
                "total_items": paginator.page.paginator.count if paginated_data else len(languages),
            }
            return Response(response)

        except Exception as e:
            logger.error(f"Error retrieving language list: {e}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Internal server error"
            })

    @swagger_auto_schema(request_body=LanguageSerializer)
    def post(self, request):
        """Create a new language"""
        try:
            required_fields = ["countryFlag", "name", "shortName", "translateName", "userID", "code", "status"]
            missing_fields = [field for field in required_fields if not request.data.get(field)]

            if missing_fields:
                logger.warning(f"Missing required fields: {missing_fields}")
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": f"Missing required fields: {' / '.join(missing_fields)}",
                    "errors": {}
                })

            serializer = LanguageSerializer(data=request.data, context={'request': request})
            if serializer.is_valid():
                serializer.save()
                return Response({
                    "statusCode": 201,
                    "status": True,
                    "message": "Language created successfully",
                    "data": serializer.data
                })
            else:
                logger.warning(f"Validation error during language creation: {serializer.errors}")
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Validation error",
                    "errors": serializer.errors
                })

        except Exception as e:
            logger.error(f"Error creating language: {e}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Internal server error"
            })
import logging
logger = logging.getLogger(__name__)  # Make sure this is at the top of your file

class LanguageRetrieveUpdateDeleteAPIView(APIView):
    """API for retrieving, updating, and deleting a language by ID"""

    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated, IsAdministrator]  

    def get_permissions(self):
        """Set different permissions for GET and POST requests"""
        if self.request.method == 'GET':
            return [AllowAny()]  # Anyone can access GET
        return [IsAuthenticated(), IsAdministrator()]  # Only authenticated admins can POST

    def get_object(self, pk):
        """Retrieve language by ID"""
        try:
            return Language.objects.get(pk=pk)
        except Language.DoesNotExist:
            logger.warning(f"Language with ID {pk} not found.")
            return None

    def get(self, request, pk):
        """Get a single language by ID"""
        try:
            language = self.get_object(pk)
            if language:
                serializer = LanguageSerializer(language)
                response = {
                    "statusCode": 200,
                    "status": True,
                    "message": "Language Found Successfully",
                    "data": serializer.data
                }
                logger.info(f"Language retrieved successfully: ID {pk}")
                return Response(response)

            response = {
                "statusCode": 404,
                "status": False,
                "message": "Language Not Found !!"
            }
            logger.info(f"Language not found: ID {pk}")
            return Response(response)

        except Exception as e:
            logger.exception(f"Error retrieving language ID {pk}: {e}")
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)

    @swagger_auto_schema(request_body=LanguageSerializer)
    def put(self, request, pk):
        """Update a language by ID""" 
        try:
            language = self.get_object(pk)
            if language:
                serializer = LanguageSerializer(language, data=request.data, partial=True, context={'request': request})
                if serializer.is_valid():
                    serializer.save()
                    response = {
                        "statusCode": 200,
                        "status": True,
                        "message": "Language Updated Successfully",
                        "data": serializer.data
                    }
                    logger.info(f"Language updated successfully: ID {pk}")
                    return Response(response)
                else:
                    logger.warning(f"Validation failed for language ID {pk}: {serializer.errors}")
                    return Response({
                        "statusCode": 500,
                        "status": False,
                        "message": serializer.errors
                    })

            response = {
                "statusCode": 400,
                "status": False,
                "message": "Language Not Found !!"
            }
            logger.warning(f"Update failed. Language not found: ID {pk}")
            return Response(response)

        except Exception as e:
            logger.exception(f"Error updating language ID {pk}: {e}")
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)

    def delete(self, request, pk):
        """Delete a language by ID"""
        try:
            language = self.get_object(pk)
            if language:
                language.delete()
                response_data = {
                    "statusCode": 200,
                    "status": True,
                    "message": "Language Deleted successfully"
                }
                logger.info(f"Language deleted successfully: ID {pk}")
                return Response(response_data)
            else:
                response_data = {
                    "statusCode": 400,
                    "status": False,
                    "message": "Language Not Found"
                }
                logger.warning(f"Delete failed. Language not found: ID {pk}")
                return Response(response_data, status=200)

        except Exception as e:
            logger.exception(f"Error deleting language ID {pk}: {e}")
            response_data = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            }
            return Response(response_data)
      
class SearchLanguageAPIView(APIView):
    """API for searching languages using query parameters"""

    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                'searchValue', openapi.IN_QUERY, description="Search term for languages",
                type=openapi.TYPE_STRING, required=True
            ),
            openapi.Parameter(
                'page', openapi.IN_QUERY, description="Page number (default: 1)",
                type=openapi.TYPE_INTEGER, required=False
            )
        ],
        responses={200: "Successful response", 400: "Bad request", 500: "Not found"}
    )
    def get(self, request):
        try:
            search_value = request.GET.get('searchValue', "").strip()
            page = int(request.GET.get('page', 1))
            per_page = 10

            logger.info(f"SearchLanguageAPIView called with searchValue='{search_value}' and page={page}")

            filter_query = Q()

            if search_value:
                filter_query |= Q(name__istartswith=search_value)
                filter_query |= Q(shortName__istartswith=search_value)
                filter_query |= Q(code__istartswith=search_value)
                filter_query |= Q(translateName__istartswith=search_value)

                if search_value.startswith('+'):
                    normalized_search = search_value
                else:
                    normalized_search = f"+{search_value}"
                
                filter_query |= Q(code__icontains=normalized_search)
                filter_query |= Q(code__icontains=search_value.replace('+', ''))

            languages = Language.objects.filter(filter_query)

            paginator = PageNumberPagination()
            paginator.page_size = per_page
            paginated_data = paginator.paginate_queryset(languages, request)

            if not paginated_data:
                logger.warning(f"No languages found for searchValue='{search_value}'")
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "No languages found"
                }, status=500)

            serializer = LanguageSerializer(paginated_data, many=True)
            total_count = languages.count()
            total_pages = (total_count + per_page - 1) // per_page

            logger.info(f"Found {total_count} languages for searchValue='{search_value}'")

            return Response({
                "status": True,
                "data": serializer.data,
                "pagination": {
                    "totalCount": total_count,
                    "totalPages": total_pages,
                    "currentPage": page,
                    "perPage": per_page,
                }
            }, status=200)

        except Exception as e:
            logger.exception(f"Error occurred in SearchLanguageAPIView: {str(e)}")
            return Response({"status": False, "message": str(e)}, status=500)     

class ConutryListCreateAPIView(APIView):
    """API for listing all countries and creating a new country"""
    
    authentication_classes = [JWTAuthentication] 
    parser_classes = (MultiPartParser, FormParser)

    def get_permissions(self):
        """Set different permissions for GET and POST requests"""
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated(), IsAdministrator()]

    def get(self, request):
        """Retrieve a list of countries with optional pagination."""
        try:
            countries = Country.objects.all().order_by("id")
            paginator = PageNumberPagination()
            paginator.page_size = request.query_params.get("page_size", 300)
            paginated_data = paginator.paginate_queryset(countries, request)

            if paginated_data is not None:
                serialized_data = CountrySerializer(paginated_data, many=True).data
                response = {
                    "statusCode": 200,
                    "status": True,
                    "message": "Country list retrieved successfully with pagination",
                    "data": serialized_data,
                    "next": paginator.get_next_link(),
                    "previous": paginator.get_previous_link(),
                    "total_pages": paginator.page.paginator.num_pages,
                    "total_items": paginator.page.paginator.count,
                }
            else:
                serialized_data = CountrySerializer(countries, many=True).data
                response = {
                    "statusCode": 200,
                    "status": True,
                    "message": "Country list retrieved successfully",
                    "data": serialized_data,
                    "total_pages": 1,
                    "total_items": len(countries),
                }

            return Response(response, status=200)

        except Exception as e:
            logger.error(f"Error retrieving country list: {e}", exc_info=True)
            return Response(
                {
                    "statusCode": 500,
                    "status": False,
                    "message": "Internal server error"
                },
                status=500
            )

    @swagger_auto_schema(request_body=CountrySerializer)
    def post(self, request):
        """Create a new country"""    
        try:
            serializer = CountrySerializer(data=request.data, context={'request': request})
            if serializer.is_valid():
                serializer.save()
                return Response({
                    "statusCode": 201,
                    "status": True,
                    "message": "Country created successfully",
                    "data": serializer.data
                }, status=201)

            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Validation error",
                "errors": serializer.errors
            }, status=400)

        except Exception as e:
            logger.error(f"Error creating country: {e}", exc_info=True)
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Internal server error"
            }, status=500)

class CountryRetrieveUpdateDeleteAPIView(APIView):
    """API for retrieving, updating, and deleting a country by ID"""

    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated, IsAdministrator]

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated(), IsAdministrator()]

    def get_object(self, pk):
        try:
            return Country.objects.get(pk=pk)
        except Country.DoesNotExist:
            return None

    def get(self, request, pk):
        try:
            country = self.get_object(pk)
            if country:
                serializer = CountrySerializer(country)
                response = {
                    "statusCode": 200,
                    "status": True,
                    "message": "Country Found Successfully",
                    "data": serializer.data
                }
                return Response(response)
            response = {
                "statusCode": 400,
                "status": False,
                "message": "Country Not Found !!"
            }
            return Response(response)
        except Exception as e:
            logger.error(f"Error retrieving country with id {pk}: {e}", exc_info=True)
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)

    @swagger_auto_schema(request_body=CountrySerializer)
    def put(self, request, pk):
        try:
            country = self.get_object(pk)
            if country:
                serializer = CountrySerializer(country, data=request.data, partial=True, context={'request': request})
                if serializer.is_valid():
                    serializer.save()
                    response = {
                        "statusCode": 200,
                        "status": True,
                        "message": "Country Updated Successfully",
                        "data": serializer.data
                    }
                    return Response(response)
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": serializer.errors
                })
            response = {
                "statusCode": 400,
                "status": False,
                "message": "Country Not Found !!"
            }
            return Response(response)
        except Exception as e:
            logger.error(f"Error updating country with id {pk}: {e}", exc_info=True)
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error {e}"
            }
            return Response(response)

    def delete(self, request, pk):
        try:
            country = self.get_object(pk)
            if country:
                country.delete()
                response_data = {
                    "statusCode": 200,
                    "status": True,
                    "message": "Country Deleted successfully"
                }
                return Response(response_data)
            response_data = {
                "statusCode": 400,
                "status": False,
                "message": "Country Not Found"
            }
            return Response(response_data)
        except Exception as e:
            logger.error(f"Error deleting country with id {pk}: {e}", exc_info=True)
            response_data = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            }
            return Response(response_data)
      
class SearchCountryAPIView(APIView):
    """API for searching countries"""
    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                'searchValue', openapi.IN_QUERY, description="Search term for country",
                type=openapi.TYPE_STRING, required=True
            ),
            openapi.Parameter(
                'page', openapi.IN_QUERY, description="Page number (default: 1)",
                type=openapi.TYPE_INTEGER, required=False
            )
        ],
        responses={200: "Successful response", 400: "Bad request", 400: "Not found"}
    )
    def get(self, request):
        try:
            search_value = request.GET.get("searchValue", "").strip() 
            page = int(request.GET.get("page", 1))  
            per_page = 10

            if not search_value:
                response = {
                    "status": False,
                    "message": "Search value is required"
                }
                logger.warning(f"SearchCountryAPIView - statusCode: 400 - {response}")
                return Response(response, status=400)
            
            normalized_search_term = search_value if search_value.startswith("+") else f"+{search_value}"

            filter_query = Q(name__istartswith=search_value) | \
                           Q(shortNAME__istartswith=search_value) | \
                           Q(slug__istartswith=search_value) | \
                           Q(dialCodes__icontains=normalized_search_term) | \
                           Q(code__istartswith=search_value) | \
                           Q(dialCodes__icontains=search_value.replace("+", ""))
                           
            countries = Country.objects.filter(filter_query)

            paginator = PageNumberPagination()
            paginator.page_size = per_page
            paginated_countries = paginator.paginate_queryset(countries, request)

            if not paginated_countries:
                response = {
                    "status": False,
                    "message": "No countries found"
                }
                logger.info(f"SearchCountryAPIView - statusCode: 400 - {response}")
                return Response(response, status=400)

            serializer = CountrySerializer(paginated_countries, many=True)
            total_count = countries.count()
            total_pages = (total_count + per_page - 1) // per_page  

            response = {
                "success": True,
                "data": serializer.data,
                "pagination": {
                    "totalCount": total_count,
                    "totalPages": total_pages,
                    "currentPage": page,
                    "perPage": per_page,
                }
            }
            logger.info(f"SearchCountryAPIView - statusCode: 200 - {response}")
            return Response(response, status=200)

        except Exception as e:
            response = {
                "status": False,
                "message": str(e)
            }
            logger.error(f"SearchCountryAPIView - statusCode: 500 - {response}")
            return Response(response, status=500)
        

        # <<<<<<<<<<<<<<<<<<<<<<<<<<Category>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
       
class CategoryListCreateAPIView(APIView):
    """API for listing all categories and creating a new category with dynamic search"""
    
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated, IsAdministrator]   
    parser_classes = (MultiPartParser, FormParser)
    pagination_class = CustomPagination  # ✅ Use custom pagination

    def get_permissions(self):
        """Set different permissions for GET and POST requests"""
        if self.request.method == 'GET':
            return [AllowAny()]  # Public access for listing categories
        return [IsAuthenticated(), IsAdministrator()] 

    def get(self, request):
        """Get a list of all categories with search and pagination"""
        try:
            search_query = request.query_params.get("search", None)
            queryset = Category.objects.all().order_by('id')

            if search_query:
                search_filters = (
                    Q(name__icontains=search_query) |
                    Q(machine_name__icontains=search_query) |
                    Q(description__icontains=search_query) |
                    Q(type__icontains=search_query) |
                    Q(distance__icontains=search_query) |
                    Q(slug__icontains=search_query)
                )
                if search_query.lower() in ["true", "false"]:
                    bool_value = search_query.lower() == "true"
                    search_filters |= (
                        Q(is_active=bool_value) |
                        Q(is_deleted=bool_value) |
                        Q(status=bool_value)
                    )

                queryset = queryset.filter(search_filters)

            paginator = self.pagination_class()
            paginated_queryset = paginator.paginate_queryset(queryset, request)
            serializer = CategorySerializer(paginated_queryset, many=True)

            response = {
                "statusCode": 200,
                "status": True,
                "message": "Categories retrieved successfully",
                "data": serializer.data,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
                "total_pages": paginator.page.paginator.num_pages,
                "total_items": paginator.page.paginator.count,
            }
            logger.info(f"CategoryListCreateAPIView GET - statusCode: 200 - {response}")
            return Response(response)

        except FieldError as fe:
            response = {
                "statusCode": 400,
                "status": False,
                "message": f"Invalid search field: {str(fe)}",
                "data": []
            }
            logger.warning(f"CategoryListCreateAPIView GET - statusCode: 400 - {response}")
            return Response(response, status=200)

        except DatabaseError as db_err:
            response = {
                "statusCode": 500,
                "status": False,
                "message": "Database error occurred. Please try again later.",
                "data": []
            }
            logger.error(f"CategoryListCreateAPIView GET - statusCode: 500 - {response}")
            return Response(response, status=500)

        except Exception as e:
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Unexpected error: {str(e)}",
                "data": []
            }
            logger.error(f"CategoryListCreateAPIView GET - statusCode: 500 - {response}")
            return Response(response, status=500)

    @swagger_auto_schema(request_body=CategorySerializer)
    def post(self, request):
        """Create a new category"""    
        try:
            serializer = CategorySerializer(data=request.data, context={'request': request})
            if serializer.is_valid():
                serializer.save()
                response = {
                    "statusCode": 200,
                    "status": True,
                    "message": "Category created successfully",
                    "data": serializer.data
                }
                logger.info(f"CategoryListCreateAPIView POST - statusCode: 200 - {response}")
            else:
                response = {
                    "statusCode": 400,
                    "status": False,
                    "message": serializer.errors
                }
                logger.warning(f"CategoryListCreateAPIView POST - statusCode: 400 - {response}")
                return Response(response)
            return Response(response)
        except Exception as e:
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            }
            logger.error(f"CategoryListCreateAPIView POST - statusCode: 500 - {response}")
            return Response(response)

# Uploads multiple images of category and subcategories
class UploadMultipleCategoryImages(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        try:
            images = request.FILES.getlist('images')
            category_ids = request.data.get('category_ids', [])

            if isinstance(category_ids, str):
                category_ids = [category_ids]

            if not images or not category_ids:
                response = {"error": "Both images and category_ids are required."}
                logger.warning(f"UploadMultipleCategoryImages - statusCode: 400 - {response}")
                return Response(response, status=status.HTTP_200_OK)

            if len(images) != len(category_ids):
                response = {"error": "Number of images must match the number of category IDs."}
                logger.warning(f"UploadMultipleCategoryImages - statusCode: 400 - {response}")
                return Response(response, status=status.HTTP_200_OK)

            updated_categories = []
            for i, category_id in enumerate(category_ids):
                try:
                    category = Category.objects.get(id=category_id)
                    category.categoriesImage = images[i]
                    category.save()
                    updated_categories.append(CategorySerializer(category).data)
                except Category.DoesNotExist:
                    response = {"error": f"Category with ID {category_id} not found."}
                    logger.warning(f"UploadMultipleCategoryImages - statusCode: 404 - {response}")
                    return Response(response, status=status.HTTP_200_OK)

            response = {
                "message": "Images uploaded successfully.",
                "categories": updated_categories
            }
            logger.info(f"UploadMultipleCategoryImages - statusCode: 200 - {response}")
            return Response(response, status=status.HTTP_200_OK)

        except Exception as e:
            response = {"error": str(e)}
            logger.error(f"UploadMultipleCategoryImages - statusCode: 500 - {response}")
            return Response(response, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class UploadMultipleCategoryImages(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        try:
            images = request.FILES.getlist('images')
            category_ids = request.data.get('category_ids', [])

            if isinstance(category_ids, str):
                category_ids = [category_ids]

            if not images or not category_ids:
                response = {"error": "Both images and category_ids are required."}
                logger.warning(f"UploadMultipleCategoryImages - statusCode: 400 - {response}")
                return Response(response, status=status.HTTP_200_OK)

            if len(images) != len(category_ids):
                response = {"error": "Number of images must match the number of category IDs."}
                logger.warning(f"UploadMultipleCategoryImages - statusCode: 400 - {response}")
                return Response(response, status=status.HTTP_200_OK)

            updated_categories = []
            for i, category_id in enumerate(category_ids):
                try:
                    category = Category.objects.get(id=category_id)
                    category.categoriesImage = images[i]
                    category.save()
                    updated_categories.append(CategorySerializer(category).data)
                except Category.DoesNotExist:
                    response = {"error": f"Category with ID {category_id} not found."}
                    logger.warning(f"UploadMultipleCategoryImages - statusCode: 404 - {response}")
                    return Response(response, status=status.HTTP_200_OK)

            response = {
                "message": "Images uploaded successfully.",
                "categories": updated_categories
            }
            logger.info(f"UploadMultipleCategoryImages - statusCode: 200 - {response}")
            return Response(response, status=status.HTTP_200_OK)

        except Exception as e:
            response = {"error": str(e)}
            logger.error(f"UploadMultipleCategoryImages - statusCode: 500 - {response}")
            return Response(response, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CategoryRetrieveUpdateDeleteAPIView(APIView):
    """API for retrieving, updating, and deleting a category by ID"""

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated(), IsAdministrator()]

    def get_object(self, pk):
        try:
            return Category.objects.get(pk=pk)
        except Category.DoesNotExist:
            logger.warning(f"Category with ID {pk} not found")
            return None

    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated, IsAdministrator]

    def get(self, request, pk):
        try:
            category = self.get_object(pk)
            if category:
                serializer = CategorySerializer(category)
                logger.info(f"Category retrieved: ID={pk}")
                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Category found successfully",
                    "data": serializer.data
                })
            logger.warning(f"Category not found: ID={pk}")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Category not found"
            })
        except Exception as e:
            logger.error(f"Error retrieving category ID={pk}: {str(e)}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            })

    @swagger_auto_schema(request_body=CategorySerializer)
    def put(self, request, pk):
        try:
            category = get_object_or_404(Category, pk=pk)
            serializer = CategorySerializer(category, data=request.data, partial=True)

            if serializer.is_valid():
                serializer.save()
                logger.info(f"Category updated successfully: ID={pk}")
                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Category updated successfully",
                    "data": serializer.data
                })
            logger.warning(f"Validation failed when updating category ID={pk}")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Validation failed",
                "errors": serializer.errors
            })
        except Exception as e:
            logger.error(f"Error updating category ID={pk}: {str(e)}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            })

    def delete(self, request, pk):
        try:
            category = self.get_object(pk)
            if category:
                category.delete()
                logger.info(f"Category deleted: ID={pk}")
                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Category deleted successfully"
                })
            logger.warning(f"Tried to delete non-existent category ID={pk}")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Category not found"
            })
        except Exception as e:
            logger.error(f"Error deleting category ID={pk}: {str(e)}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            })
           
class CategoryFilterAPIView(APIView):
    """API to filter categories based on query parameters"""

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated(), IsAdministrator()] 

    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter("name", openapi.IN_QUERY, description="Category name (partial match)", type=openapi.TYPE_STRING),
            openapi.Parameter("machine_name", openapi.IN_QUERY, description="Category machine name (partial match)", type=openapi.TYPE_STRING),
        ],
        responses={
            200: openapi.Response("Successful Response", CategorySerializer(many=True)),
            400: openapi.Response("No categories found"),
            500: openapi.Response("Server error"),
        }
    )
    def get(self, request):
        try:
            name = request.GET.get('name')
            distance_range = request.GET.get('distanceRange')

            filters = {
                "is_deleted": False,
                "is_active": True
            }

            if name:
                filters["name__icontains"] = name
                logger.info(f"Filtering categories by name: {name}")

            if distance_range:
                try:
                    min_distance, max_distance = map(float, distance_range.split('-'))
                    filters["distance__gte"] = min_distance
                    filters["distance__lte"] = max_distance
                    logger.info(f"Filtering categories by distance range: {min_distance} - {max_distance}")
                except ValueError:
                    logger.warning(f"Invalid distance range format: {distance_range}")
                    return Response({
                        "statusCode": 400,
                        "status": False,
                        "message": "Invalid distance range format. Use 'min-max'."
                    })

            categories = Category.objects.filter(**filters).order_by('id')
            paginator = PageNumberPagination()
            paginated_data = paginator.paginate_queryset(categories, request)

            if not categories.exists():
                logger.info("No categories found for applied filters")
                return Response({
                    "statusCode": 404,
                    "status": False,
                    "message": "No categories found"
                })

            serializer = CategorySerializer(paginated_data, many=True)
            logger.info(f"{categories.count()} categories retrieved")
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Categories retrieved successfully",
                "data": serializer.data,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
                "total_pages": paginator.page.paginator.num_pages,
                "total_items": paginator.page.paginator.count,
            })
        except Exception as e:
            logger.error(f"Error retrieving categories: {str(e)}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Error retrieving categories: {str(e)}"
            })

# To change order of display categories 
class CategoryReorderAPIView(APIView):
    """
    API to reorder categories using drag-and-drop functionality.
    Returns updated category list after reordering.
    """
    permission_classes = [IsAuthenticated]  # Add custom IsAdmin permission if needed

    def post(self, request):
        ordered_ids = request.data.get('ordered_ids')

        # Validate input type
        if not isinstance(ordered_ids, list):
            return Response(
                {"error": "'ordered_ids' must be a list of integers."},
                status=status.HTTP_200_OK
            )

        # Check for empty list
        if not ordered_ids:
            return Response(
                {"error": "'ordered_ids' list cannot be empty."},
                status=status.HTTP_200_OK
            )

        # Check if all IDs are integers
        if not all(isinstance(i, int) for i in ordered_ids):
            return Response(
                {"error": "All items in 'ordered_ids' must be integers."},
                status=status.HTTP_400_BADHTTP_200_OK_REQUEST
            )

        # Check that all IDs exist in the database
        existing_ids = set(Category.objects.filter(id__in=ordered_ids).values_list('id', flat=True))
        missing_ids = set(ordered_ids) - existing_ids

        if missing_ids:
            return Response(
                {"error": f"The following category IDs do not exist: {list(missing_ids)}"},
                status=status.HTTP_200_OK
            )

        try:
            with transaction.atomic():
                for index, category_id in enumerate(ordered_ids):
                    Category.objects.filter(id=category_id).update(order_by=index)

            # Fetch and serialize updated categories ordered by `order_by`
            updated_categories = Category.objects.filter(id__in=ordered_ids).order_by('order_by')
            serialized_data = CategorySerializer(updated_categories, many=True).data

            return Response(
                {
                    "message": "Categories reordered successfully.",
                    "categories": serialized_data
                },
                status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                {"error": f"An unexpected error occurred: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

# <<<<<<<<<<<<<<<<<<<<<<<<<<SubCategory>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>           


class CategorysubcategoryListAPIView(APIView):
    def get_permissions(self):
        """Set different permissions for GET and POST requests"""
        if self.request.method == 'GET':
            return [AllowAny()]  # Anyone can access GET
        return [IsAuthenticated(), IsAdministrator()] 
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated,IsAdministrator]   
    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                'name', openapi.IN_QUERY, 
                description="Filter categories by name", 
                type=openapi.TYPE_STRING
            ),
            openapi.Parameter(
                'machine_name', openapi.IN_QUERY, 
                description="Filter categories by machine_name", 
                type=openapi.TYPE_STRING
            ),
        ],
        responses={
            200: openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "statusCode": openapi.Schema(type=openapi.TYPE_INTEGER, example=200),
                    "status": openapi.Schema(type=openapi.TYPE_BOOLEAN, example=True),
                    "message": openapi.Schema(type=openapi.TYPE_STRING, example="Categories retrieved successfully"),
                    "data": openapi.Schema(
                        type=openapi.TYPE_ARRAY,
                        items=openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            properties={
                                "id": openapi.Schema(type=openapi.TYPE_INTEGER, example=8),
                                "name": openapi.Schema(type=openapi.TYPE_STRING, example="Product Purchase"),
                                "categoriesImage": openapi.Schema(type=openapi.TYPE_STRING, example="http://127.0.0.1:8000/uploads/category_image.jpg"),
                                "subcategories": openapi.Schema(
                                    type=openapi.TYPE_ARRAY,
                                    items=openapi.Schema(
                                        type=openapi.TYPE_OBJECT,
                                        properties={
                                            "id": openapi.Schema(type=openapi.TYPE_INTEGER, example=59),
                                            "name": openapi.Schema(type=openapi.TYPE_STRING, example="Electronics and Computing"),
                                            "image": openapi.Schema(type=openapi.TYPE_STRING, example="http://127.0.0.1:8000/uploads/subcategory_image.jpg"),
                                        }
                                    )
                                ),
                            }
                        )
                    )
                }
            ),
            400: openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    
                    "message": openapi.Schema(type=openapi.TYPE_STRING, example="No categories found"),
                    "status": openapi.Schema(type=openapi.TYPE_STRING, example=False)
                }
            ),
            500: openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "message": openapi.Schema(type=openapi.TYPE_STRING, example="Error retrieving categories"),
                    "error": openapi.Schema(type=openapi.TYPE_STRING, example="Detailed error message"),
                    "status": openapi.Schema(type=openapi.TYPE_STRING, example="error")
                }
            )
        }
    )
    
    def get(self, request):
        try:

            base_url =  f"https:/{settings.MEDIA_URL}"

            ids = request.GET.get('id', None)  # Fetch IDs from query params
            name = request.GET.get('name', None)
            machine_name = request.GET.get('machine_name', None)

            filters = Q(is_deleted=False, is_active=True)

            if ids:
                id_list = [int(i) for i in ids.split(',') if i.isdigit()]  # Convert to list of integers
                filters &= Q(id__in=id_list)

            if name:
                filters &= Q(name__icontains=name)

            if machine_name:
                filters &= Q(machine_name__icontains=machine_name)

            categories = Category.objects.filter(filters).prefetch_related('subcategories')

            if not categories.exists():
                return Response({
                    "statusCode": 400,
                    "message": "No categories found",
                    "status": False,
                    "data": []
                }, status=200)

            subcategories_data = []  # Single array for subcategories

            for category in categories:
                for sub in category.subcategories.all():
                    sub_data = {
                        
                        "subcategoryid": sub.id,
                        "name": sub.name,
                        "description": sub.description,
                        "status": sub.status,
                        "isActive": sub.is_active,
                        "isDeleted": sub.is_deleted,
                        "createdAt": sub.created_at,
                        "updatedAt": sub.updated_at,
                        "image": f"{base_url}{sub.subcategoriesImage}" if sub.subcategoriesImage else None,
                        "orderBy": sub.order_by,
                        "parentCategoryId": category.id,  # Include category ID for reference
                        "parentCategoryName": category.name  # Include category name for reference
                    }
                    subcategories_data.append(sub_data)

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Subcategories retrieved successfully",
                "data": subcategories_data  # Flat array of subcategories
            }, status=200)

        except Exception as error:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Internal server error",
                "error": str(error)
            }, status=500)

       

class SubcategoryListCreateAPIView(APIView):
    """API for listing all subcategories and creating a new one"""
    
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated,IsAdministrator]   
    parser_classes = (MultiPartParser, FormParser)
    
    
    def get_permissions(self):
        """Set different permissions for GET and POST requests"""
        if self.request.method == 'GET':
            return [AllowAny()]  
        return [IsAuthenticated(), IsAdministrator()] 

    def get(self, request):
        """Get a list of all subcategories with pagination and search"""
        try:
            search_query = request.query_params.get('search', None)
            
            subcategories = Subcategory.objects.filter(is_deleted=False).order_by('id')
            
            if search_query:
                # Dynamically filter by all fields
                subcategories = subcategories.filter(
                    Q(name__icontains=search_query) |
                    Q(slug__icontains=search_query) |
                    Q(description__icontains=search_query) |
                    Q(type__icontains=search_query) |
                    Q(slug__icontains=search_query) |
                    Q(parentCategoryId__name__icontains=search_query) 
                    
                )
            
            paginator = PageNumberPagination()
            paginated_data = paginator.paginate_queryset(subcategories, request)
            serializer = SubcategorySerializer(paginated_data, many=True)

            response = {
                "statusCode": 200,
                "status": True,
                "message": "Subcategories retrieved successfully",
                "data": serializer.data,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
                "total_pages": paginator.page.paginator.num_pages,
                "total_items": paginator.page.paginator.count,
            }
            return Response(response, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response(
                {"statusCode": 500, "status": False, "message": f"Internal server error: {e}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @swagger_auto_schema(request_body=SubcategorySerializer)
    def post(self, request):
        """Create a new subcategory"""
        try:
            serializer = SubcategorySerializer(data=request.data,  context={'request': request})
            if serializer.is_valid():
                serializer.save()
                response = {
                    "statusCode": 200,
                    "status": True,
                    "message": "Subcategory created successfully",
                    "data": serializer.data,
                }
                return Response(response, status=status.HTTP_201_CREATED)

            return Response(
                {"statusCode": 400, "status": False, "message": serializer.errors},
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            return Response(
                {"statusCode": 500, "status": False, "message": f"Internal server error: {e}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class SubcategoryRetrieveUpdateDeleteAPIView(APIView):
    """API for retrieving, updating, and deleting a subcategory by ID"""
    def get_permissions(self):
        """Set different permissions for GET and POST requests"""
        if self.request.method == 'GET':
            return [AllowAny()]  # Anyone can access GET
        return [IsAuthenticated(), IsAdministrator()]  # Only authenticated admins can POST


    def get_object(self, pk):
        """Retrieve subcategory by ID"""
        try:
            return Subcategory.objects.get(pk=pk, is_deleted=False)
        except Subcategory.DoesNotExist:
            return None
        
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated,IsAdministrator]  
    def get(self, request, pk):
        """Get a single subcategory by ID"""
        try:
            subcategory = self.get_object(pk)
            if subcategory:
                serializer = SubcategorySerializer(subcategory)
                response = {
                    "statusCode": 200,
                    "status": True,
                    "message": "Subcategory found successfully",
                    "data": serializer.data,
                }
                return Response(response)
            return Response({"statusCode": 400, "status": False, "message": "Subcategory not found"})
        except Exception as e:
            response = {"statusCode": 500, "status": False, "message": f"Internal server error: {str(e)}"}
            return Response(response)

    @swagger_auto_schema(request_body=SubcategorySerializer)
    def put(self, request, pk):
        """Update a subcategory by ID"""
        try:
            subcategory = self.get_object(pk)

            if not subcategory:
                return Response({"statusCode": 400, "status": False, "message": "Subcategory not found"})

            serializer = SubcategorySerializer(subcategory, data=request.data, partial=True, context={'request': request})

            if serializer.is_valid():
                # Check if any value is actually changing
                existing_data = {field: getattr(subcategory, field) for field in serializer.validated_data}
                new_data = serializer.validated_data

                if existing_data == new_data:
                    return Response({
                        "statusCode": 200,
                        "status": True,
                        "message": "Subcategory updated successfully But, No changes detected",
                        "data": serializer.data,
                    })

                # Save only if changes exist
                serializer.save()
                response = {
                    "statusCode": 200,
                    "status": True,
                    "message": "Subcategory updated successfully",
                    "data": serializer.data,
                }
                return Response(response)

            return Response({"statusCode": 400, "status": False, "message": serializer.errors})

        except Exception as e:
            response = {
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            }
            return Response(response)


    def delete(self, request, pk):
        """Soft delete a subcategory by ID"""
        try:
            subcategory = self.get_object(pk)
            if subcategory:
                subcategory.delete()
                response_data = {
                    "statusCode": 200,
                    "status": True,
                    "message": "Subcategory deleted successfully",
                }
                return Response(response_data)
            else:
                response_data = {"statusCode": 400, "status": False, "message": "Subcategory not found"}
                return Response(response_data)
        except Exception as e:
            logger.exception(f"error in api:::{str(e)}")
            response_data = {"statusCode": 500, "status": False, "message": f"Internal server error: {str(e)}"}
            return Response(response_data)
       

class GetAllSubCategoryWithFilterParents(APIView):
    
    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter("name", openapi.IN_QUERY, description="Category name (partial match)", type=openapi.TYPE_STRING),
            openapi.Parameter(
                "machine_name",
                openapi.IN_QUERY,
                description="SubCategory machine name (partial match)",
                type=openapi.TYPE_STRING
            ),
            
        ],
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["parentCategoryIds"],
            properties={
                "parentCategoryIds": openapi.Schema(
                    type=openapi.TYPE_ARRAY,
                    items=openapi.Items(type=openapi.TYPE_INTEGER),
                    description="List of parent category IDs"
                ),
            },
        ),
        responses={
            200: openapi.Response("Successful Response"),
            400: openapi.Response("Invalid request payload"),
            500: openapi.Response("Server error"),
        }
    )

    def get(self, request, parentCategoryId, *args, **kwargs):

        base_url =  f"https:/{settings.MEDIA_URL}"

        # Ensure parentCategoryId is an integer
        try:
            parent_category_id = int(parentCategoryId)
        except ValueError:
            return Response(
                {"statusCode": 400, "status": False, "message": "Invalid parentCategoryId"},
                status=status.HTTP_200_OK,
            )

        # Validate if the parent category exists
        if not Category.objects.filter(id=parent_category_id).exists():
            return Response(
                {"statusCode": 404, "status": False, "message": "Parent category not found"},
                status=status.HTTP_200_OK,
            )

        # Initialize filters
        filters = Q(is_deleted=False, is_active=True, parentCategoryId=parent_category_id)

        # Apply string filter for name
        name = request.query_params.get("name")
        if name:
            filters &= Q(name__icontains=name)

        # Apply numeric range filters
        def apply_range_filter(param, field_name):
            value_range = request.query_params.get(param)  # Expected format: "min,max"
            if value_range:
                try:
                    min_val, max_val = map(float, value_range.split(","))
                    return Q(**{f"{field_name}__gte": min_val, f"{field_name}__lte": max_val})
                except ValueError:
                    return Q()  # Skip filtering if format is incorrect
            return Q()

        filters &= apply_range_filter("distanceRange", "distance")
        filters &= apply_range_filter("priceRange", "price")  # Changed "order_by" to "price"

        # Apply boolean filters
        boolean_filters = ["onSite", "clickCollect", "halal", "handicapped", "rooftop", "freeCancellation"]
        for field in boolean_filters:
            value = request.query_params.get(field)
            if value is not None:
                if value.lower() in ["true", "false"]:
                    filters &= Q(**{field: value.lower() == "true"})
                else:
                    return Response(
                        {"statusCode": 400, "status": False, "message": f"Invalid value for {field}. Must be 'true' or 'false'"},
                        status=status.HTTP_200_OK,
                    )

        subcategories = Subcategory.objects.filter(filters).select_related("parentCategoryId").order_by("id")

        if not subcategories.exists():
            return Response(
                {"statusCode": 404, "message": "No active subcategories found", "status": False, "data": []},
                status=status.HTTP_200_OK,
            )
        grouped_data = {}
        for subcategory in subcategories:
            parent_category = subcategory.parentCategoryId
            parent_id = str(parent_category.id) if parent_category else "Unknown"
            parent_name = parent_category.name if parent_category else "Unknown"

            if parent_id not in grouped_data:
                grouped_data[parent_id] = {
                    "parentCategoryId": parent_id,
                    "parentCategoryName": parent_name,
                    "subcategories": [],
                }

            grouped_data[parent_id]["subcategories"].append({
                "id": subcategory.id,
                "name": subcategory.name,
                "machine_name": subcategory.machine_name,
                "description": subcategory.description,
                "parentCategoryId": parent_category.id if parent_category else None,
                "status": subcategory.status,
                "isActive": subcategory.is_active,
                "isDeleted": subcategory.is_deleted,
                "createdAt": subcategory.created_at,
                "updatedAt": subcategory.updated_at,
                "image": f"{base_url}{subcategory.subcategoriesImage}" if subcategory.subcategoriesImage else None,
                "orderBy": subcategory.order_by,
                "distance": subcategory.distance,
                "clickCollect": subcategory.clickCollect,
                "halal": subcategory.halal,
                "handicapped": subcategory.handicapped,
                "rooftop": subcategory.rooftop,
                "freeCancellation": subcategory.freeCancellation,
            })

        return Response(
            {"statusCode": 200, "status": True, "message": "Subcategories retrieved successfully", "data": list(grouped_data.values())},
            status=status.HTTP_200_OK,
        )
    

class SearchSubCategory(APIView):

    def get_permissions(self):
        """Set different permissions for GET and POST requests"""
        if self.request.method == 'GET':
            return [AllowAny()]  # Anyone can access GET
        return [IsAuthenticated(), IsAdministrator()]

    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter("name", openapi.IN_QUERY, description="SubCategory name (partial match)", type=openapi.TYPE_STRING),
            openapi.Parameter("machine_name", openapi.IN_QUERY, description="SubCategory machine name (partial match)", type=openapi.TYPE_STRING),
            openapi.Parameter("pro", openapi.IN_QUERY, description="Flag to indicate professional view", type=openapi.TYPE_BOOLEAN),
        ],
        responses={
            200: openapi.Response("Successful Response", SubcategorySerializer(many=True)),
            400: openapi.Response("Invalid request parameters"),
            500: openapi.Response("Server error"),
        }
    )
    def get(self, request, *args, **kwargs):
        try:

            base_url = f"https:/{settings.MEDIA_URL}"

            category_ids = request.query_params.get("category_id")
            name = request.query_params.get("name")
            machine_name = request.query_params.get("machine_name")
            pro = request.query_params.get("pro")

            filters = Q(is_deleted=False, is_active=True)

            if category_ids:
                try:
                    id_list = [int(i) for i in category_ids.split(',') if i.isdigit()]
                    filters &= Q(parentCategoryId__id__in=id_list)
                except ValueError:
                    return Response({"statusCode": 400, "message": "Invalid category_id format", "status": False, "data": []}, status=status.HTTP_200_OK)

            if name:
                filters &= Q(name__icontains=name)

            if machine_name:
                filters &= Q(machine_name__icontains=machine_name)

            subcategories = Subcategory.objects.filter(filters).order_by("id")

            if not subcategories.exists():
                return Response(
                    {"statusCode": 404, "message": "No matching subcategories found", "status": False, "data": []},
                    status=status.HTTP_200_OK,
                )

            is_professional = pro == 'true'
            data = []

            if is_professional:
                parent_category_map = {}
                for subcategory in subcategories:
                    parent_category = subcategory.parentCategoryId
                    if parent_category:
                        if parent_category.id not in parent_category_map:
                            parent_category_map[parent_category.id] = {
                                "parentCategoryId": parent_category.id,
                                "parentCategoryName": parent_category.name,
                                "status": parent_category.status,
                                "isActive": parent_category.is_active,
                                "isDeleted": parent_category.is_deleted,
                                "subcategories": []
                            }
                        parent_category_map[parent_category.id]["subcategories"].append({
                            "id": subcategory.id,
                            "name": subcategory.name,
                            "machine_name": subcategory.machine_name,
                            "description": subcategory.description,
                            "status": subcategory.status,
                            "clickCollect": subcategory.clickCollect,
                            "halal": subcategory.halal,
                            "handicapped": subcategory.handicapped,
                            "rooftop": subcategory.rooftop,
                            "freeCancellation": subcategory.freeCancellation,
                            "isActive": subcategory.is_active,
                            "isDeleted": subcategory.is_deleted,
                            "createdAt": subcategory.created_at,
                            "updatedAt": subcategory.updated_at,
                            "image": f"{base_url}{subcategory.subcategoriesImage}" if subcategory.subcategoriesImage else None,
                            "parentCategoryId": subcategory.parentCategoryId.id if subcategory.parentCategoryId else None,
                            "orderBy": subcategory.order_by,
                        })
                data = list(parent_category_map.values())
            else:
                for subcategory in subcategories:
                    data.append({
                        "id": subcategory.id,
                        "name": subcategory.name,
                        "machine_name": subcategory.machine_name,
                        "description": subcategory.description,
                        "status": subcategory.status,
                        "clickCollect": subcategory.clickCollect,
                        "halal": subcategory.halal,
                        "handicapped": subcategory.handicapped,
                        "rooftop": subcategory.rooftop,
                        "freeCancellation": subcategory.freeCancellation,
                        "isActive": subcategory.is_active,
                        "isDeleted": subcategory.is_deleted,
                        "createdAt": subcategory.created_at,
                        "updatedAt": subcategory.updated_at,
                        "image": f"{base_url}{subcategory.subcategoriesImage}" if subcategory.subcategoriesImage else None,
                        "parentCategoryId": subcategory.parentCategoryId.id if subcategory.parentCategoryId else None,
                        "orderBy": subcategory.order_by,
                    })

            return Response(
                {"statusCode": 200, "status": True, "message": "Subcategories retrieved successfully", "data": data},
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            return Response(
                {"statusCode": 500, "message": "An error occurred", "error": str(e), "status": False},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )



        
       
class SubscriptionListCreateAPIView(APIView):
    """API for listing all subscriptions and creating a new one"""
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, IsAdministrator]

    def get_permissions(self):
        """Set different permissions for GET and POST requests"""
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated(), IsAdministrator()]

    def get(self, request):
        try:
            subscription_type = request.GET.get("subscription_type", None)

            # Query for SubscriptionPlan based on subscription_type filter
            plan_queryset = SubscriptionPlan.objects.all()
            if subscription_type:
                if subscription_type not in ["Monthly", "Annual"]:
                    return Response(
                        {
                            "statusCode": 400,
                            "status": False,
                            "message": "Invalid subscription_type. Use 'Monthly' or 'Annual'.",
                        },
                        status=status.HTTP_200_OK,
                    )
                plan_queryset = plan_queryset.filter(subscription_type=subscription_type)

            # Prefetch related plans for subscriptions
            subscriptions = Subscription.objects.prefetch_related(
                Prefetch("plans", queryset=plan_queryset)
            )

            # Apply filtering on subscription level if a filter is applied
            if subscription_type:
                subscriptions = subscriptions.filter(plans__subscription_type=subscription_type).distinct()
            else:
                subscriptions = subscriptions.all()

            # Calculate total discount
            avg_discount = SubscriptionPlan.objects.filter(
                annualPlan__isnull=False, price__isnull=False
            ).annotate(
                discount_percentage=((12 * F("price") - F("annualPlan")) / (12 * F("price")) * 100)
            ).aggregate(Avg("discount_percentage"))["discount_percentage__avg"]

            total_discount = round(avg_discount, 2) if avg_discount else 0

            # Paginate the results
            paginator = PageNumberPagination()
            paginated_data = paginator.paginate_queryset(subscriptions, request, view=self)

            # Serialize the paginated data
            serialized_paginated_data = SubscriptionSerializer(paginated_data, many=True).data

            # Remove unwanted plans if filter is applied
            for subscription in serialized_paginated_data:
                if subscription_type:
                    subscription["plans"] = [
                        plan for plan in subscription.get("plans", []) if plan.get("subscription_type") == subscription_type
                    ]
                

            response = {
                "statusCode": 200,
                "status": True,
                "total_discount": total_discount,  # Added total_discount at the top
                "message": "Subscriptions retrieved successfully",
                "data": serialized_paginated_data,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
                "total_pages": paginator.page.paginator.num_pages if paginator.page else 1,
                "total_items": paginator.page.paginator.count if paginator.page else len(serialized_paginated_data),
            }
            return Response(response, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {
                    "statusCode": 500,
                    "status": False,
                    "message": f"Internal server error: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


    @swagger_auto_schema(request_body=SubscriptionSerializer)
    def post(self, request):
        """Create a new subscription, supporting both predefined and new plans."""
        try:
            serializer = SubscriptionSerializer(data=request.data)
            
            if serializer.is_valid():
                subscription = serializer.save()
                return Response(
                    {
                        "statusCode": 201,
                        "status": True,
                        "message": "Subscription created successfully",
                        "data": serializer.data,
                    },
                    status=status.HTTP_201_CREATED
                )
            
            # Validation failed
            return Response(
                {
                    "statusCode": 400,
                    "status": False,
                    "message": "Validation error",
                    "errors": serializer.errors,
                },
                status=status.HTTP_200_OK
            )
        
        except IntegrityError as e:
            logger.error(f"Database IntegrityError: {e}")
            return Response(
                {
                    "statusCode": 400,
                    "status": False,
                    "message": "A subscription with this name already exists.",
                },
                status=status.HTTP_200_OK
            )
        
        except DatabaseError as e:
            logger.critical(f"Database error: {e}")
            return Response(
                {
                    "statusCode": 500,
                    "status": False,
                    "message": "A database error occurred. Please try again later.",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        except Exception as e:
            logger.exception(f"Unexpected error: {e}")
            return Response(
                {
                    "statusCode": 500,
                    "status": False,
                    "message": "An unexpected error occurred. Please contact support.",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    


class SubscriptionRetrieveUpdateDeleteAPIView(APIView):
    """API for retrieving, updating, and deleting a subscription by ID"""

    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, IsAdministrator]

    def get_permissions(self):
        """Set different permissions for GET and other requests"""
        if self.request.method == 'GET':
            return [AllowAny()]  # Anyone can access GET
        return [IsAuthenticated(), IsAdministrator()]  # Only admins can update or delete

    def get_object(self, pk):
        """Retrieve subscription by ID"""
        try:
            return Subscription.objects.prefetch_related('plans').get(pk=pk)
        except Subscription.DoesNotExist:
            return None

    def get(self, request, pk):
        """Get a single subscription by ID"""
        try:
            subscription = self.get_object(pk)
            if subscription:
                serializer = SubscriptionSerializer(subscription)

                first_plan = subscription.plans.first()
               
                response = {
                    "statusCode": 200,
                    "status": True,
                    
                    "message": "Subscription found successfully",
                    "data": serializer.data
                }
                return Response(response, status=status.HTTP_200_OK)

            return Response({
                "statusCode": 404,
                "status": False,
                
                "message": "Subscription not found"
            }, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
               
                "message": f"Internal server error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @swagger_auto_schema(request_body=SubscriptionSerializer)
    def put(self, request, pk):
        """Update a subscription by ID"""
        try:
            subscription = self.get_object(pk)
            if not subscription:
                return Response({
                    "statusCode": 404,
                    "status": False,
                    "message": "Subscription not found"
                }, status=status.HTTP_404_NOT_FOUND)

            serializer = SubscriptionSerializer(subscription, data=request.data, partial=True)

            if serializer.is_valid():
                if serializer.validated_data:  # Check if any actual changes were made
                    subscription = serializer.save()
                    first_plan = subscription.plans.first()
                    # save_profit_percentage = first_plan.save_profit_percentage if first_plan else "0"

                    response = {
                        "statusCode": 200,
                        "status": True,
                     
                        "message": "Subscription updated successfully",
                        "data": serializer.data
                    }
                else:
                    response = {
                        "statusCode": 200,
                        "status": True,
                        "message": "Subscription updated successfully but No changes detected",
                        "data": serializer.data
                    }
                return Response(response, status=status.HTTP_200_OK)

            return Response({
                "statusCode": 400,
                "status": False,
                "message": serializer.errors
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def delete(self, request, pk):
        """Delete a subscription by ID"""
        try:
            subscription = self.get_object(pk)
            if subscription:
                subscription.delete()
                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Subscription deleted successfully"
                }, status=status.HTTP_200_OK)

            return Response({
                "statusCode": 404,
                "status": False,
                "message": "Subscription not found"
            }, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class PopularSubscriptionAPIView(APIView):
    """Fetch all popular subscriptions"""

    def get_permissions(self):
        """Set different permissions for GET and other requests"""
        if self.request.method == 'GET':
            return [AllowAny()]  # Public access for GET
        return [IsAuthenticated(), IsAdministrator()]  # Restricted for other methods

    @swagger_auto_schema(
        manual_parameters=[],
        responses={200: SubscriptionSerializer(many=True)}
    )
    def get(self, request):
        """Retrieve all popular subscriptions"""
        try:
            popular_subscriptions = Subscription.objects.filter(is_deleted=False, popular=True).prefetch_related('plans')

            if not popular_subscriptions.exists():
                return Response({
                    "statusCode": 200,
                    "status": True,
                    
                    "message": "No popular subscriptions found",
                    "data": []
                }, status=status.HTTP_200_OK)

            # Serialize the data
            serializer = SubscriptionSerializer(popular_subscriptions, many=True)

            # Extract `save_profit_percentage` from the first available plan (if any)
            first_plan = popular_subscriptions.first().plans.first() if popular_subscriptions.first() else None
            

            response = {
                "statusCode": 200,
                "status": True,
               
                "message": "Popular subscriptions fetched successfully",
                "data": serializer.data,
            }
            return Response(response, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
              
                "message": f"Error fetching popular subscriptions: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class SearchSubscriptionAPIView(APIView):
    """Search subscriptions based on query parameters"""

    def get_permissions(self):
        """Set different permissions for GET and other requests"""
        if self.request.method == 'GET':
            return [AllowAny()]  # Public access for GET
        return [IsAuthenticated(), IsAdministrator()]  # Restricted for other methods

    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter('name', openapi.IN_QUERY, description="Filter by subscription name", type=openapi.TYPE_STRING),
            openapi.Parameter('price', openapi.IN_QUERY, description="Filter by subscription price", type=openapi.TYPE_NUMBER),
            openapi.Parameter('monthlyPlan', openapi.IN_QUERY, description="Filter by monthly plan", type=openapi.TYPE_NUMBER),
            openapi.Parameter('annualPlan', openapi.IN_QUERY, description="Filter by annual plan", type=openapi.TYPE_NUMBER),
            openapi.Parameter('popular', openapi.IN_QUERY, description="Filter by popular status (true/false)", type=openapi.TYPE_BOOLEAN),
        ],
        responses={200: SubscriptionSerializer(many=True)}
    )
    def get(self, request):
        try:
            query_params = request.query_params
            search_criteria = Q(is_deleted=False)  # Base filter for Subscription model
            plan_criteria = Q()  # Filter for SubscriptionPlan model

            # **Filter by Subscription Name**
            name = query_params.get("name", "").strip()
            if name:
                search_criteria &= Q(name__icontains=name)

            # **Filter by Popularity**
            popular = query_params.get("popular", "").strip().lower()
            if popular:
                if popular in ["true", "false"]:
                    search_criteria &= Q(popular=(popular == "true"))
                else:
                    return Response(
                        {"statusCode": 400, "status": False, "message": "Invalid value for popular, use 'true' or 'false'"},
                        status=status.HTTP_200_OK,
                    )

            # **Filter by Subscription Type (Monthly or Annual) - Case Insensitive**
            subscription_type = query_params.get("subscription_type", "").strip().lower()
            if subscription_type:
                if subscription_type not in ["monthly", "annual"]:
                    return Response(
                        {"statusCode": 400, "status": False, "message": "Invalid subscription_type. Use 'Monthly' or 'Annual'."},
                        status=status.HTTP_200_OK,
                    )
                plan_criteria &= Q(subscription_type__iexact=subscription_type)

            # **Filter by Monthly or Annual Plan Price**
            monthly_plan_price = query_params.get("monthly_plan", "").strip()
            annual_plan_price = query_params.get("annual_plan", "").strip()

            if monthly_plan_price:
                try:
                    plan_criteria &= Q(price=float(monthly_plan_price))
                except ValueError:
                    return Response(
                        {"statusCode": 400, "status": False, "message": "Invalid monthly plan format"},
                        status=status.HTTP_200_OK,
                    )

            if annual_plan_price:
                try:
                    plan_criteria &= Q(price=float(annual_plan_price))
                except ValueError:
                    return Response(
                        {"statusCode": 400, "status": False, "message": "Invalid annual plan format"},
                        status=status.HTTP_200_OK,
                    )

            # **Filter by Price (inside SubscriptionPlan)**
            price = query_params.get("price", "").strip()
            if price:
                try:
                    plan_criteria &= Q(price=float(price))
                except ValueError:
                    return Response(
                        {"statusCode": 400, "status": False, "message": "Invalid price format"},
                        status=status.HTTP_200_OK,
                    )

            # **Apply Filters**
            plan_queryset = SubscriptionPlan.objects.filter(plan_criteria)
            subscriptions = Subscription.objects.filter(search_criteria).prefetch_related(
                Prefetch("plans", queryset=plan_queryset)
            )

            # **Check if Subscriptions Exist**
            filtered_subscriptions = [sub for sub in subscriptions if sub.plans.exists()]
            if not filtered_subscriptions:
                return Response(
                    {
                        "statusCode": 404,
                        "status": False,
                       
                        "message": "No subscription plans found matching the criteria",
                        "data": [],
                    },
                    status=status.HTTP_404_NOT_FOUND,
                )

            # **Serialize the Data**
            serializer = SubscriptionSerializer(filtered_subscriptions, many=True)

            # **Extract `save_profit_percentage` from the first plan (if any)**
            first_plan = filtered_subscriptions[0].plans.first() if filtered_subscriptions else None
           

            # **Return Response**
            response = {
                "statusCode": 200,
                "status": True,
                
                "message": "Subscription plans fetched successfully",
                "data": serializer.data,
            }
            return Response(response, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {
                    "statusCode": 500,
                    "status": False,
                  
                    "message": f"Error searching subscription plans: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )



class SubscriptionPlanCreateView(APIView):
    """
    API View to create a new SubscriptionPlan with proper error handling and duplicate prevention.
    """

    def post(self, request):
        try:
            subscription_id = request.data.get("subscription")
            subscription_type = request.data.get("subscription_type")

            # Check if a SubscriptionPlan with the same subscription and type already exists
            if SubscriptionPlan.objects.filter(subscription=subscription_id, subscription_type=subscription_type).exists():
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "A SubscriptionPlan with this subscription and type already exists."
                }, status=status.HTTP_200_OK)

            serializer = SubscriptionPlanSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response({
                    "statusCode": 201,
                    "status": True,
                    "message": "SubscriptionPlan created successfully",
                    "data": serializer.data
                }, status=status.HTTP_201_CREATED)

            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Error creating SubscriptionPlan",
                "errors": serializer.errors
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class SubscriptionPlanListView(APIView):
    """
    API View to list subscription plans with improved search and pagination.
    """
    def get(self, request):
        try:
            queryset = SubscriptionPlan.objects.filter(is_deleted=False, is_active=True)

            # Improved search across relevant fields
            search_query = request.GET.get('search', '').strip()
            if search_query:
                queryset = queryset.filter(
                    Q(subscription__name__icontains=search_query) |
                    Q(subscription_type__icontains=search_query) |
                    Q(status__icontains=search_query)
                )
            
            # Ordering with validation
            ordering = request.GET.get('ordering', '-created_at')
            valid_ordering_fields = ['price', 'created_at', '-price', '-created_at']
            if ordering in valid_ordering_fields:
                queryset = queryset.order_by(ordering)
            else:
                queryset = queryset.order_by('-created_at')
            
            paginator = CustomPagination()
            paginated_queryset = paginator.paginate_queryset(queryset, request, view=self)
            serializer = SubscriptionPlanSerializer(paginated_queryset, many=True)
            
            response = {
                "statusCode": 200,
                "status": True,
                "message": "Subscription plans fetched successfully",
                "data": serializer.data,
                "total_pages": paginator.page.paginator.num_pages,
                "total_items": paginator.page.paginator.count,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
            }
            return paginator.get_paginated_response(response)
        
        except ValueError as ve:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": f"Invalid request parameters: {str(ve)}"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UpdateSubscriptionPlanAPI(APIView):
    """
    API View to update an existing SubscriptionPlan with proper error handling.
    """

    def put(self, request, pk):
        try:
            subscription_plan = SubscriptionPlan.objects.get(pk=pk)
        except SubscriptionPlan.DoesNotExist:
            return Response({
                "statusCode": 404,
                "status": False,
                "message": "SubscriptionPlan not found"
            }, status=status.HTTP_200_OK)

        try:
            serializer = SubscriptionPlanSerializer(subscription_plan, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "SubscriptionPlan updated successfully",
                    "data": serializer.data
                }, status=status.HTTP_200_OK)

            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Error updating SubscriptionPlan",
                "errors": serializer.errors
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class DeleteSubscriptionPlanAPI(APIView):
    """
    API View to delete a SubscriptionPlan with proper error handling.
    """

    def delete(self, request, pk):
        try:
            subscription_plan = SubscriptionPlan.objects.get(pk=pk)
        except SubscriptionPlan.DoesNotExist:
            return Response({
                "statusCode": 404,
                "status": False,
                "message": "SubscriptionPlan not found"
            }, status=status.HTTP_200_OK)

        try:
            subscription_plan.delete()
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "SubscriptionPlan deleted successfully"
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




class SubscriptionPlanBySubscriptionView(APIView):
    def get(self, request, subscription_id=None, pk=None):  
        try:
            subscription_id = subscription_id or pk  # Handle both cases
            if not subscription_id:
                return Response({"message": "Subscription ID is required"}, status=status.HTTP_400_BAD_REQUEST)

            plans = SubscriptionPlan.objects.filter(subscription_id=subscription_id, is_deleted=False, is_active=True)

            if not plans.exists():
                return Response({"message": "No active subscription plans found"}, status=status.HTTP_404_NOT_FOUND)

            serializer = SubscriptionPlanSerializer(plans, many=True)

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Subscription plans fetched successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class RolePermissionPagination(PageNumberPagination):
    page_size = 10  
    page_size_query_param = 'page'
    max_page_size = 100

class RolePermissionViewAll(APIView):
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated,IsAdministrator]  
    def get(self, request, role_permission_id=None):
        """Retrieve role permissions for all roles or a specific role permission with pagination."""
        try:
            query = RolePermissions.objects.select_related('menu', 'role').prefetch_related('menu__submenus')

            if role_permission_id:
                query = query.filter(id=role_permission_id, is_deleted=False)
            else:
                query = query.filter(is_deleted=False)

            if not query.exists():
                return JsonResponse({"message": "Role permissions not found"}, status=400)

            formatted_response = [ 
                { 
                    "rolePermissionId": rp.id,
                    "role": {
                        "roleId": rp.role.id if rp.role else None,
                        "roleName": rp.role.name if rp.role else None,
                        "menu": {  
                            "menuId": rp.menu.id if rp.menu else None,
                            "menuName": rp.menu.menuname if rp.menu else None,
                            "permissions": {
                                "create": rp.create_permission,
                                "read": rp.read_permission,
                                "update": rp.update_permission,
                                "delete": rp.delete_permission,
                            },
                            "submenus": [
                                {
                                    "submenuId": submenu.id,
                                    "submenuName": submenu.submenuname,
                                    "permissions": {
                                        "create": rp.create_permission,  
                                        "read": rp.read_permission,
                                        "update": rp.update_permission,
                                        "delete": rp.delete_permission,
                                    }
                                }
                                for submenu in rp.menu.submenus.filter(is_deleted=False) 
                            ]
                        }
                    }
                }
                for rp in query
            ]

            if role_permission_id:
                # No pagination when fetching a specific role permission
                response = {
                    "statusCode": 200,
                    "status": True,
                    "message": "Role permission fetched successfully",
                    "data": formatted_response[0],  # Since only one record should match
                }
                return JsonResponse(response, status=200)

            # Use DRF Pagination for multiple records
            paginator = RolePermissionPagination()
            paginated_data = paginator.paginate_queryset(formatted_response, request, view=self)

            response = {
                "statusCode": 200,
                "status": True,
                "message": "Role permissions fetched successfully",
                "data": paginated_data,
            }

            if paginator.page.paginator.num_pages > 1:  # Include pagination info only if multiple pages exist
                response.update({
                    "next": paginator.get_next_link(),
                    "previous": paginator.get_previous_link(),
                    "total_pages": paginator.page.paginator.num_pages,
                    "total_items": paginator.page.paginator.count,
                })

            return paginator.get_paginated_response(response)

        except Exception as e:
            return JsonResponse({"message": "Error fetching role permissions", "error": str(e)}, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class RolePermissionsView(View):
    def get_permissions(self):
        """Set different permissions for GET and POST requests"""
        if self.request.method == 'GET':
            return [AllowAny()]  # Anyone can access GET
        return [IsAuthenticated(), IsAdministrator()]  # Only authenticated admins can POST
 
    # permission_classes = [permissions.IsAuthenticated]
    @method_decorator(csrf_exempt)
    def post(self, request):
        """Create role permissions only if they do not exist for a given role and menu."""
        try:
            data = json.loads(request.body.decode("utf-8"))

            # Extract role name and fetch role object
            role_id = data.get("role_id")
            if not role_id:
                return JsonResponse({"message": "Role id is required"}, status=400)

            role = Role.objects.filter(id=role_id).first()
            if not role:
                return JsonResponse({"message": "Role not found"}, status=400)

            role_permissions = data.get("rolePermissions", [])
            if not isinstance(role_permissions, list):
                return JsonResponse({"message": "Invalid rolePermissions format"}, status=400)

            existing_menu_ids = set(
                RolePermissions.objects.filter(role=role).values_list("menu_id", flat=True)
            )

            new_permissions = []
            for rp in role_permissions:
                menu_id = rp.get("menuId")

                if not menu_id:
                    return JsonResponse({"message": "menuId is required"}, status=400)

                # Check if menu exists in the database
                menu = Menu.objects.filter(id=menu_id).first()
                if not menu:
                    return JsonResponse({"message": f"Menu ID {menu_id} not found"}, status=400)

                # Skip creation if the RolePermissions already exists
                if menu_id in existing_menu_ids:
                    continue  # Skip this menu

                # Create new RolePermissions
                new_permissions.append(RolePermissions(
                    role=role,
                    menu=menu,
                    create_permission=rp.get("create", False),
                    read_permission=rp.get("read", False),
                    update_permission=rp.get("update", False),
                    delete_permission=rp.get("delete", False)
                ))

            if not new_permissions:
                return JsonResponse({"message": "Role permissions already exist for all provided menus"}, status=400)

            # Bulk create new permissions
            with transaction.atomic():
                RolePermissions.objects.bulk_create(new_permissions)

            # Fetch updated role permissions, including related submenus
            updated_permissions = RolePermissions.objects.filter(role=role).select_related('menu').prefetch_related('menu__submenus')

            # Organize response with menu first, then submenus
            response_data = {
                "message": "Role permissions created successfully",
                "role": {
                    "roleId": role_id,
                    "roleName": role.name
                },
                "menus": [
                    {
                        "menuId": menu.id,
                        "menuName": menu.menuname,
                        "rolePermissions": [
                            {
                                "rolePermissionId": rp.id,
                                "permissions": {
                                    "create": rp.create_permission,
                                    "read": rp.read_permission,
                                    "update": rp.update_permission,
                                    "delete": rp.delete_permission
                                }
                            }
                            for rp in updated_permissions if rp.menu == menu  # Only include role permissions for this menu
                        ],
                        "submenus": [
                            {
                                "submenuId": submenu.id,
                                "submenuName": submenu.submenuname,
                                "rolePermissions": [
                                    {
                                        "rolePermissionId": rp.id,
                                        "permissions": {
                                            "create": rp.create_permission,
                                            "read": rp.read_permission,
                                            "update": rp.update_permission,
                                            "delete": rp.delete_permission
                                        }
                                    }
                                    for rp in updated_permissions if rp.menu == submenu.menu  # Only include role permissions for this submenu's menu
                                ]
                            }
                            for submenu in menu.submenus.all()  # Get all submenus for the current menu
                        ]
                    }
                    for menu in Menu.objects.filter(id__in=[rp.menu.id for rp in updated_permissions]).distinct()  # Only the menus that have updated permissions
                ]
            }

            return JsonResponse(response_data, status=200)

        except json.JSONDecodeError:
            return JsonResponse({"message": "Invalid JSON format"}, status=400)

        except Exception as e:
            return JsonResponse({"message": "Error saving role permissions", "error": str(e)}, status=500)

    def get(self, request, role_permission_id=None):
        authentication_classes = [JWTAuthentication] 
        permission_classes = [IsAuthenticated,IsAdministrator] 
        """Retrieve role permissions for all roles or a specific role permission."""
        try:
            
            query = RolePermissions.objects.select_related('menu', 'role').prefetch_related('menu__submenus')
            if role_permission_id:              
                query = query.filter(id=role_permission_id, is_deleted=False)
            else:             
                query = query.filter(is_deleted=False)

            if not query.exists():
                return JsonResponse({"message": "Role permissions not found"}, status=400)

            formatted_response = [ 
                { 
                    "rolePermissionId": rp.id,
                    "role": {
                        "roleId": rp.role.id if rp.role else None,
                        "roleName": rp.role.name if rp.role else None,
                        "menu": {  
                            "menuId": rp.menu.id if rp.menu else None,
                            "menuName": rp.menu.menuname if rp.menu else None,
                            "permissions": {
                                "create": rp.create_permission,
                                "read": rp.read_permission,
                                "update": rp.update_permission,
                                "delete": rp.delete_permission,
                            },
                            "submenus": [
                                {
                                    "submenuId": submenu.id,
                                    "submenuName": submenu.submenuname,
                                    "permissions": {
                                        "create": rp.create_permission,  
                                        "read": rp.read_permission,
                                        "update": rp.update_permission,
                                        "delete": rp.delete_permission,
                                    }
                                }
                                for submenu in rp.menu.submenus.filter(is_deleted=False) 
                            ]
                        }
                    }
                }
                for rp in query
            ]

            return JsonResponse({"success": True, "message": "Role permissions get successfully", "data": formatted_response}, status=200)

        except Exception as e:
            return JsonResponse({"message": "Error fetching role permissions", "error": str(e)}, status=500)


    def put(self, request, role_permission_id):
        """Update role permissions by RolePermissions ID."""
        authentication_classes = [JWTAuthentication] 
        permission_classes = [IsAuthenticated,IsAdministrator] 
        try:
            data = json.loads(request.body)

            menu_id = data.get("menuId")
            if not menu_id:
                return JsonResponse({"message": "menuId is required"}, status=400)

            # Check if the menu exists
            menu = Menu.objects.filter(id=menu_id).first()
            if not menu:
                return JsonResponse({"message": f"Menu with ID {menu_id} not found"}, status=400)

            # Extract permissions
            permissions = data.get("permissions", {})
            create = permissions.get("create", False)
            read = permissions.get("read", False)
            update = permissions.get("update", False)
            delete = permissions.get("delete", False)

            # Fetch the RolePermissions object
            role_permission = RolePermissions.objects.filter(id=role_permission_id,is_deleted = False).first()
            if not role_permission:
                return JsonResponse({"message": "RolePermissions not found"}, status=400)

            # Update role permissions within a transaction
            with transaction.atomic():
                role_permission.menu = menu
                role_permission.create_permission = create
                role_permission.read_permission = read
                role_permission.update_permission = update
                role_permission.delete_permission = delete
                role_permission.save()

            # Fetch updated role permission
            updated_role_permission = RolePermissions.objects.select_related('menu', 'role').get(id=role_permission_id)

            response_data = {
                "rolePermissionId": updated_role_permission.id,
                "role":{
                    "roleId": updated_role_permission.role.id if updated_role_permission.role else None,
                    "roleName": updated_role_permission.role.name if updated_role_permission.role else None,
                    "menu":{
                        "menuId": updated_role_permission.menu.id if updated_role_permission.menu else None,
                        "menuName": updated_role_permission.menu.menuname if updated_role_permission.menu else None,
                        "permissions": {
                            "create": updated_role_permission.create_permission,
                            "read": updated_role_permission.read_permission,
                            "update": updated_role_permission.update_permission,
                            "delete": updated_role_permission.delete_permission,
                        }
                    }
                }
            }

            return JsonResponse({"message": "Role permission updated successfully", "data": response_data}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({"message": "Invalid JSON format"}, status=400)
        except Exception as e:
            return JsonResponse({"message": "Error updating role permission", "error": str(e)}, status=500)


    def delete(self, request, role_permission_id):
        """Soft delete a specific RolePermissions by its ID."""
        authentication_classes = [JWTAuthentication] 
        permission_classes = [IsAuthenticated,IsAdministrator] 
        try:
            role_permission = RolePermissions.objects.filter(id=role_permission_id,is_deleted = False).first()
            if not role_permission:
                return JsonResponse({"message": "RolePermissions not found or Already Deleted"}, status=400)

            role_permission.is_deleted = True
            role_permission.save()

            return JsonResponse({
                "message": "Role permission deleted successfully",
                "data": {
                    "rolePermissionId": role_permission.id,
                    "role": {
                        "roleId": role_permission.role.id if role_permission.role else None,
                        "roleName": role_permission.role.name if role_permission.role else None,
                    # },
                    "menu": {
                        "menuId": role_permission.menu.id if role_permission.menu else None,
                        "menuName": role_permission.menu.menuname if role_permission.menu else None,
                    # },
                    "permissions": {
                        "create": role_permission.create_permission,
                        "read": role_permission.read_permission,
                        "update": role_permission.update_permission,
                        "delete": role_permission.delete_permission
                    },
                    "isDeleted": role_permission.is_deleted
                }
                },
                },
            }, status=200)
        except Exception as e:
            return JsonResponse({"message": "Error deleting role permission", "error": str(e)}, status=500)



from rest_framework.exceptions import NotFound

class AdminTicketListView(generics.ListAPIView):
    serializer_class = SupportTicketSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    # pagination_class = CustomPagination
    filterset_fields = ['id', 'created_by_user_id']

    def get_queryset(self):
        return SupportTicket.objects.filter(is_deleted=False,).select_related().exclude(
        ticket_category='refund').order_by('-created_at') 

    def list(self, request, *args, **kwargs):
        try:
            queryset = self.filter_queryset(self.get_queryset())
            page = self.paginate_queryset(queryset)
            # serializer = self.get_serializer(page, many=True)

            paginator = CustomPagination()
            paginated_queryset = paginator.paginate_queryset(queryset, request, view=self)
            # serializer = SupportTicket(paginated_queryset, many=True)
            serializer = self.get_serializer(page, many=True)
            
            
            return self.get_paginated_response({
                "statusCode": 200,
                "status": True,
                "message": "Tickets List retrieved successfully",
                "data": serializer.data,
                "total_pages": paginator.page.paginator.num_pages,
                "total_items": paginator.page.paginator.count,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
            })
        except Exception as e:
            return Response(
                {
                    "statusCode": 500,
                    "status": False,
                    "message": "Failed to retrieve tickets",
                    "error": str(e)
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class AdminTicketDetailView(generics.RetrieveAPIView):
    serializer_class = SupportTicketSerializer
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'pk'

    def get_queryset(self):
        # Adjusting to remove `order__company` and `seller` since they are not in the model
        return SupportTicket.objects.filter(is_deleted=False).select_related()

    def retrieve(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            serializer = self.get_serializer(instance)
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Ticket retrieved successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)
        except NotFound:
            return Response({
                "statusCode": 404,
                "status": False,
                "message": "Ticket not found"
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Failed to retrieve ticket",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AdminResolveTicketView(APIView):
    authentication_classes = [JWTAuthentication] 
    # permission_classes = [IsAuthenticated, IsAdministrator]

    def post(self, request, ticket_id):
        try:
            # Fetch the ticket, ensuring it is 'escalated'
            ticket = get_object_or_404(SupportTicket, id=ticket_id )#, status="escalated")

            # Get action from the request data
            action = request.data.get("status")

            if not action:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Action is required"
                }, status=status.HTTP_200_OK)

            # Validate the action and update ticket status accordingly
            if action == "approve_refund":
                ticket.status = "resolved"
            elif action == "reject_complaint":
                ticket.status = "closed"
            else:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Invalid action"
                }, status=status.HTTP_200_OK)

            # Mark the ticket as resolved and save the timestamp
            ticket.resolved_at = timezone.now()
            ticket.save()

            # Return success response with updated ticket info
            return Response({
                "statusCode": 200,
                "status": True,
                "message": f"Action '{action}' applied successfully",
                "data": {
                    "ticket_id": ticket.id,
                    "status": ticket.status
                }
            }, status=status.HTTP_200_OK)

        except Exception as e:
            # Handle any unexpected errors
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Failed to resolve ticket or ticket status is not escalated",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AdminDeleteTicketView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = SupportTicketSerializer

    def get_object(self, ticket_id):
        # Fetch the ticket, ensuring it's not already deleted
        return get_object_or_404(SupportTicket, id=ticket_id)

    def delete(self, request, ticket_id):
        try:
            # Fetch the ticket instance
            instance = self.get_object(ticket_id)
            serializer = self.serializer_class(instance)

            if instance.is_deleted:
                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Ticket already deleted"
                }, status=status.HTTP_200_OK)

            # Mark the ticket as deleted and save
            instance.is_deleted = True
            instance.save()

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Ticket deleted successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            # Handle any unexpected errors
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Failed to delete ticket",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class GetRefundTicketsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            logger.info("Fetching refund tickets with specific_order=True and ticket_category='refund'")

            # Fetch tickets with specific_order=True and ticket_category='refund'
            tickets = SupportTicket.objects.filter(
                specific_order=True,
                ticket_category='refund',
                is_deleted=False
            ).order_by('-created_at') 

            # Serialize the data
            serializer = SupportTicketSerializer(tickets, many=True)
            data = serializer.data

            logger.info(f"{len(data)} tickets fetched successfully.")

            # Enrich data with customer_name, seller_name, and total_price
            for ticket in data:
                user_id = ticket['created_by_user_id']
                user = Users.objects.filter(id=user_id).first()

                # Log user fetching
                if user:
                    ticket['customer_name'] = user.username
                    logger.info(f"Customer name '{user.username}' added for ticket ID {ticket['id']}")
                else:
                    ticket['customer_name'] = None
                    logger.warning(f"No user found for user ID {user_id} for ticket ID {ticket['id']}")

                # Fetch seller name and total price if order is present
                order_id = ticket['order']
                if order_id:
                    try:
                        order = Order.objects.select_related('professional_user', 'company').filter(id=order_id).first()
                        if order:
                            # Get seller name
                            ticket['seller_name'] = order.professional_user.userName if order.professional_user else None
                            if order.professional_user:
                                logger.info(f"Seller name '{order.professional_user.userName}' added for order ID {order_id}")
                            else:
                                logger.warning(f"No professional user found for order ID {order_id}")

                            # Get total price
                            ticket['total_price'] = float(order.total_price) if order.total_price else 0.00
                            logger.info(f"Total price '{ticket['total_price']}' added for order ID {order_id}")
                        else:
                            ticket['seller_name'] = None
                            ticket['total_price'] = 0.00
                            logger.warning(f"No company found for order ID {order_id}")
                    except Exception as ex:
                        logger.error(f"Error fetching order details for order ID {order_id}: {str(ex)}")
                        ticket['seller_name'] = None
                        ticket['total_price'] = 0.00
                else:
                    ticket['seller_name'] = None
                    ticket['total_price'] = 0.00
                    logger.warning(f"No order ID found for ticket ID {ticket['id']}")

            logger.info("Refund tickets data enriched successfully.")
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Refund tickets fetched successfully",
                "data": data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"An error occurred while fetching refund tickets: {str(e)}", exc_info=True)
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An error occurred while fetching refund tickets",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




class VerifyProfessionalUserAPIView(APIView):
    """Admin can verify or reject professional users"""

    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, IsAdministrator]

    def post(self, request, user_id):
        """Update document statuses and verify/reject professional users"""
        professional_user = get_object_or_404(ProfessionalUser, id=user_id)

        # Get document status updates from the request data
        document_statuses = request.data.get("document_statuses", {})

        if not document_statuses:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Document status updates are required."
            }, status=status.HTTP_400_BAD_REQUEST)

        # Document fields to be updated
        document_fields = [
            "kbiss_status",
            "iban_status",
            "proofOfAddress_status",
            "identityCardFront_status",
            "identityCardBack_status",
        ]

        # Normalize status values to lowercase
        document_statuses = {k: v.lower() for k, v in document_statuses.items()}

        # Apply incoming status updates temporarily
        temp_document_values = {
            field: document_statuses.get(field, getattr(professional_user, field))
            for field in document_fields
        }

        # Determine finalDocument_status based on other document statuses
        print("temp_document_values>>>>>>>>>",temp_document_values)
        if "rejected" in temp_document_values.values():
            final_document_status = "rejected"
        elif all(status == "approved" for status in temp_document_values.values()):
            final_document_status = "approved"
        else:
            final_document_status = "pending"
        print("final_document_status>>>>>>",final_document_status)    

        # Apply the final status update
        temp_document_values["finalDocument_status"] = final_document_status

        # Update each document status and save it
        for field, value in temp_document_values.items():
            if value not in ["approved", "pending", "rejected"]:
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": f"Invalid status for {field}. Must be 'approved', 'pending', or 'rejected'."
                }, status=status.HTTP_400_BAD_REQUEST)
            setattr(professional_user, field, value)

        # Check if all documents are approved to mark user as verified
        professional_user.is_verified = (final_document_status == "approved")
        professional_user.save()
        professional_user.refresh_from_db()  # Ensure we get the latest data

        # Prepare response with updated user details
        updated_user_details = ProfessionalUserListSerializer(professional_user).data

        return Response({
            "statusCode": 200,
            "status": True,
            "message": "Document statuses updated successfully.",
            "finalDocument_status": final_document_status,
            "user_details": updated_user_details
        }, status=status.HTTP_200_OK)


class FacilityCreateView(APIView):
    def post(self, request, *args, **kwargs):
        try:
            name = request.data.get("name", "").strip()
            icon = request.FILES.get("icon", None)  # 

            if not name:
                return Response(
                    {"statusCode": 400, "status": False, "message": "Facility name is required"},
                    status=status.HTTP_200_OK,
                )

            facility = Facility.objects.create(name=name, icon=icon)
            serializer = FacilitySerializer(facility)

            return Response(
                {
                    "statusCode": 201,
                    "status": True,
                    "message": "Facility created successfully",
                    "facility": serializer.data,
                },
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response(
                {"statusCode": 500, "status": False, "message": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,)
            
class FacilityListView(APIView):
    def get(self, request):
        try:
            facilities = Facility.objects.all()
            if not facilities:
                return Response(
                    {
                        "statusCode": 404,
                        "status": False,
                        "message": "No facilities found",
                    },
                    status=status.HTTP_200_OK,
                )

            serializer = FacilitySerializer(facilities, many=True)
            return Response(
                {
                    "statusCode": 200,
                    "status": True,
                    "message": "Facilities retrieved successfully",
                    "data": serializer.data,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {
                    "statusCode": 500,
                    "status": False,
                    "message": str(e),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class FacilityDetailView(APIView):
    def get(self, request, id):
        try:
            facility = Facility.objects.get(id=id)
            serializer = FacilitySerializer(facility)
            return Response(
                {
                    "statusCode": 200,
                    "status": True,
                    "message": "Facility retrieved successfully",
                    "data": serializer.data,
                },
                status=status.HTTP_200_OK,
            )
        except Facility.DoesNotExist:
            return Response(
                {
                    "statusCode": 404,
                    "status": False,
                    "message": "Facility not found",
                },
                status=status.HTTP_200_OK,
            )


class FacilityUpdateView(APIView):
    def post(self, request, id):  # Using POST for updates
        try:
            facility = Facility.objects.get(id=id)
            facility.name = request.data.get("name", facility.name)
            facility.icon = request.data.get("icon", facility.icon)
            facility.save()
            
            serializer = FacilitySerializer(facility)
            return Response(
                {
                    "statusCode": 200,
                    "status": True,
                    "message": "Facility updated successfully",
                    "data": serializer.data,
                },
                status=status.HTTP_200_OK,
            )
        except Facility.DoesNotExist:
            return Response(
                {
                    "statusCode": 404,
                    "status": False,
                    "message": "Facility not found",
                },
                status=status.HTTP_200_OK,
            )

class FacilityDeleteView(APIView):
    def post(self, request, id):  # Using POST for deletion
        try:
            facility = Facility.objects.get(id=id)
            facility.delete()
            return Response(
                {
                    "statusCode": 200,
                    "status": True,
                    "message": "Facility deleted successfully",
                },
                status=status.HTTP_200_OK,
            )
        except Facility.DoesNotExist:
            return Response(
                {
                    "statusCode": 404,
                    "status": False,
                    "message": "Facility not found",
                },
                status=status.HTTP_200_OK,
            )

class FacilitySelectView(APIView):
    def post(self, request):
        try:
            selected_facilities = request.data.get("selected_facilities", [])

            if not isinstance(selected_facilities, list):
                return Response(
                    {
                        "statusCode": 400,
                        "status": False,
                        "message": "'selected_facilities' must be a list",
                    },
                    status=status.HTTP_200_OK,
                )

            # Validate facility IDs
            existing_facility_ids = set(Facility.objects.values_list("id", flat=True))
            invalid_ids = [fid for fid in selected_facilities if fid not in existing_facility_ids]

            if invalid_ids:
                return Response(
                    {
                        "statusCode": 400,
                        "status": False,
                        "message": f"Invalid facility IDs: {invalid_ids}",
                    },
                    status=status.HTTP_200_OK,
                )

            # Reset previous selections & assign new ones
            Facility.objects.update(is_selected=False)
            Facility.objects.filter(id__in=selected_facilities).update(is_selected=True)

            # Fetch updated selected facilities
            selected_facilities_data = Facility.objects.filter(is_selected=True)
            serializer = FacilitySerializer(selected_facilities_data, many=True)

            return Response(
                {
                    "statusCode": 200,
                    "status": True,
                    "message": "Facilities selected successfully",
                    "selected_facilities": serializer.data,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {
                    "statusCode": 500,
                    "status": False,
                    "message": str(e),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

from .models import Facility
from .serializers import FacilitySerializer

class FacilityBulkCreateView(APIView):
    def post(self, request, *args, **kwargs):
        try:
            facility_names = request.POST.getlist("name")  # Fetch facility names
            icons = request.FILES.getlist("icon")  # Fetch icons

            # Ensure both fields are lists and contain values
            if isinstance(facility_names, str):
                facility_names = [facility_names]
            if isinstance(icons, str):
                icons = [icons]

            if not facility_names or not icons or len(facility_names) != len(icons):
                return Response(
                    {
                        "statusCode": 400,
                        "status": False,
                        "message": "Facilities name and icons must be provided with the same count",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            facility_objects = []
            for i in range(len(facility_names)):
                facility_name = facility_names[i].strip()
                icon = icons[i]
                
                if not facility_name:
                    return Response(
                        {"statusCode": 400, "status": False, "message": "Facility name cannot be empty."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                facility_objects.append(Facility(name=facility_name, icon=icon))

            #  Bulk create facilities
            created_facilities = Facility.objects.bulk_create(facility_objects)

            #  Fetch the inserted data using their IDs
            created_data = Facility.objects.filter(name__in=[f.name for f in created_facilities]).values("id", "name", "icon")

            return Response(
                {
                    "statusCode": 201,
                    "status": True,
                    "message": "Facilities created successfully",
                    "data": list(created_data),
                },
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response(
                {"statusCode": 500, "status": False, "message": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


    

from django.conf import settings
from django.db.models import Q
from .models import Subcategory

class SubcategoryFilterView(APIView):
    def post(self, request, *args, **kwargs):  # Use POST instead of GET
        base_url = f"https:/{settings.MEDIA_URL}"

        # Retrieve parentCategoryIds from request body
        parent_category_ids = request.data.get("parentCategoryIds")  

        if not parent_category_ids or not isinstance(parent_category_ids, list):
            return Response(
                {"statusCode": 400, "status": False, "message": "parentCategoryIds parameter is required and must be a list"},
                status=status.HTTP_200_OK,
            )

        try:
            parent_category_ids = [int(id) for id in parent_category_ids]
        except ValueError:
            return Response(
                {"statusCode": 400, "status": False, "message": "Invalid parentCategoryIds"},
                status=status.HTTP_200_OK,
            )

        # Filter active subcategories that belong to the given parent categories
        filters = Q(is_deleted=False, is_active=True) & Q(parentCategoryId__in=parent_category_ids)
        subcategories = Subcategory.objects.filter(filters).select_related("parentCategoryId").order_by("id")

        if not subcategories.exists():
            return Response(
                {"statusCode": 404, "message": "No active subcategories found", "status": False, "data": []},
                status=status.HTTP_200_OK,
            )

        grouped_data = {}
        for subcategory in subcategories:
            parent_category = subcategory.parentCategoryId
            parent_id = str(parent_category.id) if parent_category else "Unknown"
            parent_name = parent_category.name if parent_category else "Unknown"

            if parent_id not in grouped_data:
                grouped_data[parent_id] = {
                    "parentCategoryId": parent_id,
                    "parentCategoryName": parent_name,
                    "subcategories": [],
                }

            grouped_data[parent_id]["subcategories"].append({
                "id": subcategory.id,
                "name": subcategory.name,
                "machine_name": subcategory.machine_name,
                "description": subcategory.description,
                "parentCategoryId": parent_category.id if parent_category else None,
                "status": subcategory.status,
                "isActive": subcategory.is_active,
                "isDeleted": subcategory.is_deleted,
                "createdAt": subcategory.created_at,
                "updatedAt": subcategory.updated_at,
                "image": f"{base_url}{subcategory.subcategoriesImage}" if subcategory.subcategoriesImage else None,
                "orderBy": subcategory.order_by,
                "distance": subcategory.distance,
                "clickCollect": subcategory.clickCollect,
                "halal": subcategory.halal,
                "handicapped": subcategory.handicapped,
                "rooftop": subcategory.rooftop,
                "freeCancellation": subcategory.freeCancellation,
            })

        return Response(
            {"statusCode": 200, "status": True, "message": "Subcategories retrieved successfully", "data": list(grouped_data.values())},
            status=status.HTTP_200_OK,
        )




# Function to generate a random coupon code
def generate_coupon_code():
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=10))

class CouponCreateView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, IsAdministrator]

    def post(self, request):
        try:
            data = request.data.copy()

            # Generate a random coupon code if not provided
            if "couponCode" not in data or not data["couponCode"]:
                data["couponCode"] = generate_coupon_code()

            serializer = CouponSerializer(data=data)
            serializer.is_valid(raise_exception=True)
            serializer.save()

            return Response(
                {
                    "statusCode": 201,
                    "status": True,
                    "message": "Coupon created successfully",
                    "data": serializer.data,
                },
                status=status.HTTP_201_CREATED,
            )

        except IntegrityError:
            return Response(
                {
                    "statusCode": 400,
                    "status": False,
                    "message": "A coupon with this code already exists.",
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {
                    "statusCode": 500,
                    "status": False,
                    "message": str(e),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class CouponListView(APIView):
    # authentication_classes = [JWTAuthentication]
    # permission_classes = [IsAuthenticated, IsAdministrator]

    def get(self, request):
        try:
            coupons = Coupon.objects.filter(is_deleted=False)

            if not coupons.exists():
                return Response(
                    {
                        "statusCode": 404,
                        "status": False,
                        "message": "No active coupons found",
                    },
                    status=status.HTTP_200_OK,
                )
            paginator = PageNumberPagination()
            paginator.page_size = 10  # You can change the page size here
            paginated_coupons = paginator.paginate_queryset(coupons, request)
            serializer = CouponSerializer(paginated_coupons, many=True)
            # serializer = CouponSerializer(coupons, many=True)
            return Response(
                {
                    "statusCode": 200,
                    "status": True,
                    "message": "Coupons retrieved successfully",
                    "next": paginator.get_next_link(),
                    "previous": paginator.get_previous_link(),
                    "total_pages": paginator.page.paginator.num_pages,
                    "total_items": paginator.page.paginator.count,
                    "data": serializer.data,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {
                    "statusCode": 500,
                    "status": False,
                    "message": str(e),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

class CouponRetrieveView(APIView):
    # authentication_classes = [JWTAuthentication]
    # permission_classes = [IsAuthenticated, IsAdministrator]

    def get(self, request, id):
        try:
            coupon = Coupon.objects.filter(id=id, is_deleted=False).first()

            if not coupon:
                return Response(
                    {
                        "statusCode": 404,
                        "status": False,
                        "message": "Coupon not found",
                    },
                    status=status.HTTP_200_OK,
                )

            serializer = CouponSerializer(coupon)
            return Response(
                {
                    "statusCode": 200,
                    "status": True,
                    "message": "Coupon retrieved successfully",
                    "data": serializer.data,
                },
                status=status.HTTP_200_OK,
            )
                
        except Exception as e:
            return Response(
                {
                    "statusCode": 500,
                    "status": False,
                    "message": str(e),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

#  Update a Coupon (Admin Only)
class CouponUpdateView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, IsAdministrator]

    def put(self, request, id):
        try:
            coupon = Coupon.objects.filter(id=id, is_deleted=False).first()

            if not coupon:
                return Response(
                    {
                        "statusCode": 404,
                        "status": False,
                        "message": "Coupon not found",
                    },
                    status=status.HTTP_200_OK,
                )

            serializer = CouponSerializer(coupon, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save()

            return Response(
                {
                    "statusCode": 200,
                    "status": True,
                    "message": "Coupon updated successfully",
                    "data": serializer.data,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {
                    "statusCode": 500,
                    "status": False,
                    "message": str(e),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

#  Soft Delete a Coupon (Mark as Deleted Instead of Removing)
class CouponDeleteView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, IsAdministrator]

    def delete(self, request, id):
        try:
            coupon = Coupon.objects.filter(id=id, is_deleted=False).first()

            if not coupon:
                return Response(
                    {
                        "statusCode": 404,
                        "status": False,
                        "message": "Coupon not found",
                    },
                    status=status.HTTP_200_OK,
                )

            # Mark the coupon as deleted
            coupon.is_deleted = True
            coupon.save()

            return Response(
                {
                    "statusCode": 200,
                    "status": True,
                    "message": "Coupon deleted successfully (soft delete).",
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {
                    "statusCode": 500,
                    "status": False,
                    "message": str(e),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )



from django.db.models import Q

class PaymentListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            status_filter = request.GET.get('status')
            name_filter = request.GET.get('name')

            payments = Payment.objects.select_related('user', 'subscription_plan', 'subscription_type').all()

            if status_filter:
                payments = payments.filter(status__iexact=status_filter)

            if name_filter:
                payments = payments.filter(
                    Q(user__username__icontains=name_filter) |
                    Q(user__name__icontains=name_filter)
                )

            paginator = CustomPagination()
            paginated_data = paginator.paginate_queryset(payments, request)
            serializer = PaymentSerializer(paginated_data, many=True)

            return paginator.get_paginated_response({
                "statusCode": 200,
                "status": True,
                "message": "Payments fetched successfully.",
                "data": serializer.data
            })

        except DatabaseError as e:
            logger.error(f"[Payment Fetch Error] {str(e)}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Database error while fetching payments.",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class PaymentListViewExports(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            status_filter = request.GET.get('status')
            name_filter = request.GET.get('name')
            export_format = request.GET.get("export_format", "").upper()

            payments = Payment.objects.select_related('user', 'subscription_plan', 'subscription_type').all()

            if status_filter:
                payments = payments.filter(status__iexact=status_filter)

            if name_filter:
                payments = payments.filter(
                    Q(user__username__icontains=name_filter) |
                    Q(user__name__icontains=name_filter)
                )

            # Export if format is specified
            if export_format in ["CSV", "EXCEL", "PDF"]:
                return self.export_data(payments, export_format)

            paginator = CustomPagination()
            paginated_data = paginator.paginate_queryset(payments, request)
            serializer = PaymentSerializer(paginated_data, many=True)

            return paginator.get_paginated_response({
                "statusCode": 200,
                "status": True,
                "message": "Payments fetched successfully.",
                "data": serializer.data
            })

        except DatabaseError as e:
            logger.error(f"[Payment Fetch Error] {str(e)}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Database error while fetching payments.",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def export_data(self, queryset, export_format):
        if export_format == "CSV":
            return self.export_csv(queryset)
        elif export_format == "EXCEL":
            return self.export_excel(queryset)
        elif export_format == "PDF":
            return self.export_pdf(queryset)

    def export_csv(self, queryset):
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(["ID", "Email", "Amount", "Status", "Subscription Plan", "Subscription Type", "stripe_customer_Id", "Created At"])

        for payment in queryset:
            writer.writerow([
                payment.id,
                # payment.user.username if payment.user else "",
                payment.user.email if payment.user else "",
                str(payment.amount),
                payment.status,
                payment.subscription_plan.name if payment.subscription_plan else "",
                payment.subscription_type.name if payment.subscription_type else "",
                payment.stripe_customer_id,
                payment.created_at.strftime("%Y-%m-%d %H:%M:%S")
            ])

        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=subscription_payments.csv'
        return response

    def export_excel(self, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "SubscriptionPayments"

        headers = ["ID",  "Email", "Amount", "Status", "Subscription Plan", "Subscription Type", "stripe_customer_Id", "Created At"]
        ws.append(headers)

        for payment in queryset:
            ws.append([
                payment.id,
                # payment.user.username if payment.user else "",
                payment.user.email if payment.user else "",
                str(payment.amount),
                payment.status,
                payment.subscription_plan.name if payment.subscription_plan else "",
                payment.subscription_type.name if payment.subscription_type else "",
                payment.stripe_customer_id,
                payment.created_at.strftime("%Y-%m-%d %H:%M:%S")
            ])

        for i, column_cells in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value) or "") for cell in column_cells)
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=subscription_payments.xlsx'
        wb.save(response)
        return response

    def export_pdf(self, queryset):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=subscription_payments.pdf'

        p = canvas.Canvas(response, pagesize=letter)
        width, height = letter
        x_margin = 30
        y_margin = height - 50
        line_height = 20

        # Define headers and fixed column widths
        headers = [
            "ID", "Email", "Amount", "Status", 
            "Subscription Plan", "Subscription Type", 
            "Stripe Customer ID", "Created At"
        ]
        column_widths = [30, 130, 60, 60, 100, 100, 130, 100]

        # Draw header row
        x = x_margin
        for i, header in enumerate(headers):
            p.setFont("Helvetica-Bold", 8)
            p.drawString(x, y_margin, header)
            x += column_widths[i]

        y = y_margin - line_height

        # Draw each row of data
        for payment in queryset:
            row = [
                str(payment.id),
                payment.user.email if payment.user else "",
                str(payment.amount),
                payment.status,
                payment.subscription_plan.name if payment.subscription_plan else "",
                payment.subscription_type.name if payment.subscription_type else "",
                payment.stripe_customer_id or "",
                payment.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            ]

            x = x_margin
            for i, cell in enumerate(row):
                p.setFont("Helvetica", 8)
                p.drawString(x, y, str(cell))
                x += column_widths[i]

            y -= line_height

            # Check for page break
            if y < 50:
                p.showPage()
                y = y_margin

                # Redraw headers on new page
                x = x_margin
                for i, header in enumerate(headers):
                    p.setFont("Helvetica-Bold", 8)
                    p.drawString(x, y, header)
                    x += column_widths[i]
                y -= line_height

        p.save()
        return response

class UserPaymentListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            search_query = request.GET.get("search", "")
            status_filter = request.GET.get("status", "")

            queryset = UserPayment.objects.select_related('user', 'professional_user__company').all()

            if search_query:
                queryset = queryset.filter(
                    Q(user__username__icontains=search_query) |
                    Q(professional_user__userName__icontains=search_query)
                )

            if status_filter:
                queryset = queryset.filter(status__iexact=status_filter)

            paginator = CustomPagination()
            paginated_qs = paginator.paginate_queryset(queryset, request)

            # Serialize with default serializer
            serializer = UserPaymentSerializer(paginated_qs, many=True)
            serialized_data = serializer.data

            # Add username fields manually
            for idx, obj in enumerate(paginated_qs):
                serialized_data[idx]["customer_name"] = obj.user.username if obj.user else None
                serialized_data[idx]["professional_name"] = (
                    obj.professional_user.company.managerFullName
                    if obj.professional_user and obj.professional_user.company else None
                )

            return paginator.get_paginated_response({
                "statusCode": 200,
                "status": True,
                "message": "User payments fetched successfully.",
                "data": serialized_data
            })

        except DatabaseError as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Database error while fetching user payments.",
                "error": str(e)
            }, status=500)



class UserPaymentListViewExports(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            search_query = request.GET.get("search", "")
            status_filter = request.GET.get("status", "")
            export_format = request.GET.get("export_format", "").upper()  # optional

            queryset = UserPayment.objects.select_related('user', 'professional_user__company').all()

            if search_query:
                queryset = queryset.filter(
                    Q(user__username__icontains=search_query) |
                    Q(professional_user__userName__icontains=search_query)
                )

            if status_filter:
                queryset = queryset.filter(status__iexact=status_filter)

            # Export if requested
            if export_format in ["CSV", "EXCEL", "PDF"]:
                return self.export_data(queryset, export_format)

            # Otherwise, return paginated response
            paginator = CustomPagination()
            paginated_qs = paginator.paginate_queryset(queryset, request)
            serializer = UserPaymentSerializer(paginated_qs, many=True)
            data = serializer.data

            for idx, obj in enumerate(paginated_qs):
                data[idx]["customer_name"] = obj.user.username if obj.user else None
                data[idx]["professional_name"] = (
                    obj.professional_user.company.managerFullName
                    if obj.professional_user and obj.professional_user.company else None
                )

            return paginator.get_paginated_response({
                "statusCode": 200,
                "status": True,
                "message": "User payments fetched successfully.",
                "data": data
            })

        except DatabaseError as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Database error while fetching user payments.",
                "error": str(e)
            }, status=500)

    def export_data(self, queryset, export_format):
        if export_format == "CSV":
            return self.export_csv(queryset)
        elif export_format == "EXCEL":
            return self.export_excel(queryset)
        elif export_format == "PDF":
            return self.export_pdf(queryset)

    def export_csv(self, queryset):
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(["ID", "Customer", "Professional", "Amount", "Status", "Payment Mode", "Transaction ID", "Created At"])

        for item in queryset:
            writer.writerow([
                item.id,
                item.user.username if item.user else "",
                item.professional_user.company.managerFullName if item.professional_user and item.professional_user.company else "",
                str(item.amount),
                item.status,
                item.payment_mode,
                item.transaction_id,
                item.created_at.strftime("%Y-%m-%d %H:%M:%S")
            ])

        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=user_payments.csv'
        return response

    def export_excel(self, request):
        # Query your data - for example, all user payments
        payments = UserPayment.objects.all()

        # Create an in-memory workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "User Payments"

        # Define headers based on your model fields
        headers = ["ID", "User", "Amount", "created_at", "Status"]
        ws.append(headers)

        # Fill rows with data
        for payment in payments:
            row = [
                payment.id,
                payment.user.username if payment.user else "N/A",  # adjust user field
                payment.amount,  # adjust field names
                payment.created_at.strftime("%Y-%m-%d") if payment.created_at else "",
                payment.status,
            ]
            ws.append(row)

        # Optionally, adjust column widths
        for i, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = max(len(column_title) + 2, 15)

        # Prepare HTTP response with Excel file
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response['Content-Disposition'] = 'attachment; filename="user_payments.xlsx"'

        # Save workbook to response stream
        wb.save(response)
        return response


    def export_pdf(self, queryset):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=user_payments.pdf'

        p = canvas.Canvas(response, pagesize=letter)
        width, height = letter
        y = height - 50
        line_height = 20

        # Define x-positions for each column
        columns = [
            ("ID", 40),
            ("Customer", 70),
            ("Professional", 150),
            ("Amount", 270),
            ("Status", 320),
            ("Payment Mode", 390),
            ("Transaction ID", 470),
            ("Created At", 560)
        ]

        # Draw headers
        p.setFont("Helvetica-Bold", 10)
        for title, x in columns:
            p.drawString(x, y, title)

        y -= line_height
        p.setFont("Helvetica", 9)

        # Draw rows
        for item in queryset:
            if y < 50:
                p.showPage()
                y = height - 50
                p.setFont("Helvetica-Bold", 10)
                for title, x in columns:
                    p.drawString(x, y, title)
                y -= line_height
                p.setFont("Helvetica", 9)

            row = [
                str(item.id),
                item.user.username if item.user else "",
                item.professional_user.company.managerFullName if item.professional_user and item.professional_user.company else "",
                str(item.amount),
                item.status,
                item.payment_mode,
                item.transaction_id,
                item.created_at.strftime("%Y-%m-%d %H:%M:%S")
            ]

            for idx, (value, (_, x)) in enumerate(zip(row, columns)):
                value = str(value) if value is not None else ""
                if len(value) > 25:
                    value = value[:22] + "..."
                p.drawString(x, y, value)


            y -= line_height

        p.save()
        return response

class CardListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            status_filter = request.GET.get('status')
            name_filter = request.GET.get('name')

            cards = Card.objects.select_related('user').all()

            if status_filter:
                cards = cards.filter(status__iexact=status_filter)

            if name_filter:
                cards = cards.filter(
                    Q(user__username__icontains=name_filter) |
                    Q(user__name__icontains=name_filter)
                )

            paginator = CustomPagination()
            paginated_data = paginator.paginate_queryset(cards, request)
            serializer = CardSerializer(paginated_data, many=True)

            return paginator.get_paginated_response({
                "statusCode": 200,
                "status": True,
                "message": "Cards fetched successfully.",
                "data": serializer.data
            })

        except DatabaseError as e:
            logger.error(f"[Card Fetch Error] {str(e)}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Database error while fetching cards.",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)





class CustomPagination(PageNumberPagination):
    page_size = 10 
    page_size_query_param = 'page_size'  
    max_page_size = 100  

class ProfessionalUserOrderSummaryAPIView(APIView):
    def get(self, request):
        try:
            pro_users = ProfessionalUser.objects.all()
            response_list = []
            
            paginator = CustomPagination()
            
            for pro_user in pro_users:
                company = getattr(pro_user, 'company', None)
                orders = Order.objects.filter(professional_user=pro_user)

                user_order_map = defaultdict(dict)

                for order in orders:
                    order_items = OrderItem.objects.filter(order=order)
                    
                    address = None
                    if order.user.manualAddress:
                        address = f"{order.user.manualAddress.address1}, {order.user.manualAddress.city}, {order.user.manualAddress.country}, {order.user.manualAddress.postalCode}"
                    elif order.user.automatic_address:
                        address = f"{order.user.automatic_address.address1}, {order.user.automatic_address.city}, {order.user.automatic_address.country}, {order.user.automatic_address.postalCode}"
                    
                    order_key = f"order_{order.id}"
                    user_order_map[order_key] = {
                        "id": order.id,
                        "order_id": order.id,
                        "email": order.user.email,
                        "customer_name": f"{order.user.firstName} {order.user.lastName}",
                        "address": address,
                        "status": order.orderStatus,
                        "ordered_at": order.created_at,
                        "order_type": order.order_type,
                        "company": {
                            "id": order.company.id if order.company else None,
                            "name": order.company.companyName if order.company else None
                        },
                        "productType": "Product",
                        "is_paid": order.is_paid,
                        "products": [],
                        "total_products_ordered": 0,
                        "subtotal": 0,
                        "discount": 0,
                        "total_price": 0
                    }

                    for item in order_items:
                        product = item.product
                        quantity = item.quantity
                        if not product:
                             continue

                        # Dynamic price calculation based on order type
                        if order.order_type == 'delivery':
                            price = float(product.priceDelivery or 0)
                        elif order.order_type == 'click_collect':
                            price = float(product.priceClickAndCollect or 0)
                        elif order.order_type == 'onsite':
                            price = float(product.priceOnsite or 0)
                        else:
                            price = float(product.basePrice or 0)

                        line_total = price * quantity
                        promo_total = float(product.promotionalPrice or 0) * quantity

                        user_order_map[order_key]["products"].append({
                            "price": str(line_total),
                            "product": {
                                "id": product.id,
                                "productType": product.productType,
                                "quantity": quantity,
                                "productname": product.productname,
                                "description": product.description,
                                "basePrice": str(product.basePrice),
                                "promotionalPrice": str(product.promotionalPrice),
                                "priceDelivery": str(product.priceDelivery),
                                "priceClickAndCollect": str(product.priceClickAndCollect),
                                "priceOnsite": str(product.priceOnsite),
                                "category": {
                                    "id": product.categoryId.id if product.categoryId else None,
                                    "name": product.categoryId.name if product.categoryId else None
                                },
                                "subcategory": {
                                    "id": product.subCategoryId.id if product.subCategoryId else None,
                                    "name": product.subCategoryId.name if product.subCategoryId else None
                                },
                                "ProductImage": request.build_absolute_uri(product.ProductImage.url) if product.ProductImage else None
                            }
                        })

                        user_order_map[order_key]["total_products_ordered"] += quantity
                        user_order_map[order_key]["subtotal"] += line_total
                        user_order_map[order_key]["total_price"] += (line_total - promo_total)

                    # Final discount calculation
                    user_order_map[order_key]["discount"] = user_order_map[order_key]["subtotal"] - user_order_map[order_key]["total_price"]

                sorted_orders = sorted(user_order_map.values(), key=lambda x: x["ordered_at"], reverse=True)

                response_list.append({
                    "professional_user_id": pro_user.id,
                    "professional_user_name": pro_user.userName,
                    "company_name": company.companyName if company else None,
                    "total_orders": orders.count(),
                    "orders": sorted_orders
                })
                
            paginated_data = paginator.paginate_queryset(response_list, request)
            
            if not paginated_data:
                return Response({
                    "statusCode": 404,
                    "status": False,
                    "message": "No data found for the requested page.",
                 
                }, status=status.HTTP_200_OK)

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Professional user order summary retrieved successfully",
                "total_pages": paginator.page.paginator.num_pages,
                "total_items": paginator.page.paginator.count,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
                "data": paginated_data
            }, status=status.HTTP_200_OK)
            
        except ProfessionalUser.DoesNotExist:
            return Response({
                "statusCode": 404,
                "status": False,
                "message": "Professional User not found.",
              
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"An error occurred: {str(e)}", 
              
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




class DashboardStatsAPIView(APIView):
    def get(self, request):
        filter_type = request.query_params.get('filter_type', 'all')
        graph_range = request.query_params.get('graph_range', 'year')
        today = now().date()

        if filter_type == 'today':
            graph_range = 'today'
            start_date = today
        else:
            start_date = {
                'week': today - timedelta(days=7),
                'month': today - timedelta(days=30),
                'year': today - timedelta(days=365),
            }.get(graph_range, today - timedelta(days=365))

        user_filter = {'createdAt__date': today} if filter_type == 'today' else {'createdAt__date__gte': start_date}
        prof_user_filter = {'created_at__date': today} if filter_type == 'today' else {'created_at__date__gte': start_date}
        ticket_filter = {'created_at__date': today, 'productType': 'ticket'} if filter_type == 'today' else {'created_at__date__gte': start_date, 'productType': 'ticket'}
        order_filter = {'date': today} if filter_type == 'today' else {'date__gte': start_date}

        total_users = Users.objects.filter(**user_filter).count()
        total_prof_users = ProfessionalUser.objects.filter(**prof_user_filter).count()
        total_tickets = Product.objects.filter(**ticket_filter).count()
        total_orders = Order.objects.filter(**order_filter).count()
        total_active_companies = CompanyDetails.objects.filter(isActive=True).count()

        logger.warning(f"[Dashboard] Total Orders Filtered: {total_orders}")

        category_popularity_qs = Category.objects.annotate(
            company_count=Count('companies_category', filter=Q(companies_category__isActive=True))
        ).order_by('-company_count')

        category_popularity = [
            {'id': c.id, 'name': c.name, 'company_count': c.company_count}
            for c in category_popularity_qs
        ]

        trunc_func = {
            'year': TruncYear,
            'month': TruncMonth,
            'today': TruncDay,
        }.get(graph_range, TruncDay)

        graph_data = {
            'users': self.get_graph_data(Users, 'createdAt', start_date, trunc_func, graph_range),
            'professional_users': self.get_graph_data(ProfessionalUser, 'created_at', start_date, trunc_func, graph_range),
            'tickets': self.get_graph_data(Product, 'created_at', start_date, trunc_func, graph_range, {'productType': 'ticket'}),
            'orders': self.get_graph_data(Order, 'date', start_date, trunc_func, graph_range, is_date_field=True),
        }

        order_items = OrderItem.objects.filter(order__date__gte=start_date).select_related('product__categoryId')

        # Calculate total revenue
        total_revenue = order_items.aggregate(
            revenue=Sum(F('quantity') * F('price'), output_field=FloatField())
        )['revenue'] or 0.0

        # Prepare sales and order count per category
        sales_data = defaultdict(float)
        order_count_data = defaultdict(int)

        for item in order_items:
            if item.product and item.product.categoryId:
                category_name = item.product.categoryId.name
                sales_data[category_name] += item.quantity * float(item.price)
                order_count_data[category_name] += item.quantity

        all_categories = Category.objects.all()
        category_sales = [
            {
                'name': category.name,
                'total_sales': round(sales_data.get(category.name, 0.0), 2),
                'total_orders': order_count_data.get(category.name, 0)
            }
            for category in all_categories
        ]

        return Response({
            'status': True,
            'message': "Dashboard data fetched successfully.",
            'data': {
                'totals': {
                    'total_users': total_users,
                    'total_professional_users': total_prof_users,
                    'total_tickets': total_tickets,
                    'total_orders': total_orders,
                    'total_active_companies': total_active_companies,
                    'total_revenue': round(total_revenue, 2),
                },
                'graph_data': graph_data,
                'category_popularity': category_popularity,
                'category_sales': category_sales,
                
            }
        })

    def get_graph_data(self, model, date_field, start_date, trunc_func, graph_range, extra_filters=None, is_date_field=False):
        filters = (
            {f"{date_field}": start_date} if graph_range == 'today'
            else {f"{date_field}__gte": start_date} if is_date_field
            else {f"{date_field}__date__gte": start_date}
        )

        if extra_filters:
            filters.update(extra_filters)

        queryset = model.objects.filter(**filters)

        if graph_range == 'week':
            return [{'period': 'last_7_days', 'count': queryset.count()}]

        queryset = queryset.annotate(
            period=trunc_func(date_field)
        ).values('period').annotate(
            count=Count('id')
        ).order_by('period')

        data_dict = defaultdict(int)
        for entry in queryset:
            dt = entry['period']
            if not dt:
                continue
            period_str = (
                dt.strftime('%Y') if graph_range == 'year' else
                dt.strftime('%Y-%m') if graph_range == 'month' else
                dt.strftime('%Y-%m-%d')
            )
            data_dict[period_str] += entry['count']

        return [{'period': k, 'count': v} for k, v in sorted(data_dict.items())]



class HelpCategoryCreateView(generics.CreateAPIView):
    queryset = HelpCategory.objects.all()
    serializer_class = HelpCategorySerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, IsAdministrator]

    def create(self, request, *args, **kwargs):
        try:
            # Extract faqs first, then remove from request data to avoid serializer error
            faqs_data = request.data.pop('faqs', [])

            # Step 1: Create HelpCategory
            category_serializer = self.get_serializer(data=request.data)
            category_serializer.is_valid(raise_exception=True)
            category = category_serializer.save()

            # Step 2: Create HelpFAQs
            faq_serializers = []
            for faq in faqs_data:
                faq['category'] = category.id  # Assign the category ID
                faq_serializer = HelpFAQSerializer(data=faq)
                faq_serializer.is_valid(raise_exception=True)
                faq_serializer.save()
                faq_serializers.append(faq_serializer.data)

            return Response({
                "statusCode": 201,
                "status": True,
                "message": "Help Center category with FAQs created successfully.",
                "data": {
                    "category": category_serializer.data,
                    
                }
            }, status=status.HTTP_201_CREATED)

        except ValidationError as e:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": f"Invalid data: {e.detail}"
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"An unexpected error occurred: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




class HelpCenterView(APIView):
    def get(self, request):
        try:
            categories = HelpCategory.objects.prefetch_related('faqs').all()

            # Check if 'page_size' is passed in query params
            page_size = request.query_params.get('page_size')
            if page_size is not None:
                paginator = CustomPagination()
                paginated_categories = paginator.paginate_queryset(categories, request)
                serializer = HelpCategorySerializer(paginated_categories, many=True)
                
                # Manually build the paginated response to keep your structure
                paginated_response = {
                    "statusCode": 200,
                    "status": True,
                    "message": "Help Center content retrieved successfully.",
                    "categories": serializer.data,
                    "total_items": paginator.page.paginator.count,
                    "total_pages": paginator.page.paginator.num_pages,
                    "next": paginator.get_next_link(),
                    "previous": paginator.get_previous_link()
                }
                return Response(paginated_response, status=status.HTTP_200_OK)
            
            else:
                # If no pagination requested, return full data
                serializer = HelpCategorySerializer(categories, many=True)
                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Help Center content retrieved successfully.",
                    "categories": serializer.data
                }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"An unexpected error occurred: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class HelpCategoryUpdateView(APIView):
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated, IsAdministrator]
    def put(self, request, pk):
        try:
            # Fetch the category object by ID
            category = HelpCategory.objects.get(pk=pk)

            # Serialize the data
            serializer = HelpCategorySerializer(category, data=request.data, partial=True)

            # Validate the data
            if serializer.is_valid():
                # Save the updated category
                serializer.save()

                # Return success response
                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Help Center category updated successfully.",
                    "data": serializer.data
                }, status=status.HTTP_200_OK)
            else:
                # If data is invalid, return the validation errors
                return Response({
                    "statusCode": 400,
                    "status": False,
                    "message": "Invalid data provided.",
                    "errors": serializer.errors
                }, status=status.HTTP_200_OK)
        
        except HelpCategory.DoesNotExist:
            # If category doesn't exist, return error response
            return Response({
                "statusCode": 404,
                "status": False,
                "message": "Help Center category not found."
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            # Catch any unexpected errors and return a general error response
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"An unexpected error occurred: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class HelpCategoryDeleteView(generics.DestroyAPIView):
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated, IsAdministrator]
    queryset = HelpCategory.objects.all()
    lookup_field = 'id'

    def destroy(self, request, *args, **kwargs):
        try:
            # Attempt to get the object by its ID
            category = self.get_object()
            # Delete the category and its related FAQs
            category.delete()
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Help Center category and its FAQs deleted successfully."
            }, status=status.HTTP_200_OK)
        except NotFound:
            return Response({
                "statusCode": 404,
                "status": False,
                "message": "Help category not found."
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"An unexpected error occurred: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

# help FAQs/views.ct


class HelpFAQViewSet(viewsets.ViewSet):
    def list(self, request):
        faqs = HelpFAQ.objects.all()
        serializer = HelpFAQSerializer(faqs, many=True)
        return Response({"status": True, "message": "FAQ list fetched successfully", "data": serializer.data}, status=status.HTTP_200_OK)

    def create(self, request):
        serializer = HelpFAQSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"status": True, "message": "FAQ created successfully", "data": serializer.data}, status=status.HTTP_201_CREATED)
        return Response({"status": False, "message": "Validation failed", "errors": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    def retrieve(self, request, pk=None):
        try:
            faq = HelpFAQ.objects.get(pk=pk)
        except HelpFAQ.DoesNotExist:
            raise NotFound({"status": False, "message": "FAQ not found"})
        
        serializer = HelpFAQSerializer(faq)
        return Response({"status": True, "message": "FAQ fetched successfully", "data": serializer.data}, status=status.HTTP_200_OK)

    def update(self, request, pk=None):
        try:
            faq = HelpFAQ.objects.get(pk=pk)
        except HelpFAQ.DoesNotExist:
            raise NotFound({"status": False, "message": "FAQ not found"})

        serializer = HelpFAQSerializer(faq, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"status": True, "message": "FAQ updated successfully", "data": serializer.data}, status=status.HTTP_200_OK)
        return Response({"status": False, "message": "Validation failed", "errors": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    def partial_update(self, request, pk=None):
        try:
            faq = HelpFAQ.objects.get(pk=pk)
        except HelpFAQ.DoesNotExist:
            raise NotFound({"status": False, "message": "FAQ not found"})

        serializer = HelpFAQSerializer(faq, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({"status": True, "message": "FAQ partially updated", "data": serializer.data}, status=status.HTTP_200_OK)
        return Response({"status": False, "message": "Validation failed", "errors": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None):
        try:
            faq = HelpFAQ.objects.get(pk=pk)
        except HelpFAQ.DoesNotExist:
            raise NotFound({"status": False, "message": "FAQ not found"})

        faq.delete()
        return Response({"status": True, "message": "FAQ deleted successfully"}, status=status.HTTP_200_OK)



# <<<<<<<<<<<<<<<<<<<<<<<<<<<<comapny details>>>>>>>>>>>>>>>>>>>>>>>>>>>


class CompanyDetailsListView(ListAPIView):
    serializer_class = CompanyDetailsUpdateSerializer
    permission_classes = [AllowAny]
    filter_backends = [filters.SearchFilter]
    search_fields = [
        'companyName', 'userName', 'managerFullName', 'email', 'phoneNumber',
        'siret', 'sectorofActivity', 'vatNumber'
    ]

    def get_queryset(self):
        queryset = CompanyDetails.objects.filter()

        # Apply 'isActive' filter if it's provided
        is_active = self.request.query_params.get('isActive')
        if is_active is not None:
            is_active = is_active.lower() == 'true'
            queryset = queryset.filter(isActive=is_active)

        return queryset

    def list(self, request, *args, **kwargs):
        try:
            queryset = self.filter_queryset(self.get_queryset())

            paginator = CustomPagination()
            paginated_companies = paginator.paginate_queryset(queryset, request)
            serializer = self.get_serializer(paginated_companies, many=True)

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Companies fetched successfully",
                "data": serializer.data,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
                "total_pages": paginator.page.paginator.num_pages,
                "total_items": paginator.page.paginator.count,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "message": f"Error: {str(e)}",
                "statusCode": 500,
                "status": False
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# <<<<<<<<<<get_by_id>>>>>>>>>>>>>>>>>>>   
  

from ProfessionalUser.serializers import StoreReelSerializer, StoreImageSerializer

class CompanyDetailsRetrieveView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, company_id):
        try:
            company = CompanyDetails.objects.get(id=company_id, isActive=True)
            serializer = CompanyDetailsUpdateSerializer(company)

            # Get all related products/services/tickets
            all_products = Product.objects.filter(company=company, isActive=True)
            products = all_products.filter(productType='product')
            services = all_products.filter(productType='services')
            tickets = all_products.filter(productType='ticket')

            # Store events, reels, and images
            store_events = StoreEvent.objects.filter(company=company)
            store_reels = StoreReel.objects.filter(company_id=company, isActive=True, is_deleted=False)
            store_images = StoreImage.objects.filter(company_id=company, isActive=True)

            # Orders (in order by created_at descending)
            orders = Order.objects.filter(company=company).order_by('-created_at')

            # Manually serialize orders
            order_data = []
            for order in orders:
                order_data.append({
                    "order_id": order.order_id,
                    # "count": order.count(),
                    "user_id": order.user.id if order.user else None,
                    "professional_user": order.company.managerFullName if order.company else None,
                    "company_id": order.company.id if order.company else None,
                    "order_status": order.orderStatus,
                    "order_type": order.order_type,
                    "total_price": float(order.total_price),
                    "delivery_time": order.deliveryTime.strftime("%H:%M:%S") if order.deliveryTime else None,
                    "service_duration": order.serviceDuration.strftime("%H:%M:%S") if order.serviceDuration else None,
                    "is_paid": order.is_paid,
                    "promo_code": order.promo_code.promocode if order.promo_code else None,
                    "customer_name": order.customer_name,
                    "contact_number": order.contact_number,
                    "email": order.email,
                    "created_at": order.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                })

            # Employees
            employees = Employee.objects.filter(company=company)

            # Response
            return Response({
                "message": "Company details fetched successfully",
                "statusCode": 200,
                "status": True,
                "data": {
                    "company": serializer.data,
                    "product_count": products.count(),
                    "service_count": services.count(),
                    "ticket_count": tickets.count(),
                    "order_count": orders.count(),
                    "employee_count": employees.count(),
                    "store_event_count": store_events.count(),
                    "store_reel_count": store_reels.count(),
                    "store_image_count": store_images.count(),
                    "products": ProductSerializer(products, many=True, context={'request': request}).data,
                    "services": ProductSerializer(services, many=True, context={'request': request}).data,
                    "tickets": ProductSerializer(tickets, many=True, context={'request': request}).data,
                    "store_events": StoreEventSerializer(store_events, many=True, context={'request': request}).data,
                    "store_reels": StoreReelSerializer(store_reels, many=True, context={'request': request}).data,
                    "store_images": StoreImageSerializer(store_images, many=True, context={'request': request}).data,
                    "orders": order_data,
                    "employees": EmployeeSerializer(employees, many=True, context={'request': request}).data,
                }
            }, status=status.HTTP_200_OK)

        except CompanyDetails.DoesNotExist:
            return Response({
                "message": "Company not found",
                "statusCode": 404,
                "status": False
            }, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            return Response({
                "message": f"Internal server error: {str(e)}",
                "statusCode": 500,
                "status": False
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class CompanyDetailsUpdateView(APIView):
    permission_classes = [IsAuthenticated]  # Or change if needed

    def put(self, request, company_id):
        try:
            company = CompanyDetails.objects.get(id=company_id)
        except CompanyDetails.DoesNotExist:
            return Response({
                "message": "Company not found",
                "statusCode": 404,
                "status": False
            }, status=status.HTTP_404_NOT_FOUND)

        serializer = CompanyDetailsUpdateSerializer(company, data=request.data, context={'request': request}, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "message": "Company updated successfully",
                "statusCode": 200,
                "status": True,
                "data": serializer.data
            }, status=status.HTTP_200_OK)
        
        return Response({
            "message": "Validation failed",
            "statusCode": 400,
            "status": False,
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)    

class UpdateCompanyIsActiveAPIView(APIView):
    """
    API to update the isActive status of a company with an optional warning message.
    """
    def put(self, request, company_id):
        try:
            company = get_object_or_404(CompanyDetails, id=company_id)
            previous_status = company.isActive  # Store the previous status
            is_active = request.data.get('isActive', None)
            warning_message = request.data.get('warning_message', "Company status has been updated.")

            if is_active is not None:
                # Convert the input to a boolean
                is_active = bool(is_active)
                
                # Update only if the status has actually changed
                if is_active != previous_status:
                    company.isActive = is_active
                    company.warning_message = warning_message
                    company.save()

                    return Response({
                        "message": "Company status updated successfully",
                        "statusCode": 200,
                        "status": True,
                        "data": {
                            "id": company.id,
                            "companyName": company.companyName,
                            "isActive": company.isActive,
                            "warning_message": company.warning_message
                        }
                    }, status=status.HTTP_200_OK)

                else:
                    return Response({
                        "message": "No status change detected",
                        "statusCode": 400,
                        "status": False
                    }, status=status.HTTP_400_BAD_REQUEST)

            else:
                return Response({
                    "message": "isActive field is required",
                    "statusCode": 400,
                    "status": False
                }, status=status.HTTP_400_BAD_REQUEST)

        except CompanyDetails.DoesNotExist:
            return Response({
                "message": "Company not found",
                "statusCode": 404,
                "status": False
            }, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            logger.error(f"Unexpected error updating company status: {str(e)}", exc_info=True)
            return Response({
                "message": "An unexpected error occurred",
                "statusCode": 500,
                "status": False,
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class StoreMediaListView(APIView, PageNumberPagination):
    page_size = 10

    def get(self, request):
        media_type = request.query_params.get('type')
        if media_type not in ['image', 'reel']:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Invalid type. Use 'image' or 'reel'."
            }, status=status.HTTP_200_OK)

        # Fetch appropriate queryset based on type
        if media_type == 'reel':
            queryset = StoreReel.objects.filter(is_deleted=False)
            serializer_class = StoreReelSerializer
        else:
            queryset = StoreImage.objects.filter()
            serializer_class = StoreImageSerializer

        # Apply filters if present
        company_id = request.query_params.get('company_id')
        category_id = request.query_params.get('category')
        subcategory_id = request.query_params.get('subcategory')

        if company_id:
            queryset = queryset.filter(company_id=company_id)
        if category_id:
            queryset = queryset.filter(category_id=category_id)
        if subcategory_id:
            queryset = queryset.filter(subcategory_id=subcategory_id)

        # Paginate and serialize the data
        paginator = self.paginate_queryset(queryset, request, view=self)
        serializer = serializer_class(paginator, many=True)

        return Response({
            "statusCode": 200,
            "status": True,
            "message": f"{media_type.capitalize()}s fetched successfully",
            "data": serializer.data,
            "next": self.get_next_link(),
            "previous": self.get_previous_link(),
            "total_pages": self.page.paginator.num_pages,
            "total_items": self.page.paginator.count
        }, status=status.HTTP_200_OK)

# <<<<<<<<<<<<<<<<<<<<<<<<<<Reel-Reports>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

class AllReelReportsView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            if not request.user:
                return Response({
                    "message": "Unauthorized",
                    "statusCode": 403,
                    "status": False
                }, status=status.HTTP_200_OK)

            reports = ReelReport.objects.all().order_by('-id')  # Optional: order by recent
            paginator = CustomPagination()
            paginated_reports = paginator.paginate_queryset(reports, request)
            serializer = ReelReportAdminSerializer(paginated_reports, many=True)

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "All reports fetched successfully",
                "data": serializer.data,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
                "total_pages": paginator.page.paginator.num_pages,
                "total_items": paginator.page.paginator.count,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "message": f"Internal server error: {str(e)}",
                "statusCode": 500,
                "status": False
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ReelReportDetailView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication] 

    def get(self, request, report_id):
        try:
            if not request.user:
                return Response({
                    "message": "Unauthorized",
                    "statusCode": 403,
                    "status": False
                }, status=status.HTTP_200_OK)

            report = get_object_or_404(ReelReport, id=report_id)
            serializer = ReelReportAdminSerializer(report)

            return Response({
                "message": "Report detail fetched successfully",
                "statusCode": 200,
                "status": True,
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        except ReelReport.DoesNotExist:
            return Response({
                "message": "Report not found",
                "statusCode": 404,
                "status": False
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "message": f"Internal server error: {str(e)}",
                "statusCode": 500,
                "status": False
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UpdateReelReportStatusView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication] 

    def patch(self, request, report_id):
        try:
            if not request.user:
                return Response({
                    "message": "Unauthorized",
                    "statusCode": 403,
                    "status": False
                }, status=status.HTTP_200_OK)

            report = get_object_or_404(ReelReport, id=report_id)
            new_status = request.data.get('status')

            if new_status not in ['pending', 'reviewed']:
                return Response({
                    "message": "Invalid status value",
                    "statusCode": 400,
                    "status": False
                }, status=status.HTTP_200_OK)

            report.status = new_status
            report.save()

            serializer = ReelReportAdminSerializer(report)
            return Response({
                "message": "Report status updated successfully",
                "statusCode": 200,
                "status": True,
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        except ReelReport.DoesNotExist:
            return Response({
                "message": "Report not found",
                "statusCode": 404,
                "status": False
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "message": f"Internal server error: {str(e)}",
                "statusCode": 500,
                "status": False
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class DeleteReelReportView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, report_id):
        try:
            if not request.user:
                return Response({
                    "message": "Unauthorized",
                    "statusCode": 403,
                    "status": False
                }, status=status.HTTP_200_OK)

            try:
                report = ReelReport.objects.get(id=report_id)
            except ReelReport.DoesNotExist:
                return Response({
                    "message": "Report not found",
                    "statusCode": 404,
                    "status": False
                }, status=status.HTTP_200_OK)

            report.delete()

            return Response({
                "message": "Report deleted successfully",
                "statusCode": 200,
                "status": True
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "message": f"Internal server error: {str(e)}",
                "statusCode": 500,
                "status": False
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UpdateStoreMediaStatusView(APIView):
    def put(self, request):
        media_type = request.data.get('type', '').lower()
        media_id = request.data.get('id')
        is_active = request.data.get('isActive')

        if media_type not in ['image', 'reel']:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Invalid type. Use 'image' or 'reel'."
            }, status=status.HTTP_400_BAD_REQUEST)

        # Fetch the instance based on type
        if media_type == 'reel':
            media_instance = StoreReel.objects.filter(id=media_id, is_deleted=False).first()
            serializer_class = StoreReelSerializer
        else:
            media_instance = StoreImage.objects.filter(id=media_id).first()
            serializer_class = StoreImageSerializer

        if not media_instance:
            return Response({
                "statusCode": 404,
                "status": False,
                "message": f"{media_type.capitalize()} not found."
            }, status=status.HTTP_200_OK)

        # Update isActive status
        media_instance.isActive = is_active
        media_instance.save()

        # Serialize the updated object
        serializer = serializer_class(media_instance)

        return Response({
            "statusCode": 200,
            "status": True,
            "message": f"{media_type.capitalize()} status updated successfully.",
            "data": serializer.data
        }, status=status.HTTP_200_OK)

class AdvertisementDeleteAPIView(APIView):
    def delete(self, request, pk, format=None):
        try:
            ad = Advertisement.objects.get(pk=pk)
        except Advertisement.DoesNotExist:
            raise Http404("Advertisement not found.")

        ad.delete()
        return Response({
            "status": True,
            "message": "Advertisement deleted successfully."
        }, status=status.HTTP_200_OK)

class AdvertisementListCreateView(APIView):
    def get(self, request):
        ads = Advertisement.objects.all()
        serializer = AdvertisementSerializer(ads, many=True)
        return Response({
            "message": "Advertisements fetched successfully",
            "status": True,
            "statusCode": 200,
            "data": serializer.data
        }, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = AdvertisementSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "message": "Advertisement created successfully",
                "status": True,
                "statusCode": 200,
                "data": serializer.data
            }, status=status.HTTP_200_OK)
        return Response({
            "message": "Failed to create advertisement",
            "status": False,
            "statusCode": 400,
            "errors": serializer.errors
        }, status=status.HTTP_200_OK)
    
    
    
class AdvertisementDetailView(APIView):
    def get(self, request, pk):
        ad = get_object_or_404(Advertisement, pk=pk)
        serializer = AdvertisementSerializer(ad)
        return Response({
            "message": "Advertisement fetched successfully",
            "status": True,
            "statusCode": 200,
            "data": serializer.data
        }, status=status.HTTP_200_OK)

    def put(self, request, pk):
        ad = get_object_or_404(Advertisement, pk=pk)
        serializer = AdvertisementSerializer(ad, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "message": "Advertisement updated successfully",
                "status":True,
                "statusCode": 200,
                "data": serializer.data
            }, status=status.HTTP_200_OK)
        return Response({
            "message": "Failed to update advertisement",
            "status": False,
            "statusCode": 400,
            "errors": serializer.errors
        }, status=status.HTTP_200_OK)

    def delete(self, request, pk):
        ad = get_object_or_404(Advertisement, pk=pk)
        ad.delete()
        return Response({
            "message": "Advertisement deleted successfully",
            "status": True,
            "statusCode": 200
        }, status=status.HTTP_200_OK)

class UpdateAdvertisementIsActiveAPIView(APIView):
    """
    API to update the is_active status of an Advertisement with an optional warning message.
    """
    def put(self, request, advertisement_id):
        try:
            advertisement = get_object_or_404(Advertisement, id=advertisement_id)
            previous_status = advertisement.is_active  # Store the previous status
            is_active = request.data.get('is_active', None)
            warning_message = request.data.get('warning_message', "Advertisement status has been updated.")

            if is_active is not None:
                # Convert the input to a boolean
                is_active = bool(is_active)
                
                # Update only if the status has actually changed
                if is_active != previous_status:
                    advertisement.is_active = is_active
                    advertisement.warning_message = warning_message
                    advertisement.save()

                    return Response({
                        "message": "Advertisement status updated successfully",
                        "statusCode": 200,
                        "status": True,
                        "data": {
                            "id": advertisement.id,
                            "title": advertisement.title,
                            "is_active": advertisement.is_active,
                            "warning_message": advertisement.warning_message
                        }
                    }, status=status.HTTP_200_OK)

                else:
                    return Response({
                        "message": "No status change detected",
                        "statusCode": 400,
                        "status": False
                    }, status=status.HTTP_400_BAD_REQUEST)

            else:
                return Response({
                    "message": "is_active field is required",
                    "statusCode": 400,
                    "status": False
                }, status=status.HTTP_400_BAD_REQUEST)

        except Advertisement.DoesNotExist:
            return Response({
                "message": "Advertisement not found",
                "statusCode": 404,
                "status": False
            }, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            logger.error(f"Unexpected error updating advertisement status: {str(e)}", exc_info=True)
            return Response({
                "message": "An unexpected error occurred",
                "statusCode": 500,
                "status": False,
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


 # ----------------------------
#  Bank Account API
# ----------------------------
from django.http import Http404

class BankAccountCreateView(generics.CreateAPIView):
    """
    API to Create Bank Account Details for Admin.
    """
    queryset = AdminBankAccountDetails.objects.all()
    serializer_class = AdminBankAccountDetailsSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        """
        Save the authenticated user as the owner of the bank account.
        """
        serializer.save(user=self.request.user)

    def create(self, request, *args, **kwargs):
        try:
            serializer = self.get_serializer(data=request.data)
            if serializer.is_valid():
                self.perform_create(serializer)

                # Get the serialized data and include user role & username
                data = serializer.data

                # Check if user exists and fetch role and username
                if self.request.user:
                    data["user_role"] = self.request.user.role.name if self.request.user.role else "No Role Assigned"
                    data["username"] = self.request.user.name if self.request.user.name else "No Name Provided"
                else:
                    logger.error("User is not authenticated.")
                    raise Http404("User not found")

                return Response({
                    "statusCode": 201,
                    "status": True,
                    "message": "Bank account details created successfully",
                    "data": [data]
                }, status=status.HTTP_201_CREATED)

            # If serializer validation fails
            logger.error(f"Validation Error: {serializer.errors}")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Validation failed",
                "errors": serializer.errors
            }, status=status.HTTP_200_OK)

        except Http404 as e:
            logger.error(f"User not found: {str(e)}", exc_info=True)
            return Response({
                "statusCode": 404,
                "status": False,
                "message": "User not found"
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Unexpected Error: {str(e)}", exc_info=True)
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class BankAccountRetrieveView(generics.RetrieveAPIView):
    """
    API to Retrieve Bank Account Details by ID.
    """
    queryset = AdminBankAccountDetails.objects.all()
    serializer_class = AdminBankAccountDetailsSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

    def retrieve(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            serializer = self.get_serializer(instance)
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Bank account details retrieved successfully",
                "data": [serializer.data]
            }, status=status.HTTP_200_OK)
        
        except Http404:
            logger.error(f"Bank account details not found for ID: {kwargs.get('pk')}")
            return Response({
                "statusCode": 404,
                "status": False,
                "message": f"Bank account details not found for ID: {kwargs.get('pk')}"
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error retrieving bank account details: {str(e)}", exc_info=True)
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




class BankAccountUpdateView(generics.UpdateAPIView):
    """
    API to Update Bank Account Details by ID.
    """
    queryset = AdminBankAccountDetails.objects.all()
    serializer_class = AdminBankAccountDetailsSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

    def update(self, request, *args, **kwargs):
        try:
            # Get the object - if it doesn't exist, return a 404
            instance = self.get_object()
            
            # Partial update: allows updating only the fields sent in the request
            serializer = self.get_serializer(instance, data=request.data, partial=True)

            # Check for validation
            if serializer.is_valid():
                serializer.save()
                logger.info(f"Bank account with ID {kwargs.get('pk')} updated successfully.")
                
                return Response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Bank account details updated successfully",
                    "data": [serializer.data]
                }, status=status.HTTP_200_OK)

            logger.error(f"Validation Error: {serializer.errors}")
            return Response({
                "statusCode": 400,
                "status": False,
                "message": "Validation failed",
                "errors": serializer.errors
            }, status=status.HTTP_200_OK)

        except Http404:
            logger.error(f"Bank account details not found for ID: {kwargs.get('pk')}")
            return Response({
                "statusCode": 404,
                "status": False,
                "message": f"Bank account details not found for ID: {kwargs.get('pk')}"
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Unexpected Error: {str(e)}", exc_info=True)
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

      

class BankAccountDeleteView(generics.DestroyAPIView):
    """
    API to Delete Bank Account Details by ID.
    """
    queryset = AdminBankAccountDetails.objects.all()
    serializer_class = AdminBankAccountDetailsSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

    def destroy(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            instance.delete()
            logger.info(f"Bank account with ID {kwargs.get('pk')} deleted successfully.")
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Bank account details deleted successfully"
            }, status=status.HTTP_200_OK)

        except Http404:
            logger.error(f"Bank account details not found for ID: {kwargs.get('pk')}")
            return Response({
                "statusCode": 404,
                "status": False,
                "message": f"Bank account details not found for ID: {kwargs.get('pk')}"
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error deleting bank account details for ID {kwargs.get('pk')}: {str(e)}", exc_info=True)
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class BankAccountListView(generics.ListAPIView):
    """
    API to List All Bank Account Details with Search, Filtering, and Pagination.
    """
    queryset = AdminBankAccountDetails.objects.all()
    serializer_class = AdminBankAccountDetailsSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    pagination_class = CustomPagination

    # Define fields for filtering and searching
    filterset_fields = [
        'account_holder_name', 'account_number', 'bank_name', 'branch_name',
        'ifsc_code', 'iban_number', 'swift_code', 'is_active'
    ]
    search_fields = [
        'account_holder_name', 'account_number', 'bank_name',
        'branch_name', 'ifsc_code', 'iban_number', 'swift_code'
    ]

    def list(self, request, *args, **kwargs):
        try:
            # Apply filtering and searching
            queryset = self.filter_queryset(self.get_queryset())

            if not queryset.exists():
                # If there are no matching records, return 404
                logger.info("No bank account details found for the given filters.")
                return Response({
                    "statusCode": 404,
                    "status": False,
                    "message": "No bank account details found."
                }, status=status.HTTP_200_OK)

            # Paginate the queryset
            page = self.paginate_queryset(queryset)

            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response({
                    "statusCode": 200,
                    "status": True,
                    "message": "Bank account details retrieved successfully",
                    "data": serializer.data,
                    "total_pages": self.paginator.page.paginator.num_pages,
                    "total_items": self.paginator.page.paginator.count
                })

            # If not paginated, return the complete list
            serializer = self.get_serializer(queryset, many=True)
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Bank account details retrieved successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error listing bank account details: {str(e)}", exc_info=True)
            return Response({
                "statusCode": 500,
                "status": False,
                "message": f"Internal server error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def safe_url(field, bucket_folder="uploads"):
    if field:
        path = str(field)
        if path.startswith("http"):
            return path
        filename = path.split("/")[-1]
        return f"https://markerplacemobileapp.s3.us-east-1.amazonaws.com/{bucket_folder}/{filename}"
    return ""
from datetime import datetime, timedelta
from django.utils.timezone import localtime

class AdminAllOrdersAPIView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_company_info(self, company, current_weekday, now):
        if not company:
            return {
                "company_id": None,
                "company_name": "Unknown",
                "company_profile_photo": safe_url("product_images"),
                "company_address": None,
                "company_pincode": None,
                "companyratings": 0.0,
                "is_open": False,
            }

        is_open = False
        day_hours = company.opening_hours.get(current_weekday) if company.opening_hours else None
        if day_hours:
            try:
                start = datetime.strptime(day_hours.get("start", ""), "%H:%M").time()
                end = datetime.strptime(day_hours.get("end", ""), "%H:%M").time()
                current_time = now.time()
                is_open = start <= current_time <= end if start < end else current_time >= start or current_time <= end
            except Exception as e:
                print("Opening hours error:", e)

        return {
            "company_id": company.id,
            "company_name": company.companyName,
            "company_profile_photo": safe_url(company.profilePhoto, "company_images") if company.profilePhoto else safe_url("product_images"),
            "company_address": getattr(company.manual_address, "address1", None),
            "company_pincode": getattr(company.manual_address, "postalCode", None),
            "companyratings": company.average_rating,
            "is_open": is_open,
        }
    
    def get_invoice_details(self, order):
        invoice = getattr(order, 'invoice', None)
        if not invoice:
            return None
        return {
            "invoice_id": invoice.id,
            "invoice_number": invoice.invoice_number,
            "invoice_date": invoice.created_at.strftime("%d-%b-%Y") if invoice.created_at else None,
            "due_date": invoice.due_date.strftime("%d-%b-%Y") if invoice.due_date else None,
            "amount_due": float(invoice.amount_due),
            "is_paid": invoice.is_paid,
        }
    
    def get(self, request):
        now = localtime()
        current_weekday = now.strftime('%A').lower()
        order_id = request.query_params.get('order_id')
        booking_id = request.query_params.get('booking_id')
        

        # --- Orders ---
        orders_qs = Order.objects.select_related('company', 'user') \
            .prefetch_related('order_items__product', 'company__manual_address') \
            .order_by('-created_at')
        
        search_query = request.query_params.get("search")

        if search_query:
            orders_qs = orders_qs.filter(
                Q(order_id__icontains=search_query) |
                Q(user__email__icontains=search_query) |
                Q(user__username__icontains=search_query) |
                Q(company__companyName__icontains=search_query)
            )
        
        if order_id:
            orders_qs = orders_qs.filter(order_id=order_id)
        elif booking_id:
            orders_qs = orders_qs.filter(order_id=booking_id)

        orders_data = []
        for order in orders_qs:
            prep_time = getattr(order, f"{order.order_type.lower().replace(' ', '')}PreparationTime", None)
            time_left = None
            if prep_time:
                ready_time = order.created_at + timedelta(
                    hours=prep_time.hour, minutes=prep_time.minute, seconds=prep_time.second
                )
                delta = ready_time - now
                time_left = str(delta).split('.')[0] if delta.total_seconds() > 0 else "Ready"

            products = [{
                "product_id": item.product.id,
                "product_name": item.product.productname,
                "product_image": safe_url(item.product.ProductImage, "product_images") if item.product.ProductImage else safe_url("product_images"),
                "quantity": item.quantity,
                "price": round(item.price, 2),
                "average_rating": item.product.average_rating,
                "total_ratings": item.product.total_ratings,
                "is_rating_pending": False
            } for item in order.order_items.all() if item.product]

            orders_data.append({
                "type": "Orders",
                
                "user_id": order.user.id,
                "user_email": order.user.email,
                "username": order.user.username,
                "company": self.get_company_info(order.company, current_weekday, now),
                "item_count": order.order_items.count(),
                "status": "pending" if order.orderStatus in ["new order", "processing"] else order.orderStatus,
                "order_id": order.order_id,
                "booking_id": order.order_id,
                "total_price": round(order.total_price, 2),
                "order_type": order.order_type,
                "preparation_time": str(prep_time) if prep_time else None,
                "time_left": time_left,
                "order_date": order.created_at.strftime("%d-%b-%Y"),
                "products": products,
                "invoice": self.get_invoice_details(order)
            })

        # --- Reusable Booking Builder ---
        def build_booking_data(qs, label, fields_func, search_query=None):
            if booking_id:
                qs = qs.filter(booking_id__iexact=booking_id)
            if search_query:
                qs = qs.filter(
                    Q(booking_id__icontains=search_query) |
                    Q(user__username__icontains=search_query) |
                    Q(company__companyName__icontains=search_query)
                )
            return [{
                "type": label,
                "count": qs.count(),
                **fields_func(item)
            } for item in qs]


        # Bookings
        room_data = build_booking_data(
            RoomBooking.objects.select_related('company', 'room__product', 'user').order_by('-created_at'),
            "Room_Bookings",
            lambda rb: {
                "name": rb.room.product.productname if rb.room and rb.room.product else None,
                "room_quantity": rb.room_quantity,
                "user_username": rb.user.username,
                "company": rb.company.companyName,
                "adults": rb.adults,
                "order_id": rb.id,
                "booking_id": rb.booking_id,
                "pets": rb.pets,
                "is_paid": rb.is_paid,
                "total_price": float(rb.total_price or 0),
                "status": rb.booking_status,
                "checkin_date": rb.checkin_date.strftime("%d-%b-%Y") if rb.checkin_date else None,
                "checkout_date": rb.checkout_date.strftime("%d-%b-%Y") if rb.checkout_date else None,
                "order_date": rb.booking_date.strftime("%d-%b-%Y") if rb.booking_date else None,
                "created_at": rb.created_at,
                "expires_at": rb.expires_at.strftime("%d-%b-%Y") if rb.expires_at else None
            }, search_query

        )

        event_data = build_booking_data(
            eventBooking.objects.select_related('ticket_id', 'company', 'user').order_by('-created_at'),
            "Event_Bookings",
            lambda eb: {
                "order_type": eb.ticket_type,
                "user_username": eb.user.username,
                "company": eb.company.companyName,
                "order_id": eb.ticket_id.id if eb.ticket_id else None,
                "booking_id": eb.booking_id,
                "name": eb.ticket_id.productname if eb.ticket_id else None,
                "number_of_people": eb.number_of_people,
                "total_price": float(eb.price or 0),
                "status": eb.status,
                "is_paid": eb.is_paid,
                "order_date": eb.booking_date.strftime("%d-%b-%Y") if eb.booking_date else None,
                "created_at": eb.created_at,
                "booking_time": eb.booking_time.strftime("%I:%M %p") if eb.booking_time else None,
            }, search_query
        )

        experience_data = build_booking_data(
            experienceBooking.objects.select_related('ticket_id', 'company', 'user').order_by('-created_at'),
            "Experience_Bookings",
            lambda exb: {
                "order_type": exb.ticket_type,
                "user_username": exb.user.username,
                "company": exb.company.companyName,
                "name": exb.ticket_id.productname if exb.ticket_id else None,
                "total_price": float(exb.price or 0),
                "status": exb.status,
                "order_id": exb.id,
                "booking_id": exb.booking_id,
                "is_paid": exb.is_paid,
                "full_name": exb.full_name,
                "order_date": exb.booking_date.strftime("%d-%b-%Y") if exb.booking_date else None,
                "created_at": exb.created_at,
                "booking_time": exb.booking_time.strftime("%I:%M %p") if exb.booking_time else None,
                "end_date": exb.end_date.strftime("%d-%b-%Y") if exb.end_date else None,
                "adults": exb.adult,
                "children": exb.children,
                "number_of_people": exb.number_of_people
            }, search_query
        )

        slot_data = build_booking_data(
            slotBooking.objects.select_related('user', 'company', 'Product').order_by('-created_at'),
            "Slot_Bookings",
            lambda sb: {
                "slot": sb.slot.strftime("%I:%M %p") if sb.slot else None,
                "name": sb.Product.productname if sb.Product else None,
                "total_price": float(sb.price or 0),
                "user_username": sb.user.username,
                "company": sb.company.companyName,
                "status": sb.status,
                "order_id": sb.id,
                "booking_id": sb.booking_id,
                "is_paid": sb.is_paid,
                "full_name": sb.full_name,
                "order_date": sb.booking_date.strftime("%d-%b-%Y") if sb.booking_date else None,
                "created_at": sb.created_at,
                "booking_time": sb.booking_time.strftime("%I:%M %p") if sb.booking_time else None,
                "number_of_people": sb.number_of_people
            }, search_query
        )
        asthetic_data = build_booking_data(
            aestheticsBooking.objects.select_related('user', 'company', 'Product').order_by('-created_at'),
            "Asthetic_Bookings",
            lambda sb: {
                # "slot": sb.slot.strftime("%I:%M %p") if sb.slot else None,
                "name": sb.Product.productname if sb.Product else None,
                "total_price": float(sb.price or 0),
                "user_username": sb.user.username,
                "company": sb.company.companyName,
                "status": sb.status,
                "order_id": sb.id,
                "booking_id": sb.booking_id,
                "is_paid": sb.is_paid,
                "full_name": sb.full_name,
                "order_date": sb.booking_date.strftime("%d-%b-%Y") if sb.booking_date else None,
                "created_at": sb.created_at.strftime("%d-%b-%Y") if sb.created_at else None,
                "booking_time": sb.booking_time.strftime("%I:%M %p") if sb.booking_time else None,
                "number_of_people": sb.number_of_people
            }, search_query
        )
        relaxation_data = build_booking_data(
            relaxationBooking.objects.select_related('user', 'company', 'Product').order_by('-created_at'),
            "relaxation_Bookings",
            lambda sb: {
                # "slot": sb.slot.strftime("%I:%M %p") if sb.slot else None,
                "name": sb.Product.productname if sb.Product else None,
                "total_price": float(sb.price or 0),
                "user_username": sb.user.username,
                "company": sb.company.companyName,
                "status": sb.status,
                "order_id": sb.id,
                "booking_id": sb.booking_id,
                "is_paid": sb.is_paid,
                "full_name": sb.full_name,
                "order_date": sb.booking_date.strftime("%d-%b-%Y") if sb.booking_date else None,
                "created_at": sb.created_at.strftime("%d-%b-%Y") if sb.created_at else None,
                "booking_time": sb.booking_time.strftime("%I:%M %p") if sb.booking_time else None,
                "number_of_people": sb.number_of_people
            }, search_query
        )
        artandculture_data = build_booking_data(
            artandcultureBooking.objects.select_related('user', 'company', 'Product').order_by('-created_at'),
            "artandculture_Bookings",
            lambda sb: {
                # "slot": sb.slot.strftime("%I:%M %p") if sb.slot else None,
                "name": sb.Product.productname if sb.Product else None,
                "total_price": float(sb.price or 0),
                "user_username": sb.user.username,
                "company": sb.company.companyName,
                "status": sb.status,
                "order_id": sb.id,
                "booking_id": sb.booking_id,
                "is_paid": sb.is_paid,
                "full_name": sb.full_name,
                "order_date": sb.booking_date.strftime("%d-%b-%Y") if sb.booking_date else None,
                "created_at": sb.created_at.strftime("%d-%b-%Y") if sb.created_at else None,
                "booking_time": sb.booking_time.strftime("%I:%M %p") if sb.booking_time else None,
                "number_of_people": sb.number_of_people
            }, search_query
        )

        return Response({
            "statusCode": 200,
            "status": True,
            "message": f"{'Single order/booking' if order_id or booking_id else 'All user orders and bookings'} fetched successfully",
            "total_orders": orders_qs.count(),
            "total_room_bookings": RoomBooking.objects.count(),
            "total_event_bookings": eventBooking.objects.count(),
            "total_experience_bookings": experienceBooking.objects.count(),
            "total_slot_bookings": slotBooking.objects.count(),
            "total_aesthetic_bookings": aestheticsBooking.objects.count(),
            "total_relaxation_bookings": relaxationBooking.objects.count(),
            "total_artandculture_bookings": artandcultureBooking.objects.count(),
            "orders": orders_data,
            "room_bookings": room_data,
            "event_bookings": event_data,
            "experience_bookings": experience_data,
            "slot_bookings": slot_data,
            "aesthetic_bookings":asthetic_data,
            "relaxation_bookings":relaxation_data,
            "artandculture_bookings":artandculture_data
        }, status=status.HTTP_200_OK)


class AdminUpdateUserView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def put(self, request, user_id):
        user = get_object_or_404(Users, id=user_id)
        serializer = AdminProUserSerializer(instance=user, data=request.data, context={'request': request}, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response({
                "status": True,
                "message": "User updated successfully.",
                "data": serializer.data
            }, status=status.HTTP_200_OK)
        else:
            return Response({
                "status": False,
                "message": "Validation failed.",
                "errors": serializer.errors
            }, status=status.HTTP_200_OK)




class AdminNotificationListView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self, user, notification_type=None, start_date=None, end_date=None, search=None):
        queryset = AdminNotification.objects.all().order_by('-created_at')

        if not user.is_superuser and not user.is_staff:
            queryset = queryset.filter(Q(user=user) | Q(professional_user__user=user))

        if notification_type:
            queryset = queryset.filter(notification_type=notification_type)

        if start_date:
            try:
                queryset = queryset.filter(created_at__gte=start_date)
            except ValueError:
                raise ValidationError("Invalid start_date format. Use YYYY-MM-DD")

        if end_date:
            try:
                queryset = queryset.filter(created_at__lte=end_date)
            except ValueError:
                raise ValidationError("Invalid end_date format. Use YYYY-MM-DD")

        if search:
            queryset = queryset.filter(Q(title__icontains=search) | Q(message__icontains=search))

        return queryset

    def get(self, request):
        try:
            user = request.user
            notification_type = request.query_params.get('notification_type')
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            search = request.query_params.get('search')

            queryset = self.get_queryset(user, notification_type, start_date, end_date, search)

            unread_count = queryset.filter(is_read=False).count()
            total_count = queryset.count()

            # Mark all as read
            # queryset.filter(is_read=False).update(is_read=True)

            paginator = CustomPagination()
            paginated_queryset = paginator.paginate_queryset(queryset, request)
            serializer = AdminNotificationSerializer(paginated_queryset, many=True)

            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Admin notifications retrieved successfully",
                "data": serializer.data,
                "next": paginator.get_next_link(),
                "previous": paginator.get_previous_link(),
                "total_pages": paginator.page.paginator.num_pages,
                "total_items": paginator.page.paginator.count,
                "unread_count": unread_count,
                "total_count": total_count
            }, status=status.HTTP_200_OK)

        except ValidationError as ve:
            return Response({
                "statusCode": 400,
                "status": False,
                "message": str(ve),
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error fetching admin notifications: {str(e)}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Internal server error",
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class AdminNotificationDeleteView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def delete(self, request):
        try:
            notification_id = request.query_params.get('notification_id')

            if notification_id:
                notif = AdminNotification.objects.filter(id=notification_id).first()
                if not notif:
                    return Response({
                        "statusCode": 404,
                        "status": False,
                        "message": "Notification not found."
                    }, status=status.HTTP_200_OK)

                notif.delete()
                message = "Notification deleted successfully."
            else:
                AdminNotification.objects.all().delete()
                message = "All admin notifications deleted successfully."

            return Response({
                "statusCode": 200,
                "status": True,
                "message": message
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Admin notification delete error: {str(e)}")
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "Internal server error"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class LegalDocumentCreateView(APIView):
    def post(self, request):
        serializer = LegalDocumentSerializer(data=request.data)
        if serializer.is_valid():
            document = serializer.save()
            return Response({
                "statusCode": 200,
                "status": True,
                "message": "Legal document created successfully.",
                "data": serializer.data
            }, status=status.HTTP_200_OK)
        
        return Response({
            "statusCode": 400,
            "status": False,
            "message": "Validation failed.",
            "errors": serializer.errors
        }, status=status.HTTP_200_OK)

class LegalDocumentDetailView(APIView):
    def get(self, request):
        try:
            documents = LegalDocument.objects.filter(is_active=True)
            serializer = LegalDocumentSerializer(documents, many=True)
            return Response({
                "statusCode":200,
                "status": True,
                "message": "Legal documents retrieved successfully.",
                "data": serializer.data
            }, status=status.HTTP_200_OK)
        except LegalDocument.DoesNotExist:
            return Response({
                "statusCode": 404,
                "status": False,
                "message": "Document not found.",
                "data": None
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                "statusCode": 500,
                "status": False,
                "message": "An unexpected error occurred.",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class LegalDocumentUpdateView(APIView):
    def put(self, request, id):
        try:
            document = LegalDocument.objects.get(id=id, is_deleted=False)
        except LegalDocument.DoesNotExist:
            return Response({
                "status": False,
                "message": "Document not found."
            }, status=status.HTTP_404_NOT_FOUND)

        serializer = LegalDocumentSerializer(document, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "status": True,
                "message": "Document updated successfully.",
                "data": serializer.data
            }, status=status.HTTP_200_OK)
        else:
            return Response({
                "status": False,
                "message": "Validation failed.",
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)


class LegalDocumentDeleteView(APIView):
    def delete(self, request, id):
        try:
            document = LegalDocument.objects.get(id=id, is_deleted=False)
        except LegalDocument.DoesNotExist:
            return Response({
                "status": False,
                "message": "Document not found."
            }, status=status.HTTP_404_NOT_FOUND)

        # Soft delete
        document.is_deleted = True
        document.is_active = False
        document.save()

        return Response({
            "status": True,
            "message": "Document deleted successfully."
        }, status=status.HTTP_200_OK)
    
class LegalDocumentTitleChoicesView(APIView):
    def get(self, request):
        choices = [{"title": key, "label": label} for key, label in LegalDocument.TITLE_CHOICES]
        return Response({
            "status": True,
            "message": "Title choices retrieved successfully.",
            "data": choices
        }, status=status.HTTP_200_OK)    